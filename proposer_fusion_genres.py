#!/usr/bin/env python3
"""
proposer_fusion_genres.py — Produit le TABLEAU DE PROPOSITION pour nettoyer
le champ genre (375 valeurs distinctes constatées le 2026-07-27).

PRINCIPE — même méthode que pour public_vise, adaptée au volume :
la décision revient au bibliothécaire, mais on ne va pas lui demander de
trancher 375 lignes une par une dans le chat. Ce script écrit donc un
Excel : valeur actuelle, effectif, valeur proposée, motif. Thomas le relit,
corrige la colonne « Proposé » où il n'est pas d'accord, et le script
appliquer_fusion_genres.py (à venir) exécutera le fichier VALIDÉ tel quel.

LA RÈGLE DE PROPOSITION — unifier le vocabulaire, garder les critères :
  - un genre composé reste composé (« Aventure / Humour » garde ses deux
    critères) ; on ne tronque JAMAIS ;
  - chaque composant est ramené à sa forme canonique : « Histoire » ->
    « Historique », « Amour » ou « Romance » seuls -> « Amour / Romance »
    (les paires indissociables du fonds sont traitées comme UN critère) ;
  - séparateur uniforme « / », doublons retirés, ordre d'origine conservé
    (le premier genre est le dominant).

STRICTEMENT EN LECTURE sur la base. N'écrit que le fichier Excel.

Usage :  python3 proposer_fusion_genres.py
"""
import datetime
import os
import sys

DOSSIER = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, DOSSIER)
import db  # noqa: E402

# Paires que le fonds traite comme UN SEUL critère : chaque membre isolé
# est ramené à la paire complète.
PAIRES = {
    "amour": "Amour / Romance", "romance": "Amour / Romance",
    "geographie": "Géographie / Voyage", "voyage": "Géographie / Voyage",
    "nature": "Nature / Animaux", "animaux": "Nature / Animaux",
    "conte": "Conte / Mythe", "mythe": "Conte / Mythe",
    "policier": "Policier / Mystère", "mystere": "Policier / Mystère",
}

# Synonymes simples -> forme canonique
SYNONYMES = {
    "histoire": "Historique",
    "recits de vie": "Récits de vie",
    "science fiction": "Science-fiction",
    "sciences fiction": "Science-fiction",
    "chanson pour enfants": "Chanson",
    "economie": "Économie",
}

# Composants qui dupliquent la colonne categorie : signalés « à trancher »,
# jamais convertis en douce.
DOUBLONS_CATEGORIE = {"documentaire", "comics", "roman graphique"}


def _cle(t):
    import unicodedata, re
    t = unicodedata.normalize("NFKD", t or "")
    t = "".join(c for c in t if not unicodedata.combining(c)).lower()
    return re.sub(r"\s+", " ", t).strip()


def proposer(valeur):
    """Renvoie (proposition, motif). Proposition identique = rien à faire."""
    if not valeur:
        return valeur, ""
    composants = [c.strip() for c in valeur.split("/") if c.strip()]
    sortie, vus, motifs = [], set(), []
    for c in composants:
        k = _cle(c)
        if k in DOUBLONS_CATEGORIE:
            canon = c  # conservé tel quel, mais signalé
            motifs.append(f"« {c} » double la colonne categorie : à trancher")
        elif k in PAIRES:
            canon = PAIRES[k]
            if _cle(canon) != k:
                motifs.append(f"« {c} » -> paire « {canon} »")
        elif k in SYNONYMES:
            canon = SYNONYMES[k]
            motifs.append(f"« {c} » -> « {canon} »")
        else:
            canon = c
        if _cle(canon) not in vus:
            vus.add(_cle(canon))
            sortie.append(canon)
        else:
            motifs.append("doublon retiré")
    proposition = " / ".join(sortie)
    return proposition, " ; ".join(dict.fromkeys(motifs))


def main():
    conn = db.connect()
    lignes = conn.execute(
        "SELECT genre, COUNT(*) FROM notice WHERE genre IS NOT NULL "
        "AND genre != '' GROUP BY genre ORDER BY 2 DESC").fetchall()
    conn.close()

    tableau, inchangees = [], 0
    for valeur, n in lignes:
        prop, motif = proposer(valeur)
        if prop == valeur and not motif:
            inchangees += 1
        tableau.append({
            "Genre actuel": valeur,
            "Notices": n,
            "Proposé (à corriger si besoin)": prop,
            "Changement ?": "" if prop == valeur else "OUI",
            "Motif": motif,
        })

    distinctes_apres = len({t["Proposé (à corriger si besoin)"] for t in tableau})
    print(f"{len(tableau)} valeurs actuelles -> {distinctes_apres} après fusion "
          f"({inchangees} inchangées)")

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    ws = wb.active
    ws.title = "Fusion des genres"
    colonnes = list(tableau[0].keys())
    thin = Side(style="thin", color="D9E2F0")
    bordure = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c, nom in enumerate(colonnes, 1):
        cell = ws.cell(row=1, column=c, value=nom)
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2E4A7A")
        cell.border = bordure
        ws.column_dimensions[get_column_letter(c)].width = (
            44 if "Genre" in nom or "Proposé" in nom or "Motif" in nom else 12)
    surligne = PatternFill("solid", fgColor="FFF3CD")
    for i, ligne in enumerate(tableau, start=2):
        for c, nom in enumerate(colonnes, 1):
            cell = ws.cell(row=i, column=c, value=ligne[nom])
            cell.font = Font(name="Arial", size=9.5)
            cell.border = bordure
            if ligne["Changement ?"] == "OUI":
                cell.fill = surligne
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(colonnes))}{len(tableau)+1}"

    sortie = os.path.join(
        DOSSIER, f"Fusion_genres_{datetime.date.today().isoformat()}.xlsx")
    wb.save(sortie)
    print(f"✓ Tableau : {os.path.basename(sortie)}")
    print("  Lignes jaunes = changement proposé. Corrige la colonne « Proposé »")
    print("  où tu n'es pas d'accord, puis on applique le fichier validé.")


if __name__ == "__main__":
    main()
