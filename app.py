
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
app.py — Application web (Streamlit) pour l'inventaire du fonds jeunesse.
Médiathèque d'Arcachon — Réseau COBAS.

Fichier UNIQUE volontairement : toute la logique (recherche bibliographique,
écriture Excel, statistiques, acquisitions/désherbage, catégories
personnalisées) est regroupée ici plutôt qu'éclatée en plusieurs modules,
pour qu'il n'y ait jamais qu'un seul fichier à remplacer lors d'une mise à
jour — et donc plus aucun risque d'oublier l'un des fichiers.

Lancement local (test) :
    streamlit run app.py

Lancement sur le serveur (VPS) :
    streamlit run app.py --server.port 8501 --server.address 0.0.0.0
"""

VERSION = "2026-06-25.2"  # élargissement Statut de publication (tout le fonds) + correction Nouveautés (catégorie réelle, pas juste récence)

import urllib.parse, xml.etree.ElementTree as ET
import time, os, sys, re, unicodedata, datetime, shutil, random
from collections import Counter, defaultdict

try:
    import requests
    from bs4 import BeautifulSoup
    import streamlit as st
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install",
                           "openpyxl", "requests", "beautifulsoup4", "streamlit"])
    import requests
    from bs4 import BeautifulSoup
    import streamlit as st
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════════════════
# PARTIE 1 — Recherche bibliographique (ex lib_recherche.py)
# ═══════════════════════════════════════════════════════════════


HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Accept-Language": "fr-FR,fr;q=0.9",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Referer": "https://www.google.fr/",
}

EMPTY = {
    "titre":"","auteur":"","illustrateur":"","editeur":"","annee":"",
    "type":"","public":"","genre":"","serie":"","tome":"","pegi":"",
    "collection":"","resume":"","source":"","statut":"NON TROUVÉ"
}

# ─────────────────────────────────────────────────────────
# CORRECTIONS MANUELLES — pour les cas mal gérés par les sources
# Ajouter ici les ISBN qui posent problème avec leurs données correctes
# ─────────────────────────────────────────────────────────
CORRECTIONS = {
    "9782226443410": {
        "titre":        "Archibald - Ma maison",
        "serie":        "Archibald",
        "auteur":       "Astrid Desbordes",
        "illustrateur": "Pauline Martin",
        "editeur":      "Albin Michel",
        "annee":        "2019",
        "type":         "Album",
        "public":       "Dès 3 ans",
        "genre":        "Vie quotidienne",
        "pegi":         "",
        "tome":         "",
    },
    "9789055795789": {
        "titre":        "Les Aventures de Tom Pouce",
        "auteur":       "Meyer M.",
        "editeur":      "N/C",
        "annee":        "N/C",
        "type":         "Conte / Poésie",
        "public":       "Jeunesse",
        "genre":        "Conte / Mythe",
        "pegi":         "",
        "tome":         "",
        "serie":        "",
    },
}


CHAMPS_CLES = ["titre","auteur","editeur","annee","type","public"]

# ─────────────────────────────────────────────────────────
# TABLES DE CORRESPONDANCE
# ─────────────────────────────────────────────────────────

# ─────────────────────────────────────────────────────────
# NORMALISATION DES ÉDITEURS
# Uniformise les variantes d'un même éditeur vers une forme canonique
# ─────────────────────────────────────────────────────────
NORMALISATION_EDITEURS = {
    # Gallimard
    "gallimard-jeunesse":           "Gallimard Jeunesse",
    "éditions gallimard jeunesse":  "Gallimard Jeunesse",
    "editions gallimard jeunesse":  "Gallimard Jeunesse",
    "gallimard jeunesse musique":   "Gallimard Jeunesse",
    "gallimard jeun.":              "Gallimard Jeunesse",
    "nrf gallimard":                "Gallimard",
    "folio gallimard":              "Gallimard Jeunesse",
    # École des loisirs
    "l'ecole des loisirs":          "École des Loisirs",
    "l'école des loisirs":          "École des Loisirs",
    "ecole des loisirs":            "École des Loisirs",
    "l école des loisirs":          "École des Loisirs",
    # Bayard
    "bayard jeunesse":              "Bayard",
    "bayard editions":              "Bayard",
    "bayard éditions":              "Bayard",
    # Milan
    "milan jeunesse":               "Milan",
    "éditions milan":               "Milan",
    "editions milan":               "Milan",
    # Albin Michel
    "albin michel jeunesse":        "Albin Michel",
    "editions albin michel":        "Albin Michel",
    "éditions albin michel":        "Albin Michel",
    # Hachette
    "hachette jeunesse":            "Hachette",
    "hachette livre":               "Hachette",
    "hachette pratique":            "Hachette",
    "hachette roman":               "Hachette",
    # Flammarion
    "flammarion jeunesse":          "Flammarion",
    "père castor flammarion":       "Flammarion",
    "pere castor flammarion":       "Flammarion",
    # Nathan
    "nathan jeunesse":              "Nathan",
    "editions nathan":              "Nathan",
    "éditions nathan":              "Nathan",
    # Actes Sud
    "actes sud junior":             "Actes Sud",
    "actes sud j.":                 "Actes Sud",
    # Pika
    "pika édition":                 "Pika",
    "pika edition":                 "Pika",
    "pika editions":                "Pika",
    # Kana
    "kana manga":                   "Kana",
    # Dupuis
    "editions dupuis":              "Dupuis",
    "éditions dupuis":              "Dupuis",
    # Dargaud
    "editions dargaud":             "Dargaud",
    "éditions dargaud":             "Dargaud",
    # Casterman
    "casterman jeunesse":           "Casterman",
    "editions casterman":           "Casterman",
    # Le Lombard
    "editions du lombard":          "Le Lombard",
    "éditions du lombard":          "Le Lombard",
    # Sarbacane
    "editions sarbacane":           "Sarbacane",
    "éditions sarbacane":           "Sarbacane",
    # Kaléidoscope
    "kaleidoscope":                 "Kaléidoscope",
    "editions kaleidoscope":        "Kaléidoscope",
    # Didier
    "didier jeunesse":              "Didier",
    "editions didier":              "Didier",
    # Pastel
    "l'école des loisirs / pastel": "Pastel",
    # Auzou
    "editions auzou":               "Auzou",
    "éditions auzou":               "Auzou",
    # La Martinière
    "la martinière jeunesse":       "La Martinière",
    "éditions de la martinière":    "La Martinière",
    # Glénat
    "glenat":                       "Glénat",
    "editions glenat":              "Glénat",
    "éditions glénat":              "Glénat",
    # Bamboo
    "bamboo edition":               "Bamboo",
    "bamboo édition":               "Bamboo",
    # Folio (rattaché à Gallimard)
    "folio junior":                 "Gallimard Jeunesse",
    "folio cadet":                  "Gallimard Jeunesse",
    "folio benjamin":               "Gallimard Jeunesse",
}

def normaliser_editeur(editeur_brut):
    """Normalise le nom d'un éditeur vers sa forme canonique."""
    if not editeur_brut:
        return editeur_brut
    # Nettoyage de base
    e = editeur_brut.strip()
    e = e.strip(".,;")
    # Recherche dans la table (insensible à la casse)
    e_lower = e.lower()
    if e_lower in NORMALISATION_EDITEURS:
        return NORMALISATION_EDITEURS[e_lower]
    # Recherche partielle si pas de correspondance exacte
    for variante, canonique in NORMALISATION_EDITEURS.items():
        if variante in e_lower:
            return canonique
    # Première lettre en majuscule si pas trouvé
    return e


def _sans_accents(s):
    """Normalise une chaîne en retirant les accents, pour fiabiliser les
    comparaisons (ex: 'Glénat Manga' doit matcher la clé 'glenat manga')."""
    return ''.join(c for c in unicodedata.normalize('NFKD', s or "") if not unicodedata.combining(c))

EDITEUR_TYPE = {
    "pika":         "Manga", "kana":          "Manga",
    "kurokawa":     "Manga", "tonkam":        "Manga",
    "kazé":         "Manga", "nobi nobi":     "Manga",
    "ki-oon":       "Manga", "akata":         "Manga",
    "glenat manga": "Manga", "bamboo manga":  "Manga",
    "soleil manga": "Manga", "delcourt tonkam":"Manga",
    "le lombard":   "BD",    "dupuis":        "BD",
    "dargaud":      "BD",    "casterman":     "BD",
    "bamboo":       "BD",    "glénat":        "BD",
    "lucky comics": "BD",    "grand angle":   "BD",
    "futuropolis":  "BD",    "vents d'ouest": "BD",
    "400 coups":    "Album",
    "hatier jeunesse": "Album",
    "hatier":       "Album",
    "kaléidoscope": "Album",
    "didier jeunesse": "Album",
    "pastel":       "Album",
    "le nouvel atila": "Première lecture",
    "biosphoto":    "Documentaire",
    "le rouergue":  "Roman jeunesse",
    "seuil jeunesse":"Roman jeunesse",
    "sarbacane":    "Roman jeunesse",
    "actes sud junior":"Roman jeunesse",
    "saxo":         "Roman jeunesse",
    "hélium":       "Roman jeunesse",
    "helium":       "Roman jeunesse",
    "auzou":        "Roman jeunesse",
}

EDITEUR_PUBLIC = {
    "pika":                 "Ado (12+)",
    "kana":                 "Ado (12+)",
    "kurokawa":             "Ado (12+)",
    "ki-oon":               "Ado (12+)",
    "akata":                "Ado (12+)",
    "kazé":                 "Ado (12+)",
    "glenat manga":         "Ado (12+)",
    "école des loisirs":    "Jeunesse",
    "gallimard jeunesse":   "Jeunesse",
    "gallimard":            "Jeunesse",
    "bayard jeunesse":      "Jeunesse",
    "bayard":               "Jeunesse",
    "milan":                "Jeunesse",
    "nathan":               "Jeunesse",
    "hachette jeunesse":    "Jeunesse",
    "hachette":             "Jeunesse",
    "rageot":               "Jeunesse",
    "flammarion jeunesse":  "Jeunesse",
    "actes sud junior":     "Jeunesse",
    "actes sud":            "Jeunesse",
    "sarbacane":            "Jeunesse",
    "albin michel jeunesse":"Jeunesse",
    "albin michel":         "Jeunesse",
    "le lombard":           "Jeunesse",
    "dupuis":               "Jeunesse",
    "dargaud":              "Jeunesse",
    "bayard":               "Jeunesse",
    "editions du ricochet": "Jeunesse",
    "saxo":                 "Jeunesse",
    "helium":               "Jeunesse",
    "hélium":               "Jeunesse",
    "pastel":               "Dès 3 ans",
    "kaléidoscope":         "Dès 3 ans",
    "didier jeunesse":      "Dès 3 ans",
    "nobi nobi":            "Jeunesse",
    "auzou":                "Jeunesse",
    "bamboo":               "Jeunesse",
    "casterman":            "Jeunesse",
    "glénat":               "Jeunesse",
    "hatier jeunesse":      "Jeunesse",
    "hatier":               "Jeunesse",
    "400 coups":            "Jeunesse",
    "la martinière jeunesse":"Jeunesse",
    "la martinière":        "Jeunesse",
    "bios":                 "Jeunesse",
    "biosphoto":            "Jeunesse",
    "seuil jeunesse":       "Jeunesse",
    "seuil":                "Jeunesse",
    "le rouergue":          "Jeunesse",
    "le nouvel atila":      "Première lecture",
    "helium":               "Jeunesse",
    "milan poche":          "Jeunesse",
    "folio cadet":          "Dès 6 ans",
}

COLLECTION_TYPE = {
    "pika shonen":    "Manga", "pika shônen":  "Manga",
    "pika seinen":    "Manga", "pika shojo":   "Manga",
    "kana shonen":    "Manga", "kana shojo":   "Manga",
    "pôle fiction":   "Roman ado / YA",
    "exprim":         "Roman ado / YA",
    "scripto":        "Roman ado / YA",
    "heure noire":    "Roman ado / YA",
    "black moon":     "Roman ado / YA",
    "tribal":         "Roman ado / YA",
    "folio junior":   "Roman jeunesse",
    "folio cadet":    "Roman jeunesse",
    "folio benjamin": "Roman jeunesse",
    "milan poche":    "Roman jeunesse",
    "j'aime lire":    "Première lecture",
    "premiers romans":"Première lecture",
    "les goûters philo":          "Documentaire",
    "les yeux de la découverte":  "Documentaire",
    "mes premières découvertes":  "Documentaire",
    "mes p'tites questions":      "Documentaire",
    "pocqq":                      "Documentaire",
}

COLLECTION_PUBLIC = {
    "folio cadet":          "Dès 6 ans",
    "folio benjamin":       "Dès 6 ans",
    "folio junior":         "8-12 ans",
    "folio ado":            "Ado (12+)",
    "milan poche cadet":    "Dès 6 ans",
    "milan poche junior":   "8-12 ans",
    "pôle fiction":         "Ado (12+)",
    "exprim":               "Ado (12+)",
    "scripto":              "Ado (12+)",
    "heure noire":          "Ado (12+)",
    "black moon":           "Ado (12+)",
    "tribal":               "Ado (12+)",
    "pika shonen":          "Ado (12+)",
    "pika shônen":          "Ado (12+)",
    "pika seinen":          "Ado (12+)",
    "les yeux de la découverte": "8-12 ans",
    "mes premières découvertes": "Dès 6 ans",
    "mes p'tites questions":     "Dès 6 ans",
    "les goûters philo":         "8-12 ans",
    "pocqq":                     "8-12 ans",
    "éclats de rire":            "Dès 6 ans",
    "j'aime lire":               "Dès 6 ans",
    "premiers romans":           "Dès 6 ans",
    "cascade":                   "8-12 ans",
    "folio cadet":               "Dès 6 ans",
}

# ─────────────────────────────────────────────────────────
# DÉTECTION TYPE
# ─────────────────────────────────────────────────────────

TYPES_MOTS = [
    ("Manga",            ["manga","manhwa","manhua","shonen","shônen",
                          "shojo","shôjo","seinen","josei","kodomo"]),
    ("BD",               ["bande dessinée","bande-dessinée","comics",
                          "graphic novel","album bd"]),
    ("Album",            ["album illustré","album jeunesse","album cartonné",
                          "imagier","livre-images","tout-carton"]),
    ("Première lecture", ["première lecture","premières lectures",
                          "je commence à lire","lecture débutant",
                          "facile à lire"]),
    ("Conte / Poésie",   ["conte","contes","fable","fables","poésie",
                          "poème","comptine","légende","mythe",
                          "grimm","perrault","andersen"]),
    ("Documentaire",     ["documentaire","encyclopédie","atlas",
                          "dictionnaire","sciences","découverte",
                          "géographie","goûters philo","comment ça marche",
                          "pourquoi","c'est quoi","kézako"]),
    ("Livre-jeu / Activités", ["livre-jeu","activités","coloriage",
                               "origami","autocollants","escape book",
                               "cahier de vacances","cherche et trouve",
                               "dont vous êtes le héros"]),
    ("Roman ado / YA",   ["young adult"," ya ","exprim","scripto",
                          "pôle fiction","heure noire","black moon"]),
    ("Roman jeunesse",   ["roman jeunesse","roman enfant","folio junior",
                          "milan poche","rageot"]),
]

def detecter_type(titre, collection, resume, genre_raw, support, editeur, public):
    ed = _sans_accents(editeur.lower())
    for ek, tv in EDITEUR_TYPE.items():
        if _sans_accents(ek) in ed:
            return tv
    coll = _sans_accents(collection.lower())
    for ck, tv in COLLECTION_TYPE.items():
        if _sans_accents(ck) in coll:
            return tv
    texte = " ".join([titre, collection, resume, genre_raw, support]).lower()
    for type_nom, mots in TYPES_MOTS:
        if any(m in texte for m in mots):
            if type_nom == "Roman jeunesse" and public in ("Ado (12+)","Ado","Dès 12 ans"):
                return "Roman ado / YA"
            return type_nom
    if "texte" in support.lower():
        return "Roman ado / YA" if public in ("Ado (12+)","Ado") else "Roman jeunesse"
    return ""

# ─────────────────────────────────────────────────────────
# DÉTECTION PUBLIC
# ─────────────────────────────────────────────────────────

def normaliser_public(val):
    """Normalise les valeurs de public vers nos valeurs standard."""
    if not val:
        return ""
    v = val.lower().strip()
    if re.search(r"\b[0-3]\s*[àa]\s*\d+|maternelle|tout-petit|bébé|0-3", v): return "Dès 3 ans"
    if re.search(r"\bdès\s*6\b|6\s*[àa]\s*\d+|cp\b|6-8\b|6 ans", v):         return "Dès 6 ans"
    if re.search(r"\b[6-7]\s*[àa]\s*\d+|ce1\b|7-8\b|7 ans", v):              return "Dès 6 ans"
    if re.search(r"\b(8|9)\s*[àa]\s*\d+|dès\s*8\b|ce2\b|8-10\b|8 ans", v):  return "8-12 ans"
    if re.search(r"\b(10|11)\s*[àa]\s*\d+|dès\s*10\b|10-12\b|10 ans", v):   return "8-12 ans"
    if re.search(r"\bdès\s*12\b|12\s*[àa]\s*\d+|collège|12-16\b|12 ans", v): return "Ado (12+)"
    if re.search(r"\b(13|14|15|16)\s*[àa]\s*\d+|lycée|teen|young adult", v): return "Ado (12+)"
    if re.search(r"\bado\b|ya\b|young adult", v):                              return "Ado (12+)"
    if re.search(r"\bjeunesse\b|enfant|junior", v):                            return "Jeunesse"
    if re.search(r"\d", val):
        return val.strip()
    return ""

def detecter_public(public_code, titre, collection, editeur, tranche_age):
    # 1. Tranche d'âge explicite
    if tranche_age:
        p = normaliser_public(tranche_age)
        if p: return p

    # 2. Code BnF
    CODES = {"j":"Jeunesse","y":"Ado (12+)","a":"Tout public",
             "b":"Jeunesse","c":"Ado (12+)","d":"Adulte"}
    if public_code and public_code.lower() in CODES:
        return CODES[public_code.lower()]

    # 3. Collection
    coll = _sans_accents(collection.lower())
    for ck, pv in COLLECTION_PUBLIC.items():
        if _sans_accents(ck) in coll: return pv

    # 4. Éditeur
    ed = _sans_accents(editeur.lower())
    for ek, pv in EDITEUR_PUBLIC.items():
        if _sans_accents(ek) in ed: return pv

    # 5. Mots-clés titre + collection
    texte = (titre + " " + collection).lower()
    if any(m in texte for m in ["bébé","tout-petit","0-3"]):               return "Dès 3 ans"
    if any(m in texte for m in ["dès 6","6-8","cp ","ce1"]):               return "Dès 6 ans"
    if any(m in texte for m in ["dès 8","8-10","ce2","cm1","cm2"]):        return "8-12 ans"
    if any(m in texte for m in ["dès 10","10-12"]):                         return "8-12 ans"
    if any(m in texte for m in ["dès 12","12-16","collège"]):               return "Ado (12+)"
    if any(m in texte for m in ["ado","teen"," ya ","young adult","exprim",
                                  "pôle fiction","scripto"]):               return "Ado (12+)"
    if any(m in texte for m in ["jeunesse","enfant","poche jeun"]):         return "Jeunesse"
    return ""

# ─────────────────────────────────────────────────────────
# PEGI
# ─────────────────────────────────────────────────────────

def detecter_pegi(type_doc, public, texte_brut=""):
    """Déduit le PEGI pour les mangas et BD."""
    if type_doc not in ("Manga", "BD"):
        return ""
    # Chercher PEGI explicite dans le texte
    m = re.search(r"PEGI\s*(\d+)", texte_brut, re.IGNORECASE)
    if m:
        return f"PEGI {m.group(1)}"
    # Déduire depuis le public
    if public in ("Ado (12+)", "Dès 12 ans", "12-16 ans"):
        return "PEGI 12"
    if public in ("Ado", "Ado / YA") or re.search(r"\b(16|17|18)\b", public):
        return "PEGI 16"
    if public in ("Jeunesse", "Dès 6 ans", "6-8 ans", "8-12 ans"):
        return "PEGI 7"
    if type_doc == "Manga":
        return "PEGI 12"  # Par défaut manga
    return ""

# ─────────────────────────────────────────────────────────
# GENRE
# ─────────────────────────────────────────────────────────

GENRES = [
    ("Aventure",              ["aventure","exploration","pirates","trésor","expédition","quête"]),
    ("Policier / Mystère",    ["policier","détective","enquête","mystère","crime","suspect","meurtre"]),
    ("Fantastique", ["fantasy","fantastique","magie","sorcier","dragon","elfe","vampire",
                                "zombie","loup-garou","fée","magicien","enchantement","créature","monstre"]),
    ("Science-fiction",       ["science-fiction","science fiction"," sf ","robot","extraterrestre",
                                "futur","dystopie","spatial","planète","vaisseau","cyborg"]),
    ("Humour",                ["humour","comique","drôle","rigolo","éclats de rire","farce","blague","loufoque"]),
    ("Amour / Romance",       ["amour","romance","amoureux","love","sentiments","coeur","coup de foudre"]),
    ("Historique",            ["historique","histoire vraie","guerre","moyen âge","résistance",
                                "antiquité","révolution","chevalier","viking","empire","pharaon"]),
    ("Vie quotidienne",       ["école","famille","amitié","grandir","identité","harcèlement",
                                "divorce","adolescence","émotions","confiance","différence","handicap"]),
    ("Nature / Animaux",      ["animal","animaux","nature","vétérinaire","cheval","chien","chat",
                                "écologie","forêt","mer","sauvage","espèce","dinosaure","insecte"]),
    ("Sport",                 ["sport","football","basket","tennis","judo","badminton",
                                "natation","rugby","foot","cyclisme","athlétisme","gym"]),
    ("Arts / Musique",        ["musique","peinture","art ","artiste","dessin","danse",
                                "théâtre","cinéma","sculpture","photographie"]),
    ("Philo / Société",       ["philosophie","philo","goûters philo","éthique","société",
                                "citoyenneté","liberté","justice","droits","démocratie","religion"]),
    ("Frissons",              ["horreur","frisson","peur","fantôme","maison hantée","cauchemar","terrifiant"]),
    ("Sciences",              ["sciences","physique","chimie","biologie","mathématiques",
                                "astronomie","corps humain","expérience","inventeur","médecine"]),
    ("Géographie / Voyage",   ["géographie","pays","voyage","monde","continent","atlas","carte",
                                "exploration","découverte du monde"]),
    ("Conte / Mythe",         ["conte","mythe","légende","fable","folklore","grimm","perrault",
                                "andersen","1001 nuits","récit légendaire"]),
    ("Cuisine / Activités",   ["cuisine","recette","bricolage","jardinage","origami",
                                "coloriage","jeu de société","jeu vidéo","jeux et activités",
                                "cherche et trouve","autocollant","gommette"]),
]

# Genres par défaut selon le type de document
GENRE_PAR_TYPE = {
    "Manga":             "Aventure",       # par défaut si pas détecté
    "BD":                "",               # trop varié
    "Documentaire":      "Sciences",       # par défaut documentaire général
    "Album":             "Vie quotidienne",# par défaut album
    "Première lecture":  "Aventure",
    "Conte / Poésie":    "Conte / Mythe",
}

def _mot_present(mot, texte):
    """Teste la présence d'un mot/expression en respectant les limites de mots,
    pour éviter les faux positifs du type 'chimie' trouvé dans 'alchimie'
    ou 'jeu' trouvé dans 'enjeu'/'jeune'/'déjeuner'."""
    return re.search(r"\b" + re.escape(mot.strip()) + r"\b", texte) is not None

def detecter_genre(titre, collection, resume, sujets, type_doc=""):
    texte_titre = titre.lower()
    texte_total = " ".join([titre, collection, resume, sujets]).lower()

    scores = {}
    for nom, mots in GENRES:
        score = 0
        for m in mots:
            if _mot_present(m, texte_titre):
                score += 3        # un mot-clé présent dans le TITRE est très fiable
            elif _mot_present(m, texte_total):
                score += 1        # sinon, simple occurrence dans le résumé/collection
        if score:
            scores[nom] = score

    if scores:
        # Maximum 2 thèmes, les mieux notés — pas les 2 premiers trouvés dans la liste
        meilleurs = sorted(scores.items(), key=lambda kv: -kv[1])[:2]
        return " / ".join(nom for nom, _ in meilleurs)

    # Fallback : genre par défaut selon le type
    if type_doc in GENRE_PAR_TYPE and GENRE_PAR_TYPE[type_doc]:
        return GENRE_PAR_TYPE[type_doc]
    return ""

# ─────────────────────────────────────────────────────────
# TOME ET SÉRIE
# ─────────────────────────────────────────────────────────

# ── Patterns d'extraction du numéro de tome (ordre de priorité)
TOME_PATTERNS = [
    r'[Tt]ome\s*(\d{1,3})',           # Tome 3 / tome3
    r'\bT\.\s*(\d{1,3})\b',           # T.3
    r'[Vv]ol(?:ume)?\s*(\d{1,3})',    # Vol.3 / Volume 3
    r'[Nn]°\s*(\d{1,3})',             # N°3
    r'#\s*(\d{1,3})',                  # #3
    r'\((\d{1,2})\)\s*$',             # (3) en fin
    r',\s*(\d{1,2})\s*$',             # , 3 en fin
    r'[-–]\s*(\d{1,3})\s*(?:$|[-–:])',# - 3 ou - 3 :
    r'(?<!\d)(\d{1,3})\s*$',          # chiffre seul en fin (Blue Lock 27)
]

def extraire_tome(tome_raw, titre):
    """Extrait le numéro de tome depuis les métadonnées ou le titre."""
    # 1. Depuis les métadonnées (BnF 225$v)
    if tome_raw:
        try:
            n = int(re.search(r"\d+", tome_raw).group())
            if 1 <= n <= 99:
                return str(n)
        except Exception:
            pass
    # 2. Depuis le titre
    for pattern in TOME_PATTERNS:
        m = re.search(pattern, titre)
        if m:
            try:
                n = int(m.group(1))
                if 1 <= n <= 99:
                    return str(n)
            except Exception:
                pass
    return ""

def nettoyer_titre(titre_brut, tome, serie):
    """
    Retire le numéro de tome du titre.
    Exemples :
      "Blue Box5 - BLUE BOX"                          → "BLUE BOX"
      "Blue lock27 - Blue Lock T27"                   → "Blue Lock"
      "Game over17 - Game over / Dark web"            → "Game over / Dark web"
      "Les petits foot maniacs15 - Les Footmaniacs"   → "Les Footmaniacs"
      "Album de Boule & Bill.46 - Boule & Bill T.46"  → "Boule & Bill"
      "7 - Archibald - Ma maison"                     → "Ma maison"
      "Boule & Bill - Tome 46 - Peinture à l'os"      → "Peinture à l'os"
    """
    if not titre_brut: return ""
    t = titre_brut.strip()

    supprimer_serie = True

    if tome:
        # "NomXX - Titre réel" → "Titre réel"  ex: "Blue Box5 - BLUE BOX"
        m = re.match(r'^(.+?)[\s.]?' + re.escape(tome) + r'\s*[-–]\s*(.+)$', t)
        if m and len(m.group(1)) > 0 and len(m.group(2)) > 3:
            t = m.group(2).strip()

        # "Préfixe.XX - Titre" → "Titre"  ex: "Album de Boule & Bill.46 - ..."
        m2 = re.match(r'^.+?\.' + re.escape(tome) + r'\s*[-–]\s*(.+)$', t)
        if m2 and len(m2.group(1)) > 3:
            t = m2.group(1).strip()

    # Retirer toutes les formes de numérotation
    t = re.sub(r'\s*[-–,]\s*[Tt]ome\s*\d{1,3}', '', t)
    t = re.sub(r'\s*[-–,]\s*[Tt]\.\s*\d{1,3}\b', '', t)
    t = re.sub(r'\s*[-–,]\s*[Vv]ol(?:ume)?\s*\d{1,3}', '', t)
    t = re.sub(r'\s*[-–,]\s*[Nn]°\s*\d{1,3}', '', t)
    t = re.sub(r'\s*#\s*\d{1,3}', '', t)
    if tome:
        t = re.sub(r'\s+[Tt]\.?\s*' + re.escape(tome) + r'\s*$', '', t)
        t = re.sub(r'\s+' + re.escape(tome) + r'\s*$', '', t)

    # Si la série est en doublon au début AVEC un numéro de tome
    # ex: "Blue Lock 27 - Blue Lock" → on garde "Blue Lock"
    # Mais PAS "Archibald - Ma maison" → on garde le titre complet
    if supprimer_serie and serie and len(serie) > 3 and tome:
        serie_esc = re.escape(serie)
        # Seulement si le titre contient le nom de série SUIVI d'un chiffre
        m = re.match(r'^' + serie_esc + r'\s*\d+\s*[-–:]\s*(.+)$', t, re.IGNORECASE)
        if m and len(m.group(1)) > 3:
            t = m.group(1).strip()

    # Nettoyage final
    t = re.sub(r'\s*[-–:,]+\s*$', '', t)
    t = re.sub(r'^\s*[-–:,]+\s*', '', t)
    t = re.sub(r'\s{2,}', ' ', t)
    t = t.strip(" -–:,.")
    return t if len(t) > 2 else titre_brut.strip()

COLLECTIONS_GENERIQUES = [
    "folio junior","folio cadet","folio benjamin","folio ado",
    "poche jeunesse","livre de poche","milan poche","nathan poche",
    "cascade","tribal","pôle fiction","exprim","scripto",
    "heure noire","autres mondes","black moon",
    "pika shonen","pika shônen","pika seinen",
    "les yeux de la découverte","mes premières découvertes",
    "les goûters philo","mes p'tites questions","pocqq",
    "premiers romans","j'aime lire","je commence à lire","éclats de rire",
]

def nettoyer_serie(serie_brut, tome):
    """Nettoie le nom de série en retirant le numéro de tome et les suffixes parasites."""
    if not serie_brut:
        return ""
    s = serie_brut.strip()
    if tome:
        # Retirer ".XX" ou " XX" ou "-XX" en fin ou au milieu
        s = re.sub(r'[\s.]' + re.escape(tome) + r'\s*[-–].*$', '', s)
        s = re.sub(r'[\s.,-]' + re.escape(tome) + r'\s*$', '', s)
    # Retirer "- Titre réel" si la série contient un tiret séparateur
    # ex: "Album de Boule & Bill.46 - E..." → "Album de Boule & Bill"
    if tome:
        s = re.sub(r'\.?' + re.escape(tome) + r'.*$', '', s)
    # Retirer les séparateurs résiduels en fin
    s = re.sub(r'\s*[-–:.,]+\s*$', '', s)
    s = s.strip(" -–:.,")
    return s if len(s) > 2 else serie_brut.strip()

def extraire_serie(titre, serie_raw, collection, tome=""):
    # 1. Depuis les métadonnées (BnF, Babelio...)
    if serie_raw and len(serie_raw) > 2:
        return nettoyer_serie(serie_raw.strip(" .,()"), tome)

    # 2. Depuis le titre
    patterns = [
        r'^(.+?)\s*[,\-–]\s*[Tt]ome\s*\d+',
        r'^(.+?)\s*[,\-–]\s*[Tt]\.\s*\d+',
        r'^(.+?)\s*[Tt]ome\s*\d+',
        r'^(.+?)\s+[Tt]\.\s*\d+',
        r'^(.+?)\s*[Vv]ol(?:ume)?\s*\d+',
        r'^(.+?)\s+[Nn]°\s*\d+',
        r'^(.+?)\s+#\s*\d+',
        r'^(.+?)\s+\((\d{1,2})\)\s*$',
        r'^(.+?)\s*:\s*.+[\s\-–]\s*\d{1,3}\s*$',
        # "NomXX - Titre" → série = "Nom"
        r'^(.+?)[\s.]?\d{1,3}\s*[-–]\s*.+$',
        # "Nom XX" chiffre en fin
        r'^(.+?)\s+(\d{1,3})\s*$',
    ]
    for pattern in patterns:
        m = re.match(pattern, titre, re.IGNORECASE)
        if m:
            serie = m.group(1).strip(" .,:–-()")
            serie = nettoyer_serie(serie, tome)
            if len(serie) > 2 and not serie.isdigit():
                return serie

    # 3. Depuis la collection si elle n'est pas générique
    if collection and len(collection) > 3:
        if not any(g in collection.lower() for g in COLLECTIONS_GENERIQUES):
            return collection.strip()
    return ""

# ─────────────────────────────────────────────────────────
# HELPER
# ─────────────────────────────────────────────────────────

def enrichir(base, nouveau):
    if not base:
        return nouveau
    for champ in ("auteur","illustrateur","editeur","annee","type","public",
                  "genre","serie","tome","pegi","collection","resume"):
        if not base.get(champ) and nouveau.get(champ):
            base[champ] = nouveau[champ]
    return base

def est_complet(res):
    complet = all(res.get(c) for c in CHAMPS_CLES)
    a_tome  = not res.get("serie") or res.get("tome")
    return complet and a_tome

# ─────────────────────────────────────────────────────────
# SOURCE 1 : BnF SRU
# ─────────────────────────────────────────────────────────

def bnf_lookup(isbn):
    try:
        query = f'bib.isbn adj "{isbn}"'
        url = (
            "https://catalogue.bnf.fr/api/SRU?"
            "version=1.2&operation=searchRetrieve"
            f"&query={urllib.parse.quote(query)}"
            "&maximumRecords=1&recordSchema=unimarcxchange"
        )
        r = requests.get(url, headers=HEADERS, timeout=12)
        if r.status_code != 200: return None
        root = ET.fromstring(r.content)
        ns = {"srw":"http://www.loc.gov/zing/srw/"}
        nb = root.find(".//srw:numberOfRecords", ns)
        if nb is None or nb.text == "0": return None
        rec = root.find(".//srw:recordData", ns)
        if rec is None: return None

        def sf(tag, code):
            for f in rec.iter():
                if f.get("tag") == tag:
                    for s in f:
                        if s.get("code") == code and s.text:
                            return s.text.strip(" .,:/()")
            return ""

        def sf_list(tag, code):
            vals = []
            for f in rec.iter():
                if f.get("tag") == tag:
                    for s in f:
                        if s.get("code") == code and s.text:
                            vals.append(s.text.strip(" .,:"))
            return vals

        titre_a = sf("200","a")
        titre_e = sf("200","e")
        titre_n = sf("200","n")
        titre_p = sf("200","p")
        titre = titre_a
        if titre_n and re.match(r"^\d+$", titre_n.strip()) and int(titre_n) <= 99:
            titre += f" T.{titre_n.strip()}"
        if titre_p:
            titre += f" : {titre_p}"
        elif titre_e and titre_e.lower() != titre_a.lower():
            titre += f" : {titre_e}"
        if not titre: return None

        auteurs, illustrateurs = [], []
        for tag in ("700","701","702"):
            for f in rec.iter():
                if f.get("tag") == tag:
                    nom = prenom = fonction = ""
                    for s in f:
                        if s.get("code") == "a": nom     = (s.text or "").strip(" .,")
                        if s.get("code") == "b": prenom  = (s.text or "").strip(" .,")
                        if s.get("code") == "4": fonction = s.text or ""
                    nc = f"{prenom} {nom}".strip()
                    if nc:
                        if fonction in ("110","740","080","440","230"):
                            illustrateurs.append(nc)
                        else:
                            auteurs.append(nc)
        if not auteurs:
            for tag in ("710","711"):
                for f in rec.iter():
                    if f.get("tag") == tag:
                        for s in f:
                            if s.get("code") == "a" and s.text:
                                auteurs.append(s.text.strip(" .,"))

        editeur   = sf("214","c") or sf("210","c")
        annee_raw = sf("214","d") or sf("210","d")
        m_an = re.search(r"\d{4}", annee_raw)
        annee = m_an.group(0) if m_an else ""
        collection = sf("225","a")
        tome_225   = sf("225","v")
        serie_raw  = ""
        for tag in ("410","411","440"):
            s = sf(tag,"t")
            if s: serie_raw = s; break

        support   = sf("200","b")
        genre_all = " ".join(sf_list("606","a") + sf_list("607","a") + sf_list("608","a"))
        tranche   = sf("521","a")
        public_code = ""
        for f in rec.iter():
            if f.get("tag") == "100":
                for s in f:
                    if s.get("code") == "a" and s.text and len(s.text) > 22:
                        public_code = s.text[22]
        resume = sf("330","a")

        public   = detecter_public(public_code, titre, collection, editeur, tranche)
        type_doc = detecter_type(titre, collection, resume, genre_all, support, editeur, public)
        genre    = detecter_genre(titre, collection, resume, genre_all, type_doc)
        tome     = extraire_tome(tome_225, titre)
        serie    = extraire_serie(titre, serie_raw, collection, tome)
        pegi     = detecter_pegi(type_doc, public, "")

        return {
            "titre":        titre,
            "auteur":       " / ".join(auteurs[:2]),
            "illustrateur": " / ".join(illustrateurs[:2]),
            "editeur":      editeur, "annee": annee,
            "type":         type_doc, "public": public,
            "genre":        genre, "serie": serie,
            "tome":         tome, "pegi": pegi,
            "collection":   collection,
            "resume":       resume[:400] if resume else "",
            "source":       "BnF", "statut": "trouvé",
        }
    except Exception:
        return None

# ─────────────────────────────────────────────────────────
# SOURCE 2 : Babelio
# ─────────────────────────────────────────────────────────

def babelio_lookup(isbn):
    try:
        url = f"https://www.babelio.com/recherche.php?Recherche={isbn}&recherche=isbn"
        r = requests.get(url, headers=HEADERS, timeout=12)
        if r.status_code != 200: return None
        soup = BeautifulSoup(r.text, "html.parser")
        lien = soup.select_one("a[href*='/livres/']")
        if not lien: return None
        href = lien.get("href","")
        if not href.startswith("http"): href = "https://www.babelio.com" + href
        r2 = requests.get(href, headers=HEADERS, timeout=12)
        if r2.status_code != 200: return None
        soup2 = BeautifulSoup(r2.text, "html.parser")

        titre_el = soup2.select_one("h1[itemprop='name'], h1.livre_con")
        titre = titre_el.get_text(strip=True) if titre_el else ""
        if not titre: return None
        auteur_el = soup2.select_one("span[itemprop='author'], a[href*='/auteur/']")
        auteur = auteur_el.get_text(strip=True) if auteur_el else ""

        editeur = annee = collection = genre_b = tranche = ""
        for info in soup2.select(".book-detail li, .livre_infos li, .infos span"):
            t = info.get_text(" ", strip=True)
            if re.search(r"éditeur|editeur", t, re.I):
                editeur = re.sub(r"éditeur\s*:?\s*","", t, flags=re.I).strip()
            elif re.search(r"parution|année|date", t, re.I):
                m = re.search(r"\d{4}", t)
                if m: annee = m.group(0)
            elif re.search(r"collection", t, re.I):
                collection = re.sub(r"collection\s*:?\s*","", t, flags=re.I).strip()
            elif re.search(r"genre|type", t, re.I):
                genre_b = re.sub(r"(genre|type)\s*:?\s*","", t, flags=re.I).strip()
            elif re.search(r"âge|ans\b|public", t, re.I):
                tranche = t

        resume_el = soup2.select_one("#d_bio, .livre_resume, [itemprop='description']")
        resume = resume_el.get_text(strip=True)[:400] if resume_el else ""

        public   = detecter_public("", titre, collection, editeur, tranche)
        type_doc = detecter_type(titre, collection, resume, genre_b, "", editeur, public)
        genre    = detecter_genre(titre, collection, resume, genre_b, type_doc)
        tome     = extraire_tome("", titre)
        serie    = extraire_serie(titre, "", collection, tome)
        pegi     = detecter_pegi(type_doc, public, soup2.get_text())

        return {"titre":titre,"auteur":auteur,"illustrateur":"",
                "editeur":editeur,"annee":annee,"type":type_doc,
                "public":public,"genre":genre,"serie":serie,"tome":tome,
                "pegi":pegi,"collection":collection,"resume":resume,
                "source":"Babelio","statut":"trouvé"}
    except Exception:
        return None

# ─────────────────────────────────────────────────────────
# SOURCE 3 : BDfugue
# ─────────────────────────────────────────────────────────

def bdfugue_lookup(isbn):
    try:
        url = f"https://www.bdfugue.com/catalogsearch/result/?q={isbn}"
        r = requests.get(url, headers=HEADERS, timeout=12)
        if r.status_code != 200: return None
        soup = BeautifulSoup(r.text, "html.parser")
        lien = soup.select_one("a.product-item-link, a.product-name, h2.product-name a")
        if not lien: return None
        href = lien.get("href","")
        if not href: return None
        r2 = requests.get(href, headers=HEADERS, timeout=12)
        if r2.status_code != 200: return None
        soup2 = BeautifulSoup(r2.text, "html.parser")
        titre_el = soup2.select_one("h1.page-title, h1[itemprop='name']")
        titre = titre_el.get_text(strip=True) if titre_el else ""
        if not titre: return None

        def ga(label):
            for el in soup2.select(".product-attribute,.product.attribute,tr"):
                if label.lower() in el.get_text(" ",strip=True).lower():
                    val = el.select_one(".value, td:last-child")
                    if val: return val.get_text(strip=True)
            return ""

        auteur    = ga("auteur") or ga("scénari") or ga("dessin")
        editeur   = ga("éditeur") or ga("editeur")
        annee_r   = ga("date") or ga("parution")
        m = re.search(r"\d{4}", annee_r)
        annee     = m.group(0) if m else ""
        collection = ga("collection") or ga("série")
        resume_el = soup2.select_one(".product.attribute.description .value,#description")
        resume    = resume_el.get_text(strip=True)[:400] if resume_el else ""
        texte_page = soup2.get_text()

        public   = detecter_public("", titre, collection, editeur, "")
        type_doc = detecter_type(titre, collection, resume, "", "", editeur, public)
        if not type_doc: type_doc = "BD"
        genre    = detecter_genre(titre, collection, resume, "", type_doc)
        tome     = extraire_tome("", titre)
        serie    = extraire_serie(titre, collection, collection, tome)
        pegi     = detecter_pegi(type_doc, public, texte_page)

        return {"titre":titre,"auteur":auteur,"illustrateur":"",
                "editeur":editeur,"annee":annee,"type":type_doc,
                "public":public,"genre":genre,"serie":serie,"tome":tome,
                "pegi":pegi,"collection":collection,"resume":resume,
                "source":"BDfugue","statut":"trouvé"}
    except Exception:
        return None

# ─────────────────────────────────────────────────────────
# SOURCE 4 : Manga News
# ─────────────────────────────────────────────────────────

def manganews_lookup(isbn):
    try:
        url = f"https://www.manga-news.com/index.php/recherche?search={isbn}"
        r = requests.get(url, headers=HEADERS, timeout=12)
        if r.status_code != 200: return None
        soup = BeautifulSoup(r.text, "html.parser")
        lien = soup.select_one("a.titre, .manga-title a, .result-item a, h2 a, h3 a")
        if not lien: return None
        href = lien.get("href","")
        if not href.startswith("http"): href = "https://www.manga-news.com" + href
        r2 = requests.get(href, headers=HEADERS, timeout=12)
        if r2.status_code != 200: return None
        soup2 = BeautifulSoup(r2.text, "html.parser")

        titre_el = soup2.select_one("h1.title, h1[itemprop='name'], h1")
        titre = titre_el.get_text(strip=True) if titre_el else ""
        if not titre: return None

        def gi(label):
            for el in soup2.select(".info-manga li,.fiche-info li,dl dt,.meta-item"):
                if label.lower() in el.get_text().lower():
                    nxt = el.find_next_sibling()
                    if nxt: return nxt.get_text(strip=True)
            return ""

        auteur  = gi("auteur") or gi("scénari") or gi("dessin")
        editeur = gi("éditeur") or gi("editeur")
        annee_r = gi("date") or gi("parution")
        m = re.search(r"\d{4}", annee_r)
        annee   = m.group(0) if m else ""
        tome_r  = gi("tome") or gi("volume") or gi("numéro")
        serie_r = gi("série") or gi("serie")
        resume_el = soup2.select_one(".synopsis,.resume,[itemprop='description']")
        resume  = resume_el.get_text(strip=True)[:400] if resume_el else ""
        texte_page = soup2.get_text()

        tome  = extraire_tome(tome_r, titre)
        serie = extraire_serie(titre, serie_r, "", tome)
        public = detecter_public("", titre, "", editeur, "")
        if not public: public = "Ado (12+)"
        pegi = detecter_pegi("Manga", public, texte_page)

        return {"titre":titre,"auteur":auteur,"illustrateur":"",
                "editeur":editeur,"annee":annee,"type":"Manga",
                "public":public,"genre":"","serie":serie,"tome":tome,
                "pegi":pegi,"collection":"","resume":resume,
                "source":"Manga News","statut":"trouvé"}
    except Exception:
        return None

# ─────────────────────────────────────────────────────────
# SOURCE 5 : Fnac
# ─────────────────────────────────────────────────────────

def fnac_lookup(isbn):
    try:
        url = f"https://recherche.fnac.com/SearchResult/ResultList.aspx?SCat=2&Search={isbn}"
        r = requests.get(url, headers=HEADERS, timeout=12)
        if r.status_code != 200: return None
        soup = BeautifulSoup(r.text, "html.parser")
        lien = soup.select_one("a.Article-title,.Article-itemTitle a,h3 a,.product-title a")
        if not lien: return None
        href = lien.get("href","")
        if not href.startswith("http"): href = "https://www.fnac.com" + href
        r2 = requests.get(href, headers=HEADERS, timeout=12)
        if r2.status_code != 200: return None
        soup2 = BeautifulSoup(r2.text, "html.parser")

        titre_el = soup2.select_one("h1[itemprop='name'],h1.ProductHeader-title")
        titre = titre_el.get_text(strip=True) if titre_el else ""
        if not titre: return None

        auteur_el  = soup2.select_one("[itemprop='author'],.ProductHeader-author")
        editeur_el = soup2.select_one("[itemprop='publisher'],.ProductDetails-publisher")
        annee_el   = soup2.select_one("[itemprop='datePublished'],.ProductDetails-date")
        resume_el  = soup2.select_one("[itemprop='description'],.ProductDetails-description")
        coll_el    = soup2.select_one(".ProductDetails-collection,[itemprop='isPartOf']")
        age_el     = soup2.select_one(".ProductDetails-age,[itemprop='typicalAgeRange']")

        auteur     = auteur_el.get_text(strip=True)  if auteur_el  else ""
        editeur    = editeur_el.get_text(strip=True) if editeur_el else ""
        annee_r    = annee_el.get_text(strip=True)   if annee_el   else ""
        resume     = resume_el.get_text(strip=True)[:400] if resume_el else ""
        collection = coll_el.get_text(strip=True)    if coll_el    else ""
        tranche    = age_el.get_text(strip=True)     if age_el     else ""
        texte_page = soup2.get_text()

        m = re.search(r"\d{4}", annee_r)
        annee = m.group(0) if m else ""
        public   = detecter_public("", titre, collection, editeur, tranche)
        type_doc = detecter_type(titre, collection, resume, "", "", editeur, public)
        genre    = detecter_genre(titre, collection, resume, "", type_doc)
        tome     = extraire_tome("", titre)
        serie    = extraire_serie(titre, "", collection, tome)
        pegi     = detecter_pegi(type_doc, public, texte_page)

        return {"titre":titre,"auteur":auteur,"illustrateur":"",
                "editeur":editeur,"annee":annee,"type":type_doc,
                "public":public,"genre":genre,"serie":serie,"tome":tome,
                "pegi":pegi,"collection":collection,"resume":resume,
                "source":"Fnac","statut":"trouvé"}
    except Exception:
        return None

# ─────────────────────────────────────────────────────────
# SOURCE 6 : Amazon.fr
# ─────────────────────────────────────────────────────────

def amazon_lookup(isbn):
    try:
        url = f"https://www.amazon.fr/s?k={isbn}&i=stripbooks"
        r = requests.get(url, headers=HEADERS, timeout=12)
        if r.status_code != 200: return None
        soup = BeautifulSoup(r.text, "html.parser")

        # Trouver le premier résultat
        lien = soup.select_one("h2 a.a-link-normal, .s-result-item h2 a")
        if not lien: return None
        href = lien.get("href","")
        if not href.startswith("http"): href = "https://www.amazon.fr" + href
        r2 = requests.get(href, headers=HEADERS, timeout=12)
        if r2.status_code != 200: return None
        soup2 = BeautifulSoup(r2.text, "html.parser")

        titre_el = soup2.select_one("#productTitle, h1#title")
        titre = titre_el.get_text(strip=True) if titre_el else ""
        if not titre: return None

        # Auteur
        auteur_el = soup2.select_one(".author a, #bylineInfo a")
        auteur = auteur_el.get_text(strip=True) if auteur_el else ""

        # Détails produit
        editeur = annee = collection = ""
        texte_page = soup2.get_text()

        for row in soup2.select(".a-expander-content tr, #detailBullets_feature_div li, .detail-bullet-list li"):
            t = row.get_text(" ", strip=True)
            if re.search(r"éditeur|editeur|publisher", t, re.I):
                parts = re.split(r"[:：]", t, 1)
                if len(parts) > 1:
                    editeur = parts[1].strip().split(";")[0].strip()
            elif re.search(r"date|parution|publication", t, re.I):
                m = re.search(r"(\d{1,2}\s+\w+\s+\d{4}|\d{4})", t)
                if m:
                    annee_m = re.search(r"\d{4}", m.group(0))
                    if annee_m: annee = annee_m.group(0)
            elif re.search(r"collection|série|series", t, re.I):
                parts = re.split(r"[:：]", t, 1)
                if len(parts) > 1:
                    collection = parts[1].strip()

        # Résumé
        resume_el = soup2.select_one("#bookDescription_feature_div, #productDescription")
        resume = resume_el.get_text(strip=True)[:400] if resume_el else ""

        # PEGI Amazon (souvent dans les détails)
        pegi_brut = ""
        m_pegi = re.search(r"PEGI\s*(\d+)", texte_page, re.I)
        if m_pegi: pegi_brut = f"PEGI {m_pegi.group(1)}"

        # Tranche d'âge Amazon
        tranche = ""
        m_age = re.search(r"(\d+)\s*[àa-]\s*\d+\s*ans?|à partir de\s*(\d+)\s*ans?|dès\s*(\d+)\s*ans?", texte_page, re.I)
        if m_age: tranche = m_age.group(0)

        public   = detecter_public("", titre, collection, editeur, tranche)
        type_doc = detecter_type(titre, collection, resume, "", "", editeur, public)
        genre    = detecter_genre(titre, collection, resume, "", type_doc)
        tome     = extraire_tome("", titre)
        serie    = extraire_serie(titre, "", collection, tome)
        pegi     = pegi_brut or detecter_pegi(type_doc, public, texte_page)

        return {"titre":titre,"auteur":auteur,"illustrateur":"",
                "editeur":editeur,"annee":annee,"type":type_doc,
                "public":public,"genre":genre,"serie":serie,"tome":tome,
                "pegi":pegi,"collection":collection,"resume":resume,
                "source":"Amazon","statut":"trouvé"}
    except Exception:
        return None

# ─────────────────────────────────────────────────────────
# SOURCE 7 : Cultura
# ─────────────────────────────────────────────────────────

def cultura_lookup(isbn):
    try:
        url = f"https://www.cultura.com/catalogsearch/result/?q={isbn}"
        r = requests.get(url, headers=HEADERS, timeout=12)
        if r.status_code != 200: return None
        soup = BeautifulSoup(r.text, "html.parser")
        lien = soup.select_one("a.product-item-link,.product-name a,h2 a")
        if not lien: return None
        href = lien.get("href","")
        if not href.startswith("http"): href = "https://www.cultura.com" + href
        r2 = requests.get(href, headers=HEADERS, timeout=12)
        if r2.status_code != 200: return None
        soup2 = BeautifulSoup(r2.text, "html.parser")

        titre_el = soup2.select_one("h1[itemprop='name'],h1.product-name,h1.page-title")
        titre = titre_el.get_text(strip=True) if titre_el else ""
        if not titre: return None

        auteur_el  = soup2.select_one("[itemprop='author'],.product-author,.author")
        editeur_el = soup2.select_one("[itemprop='publisher'],.publisher,.editeur")
        annee_el   = soup2.select_one("[itemprop='datePublished'],.date-publication")
        resume_el  = soup2.select_one("[itemprop='description'],.product-description,.description")
        coll_el    = soup2.select_one(".collection,[itemprop='isPartOf']")
        age_el     = soup2.select_one(".age,.public-cible,.age-conseille")

        auteur     = auteur_el.get_text(strip=True)  if auteur_el  else ""
        editeur    = editeur_el.get_text(strip=True) if editeur_el else ""
        annee_r    = annee_el.get_text(strip=True)   if annee_el   else ""
        resume     = resume_el.get_text(strip=True)[:400] if resume_el else ""
        collection = coll_el.get_text(strip=True)    if coll_el    else ""
        tranche    = age_el.get_text(strip=True)     if age_el     else ""

        m = re.search(r"\d{4}", annee_r)
        annee = m.group(0) if m else ""
        public   = detecter_public("", titre, collection, editeur, tranche)
        type_doc = detecter_type(titre, collection, resume, "", "", editeur, public)
        genre    = detecter_genre(titre, collection, resume, "", type_doc)
        tome     = extraire_tome("", titre)
        serie    = extraire_serie(titre, "", collection, tome)
        pegi     = detecter_pegi(type_doc, public, soup2.get_text())

        return {"titre":titre,"auteur":auteur,"illustrateur":"",
                "editeur":editeur,"annee":annee,"type":type_doc,
                "public":public,"genre":genre,"serie":serie,"tome":tome,
                "pegi":pegi,"collection":collection,"resume":resume,
                "source":"Cultura","statut":"trouvé"}
    except Exception:
        return None

# ─────────────────────────────────────────────────────────
# SOURCE 8 : Decitre
# ─────────────────────────────────────────────────────────

def decitre_lookup(isbn):
    try:
        url = f"https://www.decitre.fr/livres/{isbn}.html"
        r = requests.get(url, headers=HEADERS, timeout=12)
        if r.status_code != 200: return None
        soup = BeautifulSoup(r.text, "html.parser")
        titre_el = soup.select_one("h1[itemprop='name'],h1.product-name")
        titre = titre_el.get_text(strip=True) if titre_el else ""
        if not titre: return None

        auteur_el  = soup.select_one("[itemprop='author'],.product-author")
        editeur_el = soup.select_one("[itemprop='publisher'],.product-publisher")
        annee_el   = soup.select_one("[itemprop='datePublished'],.product-date")
        coll_el    = soup.select_one(".product-collection,[itemprop='isPartOf']")
        resume_el  = soup.select_one("[itemprop='description'],.product-description")
        age_el     = soup.select_one(".product-age,.age-lecteur")

        auteur     = auteur_el.get_text(strip=True)  if auteur_el  else ""
        editeur    = editeur_el.get_text(strip=True) if editeur_el else ""
        annee_r    = annee_el.get_text(strip=True)   if annee_el   else ""
        collection = coll_el.get_text(strip=True)    if coll_el    else ""
        resume     = resume_el.get_text(strip=True)[:400] if resume_el else ""
        tranche    = age_el.get_text(strip=True)     if age_el     else ""

        m = re.search(r"\d{4}", annee_r)
        annee = m.group(0) if m else ""
        public   = detecter_public("", titre, collection, editeur, tranche)
        type_doc = detecter_type(titre, collection, resume, "", "", editeur, public)
        genre    = detecter_genre(titre, collection, resume, "", type_doc)
        tome     = extraire_tome("", titre)
        serie    = extraire_serie(titre, "", collection, tome)
        pegi     = detecter_pegi(type_doc, public, soup.get_text())

        return {"titre":titre,"auteur":auteur,"illustrateur":"",
                "editeur":editeur,"annee":annee,"type":type_doc,
                "public":public,"genre":genre,"serie":serie,"tome":tome,
                "pegi":pegi,"collection":collection,"resume":resume,
                "source":"Decitre","statut":"trouvé"}
    except Exception:
        return None

# ─────────────────────────────────────────────────────────
# SOURCE 9 : Ricochet-jeunes
# ─────────────────────────────────────────────────────────

def ricochet_lookup(isbn):
    try:
        url = f"https://www.ricochet-jeunes.org/livres/recherche?isbn={isbn}"
        r = requests.get(url, headers=HEADERS, timeout=12)
        if r.status_code != 200: return None
        soup = BeautifulSoup(r.text, "html.parser")
        lien = soup.select_one("a.livre-title,.livre-item a,h2 a,h3 a")
        if not lien: return None
        href = lien.get("href","")
        if not href.startswith("http"): href = "https://www.ricochet-jeunes.org" + href
        r2 = requests.get(href, headers=HEADERS, timeout=12)
        if r2.status_code != 200: return None
        soup2 = BeautifulSoup(r2.text, "html.parser")

        titre_el = soup2.select_one("h1.titre,h1[itemprop='name'],h1")
        titre = titre_el.get_text(strip=True) if titre_el else ""
        if not titre: return None

        def gi(label):
            for el in soup2.select("dl dt,.meta,.fiche-info"):
                if label.lower() in el.get_text().lower():
                    nxt = el.find_next_sibling()
                    if nxt: return nxt.get_text(strip=True)
            return ""

        auteur     = gi("auteur") or gi("texte")
        illustr    = gi("illustrateur") or gi("illustration")
        editeur    = gi("éditeur") or gi("editeur")
        annee_r    = gi("date") or gi("parution") or gi("année")
        collection = gi("collection")
        public_r   = gi("âge") or gi("public") or gi("à partir")
        resume_el  = soup2.select_one(".resume,.description,[itemprop='description']")
        resume     = resume_el.get_text(strip=True)[:400] if resume_el else ""

        m = re.search(r"\d{4}", annee_r)
        annee = m.group(0) if m else ""
        public   = detecter_public("", titre, collection, editeur, public_r)
        type_doc = detecter_type(titre, collection, resume, "", "", editeur, public)
        genre    = detecter_genre(titre, collection, resume, "", type_doc)
        tome     = extraire_tome("", titre)
        serie    = extraire_serie(titre, "", collection, tome)
        pegi     = detecter_pegi(type_doc, public, "")

        return {"titre":titre,"auteur":auteur,"illustrateur":illustr,
                "editeur":editeur,"annee":annee,"type":type_doc,
                "public":public,"genre":genre,"serie":serie,"tome":tome,
                "pegi":pegi,"collection":collection,"resume":resume,
                "source":"Ricochet","statut":"trouvé"}
    except Exception:
        return None

# ─────────────────────────────────────────────────────────
# SOURCE 10 : Leslibraires.fr
# ─────────────────────────────────────────────────────────

def leslibraires_lookup(isbn):
    try:
        url = f"https://www.leslibraires.fr/livre/{isbn}"
        r = requests.get(url, headers=HEADERS, timeout=12)
        if r.status_code != 200: return None
        soup = BeautifulSoup(r.text, "html.parser")

        titre_el = soup.select_one("h1[itemprop='name'],h1.product-title")
        titre = titre_el.get_text(strip=True) if titre_el else ""
        if not titre: return None

        auteur_el  = soup.select_one("[itemprop='author'],.product-author")
        editeur_el = soup.select_one("[itemprop='publisher'],.product-publisher")
        annee_el   = soup.select_one("[itemprop='datePublished'],.product-date")
        resume_el  = soup.select_one("[itemprop='description'],.product-description")
        coll_el    = soup.select_one(".product-collection")
        age_el     = soup.select_one(".product-age,.public")

        auteur     = auteur_el.get_text(strip=True)  if auteur_el  else ""
        editeur    = editeur_el.get_text(strip=True) if editeur_el else ""
        annee_r    = annee_el.get_text(strip=True)   if annee_el   else ""
        resume     = resume_el.get_text(strip=True)[:400] if resume_el else ""
        collection = coll_el.get_text(strip=True)    if coll_el    else ""
        tranche    = age_el.get_text(strip=True)     if age_el     else ""

        m = re.search(r"\d{4}", annee_r)
        annee = m.group(0) if m else ""
        public   = detecter_public("", titre, collection, editeur, tranche)
        type_doc = detecter_type(titre, collection, resume, "", "", editeur, public)
        genre    = detecter_genre(titre, collection, resume, "", type_doc)
        tome     = extraire_tome("", titre)
        serie    = extraire_serie(titre, "", collection, tome)
        pegi     = detecter_pegi(type_doc, public, "")

        return {"titre":titre,"auteur":auteur,"illustrateur":"",
                "editeur":editeur,"annee":annee,"type":type_doc,
                "public":public,"genre":genre,"serie":serie,"tome":tome,
                "pegi":pegi,"collection":collection,"resume":resume,
                "source":"LesLibraires","statut":"trouvé"}
    except Exception:
        return None

# ─────────────────────────────────────────────────────────
# SOURCE 11 : Mollat
# ─────────────────────────────────────────────────────────

def mollat_lookup(isbn):
    try:
        url = f"https://www.mollat.com/livres/recherche?recherche={isbn}"
        r = requests.get(url, headers=HEADERS, timeout=12)
        if r.status_code != 200: return None
        soup = BeautifulSoup(r.text, "html.parser")
        lien = soup.select_one("a.book-title,.book-item a,h2.title a,h3 a")
        if not lien: return None
        href = lien.get("href","")
        if not href.startswith("http"): href = "https://www.mollat.com" + href
        r2 = requests.get(href, headers=HEADERS, timeout=12)
        if r2.status_code != 200: return None
        soup2 = BeautifulSoup(r2.text, "html.parser")

        titre_el = soup2.select_one("h1[itemprop='name'],h1.book-title,h1")
        titre = titre_el.get_text(strip=True) if titre_el else ""
        if not titre: return None

        auteur_el  = soup2.select_one("[itemprop='author'],.book-author")
        editeur_el = soup2.select_one("[itemprop='publisher'],.book-publisher")
        annee_el   = soup2.select_one("[itemprop='datePublished'],.book-date")
        resume_el  = soup2.select_one("[itemprop='description'],.book-description,.resume")
        coll_el    = soup2.select_one(".book-collection,.collection")
        age_el     = soup2.select_one(".book-age,.public,.age")

        auteur     = auteur_el.get_text(strip=True)  if auteur_el  else ""
        editeur    = editeur_el.get_text(strip=True) if editeur_el else ""
        annee_r    = annee_el.get_text(strip=True)   if annee_el   else ""
        resume     = resume_el.get_text(strip=True)[:400] if resume_el else ""
        collection = coll_el.get_text(strip=True)    if coll_el    else ""
        tranche    = age_el.get_text(strip=True)     if age_el     else ""

        m = re.search(r"\d{4}", annee_r)
        annee = m.group(0) if m else ""
        public   = detecter_public("", titre, collection, editeur, tranche)
        type_doc = detecter_type(titre, collection, resume, "", "", editeur, public)
        genre    = detecter_genre(titre, collection, resume, "", type_doc)
        tome     = extraire_tome("", titre)
        serie    = extraire_serie(titre, "", collection, tome)
        pegi     = detecter_pegi(type_doc, public, "")

        return {"titre":titre,"auteur":auteur,"illustrateur":"",
                "editeur":editeur,"annee":annee,"type":type_doc,
                "public":public,"genre":genre,"serie":serie,"tome":tome,
                "pegi":pegi,"collection":collection,"resume":resume,
                "source":"Mollat","statut":"trouvé"}
    except Exception:
        return None

# ─────────────────────────────────────────────────────────
# SOURCE 12 : Booknode
# ─────────────────────────────────────────────────────────

def booknode_lookup(isbn):
    try:
        url = f"https://booknode.com/recherche?q={isbn}"
        r = requests.get(url, headers=HEADERS, timeout=12)
        if r.status_code != 200: return None
        soup = BeautifulSoup(r.text, "html.parser")
        lien = soup.select_one("a.book-title,.book-item a,.book_title a,h2 a,h3 a")
        if not lien: return None
        href = lien.get("href","")
        if not href.startswith("http"): href = "https://booknode.com" + href
        r2 = requests.get(href, headers=HEADERS, timeout=12)
        if r2.status_code != 200: return None
        soup2 = BeautifulSoup(r2.text, "html.parser")

        titre_el = soup2.select_one("h1.book-title,h1[itemprop='name'],h1")
        titre = titre_el.get_text(strip=True) if titre_el else ""
        if not titre: return None

        auteur_el  = soup2.select_one("[itemprop='author'],.author-name,.book-author")
        editeur_el = soup2.select_one("[itemprop='publisher'],.publisher")
        resume_el  = soup2.select_one(".book-synopsis,[itemprop='description'],.synopsis")
        serie_el   = soup2.select_one(".book-serie,.serie-name,.collection-name")

        auteur  = auteur_el.get_text(strip=True)  if auteur_el  else ""
        editeur = editeur_el.get_text(strip=True) if editeur_el else ""
        resume  = resume_el.get_text(strip=True)[:400] if resume_el else ""
        serie_r = serie_el.get_text(strip=True)   if serie_el   else ""

        # Tome souvent dans le titre sur Booknode
        tome  = extraire_tome("", titre)
        serie = extraire_serie(titre, serie_r, "", tome)
        public   = detecter_public("", titre, "", editeur, "")
        type_doc = detecter_type(titre, "", resume, "", "", editeur, public)
        genre    = detecter_genre(titre, "", resume, "", type_doc)
        pegi     = detecter_pegi(type_doc, public, "")

        return {"titre":titre,"auteur":auteur,"illustrateur":"",
                "editeur":editeur,"annee":"","type":type_doc,
                "public":public,"genre":genre,"serie":serie,"tome":tome,
                "pegi":pegi,"collection":"","resume":resume,
                "source":"Booknode","statut":"trouvé"}
    except Exception:
        return None

# ─────────────────────────────────────────────────────────
# SOURCE 13 : Google Books
# ─────────────────────────────────────────────────────────

def google_books_lookup(isbn):
    try:
        url = f"https://www.googleapis.com/books/v1/volumes?q=isbn:{isbn}"
        r = requests.get(url, headers=HEADERS, timeout=12)
        if r.status_code != 200: return None
        data = r.json()
        if data.get("totalItems",0) == 0: return None
        info    = data["items"][0]["volumeInfo"]
        titre   = info.get("title","")
        sub     = info.get("subtitle","")
        if sub: titre += f" : {sub}"
        auteur  = " / ".join(info.get("authors",[])[:2])
        editeur = info.get("publisher","")
        annee   = info.get("publishedDate","")[:4]
        resume  = info.get("description","")[:400]
        cats    = " ".join(info.get("categories",[]))
        public   = detecter_public("", titre, "", editeur, "")
        type_doc = detecter_type(titre, "", resume, cats, "", editeur, public)
        genre    = detecter_genre(titre, "", resume, cats, type_doc)
        tome     = extraire_tome("", titre)
        serie    = extraire_serie(titre, "", "", tome)
        pegi     = detecter_pegi(type_doc, public, "")
        return {"titre":titre,"auteur":auteur,"illustrateur":"",
                "editeur":editeur,"annee":annee,"type":type_doc,
                "public":public,"genre":genre,"serie":serie,"tome":tome,
                "pegi":pegi,"collection":"","resume":resume,
                "source":"Google Books","statut":"trouvé"}
    except Exception:
        return None

# ─────────────────────────────────────────────────────────
# SOURCE 14 : Open Library
# ─────────────────────────────────────────────────────────

def openlibrary_lookup(isbn):
    try:
        url = f"https://openlibrary.org/api/books?bibkeys=ISBN:{isbn}&format=json&jscmd=data"
        r   = requests.get(url, headers=HEADERS, timeout=12)
        if r.status_code != 200: return None
        data = r.json()
        key  = f"ISBN:{isbn}"
        if key not in data: return None
        book    = data[key]
        titre   = book.get("title","")
        auteur  = " / ".join(a.get("name","") for a in book.get("authors",[])[:2])
        pubs    = book.get("publishers",[])
        editeur = pubs[0].get("name","") if pubs else ""
        annee   = str(book.get("publish_date",""))[-4:]
        sujets  = " ".join(
            s.get("name","") if isinstance(s,dict) else str(s)
            for s in book.get("subjects",[])[:6]
        )
        if not titre: return None
        public   = detecter_public("", titre, "", editeur, "")
        type_doc = detecter_type(titre, "", "", sujets, "", editeur, public)
        genre    = detecter_genre(titre, "", "", sujets, type_doc)
        tome     = extraire_tome("", titre)
        serie    = extraire_serie(titre, "", "", tome)
        pegi     = detecter_pegi(type_doc, public, "")
        return {"titre":titre,"auteur":auteur,"illustrateur":"",
                "editeur":editeur,"annee":annee,"type":type_doc,
                "public":public,"genre":genre,"serie":serie,"tome":tome,
                "pegi":pegi,"collection":"","resume":"",
                "source":"Open Library","statut":"trouvé"}
    except Exception:
        return None

# ─────────────────────────────────────────────────────────
# MOTEUR PRINCIPAL
# ─────────────────────────────────────────────────────────

SOURCES = [
    # En premier : sites avec titres complets incluant le numéro de tome
    ("Amazon",        amazon_lookup),
    ("Fnac",          fnac_lookup),
    ("BDfugue",       bdfugue_lookup),
    ("Manga News",    manganews_lookup),
    ("Booknode",      booknode_lookup),
    ("Cultura",       cultura_lookup),
    ("Decitre",       decitre_lookup),
    ("LesLibraires",  leslibraires_lookup),
    ("Mollat",        mollat_lookup),
    ("Google Books",  google_books_lookup),
    ("Open Library",  openlibrary_lookup),
    # En dernier : BnF pour compléter auteur/collection/résumé officiels
    ("BnF",           bnf_lookup),
]
# Babelio retirée : son robots.txt interdit explicitement l'accès automatisé.

# Priorité par catégorie de fichier (isbn_mangas.txt, isbn_bd.txt, ...) :
# les sources listées ici sont interrogées EN PREMIER pour cette catégorie,
# le reste de SOURCES suit dans son ordre habituel.
PRIORITE_PAR_CATEGORIE = {
    "manga":           ["Manga News"],
    "bd":              ["BDfugue"],
    # Pas de source "spécialiste roman/documentaire" dans la liste (contrairement
    # à BDfugue/Manga News) — choix de départ basé sur la nature généraliste de
    # ces sites, à ajuster avec vous selon ce que vous observerez en pratique.
    "roman_jeunesse":  ["Booknode", "Decitre"],
    "roman_ado":       ["Booknode", "Decitre"],
    "documentaire":    ["BnF", "Google Books"],
}

def sources_pour_categorie(categorie):
    noms_prio = PRIORITE_PAR_CATEGORIE.get(categorie, [])
    if not noms_prio:
        return SOURCES
    prio  = [s for s in SOURCES if s[0] in noms_prio]
    reste = [s for s in SOURCES if s[0] not in noms_prio]
    return prio + reste

# Champs sur lesquels repose l'arrêt anticipé de la recherche
CHAMPS_ARRET = ("titre", "tome", "serie", "auteur", "illustrateur", "editeur", "annee")

def chercher_isbn(isbn, categorie=None):
    """
    Interroge les sources (dans l'ordre adapté à la catégorie du fichier),
    collecte les résultats, choisit le meilleur titre par vote majoritaire,
    puis enrichit champ par champ.
    Les corrections manuelles (table CORRECTIONS) ont priorité absolue.

    Arrêt anticipé : dès que titre/tome/série/auteur/illustrateur/éditeur/
    année sont tous renseignés ET qu'au moins 2 sources distinctes sont
    d'accord sur le titre, on arrête d'interroger les sources restantes.
    """
    # ── Corrections manuelles prioritaires
    correction = CORRECTIONS.get(isbn)

    def normaliser_titre(t):
        return re.sub(r'[^\w\s]', '', t.lower()).strip()

    tous_resultats = []

    for nom, fn in sources_pour_categorie(categorie):
        try:
            res = fn(isbn)
            if res and res.get("titre"):
                res["_source"] = nom
                tous_resultats.append(res)
        except Exception:
            pass
        time.sleep(0.1)

        # ── Vérifier si on peut arrêter la recherche plus tôt
        if len(tous_resultats) >= 2:
            votes = Counter(normaliser_titre(r["titre"]) for r in tous_resultats)
            _, nb_accord = votes.most_common(1)[0]
            if nb_accord >= 2:
                fusion = {}
                for r in tous_resultats:
                    for champ in CHAMPS_ARRET:
                        if not fusion.get(champ) and r.get(champ):
                            fusion[champ] = r[champ]
                # L'illustrateur ne bloque pas indéfiniment : beaucoup de BD n'ont
                # qu'un seul auteur (texte + dessin) et n'auront jamais ce champ.
                # Après 4 sources sans illustrateur trouvé, on considère le champ clos.
                champs_obligatoires = [c for c in CHAMPS_ARRET if c != "illustrateur"]
                illustrateur_ok = bool(fusion.get("illustrateur")) or len(tous_resultats) >= 4
                if all(fusion.get(champ) for champ in champs_obligatoires) and illustrateur_ok:
                    break

    if not tous_resultats:
        return {**EMPTY}

    # ── Vote majoritaire sur le titre
    def normaliser(t):
        return re.sub(r'[^\w\s]', '', t.lower()).strip()

    def score_titre(titre):
        """
        Score de qualité d'un titre :
        - Pénalise les titres commençant par un chiffre seul (ex: "7 - Archibald...")
        - Pénalise les titres trop longs (doublons avec numéro inclus)
        - Favorise les titres propres et complets
        """
        score = 0
        # Pénalité si commence par chiffre + tiret (titre mal formaté)
        if re.match(r'^\d{1,3}\s*[-–]', titre):
            score -= 10
        # Pénalité proportionnelle à la longueur (plus court = plus propre)
        score -= len(titre) * 0.01
        return score

    # Regrouper les titres similaires
    votes = {}
    for r in tous_resultats:
        tn = normaliser(r["titre"])
        votes[tn] = votes.get(tn, 0) + 1

    # Choisir le titre avec le meilleur score combiné (fréquence + qualité)
    def score_combine(tn):
        freq = votes[tn]
        # Retrouver le meilleur titre original pour ce groupe
        titres_groupe = [r["titre"] for r in tous_resultats if normaliser(r["titre"]) == tn]
        meilleur = max(titres_groupe, key=score_titre)
        return (freq, score_titre(meilleur))

    titre_gagnant_norm = max(votes, key=score_combine)
    titres_candidats = [
        r["titre"] for r in tous_resultats
        if normaliser(r["titre"]) == titre_gagnant_norm
    ]
    # Parmi les candidats, prendre le mieux scoré
    meilleur_titre = max(titres_candidats, key=score_titre)

    # ── Partir du résultat le plus complet
    tous_resultats.sort(
        key=lambda r: sum(1 for c in CHAMPS_CLES if r.get(c)),
        reverse=True
    )
    resultat = tous_resultats[0].copy()
    resultat["titre"] = meilleur_titre
    sources_ok = [resultat["_source"]]

    # ── Enrichir avec les autres sources
    for res in tous_resultats[1:]:
        enrichi = False
        for champ in ("auteur","illustrateur","editeur","annee",
                      "type","public","genre","serie","tome",
                      "pegi","collection","resume"):
            if not resultat.get(champ) and res.get(champ):
                resultat[champ] = res[champ]
                enrichi = True
        if enrichi and res["_source"] not in sources_ok:
            sources_ok.append(res["_source"])

    # ── Nettoyage titre et série
    tome  = resultat.get("tome", "")
    serie = resultat.get("serie", "")
    if resultat.get("titre"):
        resultat["titre"] = nettoyer_titre(resultat["titre"], tome, serie)
    if resultat.get("serie"):
        s = re.sub(r'\s*\d{1,3}\s*$', '', resultat["serie"]).strip(" -\u2013:,")
        s = re.sub(r'\s*[-\u2013,]\s*[Tt]ome\s*\d{1,3}.*$', '', s).strip()
        if len(s) > 2:
            resultat["serie"] = s

    # ── Fallbacks champs vides — appliqués systématiquement
    titre      = resultat.get("titre", "")
    type_doc   = resultat.get("type", "")
    public     = resultat.get("public", "")
    collection = resultat.get("collection", "").lower()
    editeur    = resultat.get("editeur", "")

    # ── PUBLIC : 4 niveaux de fallback + défaut ultime
    if not public:
        # Niveau 1 : depuis les infos du résultat principal
        public = detecter_public("", titre, collection, editeur, "")
    if not public:
        # Niveau 2 : chercher dans toutes les sources (résumés, tranches d'âge)
        for r in tous_resultats:
            texte = " ".join([r.get("titre",""), r.get("collection",""),
                              r.get("resume",""), r.get("editeur","")])
            m = re.search(r'(\d+)\s*[àa]\s*\d+\s*ans?|dès\s*(\d+)\s*ans?|à partir de\s*(\d+)', texte, re.I)
            tranche = m.group(0) if m else ""
            p = detecter_public("", r.get("titre",""), r.get("collection","").lower(),
                                r.get("editeur",""), tranche)
            if p:
                public = p
                break
    if not public:
        # Niveau 3 : depuis le type détecté
        if type_doc in ("Manga",):
            public = "Ado (12+)"
        elif type_doc in ("Album", "Première lecture"):
            public = "Dès 6 ans"
        elif type_doc == "Roman ado / YA":
            public = "Ado (12+)"
        elif type_doc == "Documentaire":
            public = "8-12 ans"
    if not public:
        # Niveau 4 : défaut ultime — tout document jeunesse a un public
        public = "Jeunesse"
    resultat["public"] = public

    # ── TYPE : 4 niveaux de fallback + défaut ultime
    if not type_doc:
        # Niveau 1 : depuis les infos du résultat principal
        type_doc = detecter_type(titre, collection, "", "", "", editeur, public)
    if not type_doc:
        # Niveau 2 : chercher dans toutes les sources
        for r in tous_resultats:
            t = detecter_type(r.get("titre",""), r.get("collection",""),
                              r.get("resume",""), "", "", r.get("editeur",""), public)
            if t:
                type_doc = t
                break
    if not type_doc:
        # Niveau 3 : depuis le public
        if public in ("Dès 3 ans",):
            type_doc = "Album"
        elif public in ("Dès 6 ans",):
            type_doc = "Première lecture"
        elif public in ("Ado (12+)",):
            type_doc = "Roman ado / YA"
    if not type_doc:
        # Niveau 4 : défaut ultime
        type_doc = "Roman jeunesse"
    resultat["type"] = type_doc

    # ── GENRE : 3 niveaux de fallback + défaut ultime
    if not resultat.get("genre"):
        # Niveau 1 : depuis toutes les sources
        for r in tous_resultats:
            g = detecter_genre(r.get("titre",""), r.get("collection",""),
                               r.get("resume",""), "", type_doc)
            if g:
                resultat["genre"] = g
                break
    if not resultat.get("genre"):
        # Niveau 2 : depuis le résumé agrégé de toutes les sources
        resume_total = " ".join(r.get("resume","") for r in tous_resultats)
        g = detecter_genre(titre, collection, resume_total, "", type_doc)
        if g:
            resultat["genre"] = g
    if not resultat.get("genre"):
        # Niveau 3 : défaut par type
        GENRE_DEFAUT = {
            "Manga":             "Aventure",
            "BD":                "Aventure / Humour",
            "Album":             "Vie quotidienne",
            "Première lecture":  "Aventure",
            "Roman jeunesse":    "Aventure",
            "Roman ado / YA":    "Aventure",
            "Documentaire":      "Sciences",
            "Conte / Poésie":    "Conte / Mythe",
            "Livre-jeu / Activités": "Cuisine / Activités",
        }
        resultat["genre"] = GENRE_DEFAUT.get(type_doc, "Vie quotidienne")

    # ── PEGI
    if not resultat.get("pegi"):
        resultat["pegi"] = detecter_pegi(type_doc, resultat.get("public",""), "")


    # PEGI : recalcul
    if not resultat.get("pegi"):
        resultat["pegi"] = detecter_pegi(type_doc, resultat.get("public",""), "")

    resultat["source"] = sources_ok[0] if len(sources_ok) == 1 \
                         else f"{sources_ok[0]} +{len(sources_ok)-1}"
    resultat["statut"] = "trouvé"

    # ── Année manquante → "N/C"
    if not resultat.get("annee"):
        resultat["annee"] = "N/C"

    # ── Normaliser l'éditeur
    if resultat.get("editeur"):
        resultat["editeur"] = normaliser_editeur(resultat["editeur"])

    # ── Appliquer les corrections manuelles (priorité absolue)
    if correction:
        for champ, val in correction.items():
            if val:  # ne pas écraser avec une valeur vide
                resultat[champ] = val
        if "source" in resultat:
            resultat["source"] = resultat["source"] + " [corrigé]"

    return resultat





# ═══════════════════════════════════════════════════════════════
# PARTIE 2 — Lecture/écriture Excel (ex lib_excel.py)
# ═══════════════════════════════════════════════════════════════


C_HEADER = "2E4A7A"
C_WHITE  = "FFFFFF"
C_ALT    = "DCE6F1"
C_MISS   = "FFEB9C"
C_RECU   = "C6E8C6"   # vert clair : tome reçu (coché dans un onglet Commande)
C_SORTI  = "F4C7A1"   # orange clair : document sorti du fonds (coché dans un onglet Désherbage)
C_OK     = "2E7D46"
C_WARN   = "B45309"

thin = Side(style="thin", color="B8CCE4")
brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
def hfill(h): return PatternFill("solid", fgColor=h)

COLS = [
    ("ISBN",          16, "isbn",         True),
    ("Titre",         42, "titre",        False),
    ("Tome",           6, "tome",         True),
    ("Série",         22, "serie",        False),
    ("Auteur",        22, "auteur",       False),
    ("Illustrateur",  18, "illustrateur", False),
    ("Éditeur",       18, "editeur",      False),
    ("PEGI",           8, "pegi",         True),
    ("Public",        13, "public",       True),
    ("Type",          18, "type",         True),
    ("Genre",         24, "genre",        False),
    ("Année",          7, "annee",        True),
    ("Date d'ajout",  13, "date_ajout",   True),
    ("Nouveauté ?",   10, "est_nouveaute", True),
]


def _norm_serie(s):
    return _sans_accents((s or "").strip().lower())


def categorie_du_fichier(nom_fichier):
    """Déduit une catégorie (manga, bd, ...) à partir d'un nom de fichier/catégorie,
    pour prioriser les sources les plus pertinentes lors de la recherche."""
    import re
    n = _sans_accents(nom_fichier.lower())
    if "manga" in n:
        return "manga"
    if re.search(r'(^|[_\-])bd([_\-]|\.)', n) or n == "bd":
        return "bd"
    if "documentaire" in n:
        return "documentaire"
    if "roman" in n and "ado" in n:
        return "roman_ado"
    if "roman" in n:
        return "roman_jeunesse"
    return None


def charger_ou_creer_classeur(chemin):
    """Charge le classeur existant (en migrant l'en-tête si besoin) ou en crée un
    nouveau. Renvoie (wb, ws_inventaire)."""
    if os.path.exists(chemin):
        wb = load_workbook(chemin)
        ws = wb["Inventaire"] if "Inventaire" in wb.sheetnames else wb.active

        migration_faite = False
        for col, (nom, larg, _, ctr) in enumerate(COLS, 1):
            cell_entete = ws.cell(row=2, column=col)
            if not cell_entete.value:
                cell_entete.value     = nom
                cell_entete.font      = Font(name="Arial", size=10, bold=True, color="FFFFFF")
                cell_entete.fill      = hfill(C_HEADER)
                cell_entete.border    = brd
                cell_entete.alignment = Alignment(horizontal="center", vertical="center")
                ws.column_dimensions[get_column_letter(col)].width = larg
                migration_faite = True
        if migration_faite:
            for merged_range in list(ws.merged_cells.ranges):
                if str(merged_range).startswith("A1:"):
                    ws.unmerge_cells(str(merged_range))
            ws.merge_cells(f"A1:{get_column_letter(len(COLS))}1")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventaire"
        ws.merge_cells(f"A1:{get_column_letter(len(COLS))}1")
        c = ws["A1"]
        c.value     = "Inventaire bibliographique — Médiathèque d'Arcachon — Fonds Jeunesse"
        c.font      = Font(name="Arial", size=13, bold=True, color="FFFFFF")
        c.fill      = hfill(C_HEADER)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28
        for col, (nom, larg, _, ctr) in enumerate(COLS, 1):
            cell = ws.cell(row=2, column=col, value=nom)
            cell.font      = Font(name="Arial", size=10, bold=True, color="FFFFFF")
            cell.fill      = hfill(C_HEADER)
            cell.border    = brd
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[get_column_letter(col)].width = larg
        ws.row_dimensions[2].height = 20
        ws.freeze_panes = "A3"
    return wb, ws


def sauvegarder_avec_backup(wb, chemin, nb_backups_max=30):
    """Sauvegarde le classeur, en conservant d'abord une copie horodatée de
    la version précédente dans un sous-dossier 'backups' (créé si besoin) —
    pour pouvoir revenir en arrière en cas de problème. Ne conserve que les
    nb_backups_max copies les plus récentes, pour ne pas accumuler indéfiniment.
    Une sauvegarde de secours ratée n'empêche jamais l'enregistrement réel."""
    if os.path.exists(chemin):
        try:
            dossier_backups = os.path.join(os.path.dirname(os.path.abspath(chemin)) or ".", "backups")
            os.makedirs(dossier_backups, exist_ok=True)
            horodatage = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
            nom_base = os.path.splitext(os.path.basename(chemin))[0]
            dest = os.path.join(dossier_backups, f"{nom_base}_{horodatage}.xlsx")
            shutil.copy2(chemin, dest)
            fichiers = sorted(
                (os.path.join(dossier_backups, f) for f in os.listdir(dossier_backups) if f.endswith(".xlsx")),
                key=os.path.getmtime,
            )
            while len(fichiers) > nb_backups_max:
                os.remove(fichiers.pop(0))
        except OSError:
            pass
    wb.save(chemin)


def isbns_deja_complets(ws, isbns):
    """Renvoie l'ensemble des ISBN, parmi ceux donnés, déjà présents ET complets
    (Titre + Genre renseignés) dans l'inventaire — donc à ne pas retraiter."""
    existants = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row and row[0]:
            existants[str(row[0]).strip()] = row
    complets = set()
    for isbn in isbns:
        row = existants.get(isbn)
        if row and len(row) > 10 and str(row[1] or "").strip() and str(row[10] or "").strip():
            complets.add(isbn)
    return complets


def _date_depuis_nom_onglet(nom, prefixes):
    for prefixe in prefixes:
        if nom.startswith(prefixe):
            reste = nom[len(prefixe):].split(" (")[0].strip()
            try:
                return datetime.datetime.strptime(reste, "%d-%m-%Y").date()
            except ValueError:
                pass
    return None


def onglet_le_plus_proche(wb, prefixes):
    aujourdhui = datetime.date.today()
    meilleur, meilleur_ecart = None, None
    for nom in wb.sheetnames:
        d = _date_depuis_nom_onglet(nom, prefixes)
        if d is None:
            continue
        ecart = abs((d - aujourdhui).days)
        if meilleur_ecart is None or ecart < meilleur_ecart:
            meilleur, meilleur_ecart = nom, ecart
    return meilleur


def _trouver_ou_creer_section_imprevus(feuille, titre_section, entetes, couleur):
    for r in range(1, feuille.max_row + 1):
        if feuille.cell(row=r, column=1).value == titre_section:
            rr = r + 1
            while feuille.cell(row=rr, column=1).value:
                rr += 1
            return rr
    ri = feuille.max_row + 2
    feuille.merge_cells(f"A{ri}:I{ri}")
    c = feuille.cell(row=ri, column=1, value=titre_section)
    c.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    c.fill = hfill(couleur)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    feuille.row_dimensions[ri].height = 20
    ri += 1
    for ci, h in enumerate(entetes, 1):
        cell = feuille.cell(row=ri, column=ci, value=h)
        cell.font = Font(name="Arial", size=10, bold=True)
        cell.border = brd
    return ri + 1


def modifier_document(ws, isbn, valeurs):
    """Modifie en place les champs d'une ligne existante (édition manuelle depuis
    l'application). 'valeurs' est un dict {cle_colonne: nouvelle_valeur}, les clés
    correspondant à la 3e position de COLS (ex: 'titre', 'type', 'public'...).
    Renvoie True si la ligne a été trouvée et modifiée, False sinon."""
    isbn = str(isbn).strip()
    ligne = None
    for row in ws.iter_rows(min_row=3):
        if row[0].value and str(row[0].value).strip() == isbn:
            ligne = row[0].row
            break
    if ligne is None:
        return False
    for col, (_, _, cle, ctr) in enumerate(COLS, 1):
        if cle in valeurs and cle != "isbn":
            ws.cell(row=ligne, column=col, value=valeurs[cle])
    return True


class GestionnaireInventaire:
    """Encapsule un classeur ouvert et tout l'état nécessaire pour traiter un lot
    de résultats de recherche ISBN, sans dépendre de variables globales — pour
    pouvoir être instancié librement (un par requête web, par exemple)."""

    def __init__(self, wb, ws):
        self.wb = wb
        self.ws = ws
        self.isbn_lignes = {}
        for row in ws.iter_rows(min_row=3):
            if row[0].value:
                self.isbn_lignes[str(row[0].value).strip()] = row[0].row
        self.premiere_ligne_vide = ws.max_row + 1
        self.ref_serie = self._construire_reference_serie()
        self.nouveaux = 0
        self.completes = 0
        self.doublons = 0
        self.retires = []
        self.non_presents_pilon = []

    # ── Harmonisation des séries ────────────────────────────────
    def _construire_reference_serie(self):
        noms_par_serie = defaultdict(Counter)
        for row in self.ws.iter_rows(min_row=3, values_only=True):
            if not row or not row[3]:
                continue
            skey = _norm_serie(row[3])
            if len(skey) <= 2:
                continue
            noms_par_serie[skey][str(row[3]).strip()] += 1
        return {skey: c.most_common(1)[0][0] for skey, c in noms_par_serie.items()}

    def _appliquer_reference_serie(self, res):
        skey = _norm_serie(res.get("serie", ""))
        if not skey or len(skey) <= 2:
            return
        if skey in self.ref_serie:
            res["serie"] = self.ref_serie[skey]

    def _memoriser_serie(self, res):
        skey = _norm_serie(res.get("serie", ""))
        if not skey or len(skey) <= 2:
            return
        if skey not in self.ref_serie:
            nom = str(res.get("serie", "") or "").strip()
            if nom:
                self.ref_serie[skey] = nom

    # ── Pointage des commandes/désherbages planifiés ────────────
    def cocher_commande(self, serie, tome):
        """Coche Reçu=Oui dans le 1er onglet 'Commande du ...' où ce (série, tome)
        est encore en attente. Renvoie True si trouvé."""
        skey = _norm_serie(serie)
        if not skey or tome is None:
            return False
        try:
            tome_int = int(str(tome).strip())
        except (ValueError, TypeError):
            return False
        for nom_feuille in self.wb.sheetnames:
            if not nom_feuille.startswith("Commande du "):
                continue
            feuille = self.wb[nom_feuille]
            for r in range(1, feuille.max_row + 1):
                if feuille.cell(row=r, column=1).value == "Série" and feuille.cell(row=r, column=2).value == "Tome manquant":
                    rr = r + 1
                    while feuille.cell(row=rr, column=1).value:
                        if _norm_serie(str(feuille.cell(row=rr, column=1).value)) == skey:
                            try:
                                if int(str(feuille.cell(row=rr, column=2).value).strip()) == tome_int \
                                   and not feuille.cell(row=rr, column=8).value:
                                    feuille.cell(row=rr, column=8).value = "Oui"
                                    feuille.cell(row=rr, column=9).value = datetime.date.today().strftime("%d/%m/%Y")
                                    for c in range(1, 10):
                                        feuille.cell(row=rr, column=c).fill = hfill(C_RECU)
                                    return True
                            except (ValueError, TypeError):
                                pass
                        rr += 1
        return False

    def ajouter_a_commande_proche(self, res):
        """Journalise un livre reçu hors commande planifiée, dans l'onglet
        Commande dont la date est la plus proche d'aujourd'hui."""
        nom = onglet_le_plus_proche(self.wb, ["Commande du "])
        if not nom:
            return False
        feuille = self.wb[nom]
        entetes = ["Série", "Tome", "Type", "Éditeur", "Public", "Genre", "Prix", "Reçu ?", "Date de réception"]
        ri = _trouver_ou_creer_section_imprevus(feuille, "Ajouts reçus hors commande planifiée", entetes, C_OK)
        vals = [res.get("serie", ""), res.get("tome", ""), res.get("type", ""), res.get("editeur", ""),
                res.get("public", ""), res.get("genre", ""), "", "Oui", datetime.date.today().strftime("%d/%m/%Y")]
        for ci, v in enumerate(vals, 1):
            cell = feuille.cell(row=ri, column=ci, value=v)
            cell.font = Font(name="Arial", size=10)
            cell.border = brd
            cell.fill = hfill(C_RECU)
        return True

    def cocher_desherbage(self, isbn):
        if not isbn:
            return False
        isbn = str(isbn).strip()
        for nom_feuille in self.wb.sheetnames:
            if not (nom_feuille.startswith("Désherbage du ") or nom_feuille.startswith("Proposition desherbage du ")):
                continue
            feuille = self.wb[nom_feuille]
            for r in range(1, feuille.max_row + 1):
                if feuille.cell(row=r, column=1).value == "ISBN" and feuille.cell(row=r, column=2).value == "Titre":
                    rr = r + 1
                    while feuille.cell(row=rr, column=1).value:
                        if str(feuille.cell(row=rr, column=1).value).strip() == isbn \
                           and not feuille.cell(row=rr, column=8).value:
                            feuille.cell(row=rr, column=8).value = "Oui"
                            feuille.cell(row=rr, column=9).value = datetime.date.today().strftime("%d/%m/%Y")
                            for c in range(1, 10):
                                feuille.cell(row=rr, column=c).fill = hfill(C_SORTI)
                            return True
                        rr += 1
        return False

    def ajouter_a_desherbage_proche(self, isbn, titre, typ, serie, public, genre, annee):
        nom = onglet_le_plus_proche(self.wb, ["Désherbage du ", "Proposition desherbage du "])
        if not nom:
            return False
        feuille = self.wb[nom]
        entetes = ["ISBN", "Titre", "Type", "Série", "Public", "Genre", "Année", "Sorti ?", "Date de sortie"]
        ri = _trouver_ou_creer_section_imprevus(feuille, "Sorties hors désherbage planifié", entetes, C_WARN)
        vals = [isbn, titre, typ, serie, public, genre, annee, "Oui", datetime.date.today().strftime("%d/%m/%Y")]
        for ci, v in enumerate(vals, 1):
            cell = feuille.cell(row=ri, column=ci, value=v)
            cell.font = Font(name="Arial", size=10)
            cell.border = brd
            cell.fill = hfill(C_SORTI)
        return True

    # ── Ajout / mise à jour d'un résultat ───────────────────────
    def traiter_resultat(self, res, est_nouveaute=False):
        """Ajoute ou met à jour une ligne pour ce résultat de recherche.
        Renvoie 'nouveau', 'complete', 'doublon' ou 'non_trouve'."""
        if res["statut"] != "trouvé":
            return "non_trouve"

        res["est_nouveaute"] = "Oui" if est_nouveaute else ""
        self._appliquer_reference_serie(res)
        if res.get("serie") and res.get("tome") and not self.cocher_commande(res.get("serie"), res.get("tome")) \
           and est_nouveaute:
            self.ajouter_a_commande_proche(res)

        isbn_res = str(res.get("isbn", "")).strip()
        ws = self.ws

        if isbn_res in self.isbn_lignes:
            self.doublons += 1
            ligne_existante = self.isbn_lignes[isbn_res]
            modifie = False
            for col, (_, _, cle, ctr) in enumerate(COLS, 1):
                cell = ws.cell(row=ligne_existante, column=col)
                ancienne = str(cell.value or "").strip()
                nouvelle = str(res.get(cle, "") or "").strip()
                if ancienne != nouvelle and nouvelle:
                    cell.value = res[cle]
                    modifie = True
            if modifie:
                self.completes += 1
            self._memoriser_serie(res)
            return "complete" if modifie else "doublon"
        else:
            i = self.premiere_ligne_vide
            bg = C_ALT if i % 2 == 0 else C_WHITE
            res["date_ajout"] = datetime.date.today().strftime("%d/%m/%Y")
            for col, (_, _, cle, ctr) in enumerate(COLS, 1):
                val = res.get(cle, "")
                cell = ws.cell(row=i, column=col, value=val)
                cell.font = Font(name="Arial", size=10)
                cell.border = brd
                cell.fill = hfill(bg)
                cell.alignment = Alignment(horizontal="center" if ctr else "left")
            self.isbn_lignes[isbn_res] = i
            self.premiere_ligne_vide += 1
            self.nouveaux += 1
            self._memoriser_serie(res)
            return "nouveau"

    # ── Pilon ────────────────────────────────────────────────────
    def retirer_pilon(self, isbns_pilon):
        """Retire ces ISBN de l'Inventaire, les archive dans l'onglet Pilon,
        et coche/journalise les onglets Désherbage concernés."""
        ws = self.ws
        wb = self.wb
        if not isbns_pilon:
            return
        isbns_pilon_set = set(isbns_pilon)
        lignes_a_supprimer = []
        for row in ws.iter_rows(min_row=3):
            if row[0].value and str(row[0].value).strip() in isbns_pilon_set:
                lignes_a_supprimer.append((row[0].row, [c.value for c in row]))

        if lignes_a_supprimer:
            if "Pilon" in wb.sheetnames:
                ws_pilon = wb["Pilon"]
            else:
                ws_pilon = wb.create_sheet("Pilon")
                ws_pilon.cell(row=1, column=1, value="Documents désherbés (pilon)")
                entetes_pilon = [c[0] for c in COLS] + ["Date de retrait"]
                for ci, nom in enumerate(entetes_pilon, 1):
                    cell = ws_pilon.cell(row=2, column=ci, value=nom)
                    cell.font = Font(name="Arial", size=10, bold=True, color=C_WHITE)
                    cell.fill = hfill(C_HEADER)
                for ci, (_, largeur, *_r) in enumerate(COLS, 1):
                    ws_pilon.column_dimensions[get_column_letter(ci)].width = largeur
                ws_pilon.column_dimensions[get_column_letter(len(COLS) + 1)].width = 14

            date_retrait = time.strftime("%d/%m/%Y")
            ligne_pilon = ws_pilon.max_row + 1
            for row_idx, valeurs in lignes_a_supprimer:
                for ci, v in enumerate(valeurs, 1):
                    cell = ws_pilon.cell(row=ligne_pilon, column=ci, value=v)
                    cell.font = Font(name="Arial", size=10)
                    cell.border = brd
                ws_pilon.cell(row=ligne_pilon, column=len(COLS) + 1, value=date_retrait).font = Font(name="Arial", size=10)
                ligne_pilon += 1

        for row_idx, valeurs in sorted(lignes_a_supprimer, key=lambda x: -x[0]):
            ws.delete_rows(row_idx, 1)
            self.retires.append((str(valeurs[0]).strip(), valeurs[1] or ""))
            if not self.cocher_desherbage(valeurs[0]):
                self.ajouter_a_desherbage_proche(valeurs[0], valeurs[1], valeurs[9], valeurs[3], valeurs[8], valeurs[10], valeurs[11])

        trouves_pilon = {isbn_r for isbn_r, _ in self.retires}
        self.non_presents_pilon = [i for i in isbns_pilon if i not in trouves_pilon]

        # Les index de ligne ont changé après suppression : on reconstruit.
        self.isbn_lignes = {}
        for row in ws.iter_rows(min_row=3):
            if row[0].value:
                self.isbn_lignes[str(row[0].value).strip()] = row[0].row
        self.premiere_ligne_vide = ws.max_row + 1

    def sauvegarder(self, chemin):
        sauvegarder_avec_backup(self.wb, chemin)


# ═══════════════════════════════════════════════════════════════
# PARTIE 3 — Statistiques (ex lib_stats.py)
# ═══════════════════════════════════════════════════════════════


def tranche_decennie(annee):
    if annee < 1990: return "Avant 1990"
    if annee < 2000: return "1990-1999"
    if annee < 2010: return "2000-2009"
    if annee < 2020: return "2010-2019"
    return "2020 et après"


def calculer_statistiques(wb):
    """Renvoie un dictionnaire structuré avec toutes les statistiques du fonds."""
    ws = wb["Inventaire"] if "Inventaire" in wb.sheetnames else wb.active

    compteur_types, compteur_public, compteur_genre = Counter(), Counter(), Counter()
    compteur_types_recents = Counter()
    compteur_public_recents = Counter()
    compteur_genre_recents = Counter()
    compteur_types_nouveautes = Counter()
    compteur_public_nouveautes = Counter()
    compteur_genre_nouveautes = Counter()
    annee_courante = datetime.date.today().year
    aujourdhui = datetime.date.today()
    annees_valides = []
    ajouts_recents_30j = 0
    ajouts_nouveautes_30j = 0
    total_inventaire = 0

    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or not row[0]:
            continue
        total_inventaire += 1
        if row[9]:  compteur_types[str(row[9]).strip()] += 1
        if row[8]:  compteur_public[str(row[8]).strip()] += 1
        if row[10]: compteur_genre[str(row[10]).strip()] += 1
        if row[11]:
            try:
                an = int(str(row[11]).strip())
                if 1900 <= an <= annee_courante:
                    annees_valides.append(an)
            except ValueError:
                pass
        # Une « nouveauté » au sens strict : document ajouté via la catégorie
        # Nouveautés (Scanner / Rechercher) ou le fichier isbn_nouveautés.txt
        # en terminal — pas n'importe quel ajout récent, voir colonne « Nouveauté ? ».
        est_nouveaute_cat = len(row) > 13 and str(row[13]).strip().lower() == "oui"
        if len(row) > 12 and row[12]:
            try:
                d = datetime.datetime.strptime(str(row[12]).strip(), "%d/%m/%Y").date()
                if (aujourdhui - d).days <= 30:
                    ajouts_recents_30j += 1
                    if row[9]:  compteur_types_recents[str(row[9]).strip()] += 1
                    if row[8]:  compteur_public_recents[str(row[8]).strip()] += 1
                    if row[10]: compteur_genre_recents[str(row[10]).strip()] += 1
                    if est_nouveaute_cat:
                        ajouts_nouveautes_30j += 1
                        if row[9]:  compteur_types_nouveautes[str(row[9]).strip()] += 1
                        if row[8]:  compteur_public_nouveautes[str(row[8]).strip()] += 1
                        if row[10]: compteur_genre_nouveautes[str(row[10]).strip()] += 1
            except ValueError:
                pass

    age_moyen = round(sum(annee_courante - a for a in annees_valides) / len(annees_valides), 1) \
                if annees_valides else None
    compteur_decennie = Counter(tranche_decennie(a) for a in annees_valides)
    nb_annee_inconnue = total_inventaire - len(annees_valides)
    ordre_decennies = ["Avant 1990", "1990-1999", "2000-2009", "2010-2019", "2020 et après"]

    total_pilon_cumule = 0
    compteur_types_pilon, compteur_public_pilon, compteur_genre_pilon = Counter(), Counter(), Counter()
    if "Pilon" in wb.sheetnames:
        for row in wb["Pilon"].iter_rows(min_row=3, values_only=True):
            if row and row[0]:
                total_pilon_cumule += 1
                if row[9]:  compteur_types_pilon[str(row[9]).strip()] += 1
                if row[8]:  compteur_public_pilon[str(row[8]).strip()] += 1
                if row[10]: compteur_genre_pilon[str(row[10]).strip()] += 1

    return {
        "total_inventaire": total_inventaire,
        "ajouts_recents_30j": ajouts_recents_30j,
        "ajouts_nouveautes_30j": ajouts_nouveautes_30j,
        "compteur_types": compteur_types,
        "compteur_public": compteur_public,
        "compteur_genre": compteur_genre,
        "compteur_types_recents": compteur_types_recents,
        "compteur_public_recents": compteur_public_recents,
        "compteur_genre_recents": compteur_genre_recents,
        "compteur_types_nouveautes": compteur_types_nouveautes,
        "compteur_public_nouveautes": compteur_public_nouveautes,
        "compteur_genre_nouveautes": compteur_genre_nouveautes,
        "age_moyen": age_moyen,
        "compteur_decennie": compteur_decennie,
        "ordre_decennies": ordre_decennies,
        "nb_annee_inconnue": nb_annee_inconnue,
        "total_pilon_cumule": total_pilon_cumule,
        "compteur_types_pilon": compteur_types_pilon,
        "compteur_public_pilon": compteur_public_pilon,
        "compteur_genre_pilon": compteur_genre_pilon,
    }


# ═══════════════════════════════════════════════════════════════
# PARTIE 4 — Acquisitions et désherbage (ex lib_acquisition.py)
# ═══════════════════════════════════════════════════════════════


PRIX_ESTIME = {
    "Manga": 7.20, "BD": 12.50, "Roman jeunesse": 14.00, "Roman ado / YA": 15.50,
    "Album": 13.00, "Documentaire": 15.00, "Conte / Poésie": 12.00, "Première lecture": 9.00,
}
PRIX_DEFAUT = 12.00
SEUIL_SERIE_INSTALLEE = 8
SEUIL_AGE_DOCUMENTAIRE = 8
SEUIL_AGE_AUTRES = 15


def _majoritaire(valeurs):
    c = Counter(v for v in valeurs if v)
    return c.most_common(1)[0][0] if c else ""


def detecter_gaps_series(ws, types_filtre=None):
    """Renvoie {serie: {type, editeur, public, genre, nb_possedes, ratio,
    prix_unitaire, tomes_manquants}} pour toute série ayant au moins un trou.

    Si types_filtre est fourni (liste de types, ex. ["Manga", "BD"]), seules
    les séries dont le type majoritaire figure dans cette liste sont retenues.
    types_filtre=None (ou vide) = toutes catégories, comportement inchangé."""
    series = defaultdict(list)
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or not row[0]:
            continue
        serie, tome, public, typ, genre, editeur = row[3], row[2], row[8], row[9], row[10], row[6]
        if serie and tome:
            try:
                t = int(str(tome).strip())
                series[serie].append((t, typ, editeur, public, genre))
            except ValueError:
                pass

    gaps = {}
    for serie, items in series.items():
        tomes = sorted(set(t for t, *_ in items))
        if len(tomes) < 2:
            continue
        manquants = [t for t in range(tomes[0], tomes[-1] + 1) if t not in tomes]
        if manquants:
            typ = _majoritaire([i[1] for i in items]) or "Roman jeunesse"
            if types_filtre and typ not in types_filtre:
                continue
            editeur = _majoritaire([i[2] for i in items])
            public = _majoritaire([i[3] for i in items])
            genre = _majoritaire([i[4] for i in items])
            nb_possedes = len(tomes)
            ratio = nb_possedes / (nb_possedes + len(manquants))
            gaps[serie] = {
                "type": typ, "editeur": editeur, "public": public, "genre": genre,
                "nb_possedes": nb_possedes, "ratio": ratio,
                "prix_unitaire": PRIX_ESTIME.get(typ, PRIX_DEFAUT),
                "tomes_manquants": manquants,
            }
    return gaps


def selectionner_dans_budget(gaps_par_serie, budget):
    """Sélectionne les tomes à commander dans le budget, en complétant en
    priorité les séries déjà avancées et bien fournies. Renvoie
    (selection: list[dict], total: float, series_completees: list[str])."""
    series_triees = sorted(gaps_par_serie.items(), key=lambda kv: (-kv[1]["ratio"], -kv[1]["nb_possedes"]))

    selection, total, series_completees = [], 0.0, []
    series_non_retenues = []

    def ajouter(serie, info, tomes):
        for t in tomes:
            selection.append({
                "serie": serie, "tome": t, "type": info["type"], "editeur": info["editeur"],
                "public": info["public"], "genre": info["genre"], "prix": info["prix_unitaire"],
            })

    for serie, info in series_triees:
        cout = len(info["tomes_manquants"]) * info["prix_unitaire"]
        if total + cout <= budget:
            ajouter(serie, info, info["tomes_manquants"])
            total += cout
            series_completees.append(serie)
        else:
            series_non_retenues.append((serie, info))

    series_non_retenues.sort(key=lambda kv: kv[1]["prix_unitaire"])
    for serie, info in series_non_retenues:
        tomes_pris = []
        for t in info["tomes_manquants"]:
            if total + info["prix_unitaire"] > budget:
                break
            tomes_pris.append(t)
            total += info["prix_unitaire"]
        if tomes_pris:
            ajouter(serie, info, tomes_pris)

    return selection, total, series_completees


def detecter_desherbage(ws):
    """Renvoie (documentaires_anciens, autres_candidats, a_examiner_prudence),
    trois listes de dicts triées par année croissante."""
    nb_tomes_serie = Counter()
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row and row[3]:
            nb_tomes_serie[str(row[3]).strip()] += 1

    docs_avec_annee = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row and row[0] and row[11]:
            try:
                an = int(str(row[11]).strip())
                serie = str(row[3]).strip() if row[3] else ""
                docs_avec_annee.append({
                    "annee": an, "isbn": str(row[0]).strip(), "titre": row[1] or "",
                    "type": row[9] or "", "serie": serie, "public": row[8] or "", "genre": row[10] or "",
                    "taille_serie": nb_tomes_serie.get(serie, 0) if serie else 0,
                })
            except ValueError:
                pass

    annee_courante = datetime.date.today().year

    documentaires_anciens = sorted(
        [d for d in docs_avec_annee if d["type"] == "Documentaire"
         and annee_courante - d["annee"] >= SEUIL_AGE_DOCUMENTAIRE],
        key=lambda d: d["annee"])[:25]

    autres_candidats = sorted(
        [d for d in docs_avec_annee if d["type"] != "Documentaire" and d["taille_serie"] < SEUIL_SERIE_INSTALLEE
         and annee_courante - d["annee"] >= SEUIL_AGE_AUTRES],
        key=lambda d: d["annee"])[:25]

    a_examiner_prudence = sorted(
        [d for d in docs_avec_annee if d["type"] != "Documentaire" and d["taille_serie"] >= SEUIL_SERIE_INSTALLEE
         and annee_courante - d["annee"] >= SEUIL_AGE_AUTRES],
        key=lambda d: d["annee"])[:15]

    return documentaires_anciens, autres_candidats, a_examiner_prudence


# ─────────────────────────────────────────────────────────
# Signalements manuels pour désherbage (bouton 🗑️ dans Fonds total).
# Un signalement n'est qu'une PROPOSITION : le document n'est retiré du
# fonds que lorsqu'il est ensuite réellement scanné dans Scanner/Recherche,
# catégorie Pilons — jamais directement depuis ce bouton.
# ─────────────────────────────────────────────────────────

def compter_commandes_en_attente(wb):
    """Compte, sur tous les onglets 'Commande du ...', le nombre de tomes pas
    encore marqués Reçu ?."""
    total = 0
    for nom in wb.sheetnames:
        if not nom.startswith("Commande du "):
            continue
        feuille = wb[nom]
        for r in range(1, feuille.max_row + 1):
            if feuille.cell(row=r, column=1).value == "Série" and feuille.cell(row=r, column=2).value == "Tome manquant":
                rr = r + 1
                while feuille.cell(row=rr, column=1).value:
                    if not feuille.cell(row=rr, column=8).value:
                        total += 1
                    rr += 1
    return total


def compter_desherbages_en_attente(wb):
    """Compte, sur tous les onglets 'Désherbage du ...', le nombre de documents
    pas encore marqués Sorti ?."""
    total = 0
    for nom in wb.sheetnames:
        if not (nom.startswith("Désherbage du ") or nom.startswith("Proposition desherbage du ")):
            continue
        feuille = wb[nom]
        for r in range(1, feuille.max_row + 1):
            if feuille.cell(row=r, column=1).value == "ISBN" and feuille.cell(row=r, column=2).value == "Titre":
                rr = r + 1
                while feuille.cell(row=rr, column=1).value:
                    if not feuille.cell(row=rr, column=8).value:
                        total += 1
                    rr += 1
    return total


NOM_FEUILLE_SIGNALEMENTS = "Signalements_Desherbage"


def signaler_pour_desherbage(wb, isbn, titre, typ, serie, public, genre, annee):
    """Enregistre un document comme signalé pour désherbage. Renvoie False
    s'il était déjà signalé (pas de doublon), True sinon."""
    isbn = str(isbn).strip()
    if NOM_FEUILLE_SIGNALEMENTS in wb.sheetnames:
        ws = wb[NOM_FEUILLE_SIGNALEMENTS]
    else:
        ws = wb.create_sheet(NOM_FEUILLE_SIGNALEMENTS)
        ws.sheet_state = "hidden"
        ws.cell(row=1, column=1, value="Signalements manuels pour désherbage — ne pas modifier à la main")
        entetes = ["ISBN", "Titre", "Type", "Série", "Public", "Genre", "Année", "Date du signalement"]
        for ci, h in enumerate(entetes, 1):
            c = ws.cell(row=2, column=ci, value=h)
            c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
            c.fill = hfill(C_HEADER)

    for row in ws.iter_rows(min_row=3, values_only=True):
        if row and str(row[0]).strip() == isbn:
            return False

    ligne = ws.max_row + 1
    valeurs = [isbn, titre, typ, serie, public, genre, annee, datetime.date.today().strftime("%d/%m/%Y")]
    for ci, v in enumerate(valeurs, 1):
        ws.cell(row=ligne, column=ci, value=v).font = Font(name="Arial", size=10)
    return True


def charger_signalements(wb):
    """Renvoie la liste des signalements manuels en attente, sous la même
    forme que detecter_desherbage (isbn, titre, type, serie, public, genre, annee)."""
    if NOM_FEUILLE_SIGNALEMENTS not in wb.sheetnames:
        return []
    ws = wb[NOM_FEUILLE_SIGNALEMENTS]
    signalements = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row and row[0]:
            signalements.append({
                "isbn": row[0], "titre": row[1] or "", "type": row[2] or "",
                "serie": row[3] or "", "public": row[4] or "", "genre": row[5] or "",
                "annee": row[6] or "",
            })
    return signalements


def retirer_signalement(wb, isbn):
    """Retire un signalement (l'utilisateur a changé d'avis avant validation)."""
    if NOM_FEUILLE_SIGNALEMENTS not in wb.sheetnames:
        return False
    ws = wb[NOM_FEUILLE_SIGNALEMENTS]
    isbn = str(isbn).strip()
    for row in ws.iter_rows(min_row=3):
        if row[0].value and str(row[0].value).strip() == isbn:
            ws.delete_rows(row[0].row, 1)
            return True
    return False


def vider_signalements(wb):
    """Vide tous les signalements (appelé une fois qu'ils ont été repris
    dans un onglet 'Désherbage du ...' validé, pour ne pas les compter deux fois)."""
    if NOM_FEUILLE_SIGNALEMENTS in wb.sheetnames:
        wb.remove(wb[NOM_FEUILLE_SIGNALEMENTS])


# ─────────────────────────────────────────────────────────
# Création effective des onglets Commande / Désherbage dans le classeur
# (appelée uniquement quand l'utilisateur valide la proposition affichée)
# ─────────────────────────────────────────────────────────

C_OK = "2E7D46"
C_WARN = "B45309"
_thin = Side(style="thin", color="B8CCE4")
_brd = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _nom_disponible(wb, base):
    nom = base
    n = 1
    while nom in wb.sheetnames:
        n += 1
        nom = f"{base} ({n})"
    return nom


def _bandeau(feuille, ri, titre, couleur, largeur_fusion):
    feuille.merge_cells(f"A{ri}:{get_column_letter(largeur_fusion)}{ri}")
    c = feuille.cell(row=ri, column=1, value=titre)
    c.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    c.fill = hfill(couleur)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    feuille.row_dimensions[ri].height = 20
    return ri + 1


def _section_repartition(feuille, ri, titre, couleur, compteur_local, total_local, largeur_fusion):
    ri = _bandeau(feuille, ri, titre, couleur, largeur_fusion)
    for cle, n in sorted(compteur_local.items(), key=lambda x: -x[1]):
        if not cle:
            continue
        feuille.cell(row=ri, column=1, value=cle).font = Font(name="Arial", size=10)
        feuille.cell(row=ri, column=2, value=n).font = Font(name="Arial", size=10)
        pct = f"{n/total_local*100:.1f} %" if total_local else ""
        c = feuille.cell(row=ri, column=3, value=pct)
        c.font = Font(name="Arial", size=9, italic=True, color="808080")
        ri += 1
    return ri + 1


def creer_onglets_commande_et_desherbage(wb, budget, selection, total, series_completees,
                                          documentaires_anciens, autres_candidats, a_examiner_prudence):
    """Crée les deux onglets datés du jour dans le classeur (déjà ouvert),
    à partir d'une sélection déjà calculée. Renvoie (nom_onglet_commande, nom_onglet_desherbage)."""
    date_str = datetime.date.today().strftime("%d-%m-%Y")

    # ── Onglet Commande ──
    nom_cmd = _nom_disponible(wb, f"Commande du {date_str}")
    ws_cmd = wb.create_sheet(nom_cmd)
    ri = 1
    ws_cmd.merge_cells(f"A{ri}:I{ri}")
    c = ws_cmd.cell(row=ri, column=1, value=f"Proposition d'acquisitions — {date_str}")
    c.font = Font(name="Arial", size=13, bold=True, color="FFFFFF")
    c.fill = hfill(C_HEADER)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws_cmd.row_dimensions[ri].height = 26
    ri += 2

    ri = _bandeau(ws_cmd, ri, "Budget", C_HEADER, 9)
    for label, val in [("Budget alloué", f"{budget:.2f} €"),
                        ("Total estimé de la sélection", f"{total:.2f} €"),
                        ("Reliquat estimé", f"{budget-total:.2f} €"),
                        ("Budget réel (à remplir une fois les prix confirmés)", ""),
                        ("Tomes sélectionnés", len(selection)),
                        ("Séries complétées entièrement", len(series_completees))]:
        ws_cmd.cell(row=ri, column=1, value=label).font = Font(name="Arial", size=10)
        cell_val = ws_cmd.cell(row=ri, column=2, value=val)
        cell_val.font = Font(name="Arial", size=10, bold=True)
        if "Budget réel" in label:
            cell_val.fill = hfill("FFEB9C")
        ri += 1
    ri += 1

    compteur_sel_type = Counter(s["type"] for s in selection)
    compteur_sel_public = Counter(s["public"] for s in selection)
    compteur_sel_genre = Counter(s["genre"] for s in selection)
    ri = _section_repartition(ws_cmd, ri, "Sélection — répartition par type", C_HEADER, compteur_sel_type, len(selection), 9)
    ri = _section_repartition(ws_cmd, ri, "Sélection — répartition par public", C_HEADER, compteur_sel_public, len(selection), 9)
    ri = _section_repartition(ws_cmd, ri, "Sélection — répartition par genre", C_HEADER, compteur_sel_genre, len(selection), 9)

    ri = _bandeau(ws_cmd, ri, "Tomes à commander — séries déjà au catalogue", C_OK, 9)
    entetes = ["Série", "Tome manquant", "Type", "Éditeur", "Public", "Genre", "Prix estimé", "Reçu ?", "Date de réception"]

    selection_par_type = defaultdict(list)
    for s in selection:
        selection_par_type[s["type"] or "(type inconnu)"].append(s)

    if selection_par_type:
        for typ in sorted(selection_par_type.keys()):
            items_type = selection_par_type[typ]
            sous_total = sum(s["prix"] for s in items_type)
            ri = _bandeau(ws_cmd, ri, f"{typ} — {len(items_type)} tome(s), {sous_total:.2f} €", C_OK, 9)
            for ci, h in enumerate(entetes, 1):
                cell = ws_cmd.cell(row=ri, column=ci, value=h)
                cell.font = Font(name="Arial", size=10, bold=True)
                cell.border = _brd
            ri += 1
            for s in sorted(items_type, key=lambda s: (s["serie"], s["tome"])):
                vals = [s["serie"], s["tome"], s["type"], s["editeur"], s["public"], s["genre"], s["prix"], "", ""]
                for ci, v in enumerate(vals, 1):
                    cell = ws_cmd.cell(row=ri, column=ci, value=v)
                    cell.font = Font(name="Arial", size=10)
                    cell.border = _brd
                    if ci == 7:
                        cell.number_format = "0.00 €"
                ri += 1
            ri += 1
    else:
        ws_cmd.cell(row=ri, column=1, value="(aucun tome manquant détecté dans les catégories choisies)").font = \
            Font(name="Arial", size=9.5, italic=True, color="808080")
        ri += 2

    ri = _bandeau(ws_cmd, ri, "Nouveautés à ajouter — à compléter avec Claude au moment de finaliser", C_OK, 9)
    ws_cmd.cell(row=ri, column=1, value="(section vide : se remplit en discussion, selon les tendances du moment, "
                                         "le thème sciences/lecture de l'année scolaire, et le public estival)").font = \
        Font(name="Arial", size=9.5, italic=True, color="808080")

    for col, larg in zip("ABCDEFGHI", [28, 14, 16, 20, 14, 18, 12, 9, 16]):
        ws_cmd.column_dimensions[col].width = larg
    ws_cmd.freeze_panes = "A1"

    nom_desh = _creer_feuille_desherbage(wb, documentaires_anciens, autres_candidats, a_examiner_prudence)
    return nom_cmd, nom_desh


def _creer_feuille_desherbage(wb, documentaires_anciens, autres_candidats, a_examiner_prudence,
                               signales_manuellement=None):
    """Crée l'onglet 'Désherbage du JJ-MM-AAAA' seul. Renvoie son nom."""
    signales_manuellement = signales_manuellement or []
    date_str = datetime.date.today().strftime("%d-%m-%Y")
    tous_candidats_desherbage = documentaires_anciens + autres_candidats + a_examiner_prudence + signales_manuellement

    nom_desh = _nom_disponible(wb, f"Désherbage du {date_str}")
    ws_desh = wb.create_sheet(nom_desh)
    ri = 1
    ws_desh.merge_cells(f"A{ri}:I{ri}")
    c = ws_desh.cell(row=ri, column=1, value=f"Proposition de désherbage — {date_str}")
    c.font = Font(name="Arial", size=13, bold=True, color="FFFFFF")
    c.fill = hfill(C_WARN)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws_desh.row_dimensions[ri].height = 26
    ri += 2

    ri = _bandeau(ws_desh, ri, "Récapitulatif", C_WARN, 9)
    for label, val in [("Candidats détectés", len(tous_candidats_desherbage)),
                        ("Documentaires anciens (priorité)", len(documentaires_anciens)),
                        ("Autres anciens, séries peu installées", len(autres_candidats)),
                        ("Anciens mais grande série installée (prudence)", len(a_examiner_prudence)),
                        ("Signalés manuellement depuis Fonds total", len(signales_manuellement))]:
        ws_desh.cell(row=ri, column=1, value=label).font = Font(name="Arial", size=10)
        ws_desh.cell(row=ri, column=2, value=val).font = Font(name="Arial", size=10, bold=True)
        ri += 1
    ri += 1

    compteur_desh_type = Counter(d["type"] for d in tous_candidats_desherbage)
    compteur_desh_public = Counter(d["public"] for d in tous_candidats_desherbage)
    compteur_desh_genre = Counter(d["genre"] for d in tous_candidats_desherbage)
    ri = _section_repartition(ws_desh, ri, "Candidats — répartition par type", C_WARN, compteur_desh_type, len(tous_candidats_desherbage), 9)
    ri = _section_repartition(ws_desh, ri, "Candidats — répartition par public", C_WARN, compteur_desh_public, len(tous_candidats_desherbage), 9)
    ri = _section_repartition(ws_desh, ri, "Candidats — répartition par genre", C_WARN, compteur_desh_genre, len(tous_candidats_desherbage), 9)

    entetes_d = ["ISBN", "Titre", "Type", "Série", "Public", "Genre", "Année", "Sorti ?", "Date de sortie"]

    def _table(feuille, ri, titre, liste, couleur):
        ri = _bandeau(feuille, ri, titre, couleur, 9)
        for ci, h in enumerate(entetes_d, 1):
            cell = feuille.cell(row=ri, column=ci, value=h)
            cell.font = Font(name="Arial", size=10, bold=True)
            cell.border = _brd
        ri += 1
        for d in liste:
            vals = [d["isbn"], d["titre"], d["type"], d["serie"], d["public"], d["genre"], d["annee"], "", ""]
            for ci, v in enumerate(vals, 1):
                cell = feuille.cell(row=ri, column=ci, value=v)
                cell.font = Font(name="Arial", size=10)
                cell.border = _brd
            ri += 1
        return ri + 1

    ri = _table(ws_desh, ri, "Documentaires anciens (priorité : contenu factuel possiblement dépassé)",
                documentaires_anciens, C_WARN)
    ri = _table(ws_desh, ri, "Autres anciens, séries peu installées (candidats plausibles)",
                autres_candidats, C_WARN)
    ri = _table(ws_desh, ri, "Anciens mais grande série installée (à examiner avec prudence, probablement encore demandé)",
                a_examiner_prudence, C_WARN)
    if signales_manuellement:
        ri = _table(ws_desh, ri, "Signalés manuellement depuis Fonds total",
                    signales_manuellement, C_OK)

    for col, larg in zip("ABCDEFGHI", [16, 38, 16, 20, 14, 18, 8, 9, 16]):
        ws_desh.column_dimensions[col].width = larg
    ws_desh.freeze_panes = "A1"
    return nom_desh


def creer_onglet_desherbage_seul(wb, documentaires_anciens, autres_candidats, a_examiner_prudence,
                                  signales_manuellement=None):
    """Crée uniquement l'onglet 'Désherbage du JJ-MM-AAAA', sans onglet Commande
    associé — pour une demande de désherbage indépendante d'une acquisition.
    Renvoie le nom de l'onglet créé."""
    return _creer_feuille_desherbage(wb, documentaires_anciens, autres_candidats, a_examiner_prudence,
                                      signales_manuellement)


# ═══════════════════════════════════════════════════════════════
# PARTIE 5 — Catégories personnalisées (ex lib_config.py)
# ═══════════════════════════════════════════════════════════════


NOM_FEUILLE = "Config_Categories"

# Codes de recherche déjà connus et leur Type habituel (pour préremplir le
# préréglage "Fonds total" associé à une nouvelle catégorie personnalisée).
CODES_BASE = {
    "Aucune priorité particulière": (None, None),
    "BD":             ("bd", "BD"),
    "Mangas":         ("manga", "Manga"),
    "Romans jeunesse": ("roman_jeunesse", "Roman jeunesse"),
    "Romans ado":      ("roman_ado", "Roman ado / YA"),
    "Documentaires":   ("documentaire", "Documentaire"),
}


def charger_categories_custom(wb):
    """Renvoie la liste des catégories personnalisées : [{label, code_base,
    type_associe, adulte, pilon}, ...]. Liste vide si aucune n'a été ajoutée."""
    if NOM_FEUILLE not in wb.sheetnames:
        return []
    ws = wb[NOM_FEUILLE]
    categories = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row and row[0]:
            categories.append({
                "label": row[0],
                "code_base": row[1] if row[1] else None,
                "type_associe": row[2] if row[2] else None,
                "adulte": row[3] == "Oui",
                "pilon": row[4] == "Oui",
            })
    return categories


def ajouter_categorie_custom(wb, label, nom_code_base, adulte, pilon):
    """Ajoute une nouvelle catégorie personnalisée (créée depuis l'app).
    nom_code_base est une des clés de CODES_BASE."""
    code_base, type_associe = CODES_BASE.get(nom_code_base, (None, None))

    if NOM_FEUILLE in wb.sheetnames:
        ws = wb[NOM_FEUILLE]
    else:
        ws = wb.create_sheet(NOM_FEUILLE)
        ws.sheet_state = "hidden"  # pas besoin de la montrer dans Excel au quotidien
        ws.cell(row=1, column=1, value="Catégories personnalisées (Scanner) — ne pas modifier à la main")
        entetes = ["Label", "CodeBase", "TypeAssocié", "Adulte", "Pilon"]
        for ci, h in enumerate(entetes, 1):
            c = ws.cell(row=2, column=ci, value=h)
            c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
            c.fill = hfill(C_HEADER)

    # Éviter les doublons de libellé
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row and row[0] == label:
            return False

    ligne = ws.max_row + 1
    valeurs = [label, code_base, type_associe, "Oui" if adulte else "Non", "Oui" if pilon else "Non"]
    for ci, v in enumerate(valeurs, 1):
        ws.cell(row=ligne, column=ci, value=v).font = Font(name="Arial", size=10)
    return True


def supprimer_categorie_custom(wb, label):
    """Supprime une catégorie personnalisée par son libellé. Renvoie True si trouvée."""
    if NOM_FEUILLE not in wb.sheetnames:
        return False
    ws = wb[NOM_FEUILLE]
    for row in ws.iter_rows(min_row=3):
        if row[0].value == label:
            ws.delete_rows(row[0].row, 1)
            return True
    return False


# ═══════════════════════════════════════════════════════════════
# PARTIE 6 — Application Streamlit (interface)
# ═══════════════════════════════════════════════════════════════

CHEMIN_XLSX = os.environ.get(
    "INVENTAIRE_XLSX",
    os.path.join(os.path.dirname(os.path.abspath(__file__)), "inventaire_mediatheque.xlsx"),
)

MOT_DE_PASSE = os.environ.get("MEDIATHEQUE_MOT_DE_PASSE", "arcachon2026")

CATEGORIES = [
    ("BD",                "bd",             None),
    ("BD adultes",        "bd",             "Adulte"),
    ("Mangas",            "manga",          None),
    ("Mangas adultes",    "manga",          "Adulte"),
    ("Romans jeunesse",   "roman_jeunesse", None),
    ("Romans ado",        "roman_ado",      None),
    ("Romans adultes",    "roman_jeunesse", "Adulte"),
    ("Documentaires",     "documentaire",   None),
    ("Albums",            None,             None),
    ("Nouveautés",        None,             None),
    ("Retours",           None,             None),
    ("Pilons (à retirer du fonds)", None,   None),
]

st.set_page_config(page_title="Inventaire Jeunesse — Médiathèque d'Arcachon", page_icon="📚", layout="wide")

# ─────────────────────────────────────────────────────────
# AUTHENTIFICATION SIMPLE
# ─────────────────────────────────────────────────────────

def verifier_mot_de_passe():
    if st.session_state.get("authentifie"):
        return True
    st.title("📚 Inventaire Jeunesse — Médiathèque d'Arcachon")
    mdp = st.text_input("Mot de passe d'accès", type="password")
    if st.button("Entrer"):
        if mdp == MOT_DE_PASSE:
            st.session_state["authentifie"] = True
            st.rerun()
        else:
            st.error("Mot de passe incorrect.")
    return False

if not verifier_mot_de_passe():
    st.stop()

# ─────────────────────────────────────────────────────────
# NAVIGATION
# ─────────────────────────────────────────────────────────

# Préréglages de "Fonds total" : chaque entrée filtre par Type, et éventuellement
# par Public ("Adulte" étant désormais distingué via le scanner — voir CATEGORIES).
# Facile à étendre : ajouter une ligne ici suffit à faire apparaître une nouvelle
# sous-catégorie dans le menu.
PRESETS_FONDS = {
    "Tout le fonds":      {},
    "Mangas Jeunesse":    {"type": "Manga", "public_exclut": "Adulte"},
    "Mangas Adultes":     {"type": "Manga", "public": "Adulte"},
    "BD Jeunesse":        {"type": "BD", "public_exclut": "Adulte"},
    "BD Adultes":         {"type": "BD", "public": "Adulte"},
    "Romans Jeunesse":    {"type": "Roman jeunesse", "public_exclut": "Adulte"},
    "Romans Ado":         {"type": "Roman ado / YA"},
    "Romans Adultes":     {"type": "Roman jeunesse", "public": "Adulte"},
    "Documentaires":      {"type": "Documentaire"},
    "Albums":             {"type": "Album"},
    "Premières lectures": {"type": "Première lecture"},
    "Contes / Poésie":    {"type": "Conte / Poésie"},
}

def _categories_completes():
    """CATEGORIES + les catégories personnalisées ajoutées depuis l'app."""
    if not os.path.exists(CHEMIN_XLSX):
        return CATEGORIES
    wb = load_workbook(CHEMIN_XLSX, read_only=True)
    customs = charger_categories_custom(wb)
    wb.close()
    extra = [(c["label"], c["code_base"], "Adulte" if c["adulte"] else None) for c in customs if not c["pilon"]]
    extra_pilon = [(c["label"], c["code_base"], None) for c in customs if c["pilon"]]
    # Les catégories pilon personnalisées sont ajoutées à la fin, comme la catégorie pilon intégrée
    return CATEGORIES[:-1] + extra + extra_pilon + [CATEGORIES[-1]]


def _presets_fonds_complets():
    """PRESETS_FONDS + les préréglages dérivés des catégories personnalisées (hors pilon)."""
    presets = dict(PRESETS_FONDS)
    if not os.path.exists(CHEMIN_XLSX):
        return presets
    wb = load_workbook(CHEMIN_XLSX, read_only=True)
    customs = charger_categories_custom(wb)
    wb.close()
    for c in customs:
        if c["pilon"] or not c["type_associe"]:
            continue
        filtre = {"type": c["type_associe"]}
        if c["adulte"]:
            filtre["public"] = "Adulte"
        else:
            filtre["public_exclut"] = "Adulte"
        presets[c["label"]] = filtre
    return presets


OPTIONS_MENU = [
    "Accueil",
    "Scanner / Rechercher",
    "Fonds total",
    "Fiches à compléter",
    "Audit qualité",
    "Veille de parution",
    "Statistiques",
    "Pilon",
    "Médiation / Animation",
    "Nouvelle acquisition",
    "Statut de publication",
    "Désherbage seul",
    "À propos",
]

# Note technique : Streamlit interdit d'écrire dans st.session_state["nav_principal"]
# une fois ce widget affiché (c'était la cause exacte des liens cassés en page d'accueil).
# Il ignore aussi le paramètre index pour un widget déjà existant. On contourne les deux
# en donnant au widget une clé qui change à chaque redirection demandée par un bouton :
# Streamlit le traite alors comme un widget tout neuf et respecte le nouvel index.
if "nav_compteur" not in st.session_state:
    st.session_state["nav_compteur"] = 0
if "page_actuelle" not in st.session_state:
    st.session_state["page_actuelle"] = "Accueil"

def aller_a(nom_page):
    st.session_state["page_actuelle"] = nom_page
    st.session_state["nav_compteur"] += 1
    st.rerun()

st.sidebar.title("📚 Menu")
_index_depart = OPTIONS_MENU.index(st.session_state["page_actuelle"]) if st.session_state["page_actuelle"] in OPTIONS_MENU else 0
choix_principal = st.sidebar.radio(
    "Aller à :", OPTIONS_MENU, index=_index_depart,
    key=f"nav_principal_{st.session_state['nav_compteur']}",
)
st.session_state["page_actuelle"] = choix_principal

sous_choix_fonds = None
if choix_principal == "Fonds total":
    sous_choix_fonds = st.sidebar.radio("Catégorie", list(_presets_fonds_complets().keys()))

st.sidebar.markdown("---")
st.sidebar.caption("Médiathèque d'Arcachon — Réseau COBAS")
st.sidebar.caption(f"Version : {VERSION}")

# ─────────────────────────────────────────────────────────
# PAGE : SCANNER / RECHERCHER
# ─────────────────────────────────────────────────────────

def page_scanner():
    st.title("Scanner des ISBN")

    if "isbns_scannes" not in st.session_state:
        st.session_state["isbns_scannes"] = []   # liste de (isbn, categorie_affichee, est_pilon)
    if "resultats_session" not in st.session_state:
        st.session_state["resultats_session"] = None

    categories_actuelles = _categories_completes()
    cat_vers_code = {label: code for label, code, _ in categories_actuelles}
    cat_vers_public_force = {label: pf for label, _, pf in categories_actuelles}

    labels_pilon_custom = set()
    if os.path.exists(CHEMIN_XLSX):
        wb_tmp = load_workbook(CHEMIN_XLSX, read_only=True)
        labels_pilon_custom = {c["label"] for c in charger_categories_custom(wb_tmp) if c["pilon"]}
        wb_tmp.close()
    cat_vers_pilon = {label: (label.startswith("Pilons") or label in labels_pilon_custom)
                       for label, _, _ in categories_actuelles}

    NOUVELLE_CATEGORIE = "➕ Ajouter une catégorie…"
    options = [c[0] for c in categories_actuelles] + [NOUVELLE_CATEGORIE]

    col1, col2 = st.columns([1, 2])
    with col1:
        categorie_label = st.selectbox("Catégorie scannée", options)

    if categorie_label == NOUVELLE_CATEGORIE:
        st.info("Cette nouvelle catégorie apparaîtra aussi dans le menu « Fonds total » "
                "(sauf si elle sert au désherbage).")
        with st.form("form_nouvelle_categorie"):
            nom_cat = st.text_input("Nom de la catégorie (ex: Mangas seinen)")
            base_cat = st.selectbox("Type de recherche à privilégier", list(CODES_BASE.keys()))
            fc1, fc2 = st.columns(2)
            adulte_cat = fc1.checkbox("Réservée aux adultes")
            pilon_cat = fc2.checkbox("Catégorie de retrait (comme « Pilons »)")
            valider_cat = st.form_submit_button("Créer cette catégorie", type="primary")
        if valider_cat and nom_cat.strip():
            wb = load_workbook(CHEMIN_XLSX) if os.path.exists(CHEMIN_XLSX) else charger_ou_creer_classeur(CHEMIN_XLSX)[0]
            cree = ajouter_categorie_custom(wb, nom_cat.strip(), base_cat, adulte_cat, pilon_cat)
            sauvegarder_avec_backup(wb, CHEMIN_XLSX)
            if cree:
                st.success(f"Catégorie « {nom_cat.strip()} » créée. Sélectionnez-la dans la liste ci-dessus.")
            else:
                st.warning("Une catégorie avec ce nom existe déjà.")
            st.rerun()
        return

    categorie_code = cat_vers_code[categorie_label]
    est_pilon = cat_vers_pilon.get(categorie_label, False)
    est_nouveaute = categorie_label == "Nouveautés"

    with col2:
        st.write("")
        st.write("")
        if est_pilon:
            st.warning("Catégorie « Pilons » : les ISBN scannés ici seront RETIRÉS du fonds, pas ajoutés.")

    def ajouter_scan():
        valeur = st.session_state.get("champ_scan", "").strip().replace("-", "").replace(" ", "")
        if valeur and valeur.isdigit() and len(valeur) in (10, 13):
            deja = [i for i, _c, _p in st.session_state["isbns_scannes"]]
            if valeur not in deja:
                st.session_state["isbns_scannes"].append((valeur, categorie_label, est_pilon))
        st.session_state["champ_scan"] = ""

    st.text_input(
        "Scanner ou saisir un ISBN, puis Entrée",
        key="champ_scan",
        on_change=ajouter_scan,
        placeholder="Le curseur doit rester ici pendant le scan au pistolet",
    )

    if st.session_state["isbns_scannes"]:
        st.write(f"**{len(st.session_state['isbns_scannes'])} ISBN en attente :**")
        for idx, (isbn, cat_lbl, pilon) in enumerate(st.session_state["isbns_scannes"]):
            c1, c2, c3 = st.columns([3, 2, 1])
            c1.write(isbn)
            c2.write(cat_lbl)
            if c3.button("✕", key=f"suppr_{idx}"):
                st.session_state["isbns_scannes"].pop(idx)
                st.rerun()

        c1, c2 = st.columns([1, 1])
        if c1.button("🗑 Vider la liste"):
            st.session_state["isbns_scannes"] = []
            st.rerun()
        lancer = c2.button("▶ Lancer la recherche sur ces ISBN", type="primary")
    else:
        st.info("Scannez ou tapez un premier ISBN ci-dessus pour commencer.")
        lancer = False

    if lancer:
        traiter_lot(st.session_state["isbns_scannes"])
        st.session_state["isbns_scannes"] = []

    if st.session_state["resultats_session"]:
        afficher_resultats(st.session_state["resultats_session"])


def traiter_lot(liste_scans):
    """Lance la recherche/le retrait pour le lot scanné, met à jour l'Excel."""
    categories_actuelles = _categories_completes()
    cat_vers_code = {label: code for label, code, _ in categories_actuelles}
    cat_vers_public_force = {label: pf for label, _, pf in categories_actuelles}

    isbns_ajout = [(isbn, cat_vers_code.get(cat_lbl)) for isbn, cat_lbl, pilon in liste_scans if not pilon]
    isbns_retrait = [isbn for isbn, cat_lbl, pilon in liste_scans if pilon]
    nouveautes = {isbn for isbn, cat_lbl, pilon in liste_scans if cat_lbl == "Nouveautés"}
    public_force = {isbn: cat_vers_public_force[cat_lbl]
                     for isbn, cat_lbl, pilon in liste_scans if cat_vers_public_force.get(cat_lbl)}

    wb, ws = charger_ou_creer_classeur(CHEMIN_XLSX)

    deja_complets = isbns_deja_complets(ws, [i for i, _ in isbns_ajout])
    isbns_a_chercher = [(isbn, cat) for isbn, cat in isbns_ajout if isbn not in deja_complets]
    nb_deja_complets = len(isbns_ajout) - len(isbns_a_chercher)

    gest = GestionnaireInventaire(wb, ws)

    barre = st.progress(0.0, text="Préparation…")
    resultats_detail = []

    total = len(isbns_a_chercher)
    for i, (isbn, cat) in enumerate(isbns_a_chercher, 1):
        barre.progress(i / max(total, 1), text=f"[{i}/{total}] Recherche {isbn}…")
        res = chercher_isbn(isbn, cat)
        res["isbn"] = isbn
        if isbn in public_force:
            res["public"] = public_force[isbn]
        statut = gest.traiter_resultat(res, est_nouveaute=(isbn in nouveautes))
        resultats_detail.append((isbn, res.get("titre", ""), res["statut"], statut))
        time.sleep(0.1)

    if isbns_retrait:
        barre.progress(1.0, text="Retrait des pilons…")
        gest.retirer_pilon(isbns_retrait)

    gest.sauvegarder(CHEMIN_XLSX)
    barre.empty()

    st.session_state["resultats_session"] = {
        "detail": resultats_detail,
        "nb_deja_complets": nb_deja_complets,
        "nouveaux": gest.nouveaux,
        "completes": gest.completes,
        "doublons": gest.doublons,
        "retires": gest.retires,
        "non_presents_pilon": gest.non_presents_pilon,
    }
    st.rerun()


def afficher_resultats(r):
    st.success("Traitement terminé.")
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Nouveaux ajouts", r["nouveaux"])
    c2.metric("Fiches complétées", r["completes"])
    c3.metric("Déjà complets (ignorés)", r["nb_deja_complets"])
    c4.metric("Retirés du fonds", len(r["retires"]))

    if r["detail"]:
        st.write("**Détail :**")
        st.table([
            {"ISBN": isbn, "Titre": titre or "—", "Trouvé ?": statut_recherche, "Action": statut_ecriture}
            for isbn, titre, statut_recherche, statut_ecriture in r["detail"]
        ])
    if r["non_presents_pilon"]:
        st.warning(f"{len(r['non_presents_pilon'])} ISBN du pilon n'étaient pas dans l'inventaire (ignorés) : "
                   + ", ".join(r["non_presents_pilon"]))


# ─────────────────────────────────────────────────────────
# PAGE : INVENTAIRE (catalogue en cartes, pas un tableau)
# ─────────────────────────────────────────────────────────

COULEUR_TYPE = {
    "Manga":               "#E8505B",
    "BD":                  "#3A86FF",
    "Roman jeunesse":      "#5C9943",
    "Roman ado / YA":      "#8E44AD",
    "Documentaire":        "#E67E22",
    "Album":               "#16A085",
    "Première lecture":    "#2E86AB",
    "Conte / Poésie":      "#C2185B",
    "Livre-jeu / Activités": "#7F8C8D",
}
COULEUR_DEFAUT = "#999999"


@st.cache_data(ttl=30)
def _charger_livres(chemin, signature):
    wb = load_workbook(chemin, data_only=True)
    ws = wb["Inventaire"] if "Inventaire" in wb.sheetnames else wb.active
    livres = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row and row[0]:
            livres.append({
                "isbn": row[0], "titre": row[1] or "", "tome": row[2], "serie": row[3] or "",
                "auteur": row[4] or "", "illustrateur": row[5] or "", "editeur": row[6] or "",
                "pegi": row[7] or "", "public": row[8] or "", "type": row[9] or "",
                "genre": row[10] or "", "annee": row[11] or "",
                "date_ajout": row[12] if len(row) > 12 else "",
                "est_nouveaute": row[13] if len(row) > 13 else "",
            })
    return livres


def _formulaire_edition_si_actif(l):
    """Si ce document est celui actuellement sélectionné pour édition (peu importe
    la page d'où on vient), affiche le formulaire et traite Enregistrer/Annuler."""
    if st.session_state.get("isbn_edition") != l["isbn"]:
        return
    with st.form(key=f"form_{l['isbn']}"):
        fc1, fc2, fc3 = st.columns(3)
        titre_e = fc1.text_input("Titre", value=l["titre"])
        tome_e = fc2.text_input("Tome", value=str(l["tome"] or ""))
        serie_e = fc3.text_input("Série", value=l["serie"])
        fc4, fc5, fc6 = st.columns(3)
        auteur_e = fc4.text_input("Auteur", value=l["auteur"])
        illustrateur_e = fc5.text_input("Illustrateur", value=l["illustrateur"])
        editeur_e = fc6.text_input("Éditeur", value=l["editeur"])
        fc7, fc8, fc9, fc10 = st.columns(4)
        pegi_e = fc7.text_input("PEGI", value=l["pegi"])
        public_e = fc8.text_input("Public", value=l["public"])
        type_e = fc9.text_input("Type", value=l["type"])
        annee_e = fc10.text_input("Année", value=str(l["annee"] or ""))
        genre_e = st.text_input("Genre", value=l["genre"])

        cs1, cs2 = st.columns(2)
        enregistrer = cs1.form_submit_button("💾 Enregistrer", type="primary")
        annuler = cs2.form_submit_button("Annuler")

    if enregistrer:
        wb = load_workbook(CHEMIN_XLSX)
        ws = wb["Inventaire"] if "Inventaire" in wb.sheetnames else wb.active
        modifier_document(ws, l["isbn"], {
            "titre": titre_e, "tome": tome_e, "serie": serie_e, "auteur": auteur_e,
            "illustrateur": illustrateur_e, "editeur": editeur_e, "pegi": pegi_e,
            "public": public_e, "type": type_e, "genre": genre_e, "annee": annee_e,
        })
        sauvegarder_avec_backup(wb, CHEMIN_XLSX)
        st.session_state["isbn_edition"] = None
        st.success("Document modifié.")
        st.rerun()
    if annuler:
        st.session_state["isbn_edition"] = None
        st.rerun()


def page_inventaire(preset=None, titre_page="Fonds total"):
    st.title(f"📚 {titre_page}")
    preset = preset or {}

    if not os.path.exists(CHEMIN_XLSX):
        st.warning("Aucun fichier inventaire trouvé encore. Scannez d'abord quelques ISBN.")
        return

    signature = os.path.getmtime(CHEMIN_XLSX)
    livres = _charger_livres(CHEMIN_XLSX, signature)

    # ── Application du préréglage de catégorie (menu de gauche) ──
    base = livres
    if preset.get("type"):
        base = [l for l in base if l["type"] == preset["type"]]
    if preset.get("public"):
        base = [l for l in base if l["public"] == preset["public"]]
    if preset.get("public_exclut"):
        base = [l for l in base if l["public"] != preset["public_exclut"]]

    st.caption(f"{len(base)} document(s) dans cette catégorie sur {len(livres)} au total")

    # ── Bandeau figé : filtres + en-tête de colonnes ──────────────────
    # Astuce : la div ouverte ici n'est refermée qu'après l'en-tête, plus bas.
    # Tout ce qui est affiché entre les deux (filtres + en-tête) reste donc
    # ensemble dans le même bloc fixé en haut de page pendant le défilement,
    # comme un figement de volets Excel. C'est la partie la plus difficile à
    # garantir sans test réel : à vérifier une fois en conditions réelles.
    st.markdown("""
    <style>
    .badge-type {
        display: inline-block; color: white; font-size: 10px; font-weight: 600;
        padding: 1px 8px; border-radius: 9px; letter-spacing: .02em; white-space: nowrap;
    }
    .bandeau-fixe {
        position: sticky; top: 0; z-index: 100; background: #fff;
        padding-top: 4px; padding-bottom: 8px; border-bottom: 2px solid #ddd;
        box-shadow: 0 2px 4px rgba(0,0,0,0.04);
    }
    /* Lignes plus compactes dans les listes de documents, pour en voir davantage à l'écran */
    div[data-testid="stHorizontalBlock"] { gap: 0.4rem; align-items: center; }
    div[data-testid="column"] div[data-testid="stMarkdownContainer"] p {
        margin-bottom: 0; font-size: 13px; line-height: 1.25;
    }
    div[data-testid="column"] button {
        padding: 0.05rem 0.4rem; min-height: 1.5rem; font-size: 12px;
    }
    div[data-testid="stVerticalBlock"] { gap: 0.15rem; }
    </style>
    <div class="bandeau-fixe">
    """, unsafe_allow_html=True)

    recherche = st.text_input("🔍 Rechercher un titre, un auteur, une série…", placeholder="Tapez pour filtrer…")
    c2, c3, c4, c5, c6 = st.columns(5)
    with c2:
        type_f = st.selectbox("Type", ["Tous"] + sorted({l["type"] for l in base if l["type"]}))
    with c3:
        public_f = st.selectbox("Public", ["Tous"] + sorted({l["public"] for l in base if l["public"]}))
    with c4:
        genre_f = st.selectbox("Genre", ["Tous"] + sorted({l["genre"] for l in base if l["genre"]}))
    with c5:
        serie_f = st.selectbox("Série", ["Toutes"] + sorted({l["serie"] for l in base if l["serie"]}))
    with c6:
        tri = st.selectbox("Trier par", ["Titre", "Série", "Année ↓", "Année ↑"])

    resultats = base
    if recherche:
        rl = recherche.strip().lower()
        resultats = [l for l in resultats if rl in l["titre"].lower() or rl in l["auteur"].lower() or rl in l["serie"].lower()]
    if type_f != "Tous":
        resultats = [l for l in resultats if l["type"] == type_f]
    if public_f != "Tous":
        resultats = [l for l in resultats if l["public"] == public_f]
    if genre_f != "Tous":
        resultats = [l for l in resultats if l["genre"] == genre_f]
    if serie_f != "Toutes":
        resultats = [l for l in resultats if l["serie"] == serie_f]

    def _annee_num(l):
        try:
            return int(str(l["annee"]).strip())
        except (ValueError, TypeError):
            return 0

    if tri == "Titre":
        resultats = sorted(resultats, key=lambda l: l["titre"].lower())
    elif tri == "Série":
        resultats = sorted(resultats, key=lambda l: (l["serie"].lower() or "zzz", _annee_num(l)))
    elif tri == "Année ↓":
        resultats = sorted(resultats, key=_annee_num, reverse=True)
    else:
        resultats = sorted(resultats, key=_annee_num)

    st.write(f"**{len(resultats)}** résultat(s)")

    # Pagination — nécessaire pour garder l'application fluide sur un fonds
    # de plusieurs milliers de documents (tout afficher d'un coup ralentit
    # trop le navigateur).
    PAR_PAGE = 40
    nb_pages = max(1, (len(resultats) - 1) // PAR_PAGE + 1) if resultats else 1
    cle_page = "page_inventaire_num"
    if cle_page not in st.session_state:
        st.session_state[cle_page] = 1
    st.session_state[cle_page] = min(st.session_state[cle_page], nb_pages)
    page_actuelle = st.session_state[cle_page]
    debut = (page_actuelle - 1) * PAR_PAGE
    page_resultats = resultats[debut:debut + PAR_PAGE]

    # Colonnes identiques à l'onglet Inventaire du fichier Excel, ISBN compris.
    RATIOS = [1.1, 2.4, 0.5, 1.4, 1.5, 0.7, 0.9, 0.9, 1.1, 0.6, 0.9, 0.4, 0.4]
    NOMS_COLONNES = ["ISBN", "Titre", "Tome", "Série", "Auteur",
                      "PEGI", "Public", "Type", "Genre", "Année", "Date d'ajout", "", ""]
    total_ratio = sum(RATIOS)
    pourcentages = [r / total_ratio * 100 for r in RATIOS]

    entete_cellules = "".join(
        f'<div style="flex:0 0 {p:.2f}%;padding:6px 6px;font-weight:600;font-size:12px;'
        f'color:#444;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;">{nom}</div>'
        for nom, p in zip(NOMS_COLONNES, pourcentages)
    )
    st.markdown(
        f'<div style="display:flex;border-top:1px solid #eee;">{entete_cellules}</div></div>',
        unsafe_allow_html=True,
    )
    # (la balise </div> finale referme le .bandeau-fixe ouvert plus haut)

    if not page_resultats:
        st.info("Aucun document ne correspond à ces critères.")

    for l in page_resultats:
        couleur = COULEUR_TYPE.get(l["type"], COULEUR_DEFAUT)
        (col_isbn, col_titre, col_tome, col_serie, col_auteur,
         col_pegi, col_public, col_type, col_genre, col_annee, col_date,
         col_edit, col_suppr) = st.columns(RATIOS)
        with col_isbn:
            st.markdown(f"<span style='font-size:12px;color:#666;'>{l['isbn']}</span>", unsafe_allow_html=True)
        with col_titre:
            st.markdown(f"**{l['titre']}**")
        with col_tome:
            st.write(l["tome"] or "—")
        with col_serie:
            st.write(l["serie"] or "—")
        with col_auteur:
            st.write(l["auteur"] or "—")
        with col_pegi:
            st.write(l["pegi"] or "—")
        with col_public:
            st.write(l["public"] or "—")
        with col_type:
            st.markdown(f"<span class='badge-type' style='background:{couleur};'>{l['type'] or '—'}</span>", unsafe_allow_html=True)
        with col_genre:
            st.write(l["genre"] or "—")
        with col_annee:
            st.write(l["annee"] or "—")
        with col_date:
            st.write(l["date_ajout"] or "—")
        with col_edit:
            if st.button("✏️", key=f"editbtn_{l['isbn']}"):
                st.session_state["isbn_edition"] = (None if st.session_state.get("isbn_edition") == l["isbn"] else l["isbn"])
                st.rerun()
        with col_suppr:
            if st.button("🗑️", key=f"suprbtn_{l['isbn']}"):
                st.session_state["isbn_a_signaler"] = (None if st.session_state.get("isbn_a_signaler") == l["isbn"] else l["isbn"])
                st.rerun()

        if st.session_state.get("isbn_a_signaler") == l["isbn"]:
            st.warning(f"Signaler **{l['titre']}** pour désherbage ? Il ne sortira pas du fonds tout de suite : "
                       f"il faudra ensuite le scanner réellement dans Scanner / Recherche, catégorie Pilons, "
                       f"pour qu'il soit retiré pour de bon.")
            cs1, cs2 = st.columns(2)
            if cs1.button("🗑️ Confirmer le signalement", key=f"confirm_signal_{l['isbn']}", type="primary"):
                wb = load_workbook(CHEMIN_XLSX)
                signaler_pour_desherbage(wb, l["isbn"], l["titre"], l["type"], l["serie"],
                                          l["public"], l["genre"], l["annee"])
                sauvegarder_avec_backup(wb, CHEMIN_XLSX)
                st.session_state["isbn_a_signaler"] = None
                aller_a("Désherbage seul")
            if cs2.button("Annuler", key=f"annule_signal_{l['isbn']}"):
                st.session_state["isbn_a_signaler"] = None
                st.rerun()

        _formulaire_edition_si_actif(l)

        st.markdown("<hr style='margin:2px 0; border-color:#eee;'>", unsafe_allow_html=True)

    cp1, cp2, cp3 = st.columns([1, 2, 1])
    with cp1:
        if st.button("← Précédent", disabled=(page_actuelle <= 1)):
            st.session_state[cle_page] -= 1
            st.rerun()
    with cp2:
        st.markdown(f"<div style='text-align:center; color:#888;'>Page {page_actuelle} / {nb_pages}</div>", unsafe_allow_html=True)
    with cp3:
        if st.button("Suivant →", disabled=(page_actuelle >= nb_pages)):
            st.session_state[cle_page] += 1
            st.rerun()


def page_fiches_incompletes():
    st.title("🟡 Fiches à compléter")
    st.caption("Documents pour lesquels un champ important (Type, Genre, Public ou Année) "
               "n'a pas été renseigné — souvent les anciens « NON TROUVÉ » d'une recherche.")

    if not os.path.exists(CHEMIN_XLSX):
        st.warning("Aucun fichier inventaire trouvé encore.")
        return

    if st.button("🔄 Rafraîchir"):
        _charger_livres.clear()
        st.rerun()

    signature = os.path.getmtime(CHEMIN_XLSX)
    livres = _charger_livres(CHEMIN_XLSX, signature)

    CHAMPS_VERIFIES = [("type", "Type"), ("public", "Public"), ("genre", "Genre"), ("annee", "Année")]
    incomplets = []
    for l in livres:
        manquants = [nom for cle, nom in CHAMPS_VERIFIES if not str(l.get(cle, "")).strip()]
        if manquants:
            incomplets.append((l, manquants))

    if not incomplets:
        st.success("Toutes les fiches du fonds sont complètes sur ces quatre champs. 🎉")
        return

    champ_f = st.selectbox("Filtrer par champ manquant", ["Tous"] + [nom for _, nom in CHAMPS_VERIFIES])
    if champ_f != "Tous":
        incomplets = [(l, m) for l, m in incomplets if champ_f in m]

    st.write(f"**{len(incomplets)}** fiche(s) à compléter")

    PAR_PAGE = 40
    nb_pages = max(1, (len(incomplets) - 1) // PAR_PAGE + 1)
    cle_page = "page_incompletes_num"
    if cle_page not in st.session_state:
        st.session_state[cle_page] = 1
    st.session_state[cle_page] = min(st.session_state[cle_page], nb_pages)
    page_actuelle = st.session_state[cle_page]
    debut = (page_actuelle - 1) * PAR_PAGE
    page_resultats = incomplets[debut:debut + PAR_PAGE]

    st.markdown("""
    <style>
    .badge-manquant {
        display: inline-block; background: #FFEB9C; color: #7A5B00; font-size: 10px;
        font-weight: 600; padding: 1px 7px; border-radius: 8px; margin-right: 4px; white-space: nowrap;
    }
    div[data-testid="stHorizontalBlock"] { gap: 0.4rem; align-items: center; }
    div[data-testid="column"] div[data-testid="stMarkdownContainer"] p { margin-bottom: 0; font-size: 13px; }
    div[data-testid="column"] button { padding: 0.05rem 0.4rem; min-height: 1.5rem; font-size: 12px; }
    </style>
    """, unsafe_allow_html=True)

    for l, manquants in page_resultats:
        c1, c2, c3, c4 = st.columns([1.1, 3.2, 2.4, 0.5])
        with c1:
            st.markdown(f"<span style='font-size:12px;color:#666;'>{l['isbn']}</span>", unsafe_allow_html=True)
        with c2:
            st.markdown(f"**{l['titre']}**" + (f" — {l['serie']}" if l['serie'] else ""))
        with c3:
            st.markdown("".join(f"<span class='badge-manquant'>{m}</span>" for m in manquants), unsafe_allow_html=True)
        with c4:
            if st.button("✏️", key=f"editbtn_incomplet_{l['isbn']}"):
                st.session_state["isbn_edition"] = (None if st.session_state.get("isbn_edition") == l["isbn"] else l["isbn"])
                st.rerun()

        _formulaire_edition_si_actif(l)
        st.markdown("<hr style='margin:2px 0; border-color:#eee;'>", unsafe_allow_html=True)

    cp1, cp2, cp3 = st.columns([1, 2, 1])
    with cp1:
        if st.button("← Précédent", disabled=(page_actuelle <= 1), key="prev_incomplet"):
            st.session_state[cle_page] -= 1
            st.rerun()
    with cp2:
        st.markdown(f"<div style='text-align:center; color:#888;'>Page {page_actuelle} / {nb_pages}</div>", unsafe_allow_html=True)
    with cp3:
        if st.button("Suivant →", disabled=(page_actuelle >= nb_pages), key="next_incomplet"):
            st.session_state[cle_page] += 1
            st.rerun()


def page_pilon():
    st.title("🗑 Pilon — historique du désherbage")

    if not os.path.exists(CHEMIN_XLSX):
        st.warning("Aucun fichier inventaire trouvé encore.")
        return

    wb = load_workbook(CHEMIN_XLSX, data_only=True)
    if "Pilon" not in wb.sheetnames:
        st.info("Aucun document n'a encore été retiré du fonds.")
        return

    ws = wb["Pilon"]
    lignes = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row and row[0]:
            lignes.append({
                "ISBN": row[0], "Titre": row[1], "Tome": row[2], "Série": row[3],
                "Auteur": row[4], "Éditeur": row[6], "Public": row[8], "Type": row[9],
                "Genre": row[10], "Année": row[11],
                "Date de retrait": row[13] if len(row) > 13 else "",
            })
    st.caption(f"{len(lignes)} document(s) désherbé(s) depuis le début")
    recherche = st.text_input("🔍 Rechercher dans l'historique du pilon…")
    if recherche:
        rl = recherche.strip().lower()
        lignes = [l for l in lignes if rl in str(l["Titre"]).lower() or rl in str(l["Série"] or "").lower()]
    st.dataframe(lignes, use_container_width=True, hide_index=True)


def page_desherbage_seul():
    st.title("🟠 Demander une proposition de désherbage")
    st.caption("Indépendant d'une commande — pour obtenir uniquement une liste de candidats au désherbage.")

    if not os.path.exists(CHEMIN_XLSX):
        st.warning("Aucun fichier inventaire trouvé encore.")
        return

    # ── Signalements manuels (bouton 🗑️ depuis Fonds total) — toujours visibles,
    # qu'une proposition ait déjà été générée ou non ci-dessous.
    wb_lecture = load_workbook(CHEMIN_XLSX, read_only=True)
    signales = charger_signalements(wb_lecture)
    wb_lecture.close()

    if signales:
        st.markdown(f"### 🗑️ Signalés manuellement depuis Fonds total ({len(signales)})")
        st.caption("Ces documents ne sont pas encore désherbés — ils ne sortiront du fonds qu'une fois "
                   "réellement scannés dans Scanner / Recherche, catégorie Pilons.")
        for d in signales:
            c1, c2 = st.columns([5, 1])
            c1.write(f"**{d['titre']}** — {d['serie'] or d['type']} ({d['annee']})")
            if c2.button("✖ Retirer", key=f"retrait_signal_{d['isbn']}"):
                wb = load_workbook(CHEMIN_XLSX)
                retirer_signalement(wb, d["isbn"])
                sauvegarder_avec_backup(wb, CHEMIN_XLSX)
                st.rerun()
        st.markdown("---")

    if st.button("🔍 Générer la proposition de désherbage", type="primary"):
        wb = load_workbook(CHEMIN_XLSX)
        ws = wb["Inventaire"] if "Inventaire" in wb.sheetnames else wb.active
        docs, autres, prudence = detecter_desherbage(ws)
        st.session_state["proposition_desherbage_seul"] = {
            "docs": docs, "autres": autres, "prudence": prudence, "signales": signales,
        }

    prop = st.session_state.get("proposition_desherbage_seul")
    if not prop:
        st.info("Cliquez sur le bouton ci-dessus pour générer une proposition.")
        return

    nb_total = len(prop["docs"]) + len(prop["autres"]) + len(prop["prudence"]) + len(prop["signales"])
    st.success(f"{nb_total} candidat(s) au désherbage au total (signalements manuels inclus).")

    if prop["signales"]:
        with st.expander(f"🗑️ Signalés manuellement depuis Fonds total ({len(prop['signales'])})", expanded=True):
            st.dataframe([{"ISBN": d["isbn"], "Titre": d["titre"], "Série": d["serie"], "Année": d["annee"]}
                          for d in prop["signales"]], use_container_width=True, hide_index=True)
    if prop["docs"]:
        with st.expander(f"📕 Documentaires anciens — priorité ({len(prop['docs'])})", expanded=True):
            st.dataframe([{"ISBN": d["isbn"], "Titre": d["titre"], "Série": d["serie"], "Année": d["annee"]}
                          for d in prop["docs"]], use_container_width=True, hide_index=True)
    if prop["autres"]:
        with st.expander(f"📙 Autres anciens, petites séries — candidats plausibles ({len(prop['autres'])})"):
            st.dataframe([{"ISBN": d["isbn"], "Titre": d["titre"], "Série": d["serie"], "Année": d["annee"]}
                          for d in prop["autres"]], use_container_width=True, hide_index=True)
    if prop["prudence"]:
        with st.expander(f"📒 Anciens mais grande série installée — prudence ({len(prop['prudence'])})"):
            st.dataframe([{"ISBN": d["isbn"], "Titre": d["titre"], "Série": d["serie"], "Année": d["annee"]}
                          for d in prop["prudence"]], use_container_width=True, hide_index=True)

    st.markdown("---")
    cb1, cb2 = st.columns(2)
    if cb1.button("✅ Valider et créer l'onglet Désherbage", type="primary"):
        wb = load_workbook(CHEMIN_XLSX)
        nom = creer_onglet_desherbage_seul(wb, prop["docs"], prop["autres"], prop["prudence"], prop["signales"])
        if prop["signales"]:
            vider_signalements(wb)
        sauvegarder_avec_backup(wb, CHEMIN_XLSX)
        st.success(f"Onglet « {nom} » créé dans le fichier Excel.")
        del st.session_state["proposition_desherbage_seul"]
    if cb2.button("✖ Ne pas tenir compte de cette proposition"):
        del st.session_state["proposition_desherbage_seul"]
        st.rerun()


# ─────────────────────────────────────────────────────────
# PAGES SUIVANTES
# ─────────────────────────────────────────────────────────

def page_statistiques():
    st.title("📊 Statistiques du fonds")

    if not os.path.exists(CHEMIN_XLSX):
        st.warning("Aucun fichier inventaire trouvé encore. Scannez d'abord quelques ISBN.")
        return

    if st.button("🔄 Rafraîchir"):
        st.rerun()

    wb = load_workbook(CHEMIN_XLSX, data_only=True)
    s = calculer_statistiques(wb)

    def _afficher_repartition(compteur, total_pour_pct):
        if not compteur:
            st.caption("Aucune donnée pour le moment.")
            return
        tri = sorted(compteur.items(), key=lambda x: -x[1])
        max_n = max(n for _, n in tri) if tri else 1
        lignes_html = []
        for k, n in tri:
            pct = (n / total_pour_pct * 100) if total_pour_pct else 0
            largeur = (n / max_n * 100) if max_n else 0
            lignes_html.append(
                '<tr>'
                f'<td style="padding:4px 10px 4px 0;white-space:nowrap;overflow:hidden;'
                f'text-overflow:ellipsis;font-size:14px;">{k}</td>'
                '<td style="padding:4px 10px;">'
                '<div style="background:#eee;border-radius:4px;height:14px;width:100%;">'
                f'<div style="background:#3A86FF;border-radius:4px;height:14px;width:{largeur:.1f}%;"></div>'
                '</div></td>'
                f'<td style="padding:4px 10px;text-align:right;font-size:14px;">{n}</td>'
                f'<td style="padding:4px 0 4px 10px;text-align:right;color:#888;font-size:13px;white-space:nowrap;">{pct:.1f} %</td>'
                '</tr>'
            )
        # Largeurs FIXES (colgroup), identiques pour tous les tableaux de la page,
        # pour que les barres et les colonnes restent parfaitement alignées entre
        # eux quel que soit le nom de catégorie le plus long de chaque tableau.
        colgroup = (
            '<colgroup>'
            '<col style="width:180px;"><col><col style="width:55px;"><col style="width:65px;">'
            '</colgroup>'
        )
        st.markdown(
            f"<table style='width:100%;border-collapse:collapse;table-layout:fixed;'>"
            f"{colgroup}{''.join(lignes_html)}</table>",
            unsafe_allow_html=True,
        )

    c1, c2 = st.columns(2)
    c1.metric("Total documents inventoriés", s["total_inventaire"])
    c2.metric("Ajoutés ces 30 derniers jours (toutes catégories)", s["ajouts_recents_30j"])

    st.markdown("### Répartition par type")
    _afficher_repartition(s["compteur_types"], s["total_inventaire"])

    st.markdown("### Répartition par public")
    _afficher_repartition(s["compteur_public"], s["total_inventaire"])

    st.markdown("### Répartition par genre")
    _afficher_repartition(s["compteur_genre"], s["total_inventaire"])

    st.markdown("### Âge du fonds")
    st.metric("Âge moyen", f"{s['age_moyen']} ans" if s["age_moyen"] is not None else "n/d")
    if s["nb_annee_inconnue"]:
        st.caption(f"{s['nb_annee_inconnue']} document(s) sans année connue")
    decennies = {d: s["compteur_decennie"].get(d, 0) for d in s["ordre_decennies"] if s["compteur_decennie"].get(d)}
    _afficher_repartition(decennies, s["total_inventaire"])

    st.markdown(f"### 🟢 Nouveautés (30 derniers jours) — {s['ajouts_nouveautes_30j']}")
    st.caption("Documents ajoutés via la catégorie « Nouveautés » (Scanner/Rechercher ou "
               "isbn_nouveautés.txt en terminal) uniquement — pas tous les ajouts récents.")
    if s["ajouts_nouveautes_30j"]:
        st.markdown("**Par type**")
        _afficher_repartition(s["compteur_types_nouveautes"], s["ajouts_nouveautes_30j"])
        st.markdown("**Par public**")
        _afficher_repartition(s["compteur_public_nouveautes"], s["ajouts_nouveautes_30j"])
        st.markdown("**Par genre**")
        _afficher_repartition(s["compteur_genre_nouveautes"], s["ajouts_nouveautes_30j"])
    else:
        st.caption("Aucune nouveauté (catégorie Nouveautés) sur cette période pour le moment.")

    st.markdown("### 🟠 Désherbage cumulé (toutes sessions)")
    st.metric("Total documents pilonnés", s["total_pilon_cumule"])
    if s["total_pilon_cumule"]:
        st.markdown("**Par type**")
        _afficher_repartition(s["compteur_types_pilon"], s["total_pilon_cumule"])
        st.markdown("**Par public**")
        _afficher_repartition(s["compteur_public_pilon"], s["total_pilon_cumule"])
        st.markdown("**Par genre**")
        _afficher_repartition(s["compteur_genre_pilon"], s["total_pilon_cumule"])
    else:
        st.caption("Aucun désherbage enregistré pour le moment.")


def page_acquisition():
    st.title("🛒 Nouvelle acquisition")

    if not os.path.exists(CHEMIN_XLSX):
        st.warning("Aucun fichier inventaire trouvé encore. Scannez d'abord quelques ISBN.")
        return

    budget = st.number_input("Budget disponible pour cette commande (€)", min_value=0.0, value=1000.0, step=50.0)

    types_disponibles = list(PRIX_ESTIME.keys())
    st.write("Catégories concernées par cette commande :")
    cols_types = st.columns(len(types_disponibles))
    types_coches = []
    for col, typ in zip(cols_types, types_disponibles):
        if col.checkbox(typ, value=True, key=f"cat_{typ}"):
            types_coches.append(typ)

    if st.button("🔍 Générer la proposition", type="primary"):
        if not types_coches:
            st.warning("Cochez au moins une catégorie.")
            return
        wb = load_workbook(CHEMIN_XLSX)
        ws = wb["Inventaire"] if "Inventaire" in wb.sheetnames else wb.active

        # types_coches couvre toutes les catégories connues → équivalent à "toutes",
        # mais on passe quand même la liste pour rester cohérent si PRIX_ESTIME évolue.
        gaps = detecter_gaps_series(ws, types_filtre=types_coches)
        selection, total, completees = selectionner_dans_budget(gaps, budget)
        docs, autres, prudence = detecter_desherbage(ws)

        st.session_state["proposition"] = {
            "budget": budget, "selection": selection, "total": total,
            "completees": completees, "docs": docs, "autres": autres, "prudence": prudence,
            "types_coches": types_coches,
        }

    prop = st.session_state.get("proposition")
    if not prop:
        st.info("Indiquez un budget puis cliquez sur « Générer la proposition ».")
        return

    st.success(f"{len(prop['selection'])} tome(s) sélectionné(s) pour {prop['total']:.2f} € "
               f"sur {prop['budget']:.2f} € de budget — {len(prop['completees'])} série(s) complétée(s) entièrement.")

    st.markdown("### Tomes à commander")
    if prop["selection"]:
        selection_par_type = defaultdict(list)
        for s in prop["selection"]:
            selection_par_type[s["type"] or "(type inconnu)"].append(s)
        for typ in sorted(selection_par_type.keys()):
            items_type = selection_par_type[typ]
            sous_total = sum(s["prix"] for s in items_type)
            st.markdown(f"**{typ}** — {len(items_type)} tome(s), {sous_total:.2f} €")
            st.dataframe([
                {"Série": s["serie"], "Tome": s["tome"], "Éditeur": s["editeur"],
                 "Public": s["public"], "Genre": s["genre"], "Prix estimé": f"{s['prix']:.2f} €"}
                for s in sorted(items_type, key=lambda s: (s["serie"], s["tome"]))
            ], use_container_width=True, hide_index=True)
    else:
        st.caption("Aucun tome manquant détecté dans les catégories choisies.")

    st.caption("⚠️ Prix estimés par type de document, pas des prix réels — à vérifier avant commande "
               "(voir le bandeau « Nouveautés à compléter » de l'onglet créé).")

    nb_total_desherbage = len(prop["docs"]) + len(prop["autres"]) + len(prop["prudence"])
    st.markdown(f"### Candidats au désherbage ({nb_total_desherbage})")

    if prop["docs"]:
        with st.expander(f"📕 Documentaires anciens — priorité ({len(prop['docs'])})"):
            st.dataframe([{"ISBN": d["isbn"], "Titre": d["titre"], "Série": d["serie"], "Année": d["annee"]}
                          for d in prop["docs"]], use_container_width=True, hide_index=True)
    if prop["autres"]:
        with st.expander(f"📙 Autres anciens, petites séries — candidats plausibles ({len(prop['autres'])})"):
            st.dataframe([{"ISBN": d["isbn"], "Titre": d["titre"], "Série": d["serie"], "Année": d["annee"]}
                          for d in prop["autres"]], use_container_width=True, hide_index=True)
    if prop["prudence"]:
        with st.expander(f"📒 Anciens mais grande série installée — prudence ({len(prop['prudence'])})"):
            st.dataframe([{"ISBN": d["isbn"], "Titre": d["titre"], "Série": d["serie"], "Année": d["annee"]}
                          for d in prop["prudence"]], use_container_width=True, hide_index=True)

    st.markdown("---")
    st.write("Si cette proposition vous convient, validez-la pour créer les deux onglets "
              "« Commande du... » et « Désherbage du... » dans le fichier Excel.")
    cb1, cb2 = st.columns(2)
    if cb1.button("✅ Valider et créer les onglets dans Excel", type="primary"):
        wb = load_workbook(CHEMIN_XLSX)
        nom_cmd, nom_desh = creer_onglets_commande_et_desherbage(
            wb, prop["budget"], prop["selection"], prop["total"], prop["completees"],
            prop["docs"], prop["autres"], prop["prudence"],
        )
        sauvegarder_avec_backup(wb, CHEMIN_XLSX)
        st.success(f"Onglets créés : « {nom_cmd} » et « {nom_desh} ». "
                   f"Vous pouvez maintenant les ouvrir dans Excel pour les finaliser avec Claude "
                   f"(nouveautés tendance, prix réels).")
        del st.session_state["proposition"]
    if cb2.button("✖ Ne pas tenir compte de cette proposition"):
        del st.session_state["proposition"]
        st.rerun()


# ═══════════════════════════════════════════════════════════════
# PARTIE 4bis — Audit qualité de l'inventaire
# ═══════════════════════════════════════════════════════════════

# Conventions maison à faire respecter dans le champ Genre — voir GENRES_INTERDITS
# historique : pas d'anglicismes, "Frissons" plutôt que "Horreur"/"Thriller", etc.
GENRES_A_EVITER = {
    "horreur":            "Frissons",
    "thriller":           "Frissons",
    "policier/thriller":  "Policier/Mystère",
    "mystery":            "Policier/Mystère",
    "fantasy":            "Fantastique",
    "sci-fi":             "Science-fiction",
    "scifi":              "Science-fiction",
}


def auditer_qualite_inventaire(ws):
    """Scanne l'Inventaire et renvoie un dict {categorie: [liste de problèmes]}.
    Complète (sans redoublonner) la page « Fiches à compléter », qui couvre déjà
    Type/Public/Genre/Année VIDES. Ici on cherche plutôt les champs essentiels
    oubliés ailleurs (Titre, Auteur, Éditeur), les doublons, les incohérences
    Tome/Série, les valeurs hors normes (Type non standard, Année invraisemblable)
    et les écarts aux conventions de genre maison."""
    problemes = defaultdict(list)
    isbn_vus = {}
    annee_courante = datetime.date.today().year

    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or not row[0]:
            continue
        vals = list(row) + [None] * (13 - len(row))
        isbn, titre, tome, serie, auteur, illustrateur, editeur, pegi, public, typ, genre, annee, date_ajout = vals[:13]
        isbn = str(isbn).strip()
        titre_aff = (titre or "").strip() or "(titre manquant)"
        ref = {"isbn": isbn, "titre": titre_aff}

        if isbn in isbn_vus:
            problemes["Doublons d'ISBN"].append({**ref, "detail": f"déjà vu sous « {isbn_vus[isbn]} »"})
        else:
            isbn_vus[isbn] = titre_aff

        if not titre or not str(titre).strip():
            problemes["Titre manquant"].append(ref)
        if not auteur or not str(auteur).strip():
            problemes["Auteur manquant"].append(ref)
        if not editeur or not str(editeur).strip():
            problemes["Éditeur manquant"].append(ref)

        if typ and str(typ).strip() and str(typ).strip() not in PRIX_ESTIME:
            problemes["Type non standard (faute de frappe probable ?)"].append(
                {**ref, "detail": f"« {typ} » — types attendus : {', '.join(PRIX_ESTIME.keys())}"})

        if annee and str(annee).strip() and str(annee).strip() != "N/C":
            try:
                an = int(str(annee).strip())
                if an < 1900 or an > annee_courante + 1:
                    problemes["Année invraisemblable"].append({**ref, "detail": f"année = {an}"})
            except ValueError:
                problemes["Année invraisemblable"].append({**ref, "detail": f"valeur = « {annee} » (non numérique)"})

        if tome and str(tome).strip() and not (serie and str(serie).strip()):
            problemes["Tome renseigné sans Série"].append({**ref, "detail": f"tome = {tome}"})

        if genre and str(genre).strip():
            genre_norm = str(genre).strip().lower()
            for mauvais, suggestion in GENRES_A_EVITER.items():
                if mauvais in genre_norm:
                    problemes["Genre hors convention maison"].append(
                        {**ref, "detail": f"« {genre} » → suggestion : « {suggestion} »"})
                    break

    return problemes


def page_audit_qualite():
    st.title("🔎 Audit qualité de l'inventaire")
    st.caption("Complète « Fiches à compléter » : ici on cherche les doublons, les champs "
               "Titre/Auteur/Éditeur oubliés, les incohérences Tome/Série, les types ou années "
               "invraisemblables, et les écarts aux conventions de genre maison.")

    if not os.path.exists(CHEMIN_XLSX):
        st.warning("Aucun fichier inventaire trouvé encore.")
        return

    if st.button("🔄 Rafraîchir"):
        _charger_livres.clear()
        st.rerun()

    signature = os.path.getmtime(CHEMIN_XLSX)
    livres = _charger_livres(CHEMIN_XLSX, signature)
    livres_par_isbn = {str(l["isbn"]).strip(): l for l in livres}

    wb = load_workbook(CHEMIN_XLSX, data_only=True)
    ws = wb["Inventaire"] if "Inventaire" in wb.sheetnames else wb.active
    problemes = auditer_qualite_inventaire(ws)

    total_problemes = sum(len(v) for v in problemes.values())
    if not total_problemes:
        st.success("Aucune anomalie détectée sur ces contrôles. 🎉")
        return

    st.write(f"**{total_problemes}** anomalie(s) détectée(s) sur **{len(problemes)}** catégorie(s) de contrôle.")
    st.caption("Cliquez sur ✏️ pour corriger directement un document depuis cette page.")

    st.markdown("""
    <style>
    div[data-testid="stHorizontalBlock"] { gap: 0.4rem; align-items: center; }
    div[data-testid="column"] div[data-testid="stMarkdownContainer"] p { margin-bottom: 0; font-size: 13px; }
    div[data-testid="column"] button { padding: 0.05rem 0.4rem; min-height: 1.5rem; font-size: 12px; }
    </style>
    """, unsafe_allow_html=True)

    for categorie in sorted(problemes.keys(), key=lambda c: -len(problemes[c])):
        items = problemes[categorie]
        editable = categorie != "Doublons d'ISBN"  # ISBN ambigu en cas de doublon, voir modifier_document
        with st.expander(f"{categorie} — {len(items)}"):
            if not editable:
                st.caption("⚠️ Édition désactivée ici : deux lignes partagent le même ISBN, impossible de "
                           "cibler la bonne sans ambiguïté. Corrigez directement dans Excel pour ce cas.")
            for i, it in enumerate(items):
                isbn = it["isbn"]
                l = livres_par_isbn.get(isbn)
                c1, c2, c3, c4 = st.columns([1.3, 2.6, 2.6, 0.5])
                c1.markdown(f"<span style='font-size:12px;color:#666;'>{isbn}</span>", unsafe_allow_html=True)
                c2.markdown(f"**{it['titre']}**")
                c3.markdown(f"<span style='font-size:12px;color:#888;'>{it.get('detail','')}</span>", unsafe_allow_html=True)
                with c4:
                    if editable and l and st.button("✏️", key=f"editbtn_audit_{categorie}_{i}_{isbn}"):
                        st.session_state["isbn_edition"] = (None if st.session_state.get("isbn_edition") == isbn else isbn)
                        st.rerun()
                if editable and l:
                    _formulaire_edition_si_actif(l)
                st.markdown("<hr style='margin:2px 0; border-color:#eee;'>", unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════
# PARTIE 4ter — Veille de parution (sans Decalog ni ORB)
# ═══════════════════════════════════════════════════════════════

def bnf_infos_serie(nom_serie, max_records=30):
    """Interroge la BnF (gratuite, sans clé, sans quota) par titre de série.
    Renvoie {"tome_max": int|None, "annee_recente": int|None, "nb_notices": int}.
    Réutilisée par Veille de parution (tome_max) ET par le statut de
    publication avant désherbage (annee_recente). Signal indicatif seulement :
    dépend de la qualité et de la fraîcheur du catalogage BnF."""
    resultat = {"tome_max": None, "annee_recente": None, "nb_notices": 0}
    try:
        query = f'bib.title adj "{nom_serie}"'
        url = (
            "https://catalogue.bnf.fr/api/SRU?"
            "version=1.2&operation=searchRetrieve"
            f"&query={urllib.parse.quote(query)}"
            f"&maximumRecords={max_records}&recordSchema=unimarcxchange"
        )
        r = requests.get(url, headers=HEADERS, timeout=15)
        if r.status_code != 200:
            return resultat
        root = ET.fromstring(r.content)
        ns = {"srw": "http://www.loc.gov/zing/srw/"}
        nb = root.find(".//srw:numberOfRecords", ns)
        total = int(nb.text) if nb is not None and nb.text and nb.text.isdigit() else 0
        resultat["nb_notices"] = total
        if total == 0:
            return resultat

        tomes, annees = [], []
        for rec in root.findall(".//srw:recordData", ns):
            def sf(tag, code):
                for f in rec.iter():
                    if f.get("tag") == tag:
                        for s in f:
                            if s.get("code") == code and s.text:
                                return s.text.strip(" .,:/()")
                return ""
            titre_a = sf("200", "a")
            titre_n = sf("200", "n")
            if titre_a:
                tome = extraire_tome(titre_n, titre_a)
                if tome:
                    try:
                        tomes.append(int(tome))
                    except ValueError:
                        pass
            annee_raw = sf("214", "d") or sf("210", "d")
            m_an = re.search(r"\d{4}", annee_raw)
            if m_an:
                annees.append(int(m_an.group(0)))

        resultat["tome_max"] = max(tomes) if tomes else None
        resultat["annee_recente"] = max(annees) if annees else None
        return resultat
    except Exception:
        return resultat


def lister_series_avec_max_tome(ws, types_filtre=None):
    """Renvoie {serie: {type, editeur, public, genre, tome_max}} pour toute série
    identifiée dans l'inventaire, qu'elle ait un trou ou non — contrairement à
    detecter_gaps_series qui ne garde que les séries incomplètes."""
    series = defaultdict(list)
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or not row[0]:
            continue
        serie, tome, public, typ, genre, editeur = row[3], row[2], row[8], row[9], row[10], row[6]
        if serie and tome:
            try:
                t = int(str(tome).strip())
                series[serie].append((t, typ, editeur, public, genre))
            except ValueError:
                pass

    resultat = {}
    for serie, items in series.items():
        typ = _majoritaire([i[1] for i in items]) or "Roman jeunesse"
        if types_filtre and typ not in types_filtre:
            continue
        resultat[serie] = {
            "type": typ, "editeur": _majoritaire([i[2] for i in items]),
            "public": _majoritaire([i[3] for i in items]), "genre": _majoritaire([i[4] for i in items]),
            "tome_max": max(t for t, *_ in items),
        }
    return resultat


def page_veille_parution():
    st.title("📬 Veille de parution")
    st.caption("Pour chaque série déjà au catalogue, interroge la BnF (gratuite, sans clé) pour "
               "repérer si un tome plus récent que ceux possédés existe. Signal indicatif, à "
               "vérifier avant commande — pas une confirmation officielle de parution.")

    if not os.path.exists(CHEMIN_XLSX):
        st.warning("Aucun fichier inventaire trouvé encore.")
        return

    types_disponibles = list(PRIX_ESTIME.keys())
    st.write("Catégories à scanner :")
    cols_types = st.columns(len(types_disponibles))
    types_coches = []
    for col, typ in zip(cols_types, types_disponibles):
        if col.checkbox(typ, value=(typ in ("Manga", "BD")), key=f"veille_{typ}"):
            types_coches.append(typ)
    st.caption("Manga et BD cochés par défaut — ce sont les catégories qui comptent le plus de "
               "séries en cours. Le scan prend environ 1 seconde par série : limitez les "
               "catégories scannées si vous voulez aller plus vite.")

    cb1, cb2 = st.columns([2, 1])
    with cb1:
        lancer = st.button("🔍 Lancer le scan", type="primary")
    with cb2:
        if st.button("🔄 Réinitialiser"):
            st.session_state.pop("veille_resultats", None)
            st.session_state.pop("veille_date", None)
            st.session_state.pop("veille_nb_scannees", None)
            st.rerun()

    if lancer:
        if not types_coches:
            st.warning("Cochez au moins une catégorie.")
            return
        wb = load_workbook(CHEMIN_XLSX, data_only=True)
        ws = wb["Inventaire"] if "Inventaire" in wb.sheetnames else wb.active
        series = lister_series_avec_max_tome(ws, types_filtre=types_coches)

        resultats = []
        noms = list(series.keys())
        barre = st.progress(0.0, text=f"Démarrage du scan… ({len(noms)} série(s))")
        for i, nom in enumerate(noms):
            info = series[nom]
            infos_bnf = bnf_infos_serie(nom)
            tome_trouve = infos_bnf["tome_max"]
            if tome_trouve and tome_trouve > info["tome_max"]:
                resultats.append({
                    "Série": nom, "Type": info["type"], "Éditeur": info["editeur"],
                    "Tome possédé (max)": info["tome_max"], "Tome détecté (BnF)": tome_trouve,
                    "Tome(s) manquant(s) potentiel(s)": tome_trouve - info["tome_max"],
                })
            barre.progress((i + 1) / max(len(noms), 1), text=f"Scan en cours… {i+1}/{len(noms)} séries")
            time.sleep(0.3)
        barre.empty()

        st.session_state["veille_resultats"] = sorted(resultats, key=lambda r: -r["Tome(s) manquant(s) potentiel(s)"])
        st.session_state["veille_date"] = datetime.date.today().strftime("%d/%m/%Y")
        st.session_state["veille_nb_scannees"] = len(noms)

    resultats = st.session_state.get("veille_resultats")
    if resultats is None:
        st.info("Cochez vos catégories puis cliquez sur « Lancer le scan ». Patientez pendant le scan.")
        return

    date_scan = st.session_state.get("veille_date", "")
    nb_scannees = st.session_state.get("veille_nb_scannees", 0)
    st.caption(f"Dernier scan : {date_scan} — {nb_scannees} série(s) vérifiée(s)")

    if not resultats:
        st.success("Aucun nouveau tome détecté au-delà de ce que vous possédez déjà, sur les "
                   "catégories scannées. 🎉")
        return

    st.write(f"**{len(resultats)}** série(s) avec un tome potentiellement plus récent détecté à la BnF :")
    st.dataframe(resultats, use_container_width=True, hide_index=True)
    st.caption("⚠️ Signal basé sur le catalogage BnF — à confirmer (ORB, site éditeur) avant "
               "d'ajouter ces titres à une commande.")


def lister_series_anciennes(ws, seuil_age=None):
    """Renvoie {serie: {type, editeur, public, genre, tome_max, annee_recente_possedee}}
    pour TOUTES les séries du fonds dont la parution la plus récente possédée
    remonte à au moins seuil_age ans (par défaut SEUIL_AGE_AUTRES). Contrairement à
    detecter_desherbage : ne plafonne pas le nombre de résultats, n'exclut pas les
    petites séries, et ne dépend pas du type Documentaire."""
    if seuil_age is None:
        seuil_age = SEUIL_AGE_AUTRES
    series = defaultdict(list)
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or not row[0]:
            continue
        serie, tome, public, typ, genre, editeur, annee = row[3], row[2], row[8], row[9], row[10], row[6], row[11]
        if not serie:
            continue
        an = None
        if annee:
            try:
                an = int(str(annee).strip())
            except ValueError:
                an = None
        t = None
        if tome:
            try:
                t = int(str(tome).strip())
            except ValueError:
                t = None
        series[serie].append((t, typ, editeur, public, genre, an))

    annee_courante = datetime.date.today().year
    resultat = {}
    for serie, items in series.items():
        annees = [i[5] for i in items if i[5] is not None]
        if not annees:
            continue
        annee_recente = max(annees)
        if annee_courante - annee_recente < seuil_age:
            continue
        tomes = [i[0] for i in items if i[0] is not None]
        resultat[serie] = {
            "type": _majoritaire([i[1] for i in items]) or "",
            "editeur": _majoritaire([i[2] for i in items]),
            "public": _majoritaire([i[3] for i in items]),
            "genre": _majoritaire([i[4] for i in items]),
            "tome_max": max(tomes) if tomes else None,
            "annee_recente_possedee": annee_recente,
        }
    return resultat


def page_statut_publication():
    st.title("📡 Statut de publication avant désherbage")
    st.caption("Vérifie à la BnF, pour toutes les séries anciennes du fonds (15 ans ou plus "
               "depuis la dernière parution possédée), si une parution récente existe — pour "
               "éviter de désherber une série encore active. Plus large que les seuls candidats "
               "déjà retenus par « Désherbage seul ».")

    if not os.path.exists(CHEMIN_XLSX):
        st.warning("Aucun fichier inventaire trouvé encore.")
        return

    wb = load_workbook(CHEMIN_XLSX, data_only=True)
    ws = wb["Inventaire"] if "Inventaire" in wb.sheetnames else wb.active
    series = lister_series_anciennes(ws)
    series_candidates = sorted(series.keys())

    if not series_candidates:
        st.info("Aucune série de 15 ans ou plus identifiée dans le fonds actuellement.")
        return

    st.write(f"**{len(series_candidates)}** série(s) ancienne(s) identifiée(s) sur tout le fonds "
             f"— temps estimé du contrôle : ≈1 seconde par série.")

    cb1, cb2 = st.columns([2, 1])
    with cb1:
        lancer = st.button("🔍 Vérifier le statut de ces séries", type="primary")
    with cb2:
        if st.button("🔄 Réinitialiser"):
            st.session_state.pop("statut_pub_resultats", None)
            st.session_state.pop("statut_pub_date", None)
            st.rerun()

    if lancer:
        resultats = []
        annee_courante = datetime.date.today().year
        barre = st.progress(0.0, text=f"Démarrage… ({len(series_candidates)} série(s))")
        for i, nom in enumerate(series_candidates):
            infos = bnf_infos_serie(nom)
            if infos["annee_recente"] is None:
                verdict = "❓ Rien trouvé à la BnF — à vérifier manuellement"
            elif annee_courante - infos["annee_recente"] <= 3:
                verdict = "⚠️ Parution récente détectée — prudence avant désherbage"
            else:
                verdict = "✅ Rien de récent — désherbage raisonnable"
            resultats.append({
                "Série": nom,
                "Dernier tome détecté (BnF)": infos["tome_max"] if infos["tome_max"] is not None else "",
                "Dernière parution connue": infos["annee_recente"] if infos["annee_recente"] is not None else "",
                "Verdict": verdict,
            })
            barre.progress((i + 1) / len(series_candidates), text=f"Vérification… {i+1}/{len(series_candidates)}")
            time.sleep(0.3)
        barre.empty()
        st.session_state["statut_pub_resultats"] = resultats
        st.session_state["statut_pub_date"] = datetime.date.today().strftime("%d/%m/%Y")

    resultats = st.session_state.get("statut_pub_resultats")
    if resultats is None:
        st.info("Cliquez sur le bouton pour lancer la vérification.")
        return

    st.caption(f"Dernière vérification : {st.session_state.get('statut_pub_date', '')}")
    st.dataframe(resultats, use_container_width=True, hide_index=True)
    st.caption("⚠️ Signal basé sur le catalogage BnF — à confirmer avant toute décision de "
               "désherbage définitive (ORB, site éditeur, état physique du document).")


# ═══════════════════════════════════════════════════════════════
# PARTIE 4quater — Suggestions pour médiation / animation
# ═══════════════════════════════════════════════════════════════

def creer_onglet_mediation(wb, theme, selection):
    """Crée un onglet « Animation du JJ-MM-AAAA » listant la sélection de
    titres proposée pour une médiation/animation. Renvoie le nom de l'onglet."""
    date_str = datetime.date.today().strftime("%d-%m-%Y")
    nom = _nom_disponible(wb, f"Animation du {date_str}")
    ws_a = wb.create_sheet(nom)

    ri = 1
    ws_a.merge_cells(f"A{ri}:F{ri}")
    c = ws_a.cell(row=ri, column=1, value=f"Sélection médiation / animation — {date_str}")
    c.font = Font(name="Arial", size=13, bold=True, color="FFFFFF")
    c.fill = hfill(C_HEADER)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws_a.row_dimensions[ri].height = 26
    ri += 2

    if theme:
        ws_a.cell(row=ri, column=1, value="Thème / occasion :").font = Font(name="Arial", size=10, bold=True)
        ws_a.cell(row=ri, column=2, value=theme).font = Font(name="Arial", size=10)
        ri += 2

    entetes = ["Titre", "Série", "Auteur", "Éditeur", "Public", "Genre"]
    for ci, h in enumerate(entetes, 1):
        cell = ws_a.cell(row=ri, column=ci, value=h)
        cell.font = Font(name="Arial", size=10, bold=True)
        cell.border = _brd
    ri += 1
    for l in selection:
        vals = [l["titre"], l["serie"], l["auteur"], l["editeur"], l["public"], l["genre"]]
        for ci, v in enumerate(vals, 1):
            cell = ws_a.cell(row=ri, column=ci, value=v)
            cell.font = Font(name="Arial", size=10)
            cell.border = _brd
        ri += 1

    ws_a.column_dimensions["A"].width = 38
    ws_a.column_dimensions["B"].width = 22
    ws_a.column_dimensions["C"].width = 22
    ws_a.column_dimensions["D"].width = 18
    ws_a.column_dimensions["E"].width = 13
    ws_a.column_dimensions["F"].width = 22
    return nom


def page_mediation_animation():
    st.title("🎈 Suggestions pour médiation / animation")
    st.caption("Filtre votre fonds par type, public et genre, puis tire une sélection aléatoire "
               "de titres disponibles dans votre inventaire — utile pour préparer une table "
               "thématique ou une animation sans parcourir les rayons de mémoire. "
               "⚠️ Indique ce que vous possédez, pas ce qui est disponible en rayon à l'instant "
               "présent (un exemplaire peut être actuellement prêté).")

    if not os.path.exists(CHEMIN_XLSX):
        st.warning("Aucun fichier inventaire trouvé encore.")
        return

    signature = os.path.getmtime(CHEMIN_XLSX)
    livres = _charger_livres(CHEMIN_XLSX, signature)

    c1, c2, c3 = st.columns(3)
    with c1:
        types_f = st.multiselect("Type(s)", sorted({l["type"] for l in livres if l["type"]}))
    with c2:
        publics_f = st.multiselect("Public(s)", sorted({l["public"] for l in livres if l["public"]}))
    with c3:
        genres_f = st.multiselect("Genre(s)", sorted({l["genre"] for l in livres if l["genre"]}))

    recherche = st.text_input("🔍 Mot-clé dans le titre ou la série (optionnel)",
                               placeholder="ex : sorcière, Noël, espace…")

    nb_souhaite = st.slider("Nombre de titres souhaités", min_value=3, max_value=30, value=8)

    filtres = livres
    if types_f:
        filtres = [l for l in filtres if l["type"] in types_f]
    if publics_f:
        filtres = [l for l in filtres if l["public"] in publics_f]
    if genres_f:
        filtres = [l for l in filtres if l["genre"] in genres_f]
    if recherche:
        rl = recherche.strip().lower()
        filtres = [l for l in filtres if rl in l["titre"].lower() or rl in l["serie"].lower()]

    st.caption(f"{len(filtres)} document(s) correspondent à ces critères.")

    cb1, cb2, cb3 = st.columns([1.3, 1.3, 1])
    if cb1.button("🎲 Tirer une sélection", type="primary"):
        if not filtres:
            st.warning("Aucun document ne correspond à ces critères — élargissez les filtres.")
        else:
            n = min(nb_souhaite, len(filtres))
            st.session_state["mediation_selection"] = random.sample(filtres, n)
    if cb2.button("🔄 Retirer (mêmes critères)"):
        if filtres:
            n = min(nb_souhaite, len(filtres))
            st.session_state["mediation_selection"] = random.sample(filtres, n)
    if cb3.button("🗑 Réinitialiser"):
        st.session_state.pop("mediation_selection", None)
        st.rerun()

    selection = st.session_state.get("mediation_selection")
    if not selection:
        st.info("Choisissez vos critères puis cliquez sur « Tirer une sélection ».")
        return

    st.markdown("### Sélection proposée")
    st.dataframe([
        {"Titre": l["titre"], "Série": l["serie"], "Auteur": l["auteur"],
         "Éditeur": l["editeur"], "Public": l["public"], "Genre": l["genre"]}
        for l in selection
    ], use_container_width=True, hide_index=True)

    st.markdown("---")
    theme = st.text_input("Thème / occasion (pour repère dans l'onglet Excel)",
                            placeholder="ex : Halloween CE2-CM1, Semaine de l'espace…")
    if st.button("📑 Créer un onglet Excel avec cette sélection"):
        wb = load_workbook(CHEMIN_XLSX)
        nom_onglet = creer_onglet_mediation(wb, theme, selection)
        sauvegarder_avec_backup(wb, CHEMIN_XLSX)
        st.success(f"Onglet « {nom_onglet} » créé dans le fichier Excel.")


def page_accueil():
    st.title("📚 Médiathèque d'Arcachon")
    st.caption("Fonds jeunesse — Réseau COBAS")

    if not os.path.exists(CHEMIN_XLSX):
        st.info("Aucun fichier inventaire trouvé encore. Commencez par scanner quelques ISBN.")
        if st.button("📷 Aller au Scanner"):
            aller_a("Scanner / Rechercher")
        return

    wb = load_workbook(CHEMIN_XLSX, data_only=True)
    s = calculer_statistiques(wb)
    signales = charger_signalements(wb)
    nb_commandes_attente = compter_commandes_en_attente(wb)
    nb_desherbages_attente = compter_desherbages_en_attente(wb)

    livres = _charger_livres(CHEMIN_XLSX, os.path.getmtime(CHEMIN_XLSX))
    nb_incomplets = sum(
        1 for l in livres
        if not str(l.get("type", "")).strip() or not str(l.get("public", "")).strip()
        or not str(l.get("genre", "")).strip() or not str(l.get("annee", "")).strip()
    )

    st.write("")
    c1, c2 = st.columns(2)
    c1.metric("Total du fonds", s["total_inventaire"])
    c2.metric("Nouveautés (30 derniers jours)", s["ajouts_nouveautes_30j"])
    st.write("")

    a_traiter = [
        (signales, "🗑️", "signalement(s) de désherbage en attente", "Désherbage seul"),
        (nb_commandes_attente, "🛒", "tome(s) commandé(s) pas encore reçu(s)", "Nouvelle acquisition"),
        (nb_desherbages_attente, "🟠", "désherbage(s) proposé(s) pas encore sorti(s) du fonds", "Désherbage seul"),
        (nb_incomplets, "🟡", "fiche(s) à compléter", "Fiches à compléter"),
    ]
    a_traiter = [(n if isinstance(n, int) else len(n), icone, libelle, page) for n, icone, libelle, page in a_traiter]
    a_traiter = [t for t in a_traiter if t[0] > 0]

    if not a_traiter:
        st.success("Rien à signaler — tout est à jour. ✨")
    else:
        st.markdown("##### À traiter")
        for n, icone, libelle, page in a_traiter:
            c1, c2 = st.columns([5, 1.4])
            c1.write(f"{icone} **{n}** {libelle}")
            if c2.button("Voir →", key=f"accueil_voir_{page}_{libelle[:8]}"):
                aller_a(page)


def page_apropos():
    st.title("À propos")
    st.markdown(f"""
    Cette application remplace progressivement l'usage des scripts en ligne de commande
    (`recherche_isbn.py`, `maj_statistiques.py`, `nouvelle_acquisition.py`,
    `annuler_derniere_commande.py`), qui restent disponibles en parallèle.

    Le fichier `inventaire_mediatheque.xlsx` reste la source de vérité : vous pouvez
    toujours l'ouvrir directement dans Excel à tout moment.

    Toute l'application tient dans un seul fichier (`app.py`) — plus aucun fichier
    annexe à penser à remplacer lors d'une mise à jour.

    **Version installée : `{VERSION}`**
    """)


# ─────────────────────────────────────────────────────────
# ROUTAGE
# ─────────────────────────────────────────────────────────

if choix_principal == "Accueil":
    page_accueil()
elif choix_principal == "Scanner / Rechercher":
    page_scanner()
elif choix_principal == "Fonds total":
    page_inventaire(_presets_fonds_complets().get(sous_choix_fonds, {}), titre_page=sous_choix_fonds or "Fonds total")
elif choix_principal == "Fiches à compléter":
    page_fiches_incompletes()
elif choix_principal == "Audit qualité":
    page_audit_qualite()
elif choix_principal == "Veille de parution":
    page_veille_parution()
elif choix_principal == "Statistiques":
    page_statistiques()
elif choix_principal == "Pilon":
    page_pilon()
elif choix_principal == "Médiation / Animation":
    page_mediation_animation()
elif choix_principal == "Nouvelle acquisition":
    page_acquisition()
elif choix_principal == "Statut de publication":
    page_statut_publication()
elif choix_principal == "Désherbage seul":
    page_desherbage_seul()
else:
    page_apropos()
