#!/usr/bin/env python3
"""
Recalcule l'onglet Statistiques à partir de l'état ACTUEL de l'Inventaire,
sans lancer aucune recherche internet. À utiliser après des corrections
manuelles dans Excel (Type, Genre, Public...) pour rafraîchir les chiffres
sans attendre une nouvelle recherche ISBN.

Usage : python3 maj_statistiques.py
(à lancer depuis le même dossier que inventaire_mediatheque.xlsx)
"""

import os, sys, datetime
from collections import Counter

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import range_boundaries, get_column_letter
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import range_boundaries, get_column_letter

script_dir = os.path.dirname(os.path.abspath(__file__))
output_file = os.path.join(script_dir, "inventaire_mediatheque.xlsx")

print("="*65)
print("  Mise à jour des statistiques — Médiathèque d'Arcachon")
print("="*65)

if not os.path.exists(output_file):
    print(f"\n⚠  Fichier introuvable : {output_file}")
    input("\nEntrée pour quitter...")
    sys.exit(1)

wb = load_workbook(output_file)
ws = wb["Inventaire"] if "Inventaire" in wb.sheetnames else wb.active
total_inventaire = ws.max_row - 2

C_HEADER = "2E4A7A"
def hfill(h): return PatternFill("solid", fgColor=h)
thin = Side(style="thin", color="B8CCE4")
brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

COLS_ENTETES = ["ISBN","Titre","Tome","Série","Auteur","Illustrateur","Éditeur",
                "PEGI","Public","Type","Genre","Année","Date d'ajout"]
COLS_LARGEURS = [16,42,6,22,22,18,18,8,13,18,24,7,13]

migration_faite = False
for col, nom in enumerate(COLS_ENTETES, 1):
    cell_entete = ws.cell(row=2, column=col)
    if not cell_entete.value:
        cell_entete.value     = nom
        cell_entete.font      = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell_entete.fill      = hfill(C_HEADER)
        cell_entete.border    = brd
        cell_entete.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = COLS_LARGEURS[col-1]
        migration_faite = True
if migration_faite:
    for merged_range in list(ws.merged_cells.ranges):
        if str(merged_range).startswith("A1:"):
            ws.unmerge_cells(str(merged_range))
    ws.merge_cells(f"A1:{get_column_letter(len(COLS_ENTETES))}1")
    print(f"\n  ℹ  Colonne(s) manquante(s) ajoutée(s) à l'en-tête de l'Inventaire")

compteur_types  = Counter()
compteur_public = Counter()
compteur_genre  = Counter()
annee_courante  = datetime.date.today().year
annees_valides  = []
ajouts_recents_30j = 0
aujourdhui = datetime.date.today()
compteur_types_recents  = Counter()
compteur_public_recents = Counter()
compteur_genre_recents  = Counter()

for row in ws.iter_rows(min_row=3, values_only=True):
    if not row or not row[0]:
        continue
    if row[9]:  compteur_types[str(row[9]).strip()]  += 1
    if row[8]:  compteur_public[str(row[8]).strip()] += 1
    if row[10]: compteur_genre[str(row[10]).strip()] += 1
    if row[11]:
        try:
            an = int(str(row[11]).strip())
            if 1900 <= an <= annee_courante:
                annees_valides.append(an)
        except ValueError:
            pass
    if len(row) > 12 and row[12]:
        try:
            d = datetime.datetime.strptime(str(row[12]).strip(), "%d/%m/%Y").date()
            if (aujourdhui - d).days <= 30:
                ajouts_recents_30j += 1
                if row[9]:  compteur_types_recents[str(row[9]).strip()]  += 1
                if row[8]:  compteur_public_recents[str(row[8]).strip()] += 1
                if row[10]: compteur_genre_recents[str(row[10]).strip()] += 1
        except ValueError:
            pass

def frac(n, total):
    return (n/total) if total else None

def tranche_decennie(annee):
    if annee < 1990: return "Avant 1990"
    if annee < 2000: return "1990-1999"
    if annee < 2010: return "2000-2009"
    if annee < 2020: return "2010-2019"
    return "2020 et après"

age_moyen = round(sum(annee_courante - a for a in annees_valides) / len(annees_valides), 1) \
            if annees_valides else None
compteur_decennie  = Counter(tranche_decennie(a) for a in annees_valides)
nb_annee_inconnue  = total_inventaire - len(annees_valides)
ordre_decennies    = ["Avant 1990", "1990-1999", "2000-2009", "2010-2019", "2020 et après"]

total_pilon_cumule    = 0
compteur_types_pilon  = Counter()
compteur_public_pilon = Counter()
compteur_genre_pilon  = Counter()
if "Pilon" in wb.sheetnames:
    for row in wb["Pilon"].iter_rows(min_row=3, values_only=True):
        if row and row[0]:
            total_pilon_cumule += 1
            if row[9]:  compteur_types_pilon[str(row[9]).strip()]  += 1
            if row[8]:  compteur_public_pilon[str(row[8]).strip()] += 1
            if row[10]: compteur_genre_pilon[str(row[10]).strip()] += 1

if "Statistiques" in wb.sheetnames:
    ws2 = wb["Statistiques"]
    # Démerge d'abord (les bandeaux de section sont fusionnés) avant de vider
    for merged_range in list(ws2.merged_cells.ranges):
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        if min_col <= 3:
            ws2.unmerge_cells(str(merged_range))
    # On ne vide QUE les colonnes A à C — les notes en colonne E+ restent intactes
    for row in ws2.iter_rows(min_row=1, max_row=max(ws2.max_row, 200), max_col=3):
        for cell in row:
            cell.value = None
            cell.font  = Font(name="Arial", size=10)
            cell.fill  = PatternFill(fill_type=None)
else:
    ws2 = wb.create_sheet("Statistiques")

THEME = {
    "global":    "2E4A7A",
    "nouveaute": "2E7D46",
    "pilon":     "B45309",
    "commande":  "44546A",
}

sections = [
    ("commande", "Commandes utiles", [
        ("python3 recherche_isbn.py",   "lancer une nouvelle recherche", None),
        ("python3 maj_statistiques.py", "mettre à jour les stats sans recherche", None),
        ("python3 nouvelle_acquisition.py", "préparer une nouvelle commande", None),
        ("python3 annuler_derniere_commande.py", "annuler la dernière commande/désherbage", None),
        ("caffeinate",                  "empêcher le Mac de dormir (autre onglet Terminal)", None),
    ]),
    ("global", "Total et activité récente", [
        ("Total documents inventoriés",  total_inventaire, None),
        ("Ajoutés ces 30 derniers jours", ajouts_recents_30j, frac(ajouts_recents_30j, total_inventaire)),
    ]),
    ("global", "Répartition par type",
        [(t, n, frac(n, total_inventaire)) for t, n in sorted(compteur_types.items(), key=lambda x: -x[1])]),
    ("global", "Répartition par public",
        [(p, n, frac(n, total_inventaire)) for p, n in sorted(compteur_public.items(), key=lambda x: -x[1])]),
    ("global", "Répartition par genre",
        [(g, n, frac(n, total_inventaire)) for g, n in sorted(compteur_genre.items(), key=lambda x: -x[1])]),
    ("global", "Âge du fonds",
        [("Âge moyen des documents", f"{age_moyen} ans" if age_moyen is not None else "n/d", None)]
        + [(d, compteur_decennie[d], frac(compteur_decennie[d], total_inventaire))
           for d in ordre_decennies if compteur_decennie.get(d)]
        + ([("Année inconnue", nb_annee_inconnue, frac(nb_annee_inconnue, total_inventaire))]
           if nb_annee_inconnue else [])),
    ("nouveaute", "Nouveautés (30 derniers jours) — cumulé",
        [("Total ajoutés ces 30 derniers jours", ajouts_recents_30j, frac(ajouts_recents_30j, total_inventaire))]),
    ("nouveaute", "Nouveautés (30 derniers jours) — par type",
        [(t, n, frac(n, ajouts_recents_30j)) for t, n in sorted(compteur_types_recents.items(), key=lambda x: -x[1])]),
    ("nouveaute", "Nouveautés (30 derniers jours) — par public",
        [(p, n, frac(n, ajouts_recents_30j)) for p, n in sorted(compteur_public_recents.items(), key=lambda x: -x[1])]),
    ("nouveaute", "Nouveautés (30 derniers jours) — par genre",
        [(g, n, frac(n, ajouts_recents_30j)) for g, n in sorted(compteur_genre_recents.items(), key=lambda x: -x[1])]),
    ("pilon", "Désherbage (pilon) — cumulé toutes sessions",
        [("Total documents pilonnés", total_pilon_cumule, None)]),
    ("pilon", "Désherbage (pilon) — par type",
        [(t, n, frac(n, total_pilon_cumule)) for t, n in sorted(compteur_types_pilon.items(), key=lambda x: -x[1])]),
    ("pilon", "Désherbage (pilon) — par public",
        [(p, n, frac(n, total_pilon_cumule)) for p, n in sorted(compteur_public_pilon.items(), key=lambda x: -x[1])]),
    ("pilon", "Désherbage (pilon) — par genre",
        [(g, n, frac(n, total_pilon_cumule)) for g, n in sorted(compteur_genre_pilon.items(), key=lambda x: -x[1])]),
    ("global", "Dernière mise à jour",
        [("Date", datetime.date.today().strftime("%d/%m/%Y"), None)]),
]

ri = 1
ws2.merge_cells(f"A{ri}:C{ri}")
c = ws2.cell(row=ri, column=1, value="Inventaire Médiathèque d'Arcachon")
c.font = Font(name="Arial", size=13, bold=True, color="FFFFFF")
c.fill = hfill(C_HEADER)
c.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[ri].height = 26
ri += 2

for theme, titre, lignes in sections:
    if not lignes:
        lignes = [("(aucun document pour le moment)", "", None)]
    ws2.merge_cells(f"A{ri}:C{ri}")
    c = ws2.cell(row=ri, column=1, value=titre)
    c.font = Font(name="Arial", size=10.5, bold=True, color="FFFFFF")
    c.fill = hfill(THEME[theme])
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws2.row_dimensions[ri].height = 19
    ri += 1
    for label, val, p in lignes:
        ws2.cell(row=ri, column=1, value=label).font = Font(name="Arial", size=10)
        cb = ws2.cell(row=ri, column=2, value=val)
        cb.font = Font(name="Arial", size=10)
        if isinstance(val, (int, float)) and not isinstance(val, bool):
            cb.number_format = "#,##0"
            cb.alignment = Alignment(horizontal="right")
        else:
            cb.alignment = Alignment(horizontal="left")
        if p is not None:
            cc = ws2.cell(row=ri, column=3, value=p)
            cc.font = Font(name="Arial", size=10, color="808080")
            cc.number_format = "0.0%"
            cc.alignment = Alignment(horizontal="right")
        ri += 1
    ri += 1

ws2.column_dimensions["A"].width = 38
ws2.column_dimensions["B"].width = 13
ws2.column_dimensions["C"].width = 10

wb.save(output_file)

print(f"\n✓ Statistiques mises à jour : {total_inventaire} documents inventoriés")
print(f"  (colonnes A à C régénérées ; vos notes en colonne E ou plus n'ont pas été touchées)")
print("="*65)
print(f"\n  Commandes utiles :")
print(f"    python3 recherche_isbn.py     → lancer une nouvelle recherche")
print(f"    python3 maj_statistiques.py   → mettre à jour les stats sans recherche")
print(f"    python3 nouvelle_acquisition.py → préparer une nouvelle commande")
print(f"    python3 annuler_derniere_commande.py → annuler la dernière commande/désherbage")
print(f"    caffeinate                    → empêcher le Mac de dormir (autre onglet)")
print("="*65)
input("\nEntrée pour quitter...")
