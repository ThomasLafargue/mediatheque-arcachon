#!/usr/bin/env python3
"""
appliquer_fusion_genres.py — Applique le tableau de fusion des genres VALIDÉ
par le bibliothécaire (Fusion_genres_AAAA-MM-JJ.xlsx).

CONTRAT : ce script exécute le FICHIER, pas la proposition d'origine. Si
Thomas a corrigé la colonne « Proposé (à corriger si besoin) », c'est sa
version qui fait foi — le fichier Excel est l'acte de décision, le script
n'est que l'exécutant. C'est le même circuit de validation que pour
public_vise, adapté à un volume (374 lignes) impossible à valider en chat.

Ne touche qu'aux lignes où « Proposé » diffère de « Genre actuel ».
Relançable sans danger : une valeur déjà convertie ne matche plus rien.

Usage :
    python3 appliquer_fusion_genres.py                (simulation)
    python3 appliquer_fusion_genres.py --appliquer    (conversion réelle)
    python3 appliquer_fusion_genres.py --fichier Fusion_genres_2026-07-27.xlsx
"""
import glob
import os
import sys
import time

DOSSIER = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, DOSSIER)
import db  # noqa: E402


def charger_tableau(chemin):
    from openpyxl import load_workbook
    wb = load_workbook(chemin, read_only=True)
    ws = wb.active
    lignes = list(ws.iter_rows(min_row=2, values_only=True))
    # colonnes : Genre actuel | Notices | Proposé | Changement ? | Motif
    conversions = []
    for r in lignes:
        actuel, propose = r[0], r[2]
        if actuel and propose and str(actuel) != str(propose):
            conversions.append((str(actuel), str(propose)))
    return conversions


def main():
    appliquer = "--appliquer" in sys.argv
    if "--fichier" in sys.argv:
        chemin = os.path.join(
            DOSSIER, sys.argv[sys.argv.index("--fichier") + 1])
    else:
        candidats = sorted(glob.glob(os.path.join(DOSSIER, "Fusion_genres_*.xlsx")))
        if not candidats:
            print("Aucun fichier Fusion_genres_*.xlsx trouvé.")
            return
        chemin = candidats[-1]

    conversions = charger_tableau(chemin)
    print("=" * 68)
    print("FUSION DES GENRES — " + time.strftime("%Y-%m-%d %H:%M"))
    print(f"Fichier de décision : {os.path.basename(chemin)}")
    print("MODE : " + ("APPLICATION RÉELLE" if appliquer else
                       "SIMULATION (rien ne sera écrit)"))
    print("=" * 68)
    print(f"\n{len(conversions)} conversion(s) listée(s) dans le fichier.\n")

    conn = db.connect()
    total = 0
    for actuel, propose in conversions:
        n = conn.execute(
            "SELECT COUNT(*) FROM notice WHERE genre = ?", (actuel,)
        ).fetchone()[0]
        if n == 0:
            continue
        if appliquer:
            conn.execute(
                "UPDATE notice SET genre = ? WHERE genre = ?",
                (propose, actuel))
            conn.commit()
        total += n
        print(f"  {n:>6}  {actuel[:38]:40} -> {propose[:38]}")

    if appliquer:
        restantes = conn.execute(
            "SELECT COUNT(DISTINCT genre) FROM notice "
            "WHERE genre IS NOT NULL AND genre != ''").fetchone()[0]
        print(f"\n  {total} notice(s) converties. "
              f"{restantes} valeur(s) de genre distinctes restantes.")
    else:
        print(f"\n  {total} notice(s) seraient converties.")
        print("  Pour appliquer :  python3 appliquer_fusion_genres.py --appliquer")
    conn.close()


if __name__ == "__main__":
    main()
