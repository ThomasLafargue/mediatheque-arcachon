#!/usr/bin/env python3
"""
charger_supports_en_base.py — Charge le dernier Supports_a_corriger_*.xlsx
dans la table `support_a_corriger`, pour que le CHAT puisse servir la liste
aux agents (« quels supports reste-t-il à corriger dans Decalog ? »).

CONTEXTE (2026-07-30) : l'audit DVD a révélé 11 LIVRES saisis à tort en
support DVD dans Decalog (albums, BD, romans dont gros caractères) + 1 DVD
mal coté. Chaque cas a été vérifié sur internet (librairies, éditeurs).
Le fichier Excel ne vivait que sur le Mac mini ; les agents passent par le
chat — la base est leur seul point d'accès commun.

Même mécanique que charger_ean_en_base.py : table ENTIÈREMENT REMPLACÉE à
chaque chargement, mais `statut_saisie` préservé quand le code-barres se
retrouve (une ligne marquée « saisie » ne redevient pas « à faire »).

Usage :  python3 charger_supports_en_base.py
"""
import glob
import os
import sys

DOSSIER = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, DOSSIER)
import db  # noqa: E402


def main():
    candidats = sorted(glob.glob(os.path.join(DOSSIER, "Supports_a_corriger_*.xlsx")))
    if not candidats:
        print("Aucun fichier Supports_a_corriger_*.xlsx trouvé.")
        return
    chemin = candidats[-1]
    print(f"Chargement de {os.path.basename(chemin)}...")

    from openpyxl import load_workbook
    ws = load_workbook(chemin, read_only=True).active
    lignes = list(ws.iter_rows(min_row=1, values_only=True))
    entetes = [str(c) for c in lignes[0]]
    idx = {nom: i for i, nom in enumerate(entetes)}

    conn = db.connect()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS support_a_corriger (
            code_barres TEXT PRIMARY KEY,
            identifiant TEXT,          -- EAN/ISBN
            cote_actuelle TEXT, titre TEXT, auteur_realisateur TEXT,
            editeur TEXT,
            action TEXT,               -- quoi faire dans Decalog
            verification TEXT,         -- preuve (source internet consultée)
            statut_saisie TEXT NOT NULL DEFAULT 'à faire',
            date_chargement TEXT DEFAULT (date('now'))
        )
    """)
    statuts = dict(conn.execute(
        "SELECT code_barres, statut_saisie FROM support_a_corriger "
        "WHERE statut_saisie != 'à faire'").fetchall())
    conn.execute("DELETE FROM support_a_corriger")

    n = 0
    for l in lignes[1:]:
        cb = str(l[idx["Code-barres"]] or "").strip()
        if not cb:
            continue
        conn.execute(
            "INSERT OR REPLACE INTO support_a_corriger "
            "(code_barres, identifiant, cote_actuelle, titre, "
            " auteur_realisateur, editeur, action, verification, statut_saisie) "
            "VALUES (?,?,?,?,?,?,?,?,?)",
            (cb, str(l[idx["Identifiant"]] or ""), l[idx["Cote actuelle"]],
             l[idx["Titre"]], l[idx["Réalisateur-Auteur"]], l[idx["Éditeur"]],
             l[idx["Action Decalog"]], l[idx["Vérification"]],
             statuts.get(cb, "à faire")))
        n += 1
    conn.commit()

    for action, nb in conn.execute(
            "SELECT action, COUNT(*) FROM support_a_corriger "
            "GROUP BY action ORDER BY 2 DESC").fetchall():
        print(f"  {nb:>5}  {action}")
    print(f"✓ {n} ligne(s) chargées dans la table support_a_corriger.")
    conn.close()


if __name__ == "__main__":
    main()
