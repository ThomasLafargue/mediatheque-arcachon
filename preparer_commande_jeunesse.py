#!/usr/bin/env python3
"""
preparer_commande_jeunesse.py — Construit une liste d'acquisition JEUNESSE/ADO
d'environ 10 000 € (prix public), prête à importer dans ORB.

COMMANDE (Thomas, 2026-08-01) :
  - publics : Jeunesse et Adolescent uniquement ;
  - catégories : BD, Manga, Album, Roman jeunesse, Roman ado / YA,
    Première lecture, Documentaire ;
  - exclus : jeux de société, jeux vidéo, DVD, CD audio ;
  - priorité 1 : TOMES MANQUANTS des séries entamées du fonds (séries les
    plus empruntées d'abord) ; ensuite : nouveautés/valeurs sûres
    (suggestions de la veille, puis sélections jeunesse Place des Libraires) ;
  - JAMAIS un document déjà en fonds (contrôle par EAN, par série+tome et
    par titre+auteur) ;
  - budget serré : 10 000 € prix public éditeur.

SORTIES (datées du jour) :
  - Commande_jeunesse_AAAA-MM-JJ.xlsx : liste complète et motifs ;
  - commande_orb_AAAA-MM-JJ.txt      : un EAN par ligne ;
  - commande_orb_AAAA-MM-JJ.csv      : EAN;quantité (1 par titre).

FONCTIONNEMENT :
  - STRICTEMENT EN LECTURE sur le catalogue ; n'écrit que des fichiers.
  - Scrape Place des Libraires (fiches Dilicom, prix inclus) avec une pause
    entre chaque requête. ⚠ NE PAS lancer en même temps qu'un autre scraper
    Place des Libraires (veille, enrichissement).
  - Cache _commande_cache.json : interrompu puis relancé, il reprend où il
    en était sans re-scraper.

Usage :
    python3 preparer_commande_jeunesse.py --test        # vérifie le prix sur 3 fiches
    python3 -u preparer_commande_jeunesse.py > commande.log 2>&1 &
    tail -f commande.log
"""
import argparse
import csv
import datetime
import json
import os
import re
import sys
import time
import unicodedata
from collections import defaultdict

import requests
from bs4 import BeautifulSoup

DOSSIER = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, DOSSIER)
import db  # noqa: E402

BUDGET = 10000.0
PUBLICS = ("Jeunesse", "Adolescent")
CATEGORIES = ("BD", "Manga", "Album", "Roman jeunesse", "Roman ado / YA",
              "Première lecture", "Documentaire")
SUPPORTS_EXCLUS = ("cd", "dvd", "jeu", "vinyle", "blu-ray", "coffret dvd")
# éditions à écarter : packs libraires, coffrets, produits dérivés
MOTS_EXCLUS_TITRE = ("coffret", "integrale", "présentoir", "presentoir",
                     "sabot", "calendrier", "agenda", "puzzle", "pack ",
                     "marque-page", "marque page")
MOTIF_LOT = re.compile(r"\b\d+\s*ex\b", re.I)
PRIX_MAX_GENERAL = 40.0     # au-delà : édition spéciale, pas un achat courant
PRIX_MAX_TOME = 28.0        # un tome de série ne coûte pas plus cher

BASE = "https://www.placedeslibraires.fr"
HEADERS = {"User-Agent": ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                          "AppleWebKit/605.1.15 (KHTML, like Gecko) "
                          "Version/17.4 Safari/605.1.15")}
PAUSE = 0.6
FICHIER_CACHE = os.path.join(DOSSIER, "_commande_cache.json")

_cache = {"fiche": {}, "recherche": {}}
_nb_requetes = 0


# ── utilitaires ──────────────────────────────────────────────────────────────

def _sans_accents(t):
    t = unicodedata.normalize("NFKD", t or "")
    return "".join(c for c in t if not unicodedata.combining(c))


def _norm(t):
    return re.sub(r"[^a-z0-9 ]", "", _sans_accents(t or "").lower()).strip()


def _similarite(a, b):
    ma, mb = set(_norm(a).split()), set(_norm(b).split())
    if not ma or not mb:
        return 0.0
    return len(ma & mb) / len(ma | mb)


def charger_cache():
    global _cache
    if os.path.exists(FICHIER_CACHE):
        try:
            _cache = json.load(open(FICHIER_CACHE, encoding="utf-8"))
            print(f"Cache repris : {len(_cache['fiche'])} fiches, "
                  f"{len(_cache['recherche'])} recherches déjà faites.")
        except Exception:
            pass


def sauver_cache():
    json.dump(_cache, open(FICHIER_CACHE, "w", encoding="utf-8"),
              ensure_ascii=False)


def _get(url):
    global _nb_requetes
    time.sleep(PAUSE)
    _nb_requetes += 1
    if _nb_requetes % 50 == 0:
        sauver_cache()
    try:
        r = requests.get(url, headers=HEADERS, timeout=(5, 15),
                         allow_redirects=True)
        return r.text if r.status_code == 200 else None
    except Exception:
        return None


# ── Place des Libraires : fiche (titre, support, PRIX) et recherche ─────────

def fiche(isbn):
    """Fiche PdL par ISBN : titre, auteur, éditeur, support, série, tome, prix.
    Renvoie None si introuvable."""
    isbn = str(isbn).strip()
    if isbn in _cache["fiche"]:
        return _cache["fiche"][isbn]
    resultat = None
    html = _get(f"{BASE}/livre/{isbn}/")
    if html and "Place des Libraires" in html:
        soup = BeautifulSoup(html, "html.parser")
        el = soup.find("title")
        brut = el.get_text(strip=True) if el else ""
        corps = re.sub(r"\s*-\s*Place des Libraires\s*$", "", brut).strip()
        parts = [p.strip() for p in corps.split(" - ") if p.strip()]
        if parts:
            titre = parts[0]
            serie = tome = ""
            m = re.match(r"^(.*?)\s+Tome\s+(\d+)\s*$", titre, re.I)
            if m:
                serie, tome = m.group(1).strip(), m.group(2)
            # prix : plusieurs motifs, du plus fiable au plus générique
            prix = None
            for motif in (r'itemprop="price"[^>]*content="([\d.,]+)"',
                          r'"price"\s*:\s*"?([\d.,]+)',
                          r'og:price:amount"[^>]*content="([\d.,]+)"',
                          r'([\d]{1,3},\d{2})\s*€'):
                m2 = re.search(motif, html)
                if m2:
                    try:
                        prix = float(m2.group(1).replace(",", "."))
                        break
                    except ValueError:
                        pass
            resultat = {
                "titre": titre, "serie": serie, "tome": tome,
                "auteur": parts[1] if len(parts) >= 2 else "",
                "editeur": parts[2] if len(parts) >= 3 else "",
                "support": parts[3] if len(parts) >= 4 else "",
                "prix": prix,
            }
    _cache["fiche"][isbn] = resultat
    return resultat


def rechercher(requete):
    """Recherche PdL → liste d'ISBN candidats (max 5)."""
    cle = _norm(requete)[:120]
    if cle in _cache["recherche"]:
        return _cache["recherche"][cle]
    import urllib.parse
    html = _get(f"{BASE}/listeliv.php?mots_recherche="
                + urllib.parse.quote(requete[:120]))
    candidats = []
    if html:
        for m in re.finditer(r"/livre/(\d{9,13})", html):
            if m.group(1) not in candidats:
                candidats.append(m.group(1))
            if len(candidats) >= 5:
                break
    _cache["recherche"][cle] = candidats
    return candidats


def support_exclu(support):
    s = (support or "").lower()
    return any(x in s for x in SUPPORTS_EXCLUS)


# ── chargement du fonds (dédoublonnage) ─────────────────────────────────────

def charger_fonds(conn):
    isbns = set()
    for (ident,) in conn.execute("SELECT identifiant FROM notice").fetchall():
        if ident and re.fullmatch(r"\d{9,13}", str(ident)):
            isbns.add(str(ident))

    series = {}     # serie normalisée -> {"tomes": {int}, "prets": n,
                    #                      "categorie", "serie", "auteur"}
    for serie, tome, cat, prets, createurs in conn.execute(
            "SELECT serie, tome, categorie, COALESCE(nb_prets_total, 0), createurs "
            "FROM notice WHERE serie IS NOT NULL AND serie != '' "
            "AND public_vise IN (?, ?) AND categorie IN ({})".format(
                ",".join("?" * len(CATEGORIES))),
            PUBLICS + CATEGORIES).fetchall():
        cle = _norm(serie)
        if not cle:
            continue
        d = series.setdefault(cle, {"tomes": set(), "prets": 0,
                                    "categorie": cat, "serie": serie,
                                    "auteur": (createurs or "").split(";")[0].strip()})
        d["prets"] += prets or 0
        m = re.match(r"^\s*(\d+)", str(tome or ""))
        if m:
            d["tomes"].add(int(m.group(1)))

    titres = set()
    for titre, createurs in conn.execute(
            "SELECT titre, createurs FROM notice").fetchall():
        auteur = (createurs or "").split(";")[0].split(",")[0]
        titres.add(_norm(titre) + "|" + _norm(auteur))
    return isbns, series, titres


# ── construction de la liste ────────────────────────────────────────────────

class Panier:
    def __init__(self, isbns_fonds, series_fonds, titres_fonds):
        self.lignes = []
        self.total = 0.0
        self.isbns = set(isbns_fonds)
        self.serie_tome = {(cle, t) for cle, d in series_fonds.items()
                           for t in d["tomes"]}
        self.titres = set(titres_fonds)

    def deja_vu(self, isbn, serie, tome, titre, auteur):
        if isbn in self.isbns:
            return True
        m = re.match(r"^\s*(\d+)", str(tome or ""))
        if serie and m and (_norm(serie), int(m.group(1))) in self.serie_tome:
            return True
        if (_norm(titre) + "|" + _norm((auteur or "").split(",")[0])) in self.titres:
            return True
        return False

    def ajouter(self, categorie, f, isbn, origine, motif):
        prix = f.get("prix")
        plafond = PRIX_MAX_TOME if origine.startswith("Tome") else PRIX_MAX_GENERAL
        if not prix or prix <= 0 or prix > plafond:
            return False
        titre_n = _sans_accents(f.get("titre") or "").lower()
        if any(m in titre_n for m in MOTS_EXCLUS_TITRE) or MOTIF_LOT.search(titre_n):
            return False
        if self.total + prix > BUDGET:
            return False
        if self.deja_vu(isbn, f.get("serie"), f.get("tome"),
                        f.get("titre"), f.get("auteur")):
            return False
        if support_exclu(f.get("support")):
            return False
        self.lignes.append({
            "categorie": categorie, "serie": f.get("serie", ""),
            "tome": f.get("tome", ""), "titre": f.get("titre", ""),
            "auteur": f.get("auteur", ""), "editeur": f.get("editeur", ""),
            "ean": isbn, "prix": prix, "origine": origine, "motif": motif,
        })
        self.total += prix
        self.isbns.add(isbn)
        if f.get("serie") and str(f.get("tome", "")).isdigit():
            self.serie_tome.add((_norm(f["serie"]), int(f["tome"])))
        self.titres.add(_norm(f.get("titre")) + "|"
                        + _norm((f.get("auteur") or "").split(",")[0]))
        return True


def phase_tomes_manquants(panier, series_fonds):
    print("\n─── PHASE 1 : tomes manquants des séries du fonds ───")
    ordre = [d for d in sorted(series_fonds.values(), key=lambda d: -d["prets"])
             if len(d["tomes"]) >= 2]
    a_chercher = sum(len(set(range(1, max(d["tomes"]) + 1)) - d["tomes"])
                     for d in ordre)
    print(f"  {len(ordre)} séries à examiner, {a_chercher} tomes manquants "
          f"à rechercher.")
    nb = 0
    for num_serie, d in enumerate(ordre, 1):
        if num_serie % 20 == 0:
            print(f"  ... série {num_serie}/{len(ordre)} "
                  f"({d['serie'][:30]!r}) — {nb} tomes retenus — "
                  f"{panier.total:.0f} € — {_nb_requetes} requêtes")
        tomes = d["tomes"]
        manquants = sorted(set(range(1, max(tomes) + 1)) - tomes)
        for n in manquants:
            if panier.total >= BUDGET * 0.995:
                return
            for isbn in rechercher(f"{d['serie']} tome {n}"):
                f = fiche(isbn)
                if (f and str(f.get("tome")) == str(n)
                        and _similarite(f.get("serie", ""), d["serie"]) >= 0.6
                        and panier.ajouter(
                            d["categorie"], f, isbn, "Tome manquant",
                            f"Série « {d['serie']} » ({d['prets']} prêts) : "
                            f"tome {n} absent du fonds")):
                    nb += 1
                    if nb % 20 == 0:
                        print(f"  {nb} tomes ajoutés — {panier.total:.0f} €")
                    break
    print(f"  Phase 1 terminée : {nb} tomes, total {panier.total:.2f} €")


def phase_suggestions(panier, conn):
    print("\n─── PHASE 2 : suggestions de la veille ───")
    nb = 0
    for titre, auteur, isbn, prix, motif, source in conn.execute(
            "SELECT titre, auteur, isbn, prix, motif, source "
            "FROM suggestion_acquisition "
            "WHERE statut IN ('à étudier', 'à commander') "
            "ORDER BY date_ajout DESC").fetchall():
        if panier.total >= BUDGET * 0.995:
            break
        texte = f"{motif or ''} {source or ''}".lower()
        if not any(m in texte for m in (
                "jeunesse", "ado", "enfant", "album", "bd", "manga",
                "première lecture", "premiere lecture", "kibookin")):
            continue
        candidats = ([str(isbn)] if isbn and re.fullmatch(r"\d{9,13}", str(isbn))
                     else rechercher(f"{titre} {(auteur or '').split(',')[0]}"))
        for cand in candidats:
            f = fiche(cand)
            if not f or _similarite(f.get("titre", ""), titre) < 0.5:
                continue
            if f.get("prix") is None and prix:
                f["prix"] = float(prix)
            if panier.ajouter("(veille)", f, cand, "Suggestion veille",
                              (motif or "")[:250]):
                nb += 1
                break
    print(f"  Phase 2 terminée : +{nb} titres, total {panier.total:.2f} €")


def phase_selections(panier):
    print("\n─── PHASE 3 : sélections jeunesse Place des Libraires ───")
    try:
        import veille_place_des_libraires as veille
        selections = veille.recuperer_selections()
    except Exception as e:
        print(f"  (sélections indisponibles : {e})")
        return
    jeunesse = [(n, u) for n, u in selections
                if re.search(r"jeunesse|enfant|\bado", n, re.I)]
    nb = 0
    for nom, url in jeunesse:
        if panier.total >= BUDGET * 0.995:
            break
        print(f"  • {nom}")
        try:
            isbns = veille.isbns_de_page(url)
        except Exception:
            continue
        for isbn in isbns:
            if panier.total >= BUDGET * 0.995:
                break
            f = fiche(isbn)
            if f and panier.ajouter("(sélection)", f, isbn,
                                    "Sélection libraires", nom[:250]):
                nb += 1
    print(f"  Phase 3 terminée : +{nb} titres, total {panier.total:.2f} €")


def phase_tomes_suivants(panier, series_fonds):
    """Complément budget : la SUITE des séries (tome max+1, max+2) que le
    fonds possède déjà, séries les plus empruntées d'abord."""
    print("\n─── PHASE 4 : tomes suivants des séries les plus empruntées ───")
    ordre = [d for d in sorted(series_fonds.values(), key=lambda d: -d["prets"])
             if len(d["tomes"]) >= 3 and d["prets"] > 0]
    nb = 0
    for num_serie, d in enumerate(ordre, 1):
        if panier.total >= BUDGET * 0.995:
            break
        if num_serie % 25 == 0:
            print(f"  ... série {num_serie}/{len(ordre)} — {nb} suites "
                  f"retenues — {panier.total:.0f} €")
        dernier = max(d["tomes"])
        for n in range(dernier + 1, dernier + 7):   # jusqu'à 6 tomes d'avance
            if panier.total >= BUDGET * 0.995:
                break
            trouve = False
            for isbn in rechercher(f"{d['serie']} tome {n}"):
                f = fiche(isbn)
                if (f and str(f.get("tome")) == str(n)
                        and _similarite(f.get("serie", ""), d["serie"]) >= 0.6
                        and panier.ajouter(
                            d["categorie"], f, isbn, "Tome suivant",
                            f"Suite de « {d['serie']} » ({d['prets']} prêts), "
                            f"dernier tome en fonds : {dernier}")):
                    nb += 1
                    trouve = True
                    break
            if not trouve:
                break   # le tome n n'existe pas : inutile de chercher n+1
    print(f"  Phase 4 terminée : +{nb} suites, total {panier.total:.2f} €")


# ── sorties ─────────────────────────────────────────────────────────────────

def ecrire_sorties(panier):
    jour = datetime.date.today().isoformat()
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Commande"
    entetes = ["Catégorie", "Série", "Tome", "Titre", "Auteur", "Éditeur",
               "EAN", "Prix (€)", "Origine", "Motif"]
    thin = Side(style="thin", color="D9E2F0")
    bord = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c, nom in enumerate(entetes, 1):
        cell = ws.cell(row=1, column=c, value=nom)
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2E4A7A")
        cell.border = bord
        ws.column_dimensions[get_column_letter(c)].width = (
            40 if nom in ("Titre", "Motif") else 16)
    lignes = sorted(panier.lignes,
                    key=lambda l: (l["origine"], l["categorie"],
                                   _norm(l["serie"]), str(l["tome"]).zfill(3)))
    for i, l in enumerate(lignes, start=2):
        for c, cle in enumerate(("categorie", "serie", "tome", "titre",
                                 "auteur", "editeur", "ean", "prix",
                                 "origine", "motif"), 1):
            cell = ws.cell(row=i, column=c, value=l[cle])
            cell.font = Font(name="Arial", size=9.5)
            cell.border = bord
    ws.freeze_panes = "A2"
    ws.cell(row=len(lignes) + 3, column=7, value="TOTAL").font = Font(bold=True)
    ws.cell(row=len(lignes) + 3, column=8, value=round(panier.total, 2)).font = Font(bold=True)
    chemin_xlsx = os.path.join(DOSSIER, f"Commande_jeunesse_{jour}.xlsx")
    wb.save(chemin_xlsx)

    chemin_txt = os.path.join(DOSSIER, f"commande_orb_{jour}.txt")
    with open(chemin_txt, "w") as f:
        for l in lignes:
            f.write(l["ean"] + "\n")

    chemin_csv = os.path.join(DOSSIER, f"commande_orb_{jour}.csv")
    with open(chemin_csv, "w", newline="") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(["EAN", "Quantité"])
        for l in lignes:
            w.writerow([l["ean"], 1])

    print("\n─── RÉCAPITULATIF ───")
    par_origine = defaultdict(lambda: [0, 0.0])
    par_categorie = defaultdict(lambda: [0, 0.0])
    for l in panier.lignes:
        par_origine[l["origine"]][0] += 1
        par_origine[l["origine"]][1] += l["prix"]
        par_categorie[l["categorie"]][0] += 1
        par_categorie[l["categorie"]][1] += l["prix"]
    for o, (n, t) in sorted(par_origine.items()):
        print(f"  {n:>4} titres  {t:>9.2f} €  {o}")
    print("  par catégorie :")
    for cat, (n, t) in sorted(par_categorie.items()):
        print(f"    {n:>4}  {t:>9.2f} €  {cat}")
    print(f"\n  TOTAL : {len(panier.lignes)} titres, {panier.total:.2f} € "
          f"(budget {BUDGET:.0f} €)")
    print(f"\n✓ {os.path.basename(chemin_xlsx)}")
    print(f"✓ {os.path.basename(chemin_txt)} (un EAN par ligne)")
    print(f"✓ {os.path.basename(chemin_csv)} (EAN;quantité)")


# ── programme ───────────────────────────────────────────────────────────────

def mode_test():
    print("Test d'extraction du prix sur 3 fiches connues :")
    for isbn in ("9782344042786", "9782278097838", "9782330132576"):
        f = fiche(isbn)
        if f:
            print(f"  {isbn} : {f['titre'][:40]!r} — support {f['support']!r} "
                  f"— prix {f['prix']} €")
        else:
            print(f"  {isbn} : fiche introuvable !")
    print("\nSi les prix sont corrects (compare avec le site), lance le vrai run.")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--test", action="store_true",
                        help="vérifie l'extraction du prix sur 3 fiches puis s'arrête")
    args = parser.parse_args()

    charger_cache()
    if args.test:
        mode_test()
        sauver_cache()
        return

    conn = db.connect()
    print("Chargement du fonds (dédoublonnage)...")
    isbns, series, titres = charger_fonds(conn)
    print(f"  {len(isbns)} EAN en fonds, {len(series)} séries jeunesse/ado, "
          f"{len(titres)} titres.")

    # Documents possédés mais pas encore catalogués (scannés par Thomas) :
    # tout fichier exclusions_*.txt ou « a cataloguer - a couvrir.txt »
    # (un EAN par ligne) est ajouté au dédoublonnage.
    import glob as _glob
    for chemin in _glob.glob(os.path.join(DOSSIER, "exclusions_*.txt")) + \
            _glob.glob(os.path.join(DOSSIER, "a cataloguer*.txt")):
        avec = {l.strip() for l in open(chemin) if l.strip()}
        isbns |= avec
        print(f"  + {len(avec)} EAN exclus via {os.path.basename(chemin)}")

    panier = Panier(isbns, series, titres)
    try:
        phase_tomes_manquants(panier, series)
        phase_suggestions(panier, conn)
        phase_selections(panier)
        phase_tomes_suivants(panier, series)
    finally:
        sauver_cache()
        conn.close()
        if panier.lignes:
            ecrire_sorties(panier)


if __name__ == "__main__":
    main()
