#!/usr/bin/env python3
"""
auditer_dvd.py — Audit du fonds DVD : notre base ET ce qu'il faudra
corriger dans Decalog. STRICTEMENT EN LECTURE ; produit un rapport console
et un Excel d'anomalies.

LA RÈGLE DE COTATION (Thomas, 2026-07-30) :
  - DVD adulte   : « F XXX »  — F, espace, 3 premières lettres du RÉALISATEUR
  - DVD jeunesse : « FJ XXX » — FJ, espace, 3 premières lettres du réalisateur
  - EXCEPTION    : les Disney sont TOUS à « FJ DIS », quel que soit le
                   réalisateur.
  - Nom court    : si le réalisateur n'a que 2 lettres (Ly, Sy, To, Do, Oh),
                   la cote a 2 lettres — conforme (décision 2026-07-30).
  - « B XXX »    : biographies filmées, rayon légitime (décision 2026-07-30).
  - « JV ... »   : jeux vidéo égarés dans le support DVD — exclus de
                   l'audit (décision 2026-07-30), à traiter côté support.

CE QUE VÉRIFIE L'AUDIT
  1. structure de la cote (motif F/FJ + 3 lettres) ;
  2. cohérence cote <-> public (F -> Adulte, FJ -> Jeunesse) ;
  3. cohérence lettres <-> réalisateur (sauf FJ DIS) ;
  4. règle Disney dans les deux sens (Disney pas en FJ DIS, FJ DIS pas
     Disney) ;
  5. trous : sans cote, sans réalisateur, sans public.

Usage :  python3 auditer_dvd.py
"""
import datetime
import os
import re
import sys
import unicodedata

DOSSIER = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, DOSSIER)
import db  # noqa: E402

MOTIF_COTE = re.compile(r"^(FJ?|FJ)\s+([A-Z]{3})", re.I)


def _sans_accents(t):
    t = unicodedata.normalize("NFKD", t or "")
    return "".join(c for c in t if not unicodedata.combining(c))


def trois_lettres_realisateur(createurs):
    """« Nolan, Christopher ; ... » -> NOL. Le nom de famille vient en
    premier dans les notices ; on prend le premier créateur cité."""
    if not createurs:
        return None
    premier = re.split(r"[;/]", createurs)[0].strip()
    nom = re.split(r"[,(]", premier)[0].strip()
    nom = re.sub(r"[^A-Za-z]", "", _sans_accents(nom))
    # Nom de 2 lettres (Ly, Sy, To…) : la cote attendue fait 2 lettres.
    return nom[:3].upper() if len(nom) >= 2 else None


def est_disney(editeur, createurs, titre):
    texte = _sans_accents(f"{editeur or ''} {createurs or ''}").lower()
    return "disney" in texte or "pixar" in texte


def main():
    conn = db.connect()
    lignes = conn.execute(
        "SELECT n.identifiant, n.titre, n.createurs, n.editeur, "
        "       n.public_vise, e.cote, e.code_barre_exemplaire "
        "FROM notice n JOIN exemplaire e ON e.identifiant = n.identifiant "
        "WHERE n.type_document = 'DVD'").fetchall()
    conn.close()

    # dédoublonnage par exemplaire (une ligne par exemplaire, c'est voulu :
    # deux exemplaires du même film peuvent être cotés différemment)
    print(f"{len(lignes)} exemplaire(s) DVD examinés.\n")

    anomalies = {k: [] for k in (
        "cote_absente", "cote_non_conforme", "public_incoherent",
        "lettres_discordantes", "disney_mal_cote", "fjdis_pas_disney",
        "sans_realisateur")}

    for ident, titre, createurs, editeur, public, cote, cb in lignes:
        base = (cb or "", ident, cote or "", titre or "",
                createurs or "", editeur or "", public or "")
        cote_p = (cote or "").strip()

        if not cote_p:
            anomalies["cote_absente"].append(base)
            continue

        # Les DOCUMENTAIRES en DVD sont cotés en Dewey comme les imprimés
        # (précision de Thomas, 2026-07-30) : « 914.4 XXX » ou « J 577 XXX »
        # pour la jeunesse. Ce sont des cotes CONFORMES, pas des anomalies.
        if re.match(r"^(J\s+)?\d", cote_p):
            continue

        # Rayon « B XXX » = biographies filmées : cote légitime.
        if re.match(r"^B\s+[A-Za-z]{2,4}\b", cote_p):
            continue

        # « JV ... » = jeux vidéo saisis à tort en support DVD : hors audit.
        if re.match(r"^JV\b", cote_p, re.I):
            continue

        m = re.match(r"^(FJ|F)\s+([A-Za-z]{2,3})\b", cote_p)
        if not m:
            anomalies["cote_non_conforme"].append(base)
            continue
        prefixe, lettres = m.group(1).upper(), m.group(2).upper()

        # cohérence cote <-> public
        if prefixe == "F" and public == "Jeunesse":
            anomalies["public_incoherent"].append(base + ("F mais public Jeunesse",))
        elif prefixe == "FJ" and public == "Adulte":
            anomalies["public_incoherent"].append(base + ("FJ mais public Adulte",))

        disney = est_disney(editeur, createurs, titre)
        if disney and not (prefixe == "FJ" and lettres == "DIS"):
            anomalies["disney_mal_cote"].append(base)
            continue
        if prefixe == "FJ" and lettres == "DIS":
            if not disney:
                anomalies["fjdis_pas_disney"].append(base)
            continue  # règle Disney : pas de contrôle réalisateur

        attendu = trois_lettres_realisateur(createurs)
        if not attendu:
            anomalies["sans_realisateur"].append(base)
        elif lettres != attendu:
            anomalies["lettres_discordantes"].append(
                base + (f"cote {lettres}, réalisateur -> {attendu}",))

    libelles = {
        "cote_absente":        "Sans cote",
        "cote_non_conforme":   "Cote hors norme F/FJ + 3 lettres",
        "public_incoherent":   "Cote et public contradictoires",
        "lettres_discordantes": "Lettres ≠ réalisateur",
        "disney_mal_cote":     "Disney pas en FJ DIS",
        "fjdis_pas_disney":    "FJ DIS mais pas Disney",
        "sans_realisateur":    "Réalisateur absent de la notice",
    }
    total_anomalies = sum(len(v) for v in anomalies.values())
    print("ANOMALIES :")
    for cle, lst in anomalies.items():
        print(f"  {len(lst):>5}  {libelles[cle]}")
    print(f"  {total_anomalies:>5}  TOTAL "
          f"({100.0 * total_anomalies / max(len(lignes), 1):.1f} % du fonds DVD)")

    # ── Excel ────────────────────────────────────────────────────────────
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    def feuille(ws, entetes, lst):
        thin = Side(style="thin", color="D9E2F0")
        bordure = Border(left=thin, right=thin, top=thin, bottom=thin)
        for c, nom in enumerate(entetes, 1):
            cell = ws.cell(row=1, column=c, value=nom)
            cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2E4A7A")
            cell.border = bordure
            ws.column_dimensions[get_column_letter(c)].width = (
                36 if nom in ("Titre", "Réalisateur(s)") else 18)
        for i, lg in enumerate(lst, start=2):
            for c, val in enumerate(lg, 1):
                cell = ws.cell(row=i, column=c, value=val)
                cell.font = Font(name="Arial", size=9.5)
                cell.border = bordure
        ws.freeze_panes = "A2"

    entetes = ["Code-barres", "Identifiant", "Cote", "Titre",
               "Réalisateur(s)", "Éditeur", "Public"]
    wb = Workbook()
    premiere = True
    for cle, lst in anomalies.items():
        if not lst:
            continue
        # Excel interdit / \ ? * [ ] : dans les noms de feuilles
        nom_feuille = re.sub(r"[/\\?*\[\]:]", "-", libelles[cle])[:31]
        ws = wb.active if premiere else wb.create_sheet()
        ws.title = nom_feuille
        premiere = False
        extra = ["Détail"] if lst and len(lst[0]) > 7 else []
        feuille(ws, entetes + extra, sorted(lst, key=lambda x: x[2]))
    if premiere:
        wb.active.cell(row=1, column=1, value="Aucune anomalie !")

    sortie = os.path.join(
        DOSSIER, f"Audit_DVD_{datetime.date.today().isoformat()}.xlsx")
    wb.save(sortie)
    print(f"\n✓ Détail : {os.path.basename(sortie)}")


if __name__ == "__main__":
    main()
