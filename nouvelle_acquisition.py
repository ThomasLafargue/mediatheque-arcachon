#!/usr/bin/env python3
"""
Prépare une proposition d'acquisitions ET une proposition de désherbage à
partir de l'état actuel de l'Inventaire :
- détecte les tomes manquants dans les séries déjà présentes au catalogue
- sélectionne une combinaison dans le budget donné, sans aucun doublon
- priorise les séries presque complètes et déjà bien fournies (signal de popularité)
- détecte des candidats au désherbage (âge croisé avec l'intérêt probable)
- écrit deux nouveaux onglets datés du jour :
    "Commande du JJ-MM-AAAA"               (à valider, suivi des réceptions)
    "Proposition desherbage du JJ-MM-AAAA" (à valider, suivi des sorties)

Ces deux onglets ont chacun une colonne de suivi (Reçu / Sorti + date), que
recherche_isbn.py coche automatiquement quand vous traitez ensuite
isbn_nouveautés.txt (réception) ou isbn_pilons.txt (sortie du fonds).

Limite connue : les prix sont des ESTIMATIONS par type de document (aucune
source de prix fiable n'est accessible automatiquement). Les suggestions de
nouveautés "tendance" et la vérification des prix réels se font en discussion
directe, pas via ce script.

Usage : python3 nouvelle_acquisition.py
"""

import os, sys, datetime
from collections import defaultdict, Counter

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

script_dir = os.path.dirname(os.path.abspath(__file__))
output_file = os.path.join(script_dir, "inventaire_mediatheque.xlsx")

PRIX_ESTIME = {
    "Manga": 7.20, "BD": 12.50, "Roman jeunesse": 14.00, "Roman ado / YA": 15.50,
    "Album": 13.00, "Documentaire": 15.00, "Conte / Poésie": 12.00, "Première lecture": 9.00,
}
PRIX_DEFAUT = 12.00

C_HEADER = "2E4A7A"
C_OK     = "2E7D46"
C_WARN   = "B45309"
def hfill(h): return PatternFill("solid", fgColor=h)
thin = Side(style="thin", color="B8CCE4")
brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

print("="*65)
print("  Nouvelle vague d'acquisitions — Médiathèque d'Arcachon — v2.0")
print("="*65)

if not os.path.exists(output_file):
    print(f"\n⚠  Fichier introuvable : {output_file}")
    input("\nEntrée pour quitter...")
    sys.exit(1)

wb = load_workbook(output_file)
ws = wb["Inventaire"] if "Inventaire" in wb.sheetnames else wb.active

def majoritaire(valeurs):
    c = Counter(v for v in valeurs if v)
    return c.most_common(1)[0][0] if c else ""

# ── 1. Détection des séries à trous (avec Public/Genre majoritaires) ──
series = defaultdict(list)
for row in ws.iter_rows(min_row=3, values_only=True):
    if not row or not row[0]:
        continue
    serie, tome, public, typ, genre, editeur = row[3], row[2], row[8], row[9], row[10], row[6]
    if serie and tome:
        try:
            t = int(str(tome).strip())
            series[serie].append((t, typ, editeur, public, genre))
        except ValueError:
            pass

gaps_par_serie = {}
for serie, items in series.items():
    tomes = sorted(set(t for t, *_ in items))
    if len(tomes) < 2:
        continue
    manquants = [t for t in range(tomes[0], tomes[-1] + 1) if t not in tomes]
    if manquants:
        typ     = majoritaire([i[1] for i in items]) or "Roman jeunesse"
        editeur = majoritaire([i[2] for i in items])
        public  = majoritaire([i[3] for i in items])
        genre   = majoritaire([i[4] for i in items])
        nb_possedes = len(tomes)
        ratio = nb_possedes / (nb_possedes + len(manquants))
        prix_unitaire = PRIX_ESTIME.get(typ, PRIX_DEFAUT)
        gaps_par_serie[serie] = {
            "type": typ, "editeur": editeur, "public": public, "genre": genre,
            "nb_possedes": nb_possedes, "ratio": ratio, "prix_unitaire": prix_unitaire,
            "tomes_manquants": manquants,
        }

total_tomes_manquants = sum(len(g["tomes_manquants"]) for g in gaps_par_serie.values())
print(f"\n{len(gaps_par_serie)} série(s) déjà au catalogue avec des tomes manquants")
print(f"({total_tomes_manquants} tomes manquants au total, prix estimés par type).\n")

# ── 2. Budget ──
while True:
    saisie = input("Budget disponible pour cette commande (€) : ").strip().replace(",", ".")
    try:
        budget = float(saisie)
        break
    except ValueError:
        print("  Merci d'indiquer un nombre, par exemple : 1000")

# ── 3. Sélection : compléter des séries ENTIÈRES en priorité ──
series_triees = sorted(gaps_par_serie.items(), key=lambda kv: (-kv[1]["ratio"], -kv[1]["nb_possedes"]))

selection = []
total = 0.0
series_completees = []
series_non_retenues = []

def ajouter_selection(serie, info, tomes):
    for t in tomes:
        selection.append({
            "serie": serie, "tome": t, "type": info["type"], "editeur": info["editeur"],
            "public": info["public"], "genre": info["genre"], "prix": info["prix_unitaire"],
        })

for serie, info in series_triees:
    cout = len(info["tomes_manquants"]) * info["prix_unitaire"]
    if total + cout <= budget:
        ajouter_selection(serie, info, info["tomes_manquants"])
        total += cout
        series_completees.append(serie)
    else:
        series_non_retenues.append((serie, info))

series_non_retenues.sort(key=lambda kv: kv[1]["prix_unitaire"])
for serie, info in series_non_retenues:
    tomes_pris = []
    for t in info["tomes_manquants"]:
        if total + info["prix_unitaire"] > budget:
            break
        tomes_pris.append(t)
        total += info["prix_unitaire"]
    if tomes_pris:
        ajouter_selection(serie, info, tomes_pris)

print(f"✓ {len(selection)} tome(s) sélectionné(s) pour {total:.2f} € sur {budget:.2f} € de budget")
print(f"  ({len(series_completees)} série(s) complétée(s) entièrement)\n")

# ── 4. Stats de l'inventaire actuel ──
compteur_inventaire = Counter()
total_inventaire = 0
for row in ws.iter_rows(min_row=3, values_only=True):
    if row and row[0]:
        total_inventaire += 1
        if row[9]:
            compteur_inventaire[str(row[9]).strip()] += 1

# ── 5. Candidats au désherbage ──
SEUIL_SERIE_INSTALLEE = 8
nb_tomes_serie = Counter()
for row in ws.iter_rows(min_row=3, values_only=True):
    if row and row[3]:
        nb_tomes_serie[str(row[3]).strip()] += 1

docs_avec_annee = []
for row in ws.iter_rows(min_row=3, values_only=True):
    if row and row[0] and row[11]:
        try:
            an = int(str(row[11]).strip())
            serie = str(row[3]).strip() if row[3] else ""
            docs_avec_annee.append({
                "annee": an, "isbn": str(row[0]).strip(), "titre": row[1] or "",
                "type": row[9] or "", "serie": serie, "public": row[8] or "", "genre": row[10] or "",
                "taille_serie": nb_tomes_serie.get(serie, 0) if serie else 0,
            })
        except ValueError:
            pass

annee_courante_acq = datetime.date.today().year
SEUIL_AGE_DOCUMENTAIRE = 8
SEUIL_AGE_AUTRES       = 15

documentaires_anciens = sorted(
    [d for d in docs_avec_annee if d["type"] == "Documentaire"
     and annee_courante_acq - d["annee"] >= SEUIL_AGE_DOCUMENTAIRE],
    key=lambda d: d["annee"])[:25]

autres_candidats = sorted(
    [d for d in docs_avec_annee if d["type"] != "Documentaire" and d["taille_serie"] < SEUIL_SERIE_INSTALLEE
     and annee_courante_acq - d["annee"] >= SEUIL_AGE_AUTRES],
    key=lambda d: d["annee"])[:25]

a_examiner_prudence = sorted(
    [d for d in docs_avec_annee if d["type"] != "Documentaire" and d["taille_serie"] >= SEUIL_SERIE_INSTALLEE
     and annee_courante_acq - d["annee"] >= SEUIL_AGE_AUTRES],
    key=lambda d: d["annee"])[:15]

tous_candidats_desherbage = documentaires_anciens + autres_candidats + a_examiner_prudence

# ── 6. Fonctions communes d'écriture ──
date_str = datetime.date.today().strftime("%d-%m-%Y")

def nom_disponible(base):
    nom = base
    n = 1
    while nom in wb.sheetnames:
        n += 1
        nom = f"{base} ({n})"
    return nom

def bandeau(feuille, ri, titre, couleur, largeur_fusion):
    feuille.merge_cells(f"A{ri}:{get_column_letter(largeur_fusion)}{ri}")
    c = feuille.cell(row=ri, column=1, value=titre)
    c.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    c.fill = hfill(couleur)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    feuille.row_dimensions[ri].height = 20
    return ri + 1

def section_repartition(feuille, ri, titre, couleur, compteur_local, total_local, largeur_fusion):
    ri = bandeau(feuille, ri, titre, couleur, largeur_fusion)
    for cle, n in sorted(compteur_local.items(), key=lambda x: -x[1]):
        if not cle:
            continue
        feuille.cell(row=ri, column=1, value=cle).font = Font(name="Arial", size=10)
        feuille.cell(row=ri, column=2, value=n).font = Font(name="Arial", size=10)
        pct = f"{n/total_local*100:.1f} %" if total_local else ""
        c = feuille.cell(row=ri, column=3, value=pct)
        c.font = Font(name="Arial", size=9, italic=True, color="808080")
        ri += 1
    return ri + 1

# ══════════════════════════════════════════════════════════════
# ONGLET 1 : Commande du JJ-MM-AAAA
# ══════════════════════════════════════════════════════════════
nom_cmd = nom_disponible(f"Commande du {date_str}")
ws_cmd = wb.create_sheet(nom_cmd)

ri = 1
ws_cmd.merge_cells(f"A{ri}:I{ri}")
c = ws_cmd.cell(row=ri, column=1, value=f"Proposition d'acquisitions — {date_str}")
c.font = Font(name="Arial", size=13, bold=True, color="FFFFFF")
c.fill = hfill(C_HEADER)
c.alignment = Alignment(horizontal="center", vertical="center")
ws_cmd.row_dimensions[ri].height = 26
ri += 2

ri = bandeau(ws_cmd, ri, "Budget", C_HEADER, 9)
for label, val in [("Budget alloué", f"{budget:.2f} €"),
                    ("Total estimé de la sélection", f"{total:.2f} €"),
                    ("Reliquat estimé", f"{budget-total:.2f} €"),
                    ("Budget réel (à remplir une fois les prix confirmés)", ""),
                    ("Tomes sélectionnés", len(selection)),
                    ("Séries complétées entièrement", len(series_completees))]:
    ws_cmd.cell(row=ri, column=1, value=label).font = Font(name="Arial", size=10)
    cell_val = ws_cmd.cell(row=ri, column=2, value=val)
    cell_val.font = Font(name="Arial", size=10, bold=True)
    if "Budget réel" in label:
        cell_val.fill = hfill("FFEB9C")
    ri += 1
ri += 1

compteur_sel_type = Counter(s["type"] for s in selection)
compteur_sel_public = Counter(s["public"] for s in selection)
compteur_sel_genre = Counter(s["genre"] for s in selection)
ri = section_repartition(ws_cmd, ri, "Sélection — répartition par type", C_HEADER, compteur_sel_type, len(selection), 9)
ri = section_repartition(ws_cmd, ri, "Sélection — répartition par public", C_HEADER, compteur_sel_public, len(selection), 9)
ri = section_repartition(ws_cmd, ri, "Sélection — répartition par genre", C_HEADER, compteur_sel_genre, len(selection), 9)

ri = bandeau(ws_cmd, ri, "Tomes à commander — séries déjà au catalogue", C_OK, 9)
entetes = ["Série", "Tome manquant", "Type", "Éditeur", "Public", "Genre", "Prix estimé", "Reçu ?", "Date de réception"]
for ci, h in enumerate(entetes, 1):
    cell = ws_cmd.cell(row=ri, column=ci, value=h)
    cell.font = Font(name="Arial", size=10, bold=True)
    cell.border = brd
ri += 1
for s in sorted(selection, key=lambda s: (s["serie"], s["tome"])):
    vals = [s["serie"], s["tome"], s["type"], s["editeur"], s["public"], s["genre"], s["prix"], "", ""]
    for ci, v in enumerate(vals, 1):
        cell = ws_cmd.cell(row=ri, column=ci, value=v)
        cell.font = Font(name="Arial", size=10)
        cell.border = brd
        if ci == 7:
            cell.number_format = "0.00 €"
    ri += 1
ri += 1

ri = bandeau(ws_cmd, ri, "Nouveautés à ajouter — à compléter avec Claude au moment de finaliser", C_OK, 9)
ws_cmd.cell(row=ri, column=1, value="(section vide : se remplit en discussion, selon les tendances du moment, "
                                     "le thème sciences/lecture de l'année scolaire, et le public estival)").font = \
    Font(name="Arial", size=9.5, italic=True, color="808080")
ri += 2

for col, larg in zip("ABCDEFGHI", [28, 14, 16, 20, 14, 18, 12, 9, 16]):
    ws_cmd.column_dimensions[col].width = larg
ws_cmd.freeze_panes = "A1"

# ══════════════════════════════════════════════════════════════
# ONGLET 2 : Proposition desherbage du JJ-MM-AAAA
# ══════════════════════════════════════════════════════════════
nom_desh = nom_disponible(f"Désherbage du {date_str}")
ws_desh = wb.create_sheet(nom_desh)

ri = 1
ws_desh.merge_cells(f"A{ri}:I{ri}")
c = ws_desh.cell(row=ri, column=1, value=f"Proposition de désherbage — {date_str}")
c.font = Font(name="Arial", size=13, bold=True, color="FFFFFF")
c.fill = hfill(C_WARN)
c.alignment = Alignment(horizontal="center", vertical="center")
ws_desh.row_dimensions[ri].height = 26
ri += 2

ri = bandeau(ws_desh, ri, "Récapitulatif", C_WARN, 9)
for label, val in [("Candidats détectés", len(tous_candidats_desherbage)),
                    ("Documentaires anciens (priorité)", len(documentaires_anciens)),
                    ("Autres anciens, séries peu installées", len(autres_candidats)),
                    ("Anciens mais grande série installée (prudence)", len(a_examiner_prudence))]:
    ws_desh.cell(row=ri, column=1, value=label).font = Font(name="Arial", size=10)
    ws_desh.cell(row=ri, column=2, value=val).font = Font(name="Arial", size=10, bold=True)
    ri += 1
ri += 1

compteur_desh_type = Counter(d["type"] for d in tous_candidats_desherbage)
compteur_desh_public = Counter(d["public"] for d in tous_candidats_desherbage)
compteur_desh_genre = Counter(d["genre"] for d in tous_candidats_desherbage)
ri = section_repartition(ws_desh, ri, "Candidats — répartition par type", C_WARN, compteur_desh_type, len(tous_candidats_desherbage), 9)
ri = section_repartition(ws_desh, ri, "Candidats — répartition par public", C_WARN, compteur_desh_public, len(tous_candidats_desherbage), 9)
ri = section_repartition(ws_desh, ri, "Candidats — répartition par genre", C_WARN, compteur_desh_genre, len(tous_candidats_desherbage), 9)

entetes_d = ["ISBN", "Titre", "Type", "Série", "Public", "Genre", "Année", "Sorti ?", "Date de sortie"]

def table_desherbage(feuille, ri, titre, liste, couleur):
    ri = bandeau(feuille, ri, titre, couleur, 9)
    for ci, h in enumerate(entetes_d, 1):
        cell = feuille.cell(row=ri, column=ci, value=h)
        cell.font = Font(name="Arial", size=10, bold=True)
        cell.border = brd
    ri += 1
    for d in liste:
        vals = [d["isbn"], d["titre"], d["type"], d["serie"], d["public"], d["genre"], d["annee"], "", ""]
        for ci, v in enumerate(vals, 1):
            cell = feuille.cell(row=ri, column=ci, value=v)
            cell.font = Font(name="Arial", size=10)
            cell.border = brd
        ri += 1
    return ri + 1

ri = table_desherbage(ws_desh, ri, "Documentaires anciens (priorité : contenu factuel possiblement dépassé)",
                       documentaires_anciens, C_WARN)
ri = table_desherbage(ws_desh, ri, "Autres anciens, séries peu installées (candidats plausibles)",
                       autres_candidats, C_WARN)
ri = table_desherbage(ws_desh, ri, "Anciens mais grande série installée (à examiner avec prudence, probablement encore demandé)",
                       a_examiner_prudence, C_WARN)

for col, larg in zip("ABCDEFGHI", [16, 38, 16, 20, 14, 18, 8, 9, 16]):
    ws_desh.column_dimensions[col].width = larg
ws_desh.freeze_panes = "A1"

wb.save(output_file)

print(f"✓ Onglets « {nom_cmd} » et « {nom_desh} » créés.")
print("="*65)
print(f"\n  Commandes utiles :")
print(f"    python3 recherche_isbn.py        → lancer une nouvelle recherche")
print(f"    python3 maj_statistiques.py      → mettre à jour les stats sans recherche")
print(f"    python3 nouvelle_acquisition.py  → préparer une nouvelle commande")
print(f"    python3 annuler_derniere_commande.py → annuler la dernière commande/désherbage")
print(f"    caffeinate                       → empêcher le Mac de dormir (autre onglet)")
print("="*65)
input("\nEntrée pour quitter...")
