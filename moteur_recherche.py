
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Outil de recherche bibliographique par ISBN v8.0
Médiathèque d'Arcachon - Réseau COBAS

Sources (dans l'ordre) :
  1.  BnF           - Notices officielles françaises
  2.  Babelio        - Littérature française/jeunesse
  3.  BDfugue        - Spécialiste BD/Manga
  4.  Manga News     - Spécialiste manga + PEGI
  5.  Fnac           - Large couverture + public cible
  6.  Amazon.fr      - Très large + résumés + PEGI
  7.  Cultura        - Jeunesse + public cible
  8.  Decitre        - Librairie française complète
  9.  Ricochet       - Spécialiste littérature jeunesse
 10.  Leslibraires   - Notices complètes
 11.  Mollat         - Librairie française très complète
 12.  Booknode       - Séries/tomes très bien suivis
 13.  Google Books   - Couverture internationale
 14.  Open Library   - Fallback international

Logique : cherche sur TOUTES les sources, enrichit champ par champ
          s'arrête seulement quand tout est rempli
"""

import urllib.parse, xml.etree.ElementTree as ET
import time, os, sys, re, unicodedata, datetime
from collections import Counter, defaultdict

try:
    import requests
    from bs4 import BeautifulSoup
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install",
                           "openpyxl", "requests", "beautifulsoup4"])
    import requests
    from bs4 import BeautifulSoup
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Accept-Language": "fr-FR,fr;q=0.9",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Referer": "https://www.google.fr/",
}

EMPTY = {
    "titre":"","auteur":"","illustrateur":"","editeur":"","annee":"",
    "type":"","public":"","genre":"","serie":"","tome":"","pegi":"",
    "collection":"","resume":"","source":"","statut":"NON TROUVÉ"
}

# ─────────────────────────────────────────────────────────
# CORRECTIONS MANUELLES — pour les cas mal gérés par les sources
# Ajouter ici les ISBN qui posent problème avec leurs données correctes
# ─────────────────────────────────────────────────────────
CORRECTIONS = {
    "9782226443410": {
        "titre":        "Archibald - Ma maison",
        "serie":        "Archibald",
        "auteur":       "Astrid Desbordes",
        "illustrateur": "Pauline Martin",
        "editeur":      "Albin Michel",
        "annee":        "2019",
        "type":         "Album",
        "public":       "Dès 3 ans",
        "genre":        "Vie quotidienne",
        "pegi":         "",
        "tome":         "",
    },
    "9789055795789": {
        "titre":        "Les Aventures de Tom Pouce",
        "auteur":       "Meyer M.",
        "editeur":      "N/C",
        "annee":        "N/C",
        "type":         "Conte / Poésie",
        "public":       "Jeunesse",
        "genre":        "Conte / Mythe",
        "pegi":         "",
        "tome":         "",
        "serie":        "",
    },
}


CHAMPS_CLES = ["titre","auteur","editeur","annee","type","public"]

# ─────────────────────────────────────────────────────────
# TABLES DE CORRESPONDANCE
# ─────────────────────────────────────────────────────────

# ─────────────────────────────────────────────────────────
# NORMALISATION DES ÉDITEURS
# Uniformise les variantes d'un même éditeur vers une forme canonique
# ─────────────────────────────────────────────────────────
NORMALISATION_EDITEURS = {
    # Gallimard
    "gallimard-jeunesse":           "Gallimard Jeunesse",
    "éditions gallimard jeunesse":  "Gallimard Jeunesse",
    "editions gallimard jeunesse":  "Gallimard Jeunesse",
    "gallimard jeunesse musique":   "Gallimard Jeunesse",
    "gallimard jeun.":              "Gallimard Jeunesse",
    "nrf gallimard":                "Gallimard",
    "folio gallimard":              "Gallimard Jeunesse",
    # École des loisirs
    "l'ecole des loisirs":          "École des Loisirs",
    "l'école des loisirs":          "École des Loisirs",
    "ecole des loisirs":            "École des Loisirs",
    "l école des loisirs":          "École des Loisirs",
    # Bayard
    "bayard jeunesse":              "Bayard",
    "bayard editions":              "Bayard",
    "bayard éditions":              "Bayard",
    # Milan
    "milan jeunesse":               "Milan",
    "éditions milan":               "Milan",
    "editions milan":               "Milan",
    # Albin Michel
    "albin michel jeunesse":        "Albin Michel",
    "editions albin michel":        "Albin Michel",
    "éditions albin michel":        "Albin Michel",
    # Hachette
    "hachette jeunesse":            "Hachette",
    "hachette livre":               "Hachette",
    "hachette pratique":            "Hachette",
    "hachette roman":               "Hachette",
    # Flammarion
    "flammarion jeunesse":          "Flammarion",
    "père castor flammarion":       "Flammarion",
    "pere castor flammarion":       "Flammarion",
    # Nathan
    "nathan jeunesse":              "Nathan",
    "editions nathan":              "Nathan",
    "éditions nathan":              "Nathan",
    # Actes Sud
    "actes sud junior":             "Actes Sud",
    "actes sud j.":                 "Actes Sud",
    # Pika
    "pika édition":                 "Pika",
    "pika edition":                 "Pika",
    "pika editions":                "Pika",
    # Kana
    "kana manga":                   "Kana",
    # Dupuis
    "editions dupuis":              "Dupuis",
    "éditions dupuis":              "Dupuis",
    # Dargaud
    "editions dargaud":             "Dargaud",
    "éditions dargaud":             "Dargaud",
    # Casterman
    "casterman jeunesse":           "Casterman",
    "editions casterman":           "Casterman",
    # Le Lombard
    "editions du lombard":          "Le Lombard",
    "éditions du lombard":          "Le Lombard",
    # Sarbacane
    "editions sarbacane":           "Sarbacane",
    "éditions sarbacane":           "Sarbacane",
    # Kaléidoscope
    "kaleidoscope":                 "Kaléidoscope",
    "editions kaleidoscope":        "Kaléidoscope",
    # Didier
    "didier jeunesse":              "Didier",
    "editions didier":              "Didier",
    # Pastel
    "l'école des loisirs / pastel": "Pastel",
    # Auzou
    "editions auzou":               "Auzou",
    "éditions auzou":               "Auzou",
    # La Martinière
    "la martinière jeunesse":       "La Martinière",
    "éditions de la martinière":    "La Martinière",
    # Glénat
    "glenat":                       "Glénat",
    "editions glenat":              "Glénat",
    "éditions glénat":              "Glénat",
    # Bamboo
    "bamboo edition":               "Bamboo",
    "bamboo édition":               "Bamboo",
    # Folio (rattaché à Gallimard)
    "folio junior":                 "Gallimard Jeunesse",
    "folio cadet":                  "Gallimard Jeunesse",
    "folio benjamin":               "Gallimard Jeunesse",
}

def normaliser_editeur(editeur_brut):
    """Normalise le nom d'un éditeur vers sa forme canonique."""
    if not editeur_brut:
        return editeur_brut
    # Nettoyage de base
    e = editeur_brut.strip()
    e = e.strip(".,;")
    # Recherche dans la table (insensible à la casse)
    e_lower = e.lower()
    if e_lower in NORMALISATION_EDITEURS:
        return NORMALISATION_EDITEURS[e_lower]
    # Recherche partielle si pas de correspondance exacte
    for variante, canonique in NORMALISATION_EDITEURS.items():
        if variante in e_lower:
            return canonique
    # Première lettre en majuscule si pas trouvé
    return e


def _sans_accents(s):
    """Normalise une chaîne en retirant les accents, pour fiabiliser les
    comparaisons (ex: 'Glénat Manga' doit matcher la clé 'glenat manga')."""
    return ''.join(c for c in unicodedata.normalize('NFKD', s or "") if not unicodedata.combining(c))

EDITEUR_TYPE = {
    "pika":         "Manga", "kana":          "Manga",
    "kurokawa":     "Manga", "tonkam":        "Manga",
    "kazé":         "Manga", "nobi nobi":     "Manga",
    "ki-oon":       "Manga", "akata":         "Manga",
    "glenat manga": "Manga", "bamboo manga":  "Manga",
    "soleil manga": "Manga", "delcourt tonkam":"Manga",
    "le lombard":   "BD",    "dupuis":        "BD",
    "dargaud":      "BD",    "casterman":     "BD",
    "bamboo":       "BD",    "glénat":        "BD",
    "lucky comics": "BD",    "grand angle":   "BD",
    "futuropolis":  "BD",    "vents d'ouest": "BD",
    "400 coups":    "Album",
    "hatier jeunesse": "Album",
    "hatier":       "Album",
    "kaléidoscope": "Album",
    "didier jeunesse": "Album",
    "pastel":       "Album",
    "le nouvel atila": "Première lecture",
    "biosphoto":    "Documentaire",
    "le rouergue":  "Roman jeunesse",
    "seuil jeunesse":"Roman jeunesse",
    "sarbacane":    "Roman jeunesse",
    "actes sud junior":"Roman jeunesse",
    "saxo":         "Roman jeunesse",
    "hélium":       "Roman jeunesse",
    "helium":       "Roman jeunesse",
    "auzou":        "Roman jeunesse",
}

EDITEUR_PUBLIC = {
    "pika":                 "Ado (12+)",
    "kana":                 "Ado (12+)",
    "kurokawa":             "Ado (12+)",
    "ki-oon":               "Ado (12+)",
    "akata":                "Ado (12+)",
    "kazé":                 "Ado (12+)",
    "glenat manga":         "Ado (12+)",
    "école des loisirs":    "Jeunesse",
    "gallimard jeunesse":   "Jeunesse",
    "gallimard":            "Jeunesse",
    "bayard jeunesse":      "Jeunesse",
    "bayard":               "Jeunesse",
    "milan":                "Jeunesse",
    "nathan":               "Jeunesse",
    "hachette jeunesse":    "Jeunesse",
    "hachette":             "Jeunesse",
    "rageot":               "Jeunesse",
    "flammarion jeunesse":  "Jeunesse",
    "actes sud junior":     "Jeunesse",
    "actes sud":            "Jeunesse",
    "sarbacane":            "Jeunesse",
    "albin michel jeunesse":"Jeunesse",
    "albin michel":         "Jeunesse",
    "le lombard":           "Jeunesse",
    "dupuis":               "Jeunesse",
    "dargaud":              "Jeunesse",
    "bayard":               "Jeunesse",
    "editions du ricochet": "Jeunesse",
    "saxo":                 "Jeunesse",
    "helium":               "Jeunesse",
    "hélium":               "Jeunesse",
    "pastel":               "Dès 3 ans",
    "kaléidoscope":         "Dès 3 ans",
    "didier jeunesse":      "Dès 3 ans",
    "nobi nobi":            "Jeunesse",
    "auzou":                "Jeunesse",
    "bamboo":               "Jeunesse",
    "casterman":            "Jeunesse",
    "glénat":               "Jeunesse",
    "hatier jeunesse":      "Jeunesse",
    "hatier":               "Jeunesse",
    "400 coups":            "Jeunesse",
    "la martinière jeunesse":"Jeunesse",
    "la martinière":        "Jeunesse",
    "bios":                 "Jeunesse",
    "biosphoto":            "Jeunesse",
    "seuil jeunesse":       "Jeunesse",
    "seuil":                "Jeunesse",
    "le rouergue":          "Jeunesse",
    "le nouvel atila":      "Première lecture",
    "helium":               "Jeunesse",
    "milan poche":          "Jeunesse",
    "folio cadet":          "Dès 6 ans",
}

COLLECTION_TYPE = {
    "pika shonen":    "Manga", "pika shônen":  "Manga",
    "pika seinen":    "Manga", "pika shojo":   "Manga",
    "kana shonen":    "Manga", "kana shojo":   "Manga",
    "pôle fiction":   "Roman ado / YA",
    "exprim":         "Roman ado / YA",
    "scripto":        "Roman ado / YA",
    "heure noire":    "Roman ado / YA",
    "black moon":     "Roman ado / YA",
    "tribal":         "Roman ado / YA",
    "folio junior":   "Roman jeunesse",
    "folio cadet":    "Roman jeunesse",
    "folio benjamin": "Roman jeunesse",
    "milan poche":    "Roman jeunesse",
    "j'aime lire":    "Première lecture",
    "premiers romans":"Première lecture",
    "les goûters philo":          "Documentaire",
    "les yeux de la découverte":  "Documentaire",
    "mes premières découvertes":  "Documentaire",
    "mes p'tites questions":      "Documentaire",
    "pocqq":                      "Documentaire",
}

COLLECTION_PUBLIC = {
    "folio cadet":          "Dès 6 ans",
    "folio benjamin":       "Dès 6 ans",
    "folio junior":         "8-12 ans",
    "folio ado":            "Ado (12+)",
    "milan poche cadet":    "Dès 6 ans",
    "milan poche junior":   "8-12 ans",
    "pôle fiction":         "Ado (12+)",
    "exprim":               "Ado (12+)",
    "scripto":              "Ado (12+)",
    "heure noire":          "Ado (12+)",
    "black moon":           "Ado (12+)",
    "tribal":               "Ado (12+)",
    "pika shonen":          "Ado (12+)",
    "pika shônen":          "Ado (12+)",
    "pika seinen":          "Ado (12+)",
    "les yeux de la découverte": "8-12 ans",
    "mes premières découvertes": "Dès 6 ans",
    "mes p'tites questions":     "Dès 6 ans",
    "les goûters philo":         "8-12 ans",
    "pocqq":                     "8-12 ans",
    "éclats de rire":            "Dès 6 ans",
    "j'aime lire":               "Dès 6 ans",
    "premiers romans":           "Dès 6 ans",
    "cascade":                   "8-12 ans",
    "folio cadet":               "Dès 6 ans",
}

# ─────────────────────────────────────────────────────────
# DÉTECTION TYPE
# ─────────────────────────────────────────────────────────

TYPES_MOTS = [
    ("Manga",            ["manga","manhwa","manhua","shonen","shônen",
                          "shojo","shôjo","seinen","josei","kodomo"]),
    ("BD",               ["bande dessinée","bande-dessinée","comics",
                          "graphic novel","album bd"]),
    ("Album",            ["album illustré","album jeunesse","album cartonné",
                          "imagier","livre-images","tout-carton"]),
    ("Première lecture", ["première lecture","premières lectures",
                          "je commence à lire","lecture débutant",
                          "facile à lire"]),
    ("Conte / Poésie",   ["conte","contes","fable","fables","poésie",
                          "poème","comptine","légende","mythe",
                          "grimm","perrault","andersen"]),
    ("Documentaire",     ["documentaire","encyclopédie","atlas",
                          "dictionnaire","sciences","découverte",
                          "géographie","goûters philo","comment ça marche",
                          "pourquoi","c'est quoi","kézako"]),
    ("Livre-jeu / Activités", ["livre-jeu","activités","coloriage",
                               "origami","autocollants","escape book",
                               "cahier de vacances","cherche et trouve",
                               "dont vous êtes le héros"]),
    ("Roman ado / YA",   ["young adult"," ya ","exprim","scripto",
                          "pôle fiction","heure noire","black moon"]),
    ("Roman jeunesse",   ["roman jeunesse","roman enfant","folio junior",
                          "milan poche","rageot"]),
]

def detecter_type(titre, collection, resume, genre_raw, support, editeur, public):
    ed = _sans_accents(editeur.lower())
    for ek, tv in EDITEUR_TYPE.items():
        if _sans_accents(ek) in ed:
            return tv
    coll = _sans_accents(collection.lower())
    for ck, tv in COLLECTION_TYPE.items():
        if _sans_accents(ck) in coll:
            return tv
    texte = " ".join([titre, collection, resume, genre_raw, support]).lower()
    for type_nom, mots in TYPES_MOTS:
        if any(m in texte for m in mots):
            if type_nom == "Roman jeunesse" and public in ("Ado (12+)","Ado","Dès 12 ans"):
                return "Roman ado / YA"
            return type_nom
    if "texte" in support.lower():
        return "Roman ado / YA" if public in ("Ado (12+)","Ado") else "Roman jeunesse"
    return ""

# ─────────────────────────────────────────────────────────
# DÉTECTION PUBLIC
# ─────────────────────────────────────────────────────────

def normaliser_public(val):
    """Normalise les valeurs de public vers nos valeurs standard."""
    if not val:
        return ""
    v = val.lower().strip()
    if re.search(r"\b[0-3]\s*[àa]\s*\d+|maternelle|tout-petit|bébé|0-3", v): return "Dès 3 ans"
    if re.search(r"\bdès\s*6\b|6\s*[àa]\s*\d+|cp\b|6-8\b|6 ans", v):         return "Dès 6 ans"
    if re.search(r"\b[6-7]\s*[àa]\s*\d+|ce1\b|7-8\b|7 ans", v):              return "Dès 6 ans"
    if re.search(r"\b(8|9)\s*[àa]\s*\d+|dès\s*8\b|ce2\b|8-10\b|8 ans", v):  return "8-12 ans"
    if re.search(r"\b(10|11)\s*[àa]\s*\d+|dès\s*10\b|10-12\b|10 ans", v):   return "8-12 ans"
    if re.search(r"\bdès\s*12\b|12\s*[àa]\s*\d+|collège|12-16\b|12 ans", v): return "Ado (12+)"
    if re.search(r"\b(13|14|15|16)\s*[àa]\s*\d+|lycée|teen|young adult", v): return "Ado (12+)"
    if re.search(r"\bado\b|ya\b|young adult", v):                              return "Ado (12+)"
    if re.search(r"\bjeunesse\b|enfant|junior", v):                            return "Jeunesse"
    if re.search(r"\d", val):
        return val.strip()
    return ""

def detecter_public(public_code, titre, collection, editeur, tranche_age):
    # 1. Tranche d'âge explicite
    if tranche_age:
        p = normaliser_public(tranche_age)
        if p: return p

    # 2. Code BnF
    CODES = {"j":"Jeunesse","y":"Ado (12+)","a":"Tout public",
             "b":"Jeunesse","c":"Ado (12+)","d":"Adulte"}
    if public_code and public_code.lower() in CODES:
        return CODES[public_code.lower()]

    # 3. Collection
    coll = _sans_accents(collection.lower())
    for ck, pv in COLLECTION_PUBLIC.items():
        if _sans_accents(ck) in coll: return pv

    # 4. Éditeur
    ed = _sans_accents(editeur.lower())
    for ek, pv in EDITEUR_PUBLIC.items():
        if _sans_accents(ek) in ed: return pv

    # 5. Mots-clés titre + collection
    texte = (titre + " " + collection).lower()
    if any(m in texte for m in ["bébé","tout-petit","0-3"]):               return "Dès 3 ans"
    if any(m in texte for m in ["dès 6","6-8","cp ","ce1"]):               return "Dès 6 ans"
    if any(m in texte for m in ["dès 8","8-10","ce2","cm1","cm2"]):        return "8-12 ans"
    if any(m in texte for m in ["dès 10","10-12"]):                         return "8-12 ans"
    if any(m in texte for m in ["dès 12","12-16","collège"]):               return "Ado (12+)"
    if any(m in texte for m in ["ado","teen"," ya ","young adult","exprim",
                                  "pôle fiction","scripto"]):               return "Ado (12+)"
    if any(m in texte for m in ["jeunesse","enfant","poche jeun"]):         return "Jeunesse"
    return ""

# ─────────────────────────────────────────────────────────
# PEGI
# ─────────────────────────────────────────────────────────

def detecter_pegi(type_doc, public, texte_brut=""):
    """Déduit le PEGI pour les mangas et BD."""
    if type_doc not in ("Manga", "BD"):
        return ""
    # Chercher PEGI explicite dans le texte
    m = re.search(r"PEGI\s*(\d+)", texte_brut, re.IGNORECASE)
    if m:
        return m.group(1)  # nombre seul : CAST SQL et tris fonctionnent (2026-07-27)
    # Déduire depuis le public
    if public in ("Ado (12+)", "Dès 12 ans", "12-16 ans"):
        return "12"
    if public in ("Ado", "Ado / YA") or re.search(r"\b(16|17|18)\b", public):
        return "16"
    if public in ("Jeunesse", "Dès 6 ans", "6-8 ans", "8-12 ans"):
        return "7"
    if type_doc == "Manga":
        return "12"  # Par défaut manga
    return ""

# ─────────────────────────────────────────────────────────
# GENRE
# ─────────────────────────────────────────────────────────

GENRES = [
    ("Aventure",              ["aventure","exploration","pirates","trésor","expédition","quête"]),
    ("Policier / Mystère",    ["policier","détective","enquête","mystère","crime","suspect","meurtre"]),
    ("Fantastique", ["fantasy","fantastique","magie","sorcier","dragon","elfe","vampire",
                                "zombie","loup-garou","fée","magicien","enchantement","créature","monstre"]),
    ("Science-fiction",       ["science-fiction","science fiction"," sf ","robot","extraterrestre",
                                "futur","dystopie","spatial","planète","vaisseau","cyborg"]),
    ("Humour",                ["humour","comique","drôle","rigolo","éclats de rire","farce","blague","loufoque"]),
    ("Amour / Romance",       ["amour","romance","amoureux","love","sentiments","coeur","coup de foudre"]),
    ("Historique",            ["historique","histoire vraie","guerre","moyen âge","résistance",
                                "antiquité","révolution","chevalier","viking","empire","pharaon"]),
    ("Vie quotidienne",       ["école","famille","amitié","grandir","identité","harcèlement",
                                "divorce","adolescence","émotions","confiance","différence","handicap"]),
    ("Nature / Animaux",      ["animal","animaux","nature","vétérinaire","cheval","chien","chat",
                                "écologie","forêt","mer","sauvage","espèce","dinosaure","insecte"]),
    ("Sport",                 ["sport","football","basket","tennis","judo","badminton",
                                "natation","rugby","foot","cyclisme","athlétisme","gym"]),
    ("Arts / Musique",        ["musique","peinture","art ","artiste","dessin","danse",
                                "théâtre","cinéma","sculpture","photographie"]),
    ("Philo / Société",       ["philosophie","philo","goûters philo","éthique","société",
                                "citoyenneté","liberté","justice","droits","démocratie","religion"]),
    ("Frissons",              ["horreur","frisson","peur","fantôme","maison hantée","cauchemar","terrifiant"]),
    ("Sciences",              ["sciences","physique","chimie","biologie","mathématiques",
                                "astronomie","corps humain","expérience","inventeur","médecine"]),
    ("Géographie / Voyage",   ["géographie","pays","voyage","monde","continent","atlas","carte",
                                "exploration","découverte du monde"]),
    ("Conte / Mythe",         ["conte","mythe","légende","fable","folklore","grimm","perrault",
                                "andersen","1001 nuits","récit légendaire"]),
    ("Cuisine / Activités",   ["cuisine","recette","bricolage","jardinage","origami",
                                "coloriage","jeu de société","jeu vidéo","jeux et activités",
                                "cherche et trouve","autocollant","gommette"]),
]

# Genres par défaut selon le type de document
GENRE_PAR_TYPE = {
    "Manga":             "Aventure",       # par défaut si pas détecté
    "BD":                "",               # trop varié
    "Documentaire":      "Sciences",       # par défaut documentaire général
    "Album":             "Vie quotidienne",# par défaut album
    "Première lecture":  "Aventure",
    "Conte / Poésie":    "Conte / Mythe",
}

def _mot_present(mot, texte):
    """Teste la présence d'un mot/expression en respectant les limites de mots,
    pour éviter les faux positifs du type 'chimie' trouvé dans 'alchimie'
    ou 'jeu' trouvé dans 'enjeu'/'jeune'/'déjeuner'."""
    return re.search(r"\b" + re.escape(mot.strip()) + r"\b", texte) is not None

def detecter_genre(titre, collection, resume, sujets, type_doc=""):
    texte_titre = titre.lower()
    texte_total = " ".join([titre, collection, resume, sujets]).lower()

    scores = {}
    for nom, mots in GENRES:
        score = 0
        for m in mots:
            if _mot_present(m, texte_titre):
                score += 3        # un mot-clé présent dans le TITRE est très fiable
            elif _mot_present(m, texte_total):
                score += 1        # sinon, simple occurrence dans le résumé/collection
        if score:
            scores[nom] = score

    if scores:
        # Maximum 2 thèmes, les mieux notés — pas les 2 premiers trouvés dans la liste
        meilleurs = sorted(scores.items(), key=lambda kv: -kv[1])[:2]
        return " / ".join(nom for nom, _ in meilleurs)

    # Fallback : genre par défaut selon le type
    if type_doc in GENRE_PAR_TYPE and GENRE_PAR_TYPE[type_doc]:
        return GENRE_PAR_TYPE[type_doc]
    return ""

# ─────────────────────────────────────────────────────────
# TOME ET SÉRIE
# ─────────────────────────────────────────────────────────

# ── Patterns d'extraction du numéro de tome (ordre de priorité)
TOME_PATTERNS = [
    r'[Tt]ome\s*(\d{1,3})',           # Tome 3 / tome3
    r'\bT\.\s*(\d{1,3})\b',           # T.3
    r'[Vv]ol(?:ume)?\s*(\d{1,3})',    # Vol.3 / Volume 3
    r'[Nn]°\s*(\d{1,3})',             # N°3
    r'#\s*(\d{1,3})',                  # #3
    r'\((\d{1,2})\)\s*$',             # (3) en fin
    r',\s*(\d{1,2})\s*$',             # , 3 en fin
    r'[-–]\s*(\d{1,3})\s*(?:$|[-–:])',# - 3 ou - 3 :
    r'(?<!\d)(\d{1,3})\s*$',          # chiffre seul en fin (Blue Lock 27)
]

def extraire_tome(tome_raw, titre):
    """Extrait le numéro de tome depuis les métadonnées ou le titre."""
    # 1. Depuis les métadonnées (BnF 225$v)
    if tome_raw:
        try:
            n = int(re.search(r"\d+", tome_raw).group())
            if 1 <= n <= 99:
                return str(n)
        except Exception:
            pass
    # 2. Depuis le titre
    for pattern in TOME_PATTERNS:
        m = re.search(pattern, titre)
        if m:
            try:
                n = int(m.group(1))
                if 1 <= n <= 99:
                    return str(n)
            except Exception:
                pass
    return ""

def nettoyer_titre(titre_brut, tome, serie):
    """
    Retire le numéro de tome du titre.
    Exemples :
      "Blue Box5 - BLUE BOX"                          → "BLUE BOX"
      "Blue lock27 - Blue Lock T27"                   → "Blue Lock"
      "Game over17 - Game over / Dark web"            → "Game over / Dark web"
      "Les petits foot maniacs15 - Les Footmaniacs"   → "Les Footmaniacs"
      "Album de Boule & Bill.46 - Boule & Bill T.46"  → "Boule & Bill"
      "7 - Archibald - Ma maison"                     → "Ma maison"
      "Boule & Bill - Tome 46 - Peinture à l'os"      → "Peinture à l'os"
    """
    if not titre_brut: return ""
    t = titre_brut.strip()

    supprimer_serie = True

    if tome:
        # "NomXX - Titre réel" → "Titre réel"  ex: "Blue Box5 - BLUE BOX"
        m = re.match(r'^(.+?)[\s.]?' + re.escape(tome) + r'\s*[-–]\s*(.+)$', t)
        if m and len(m.group(1)) > 0 and len(m.group(2)) > 3:
            t = m.group(2).strip()

        # "Préfixe.XX - Titre" → "Titre"  ex: "Album de Boule & Bill.46 - ..."
        m2 = re.match(r'^.+?\.' + re.escape(tome) + r'\s*[-–]\s*(.+)$', t)
        if m2 and len(m2.group(1)) > 3:
            t = m2.group(1).strip()

    # Retirer toutes les formes de numérotation
    t = re.sub(r'\s*[-–,]\s*[Tt]ome\s*\d{1,3}', '', t)
    t = re.sub(r'\s*[-–,]\s*[Tt]\.\s*\d{1,3}\b', '', t)
    t = re.sub(r'\s*[-–,]\s*[Vv]ol(?:ume)?\s*\d{1,3}', '', t)
    t = re.sub(r'\s*[-–,]\s*[Nn]°\s*\d{1,3}', '', t)
    t = re.sub(r'\s*#\s*\d{1,3}', '', t)
    if tome:
        t = re.sub(r'\s+[Tt]\.?\s*' + re.escape(tome) + r'\s*$', '', t)
        t = re.sub(r'\s+' + re.escape(tome) + r'\s*$', '', t)

    # Si la série est en doublon au début AVEC un numéro de tome
    # ex: "Blue Lock 27 - Blue Lock" → on garde "Blue Lock"
    # Mais PAS "Archibald - Ma maison" → on garde le titre complet
    if supprimer_serie and serie and len(serie) > 3 and tome:
        serie_esc = re.escape(serie)
        # Seulement si le titre contient le nom de série SUIVI d'un chiffre
        m = re.match(r'^' + serie_esc + r'\s*\d+\s*[-–:]\s*(.+)$', t, re.IGNORECASE)
        if m and len(m.group(1)) > 3:
            t = m.group(1).strip()

    # Nettoyage final
    t = re.sub(r'\s*[-–:,]+\s*$', '', t)
    t = re.sub(r'^\s*[-–:,]+\s*', '', t)
    t = re.sub(r'\s{2,}', ' ', t)
    t = t.strip(" -–:,.")
    return t if len(t) > 2 else titre_brut.strip()

COLLECTIONS_GENERIQUES = [
    "folio junior","folio cadet","folio benjamin","folio ado",
    "poche jeunesse","livre de poche","milan poche","nathan poche",
    "cascade","tribal","pôle fiction","exprim","scripto",
    "heure noire","autres mondes","black moon",
    "pika shonen","pika shônen","pika seinen",
    "les yeux de la découverte","mes premières découvertes",
    "les goûters philo","mes p'tites questions","pocqq",
    "premiers romans","j'aime lire","je commence à lire","éclats de rire",
]

def nettoyer_serie(serie_brut, tome):
    """Nettoie le nom de série en retirant le numéro de tome et les suffixes parasites."""
    if not serie_brut:
        return ""
    s = serie_brut.strip()
    if tome:
        # Retirer ".XX" ou " XX" ou "-XX" en fin ou au milieu
        s = re.sub(r'[\s.]' + re.escape(tome) + r'\s*[-–].*$', '', s)
        s = re.sub(r'[\s.,-]' + re.escape(tome) + r'\s*$', '', s)
    # Retirer "- Titre réel" si la série contient un tiret séparateur
    # ex: "Album de Boule & Bill.46 - E..." → "Album de Boule & Bill"
    if tome:
        s = re.sub(r'\.?' + re.escape(tome) + r'.*$', '', s)
    # Retirer les séparateurs résiduels en fin
    s = re.sub(r'\s*[-–:.,]+\s*$', '', s)
    s = s.strip(" -–:.,")
    return s if len(s) > 2 else serie_brut.strip()

def extraire_serie(titre, serie_raw, collection, tome=""):
    # 1. Depuis les métadonnées (BnF, Babelio...)
    if serie_raw and len(serie_raw) > 2:
        return nettoyer_serie(serie_raw.strip(" .,()"), tome)

    # 2. Depuis le titre
    patterns = [
        r'^(.+?)\s*[,\-–]\s*[Tt]ome\s*\d+',
        r'^(.+?)\s*[,\-–]\s*[Tt]\.\s*\d+',
        r'^(.+?)\s*[Tt]ome\s*\d+',
        r'^(.+?)\s+[Tt]\.\s*\d+',
        r'^(.+?)\s*[Vv]ol(?:ume)?\s*\d+',
        r'^(.+?)\s+[Nn]°\s*\d+',
        r'^(.+?)\s+#\s*\d+',
        r'^(.+?)\s+\((\d{1,2})\)\s*$',
        r'^(.+?)\s*:\s*.+[\s\-–]\s*\d{1,3}\s*$',
        # "NomXX - Titre" → série = "Nom"
        r'^(.+?)[\s.]?\d{1,3}\s*[-–]\s*.+$',
        # "Nom XX" chiffre en fin
        r'^(.+?)\s+(\d{1,3})\s*$',
    ]
    for pattern in patterns:
        m = re.match(pattern, titre, re.IGNORECASE)
        if m:
            serie = m.group(1).strip(" .,:–-()")
            serie = nettoyer_serie(serie, tome)
            if len(serie) > 2 and not serie.isdigit():
                return serie

    # 3. Depuis la collection si elle n'est pas générique
    if collection and len(collection) > 3:
        if not any(g in collection.lower() for g in COLLECTIONS_GENERIQUES):
            return collection.strip()
    return ""

# ─────────────────────────────────────────────────────────
# HELPER
# ─────────────────────────────────────────────────────────

def enrichir(base, nouveau):
    if not base:
        return nouveau
    for champ in ("auteur","illustrateur","editeur","annee","type","public",
                  "genre","serie","tome","pegi","collection","resume"):
        if not base.get(champ) and nouveau.get(champ):
            base[champ] = nouveau[champ]
    return base

def est_complet(res):
    complet = all(res.get(c) for c in CHAMPS_CLES)
    a_tome  = not res.get("serie") or res.get("tome")
    return complet and a_tome

# ─────────────────────────────────────────────────────────
# SOURCE 1 : BnF SRU
# ─────────────────────────────────────────────────────────

def bnf_lookup(isbn):
    try:
        query = f'bib.isbn adj "{isbn}"'
        url = (
            "https://catalogue.bnf.fr/api/SRU?"
            "version=1.2&operation=searchRetrieve"
            f"&query={urllib.parse.quote(query)}"
            "&maximumRecords=1&recordSchema=unimarcxchange"
        )
        r = requests.get(url, headers=HEADERS, timeout=(5, 10))
        if r.status_code != 200: return None
        root = ET.fromstring(r.content)
        ns = {"srw":"http://www.loc.gov/zing/srw/"}
        nb = root.find(".//srw:numberOfRecords", ns)
        if nb is None or nb.text == "0": return None
        rec = root.find(".//srw:recordData", ns)
        if rec is None: return None

        def sf(tag, code):
            for f in rec.iter():
                if f.get("tag") == tag:
                    for s in f:
                        if s.get("code") == code and s.text:
                            return s.text.strip(" .,:/()")
            return ""

        def sf_list(tag, code):
            vals = []
            for f in rec.iter():
                if f.get("tag") == tag:
                    for s in f:
                        if s.get("code") == code and s.text:
                            vals.append(s.text.strip(" .,:"))
            return vals

        titre_a = sf("200","a")
        titre_e = sf("200","e")
        titre_n = sf("200","n")
        titre_p = sf("200","p")
        titre = titre_a
        if titre_n and re.match(r"^\d+$", titre_n.strip()) and int(titre_n) <= 99:
            titre += f" T.{titre_n.strip()}"
        if titre_p:
            titre += f" : {titre_p}"
        elif titre_e and titre_e.lower() != titre_a.lower():
            titre += f" : {titre_e}"
        if not titre: return None

        auteurs, illustrateurs = [], []
        for tag in ("700","701","702"):
            for f in rec.iter():
                if f.get("tag") == tag:
                    nom = prenom = fonction = ""
                    for s in f:
                        if s.get("code") == "a": nom     = (s.text or "").strip(" .,")
                        if s.get("code") == "b": prenom  = (s.text or "").strip(" .,")
                        if s.get("code") == "4": fonction = s.text or ""
                    nc = f"{prenom} {nom}".strip()
                    if nc:
                        if fonction in ("110","740","080","440","230"):
                            illustrateurs.append(nc)
                        else:
                            auteurs.append(nc)
        if not auteurs:
            for tag in ("710","711"):
                for f in rec.iter():
                    if f.get("tag") == tag:
                        for s in f:
                            if s.get("code") == "a" and s.text:
                                auteurs.append(s.text.strip(" .,"))

        editeur   = sf("214","c") or sf("210","c")
        annee_raw = sf("214","d") or sf("210","d")
        m_an = re.search(r"\d{4}", annee_raw)
        annee = m_an.group(0) if m_an else ""
        collection = sf("225","a")
        tome_225   = sf("225","v")
        serie_raw  = ""
        for tag in ("410","411","440"):
            s = sf(tag,"t")
            if s: serie_raw = s; break

        support   = sf("200","b")
        genre_all = " ".join(sf_list("606","a") + sf_list("607","a") + sf_list("608","a"))
        tranche   = sf("521","a")
        public_code = ""
        for f in rec.iter():
            if f.get("tag") == "100":
                for s in f:
                    if s.get("code") == "a" and s.text and len(s.text) > 22:
                        public_code = s.text[22]
        resume = sf("330","a")

        public   = detecter_public(public_code, titre, collection, editeur, tranche)
        type_doc = detecter_type(titre, collection, resume, genre_all, support, editeur, public)
        genre    = detecter_genre(titre, collection, resume, genre_all, type_doc)
        tome     = extraire_tome(tome_225, titre)
        serie    = extraire_serie(titre, serie_raw, collection, tome)
        pegi     = detecter_pegi(type_doc, public, "")

        return {
            "titre":        titre,
            "auteur":       " / ".join(auteurs[:2]),
            "illustrateur": " / ".join(illustrateurs[:2]),
            "editeur":      editeur, "annee": annee,
            "type":         type_doc, "public": public,
            "genre":        genre, "serie": serie,
            "tome":         tome, "pegi": pegi,
            "collection":   collection,
            "resume":       resume[:400] if resume else "",
            "source":       "BnF", "statut": "trouvé",
        }
    except Exception:
        return None

# ─────────────────────────────────────────────────────────
# SOURCE 2 : Babelio
# ─────────────────────────────────────────────────────────

def babelio_lookup(isbn):
    try:
        url = f"https://www.babelio.com/recherche.php?Recherche={isbn}&recherche=isbn"
        r = requests.get(url, headers=HEADERS, timeout=(5, 10))
        if r.status_code != 200: return None
        soup = BeautifulSoup(r.text, "html.parser")
        lien = soup.select_one("a[href*='/livres/']")
        if not lien: return None
        href = lien.get("href","")
        if not href.startswith("http"): href = "https://www.babelio.com" + href
        r2 = requests.get(href, headers=HEADERS, timeout=(5, 10))
        if r2.status_code != 200: return None
        soup2 = BeautifulSoup(r2.text, "html.parser")

        titre_el = soup2.select_one("h1[itemprop='name'], h1.livre_con")
        titre = titre_el.get_text(strip=True) if titre_el else ""
        if not titre: return None
        auteur_el = soup2.select_one("span[itemprop='author'], a[href*='/auteur/']")
        auteur = auteur_el.get_text(strip=True) if auteur_el else ""

        editeur = annee = collection = genre_b = tranche = ""
        for info in soup2.select(".book-detail li, .livre_infos li, .infos span"):
            t = info.get_text(" ", strip=True)
            if re.search(r"éditeur|editeur", t, re.I):
                editeur = re.sub(r"éditeur\s*:?\s*","", t, flags=re.I).strip()
            elif re.search(r"parution|année|date", t, re.I):
                m = re.search(r"\d{4}", t)
                if m: annee = m.group(0)
            elif re.search(r"collection", t, re.I):
                collection = re.sub(r"collection\s*:?\s*","", t, flags=re.I).strip()
            elif re.search(r"genre|type", t, re.I):
                genre_b = re.sub(r"(genre|type)\s*:?\s*","", t, flags=re.I).strip()
            elif re.search(r"âge|ans\b|public", t, re.I):
                tranche = t

        resume_el = soup2.select_one("#d_bio, .livre_resume, [itemprop='description']")
        resume = resume_el.get_text(strip=True)[:400] if resume_el else ""

        public   = detecter_public("", titre, collection, editeur, tranche)
        type_doc = detecter_type(titre, collection, resume, genre_b, "", editeur, public)
        genre    = detecter_genre(titre, collection, resume, genre_b, type_doc)
        tome     = extraire_tome("", titre)
        serie    = extraire_serie(titre, "", collection, tome)
        pegi     = detecter_pegi(type_doc, public, soup2.get_text())

        return {"titre":titre,"auteur":auteur,"illustrateur":"",
                "editeur":editeur,"annee":annee,"type":type_doc,
                "public":public,"genre":genre,"serie":serie,"tome":tome,
                "pegi":pegi,"collection":collection,"resume":resume,
                "source":"Babelio","statut":"trouvé"}
    except Exception:
        return None

# ─────────────────────────────────────────────────────────
# SOURCE 3 : BDfugue
# ─────────────────────────────────────────────────────────

def bdfugue_lookup(isbn):
    try:
        url = f"https://www.bdfugue.com/catalogsearch/result/?q={isbn}"
        r = requests.get(url, headers=HEADERS, timeout=(5, 10))
        if r.status_code != 200: return None
        soup = BeautifulSoup(r.text, "html.parser")
        lien = soup.select_one("a.product-item-link, a.product-name, h2.product-name a")
        if not lien: return None
        href = lien.get("href","")
        if not href: return None
        r2 = requests.get(href, headers=HEADERS, timeout=(5, 10))
        if r2.status_code != 200: return None
        soup2 = BeautifulSoup(r2.text, "html.parser")
        titre_el = soup2.select_one("h1.page-title, h1[itemprop='name']")
        titre = titre_el.get_text(strip=True) if titre_el else ""
        if not titre: return None

        def ga(label):
            for el in soup2.select(".product-attribute,.product.attribute,tr"):
                if label.lower() in el.get_text(" ",strip=True).lower():
                    val = el.select_one(".value, td:last-child")
                    if val: return val.get_text(strip=True)
            return ""

        auteur    = ga("auteur") or ga("scénari") or ga("dessin")
        editeur   = ga("éditeur") or ga("editeur")
        annee_r   = ga("date") or ga("parution")
        m = re.search(r"\d{4}", annee_r)
        annee     = m.group(0) if m else ""
        collection = ga("collection") or ga("série")
        resume_el = soup2.select_one(".product.attribute.description .value,#description")
        resume    = resume_el.get_text(strip=True)[:400] if resume_el else ""
        texte_page = soup2.get_text()

        public   = detecter_public("", titre, collection, editeur, "")
        type_doc = detecter_type(titre, collection, resume, "", "", editeur, public)
        if not type_doc: type_doc = "BD"
        genre    = detecter_genre(titre, collection, resume, "", type_doc)
        tome     = extraire_tome("", titre)
        serie    = extraire_serie(titre, collection, collection, tome)
        pegi     = detecter_pegi(type_doc, public, texte_page)

        return {"titre":titre,"auteur":auteur,"illustrateur":"",
                "editeur":editeur,"annee":annee,"type":type_doc,
                "public":public,"genre":genre,"serie":serie,"tome":tome,
                "pegi":pegi,"collection":collection,"resume":resume,
                "source":"BDfugue","statut":"trouvé"}
    except Exception:
        return None

# ─────────────────────────────────────────────────────────
# SOURCE 4 : Manga News
# ─────────────────────────────────────────────────────────

def manganews_lookup(isbn):
    try:
        url = f"https://www.manga-news.com/index.php/recherche?search={isbn}"
        r = requests.get(url, headers=HEADERS, timeout=(5, 10))
        if r.status_code != 200: return None
        soup = BeautifulSoup(r.text, "html.parser")
        lien = soup.select_one("a.titre, .manga-title a, .result-item a, h2 a, h3 a")
        if not lien: return None
        href = lien.get("href","")
        if not href.startswith("http"): href = "https://www.manga-news.com" + href
        r2 = requests.get(href, headers=HEADERS, timeout=(5, 10))
        if r2.status_code != 200: return None
        soup2 = BeautifulSoup(r2.text, "html.parser")

        titre_el = soup2.select_one("h1.title, h1[itemprop='name'], h1")
        titre = titre_el.get_text(strip=True) if titre_el else ""
        if not titre: return None

        def gi(label):
            for el in soup2.select(".info-manga li,.fiche-info li,dl dt,.meta-item"):
                if label.lower() in el.get_text().lower():
                    nxt = el.find_next_sibling()
                    if nxt: return nxt.get_text(strip=True)
            return ""

        auteur  = gi("auteur") or gi("scénari") or gi("dessin")
        editeur = gi("éditeur") or gi("editeur")
        annee_r = gi("date") or gi("parution")
        m = re.search(r"\d{4}", annee_r)
        annee   = m.group(0) if m else ""
        tome_r  = gi("tome") or gi("volume") or gi("numéro")
        serie_r = gi("série") or gi("serie")
        resume_el = soup2.select_one(".synopsis,.resume,[itemprop='description']")
        resume  = resume_el.get_text(strip=True)[:400] if resume_el else ""
        texte_page = soup2.get_text()

        tome  = extraire_tome(tome_r, titre)
        serie = extraire_serie(titre, serie_r, "", tome)
        public = detecter_public("", titre, "", editeur, "")
        if not public: public = "Ado (12+)"
        pegi = detecter_pegi("Manga", public, texte_page)

        return {"titre":titre,"auteur":auteur,"illustrateur":"",
                "editeur":editeur,"annee":annee,"type":"Manga",
                "public":public,"genre":"","serie":serie,"tome":tome,
                "pegi":pegi,"collection":"","resume":resume,
                "source":"Manga News","statut":"trouvé"}
    except Exception:
        return None

# ─────────────────────────────────────────────────────────
# SOURCE 5 : Fnac
# ─────────────────────────────────────────────────────────

def fnac_lookup(isbn):
    try:
        url = f"https://recherche.fnac.com/SearchResult/ResultList.aspx?SCat=2&Search={isbn}"
        r = requests.get(url, headers=HEADERS, timeout=(5, 10))
        if r.status_code != 200: return None
        soup = BeautifulSoup(r.text, "html.parser")
        lien = soup.select_one("a.Article-title,.Article-itemTitle a,h3 a,.product-title a")
        if not lien: return None
        href = lien.get("href","")
        if not href.startswith("http"): href = "https://www.fnac.com" + href
        r2 = requests.get(href, headers=HEADERS, timeout=(5, 10))
        if r2.status_code != 200: return None
        soup2 = BeautifulSoup(r2.text, "html.parser")

        titre_el = soup2.select_one("h1[itemprop='name'],h1.ProductHeader-title")
        titre = titre_el.get_text(strip=True) if titre_el else ""
        if not titre: return None

        auteur_el  = soup2.select_one("[itemprop='author'],.ProductHeader-author")
        editeur_el = soup2.select_one("[itemprop='publisher'],.ProductDetails-publisher")
        annee_el   = soup2.select_one("[itemprop='datePublished'],.ProductDetails-date")
        resume_el  = soup2.select_one("[itemprop='description'],.ProductDetails-description")
        coll_el    = soup2.select_one(".ProductDetails-collection,[itemprop='isPartOf']")
        age_el     = soup2.select_one(".ProductDetails-age,[itemprop='typicalAgeRange']")

        auteur     = auteur_el.get_text(strip=True)  if auteur_el  else ""
        editeur    = editeur_el.get_text(strip=True) if editeur_el else ""
        annee_r    = annee_el.get_text(strip=True)   if annee_el   else ""
        resume     = resume_el.get_text(strip=True)[:400] if resume_el else ""
        collection = coll_el.get_text(strip=True)    if coll_el    else ""
        tranche    = age_el.get_text(strip=True)     if age_el     else ""
        texte_page = soup2.get_text()

        m = re.search(r"\d{4}", annee_r)
        annee = m.group(0) if m else ""
        public   = detecter_public("", titre, collection, editeur, tranche)
        type_doc = detecter_type(titre, collection, resume, "", "", editeur, public)
        genre    = detecter_genre(titre, collection, resume, "", type_doc)
        tome     = extraire_tome("", titre)
        serie    = extraire_serie(titre, "", collection, tome)
        pegi     = detecter_pegi(type_doc, public, texte_page)

        return {"titre":titre,"auteur":auteur,"illustrateur":"",
                "editeur":editeur,"annee":annee,"type":type_doc,
                "public":public,"genre":genre,"serie":serie,"tome":tome,
                "pegi":pegi,"collection":collection,"resume":resume,
                "source":"Fnac","statut":"trouvé"}
    except Exception:
        return None

# ─────────────────────────────────────────────────────────
# SOURCE 6 : Amazon.fr
# ─────────────────────────────────────────────────────────

def amazon_lookup(isbn):
    try:
        url = f"https://www.amazon.fr/s?k={isbn}&i=stripbooks"
        r = requests.get(url, headers=HEADERS, timeout=(5, 10))
        if r.status_code != 200: return None
        soup = BeautifulSoup(r.text, "html.parser")

        # Trouver le premier résultat
        lien = soup.select_one("h2 a.a-link-normal, .s-result-item h2 a")
        if not lien: return None
        href = lien.get("href","")
        if not href.startswith("http"): href = "https://www.amazon.fr" + href
        r2 = requests.get(href, headers=HEADERS, timeout=(5, 10))
        if r2.status_code != 200: return None
        soup2 = BeautifulSoup(r2.text, "html.parser")

        titre_el = soup2.select_one("#productTitle, h1#title")
        titre = titre_el.get_text(strip=True) if titre_el else ""
        if not titre: return None

        # Auteur
        auteur_el = soup2.select_one(".author a, #bylineInfo a")
        auteur = auteur_el.get_text(strip=True) if auteur_el else ""

        # Détails produit
        editeur = annee = collection = ""
        texte_page = soup2.get_text()

        for row in soup2.select(".a-expander-content tr, #detailBullets_feature_div li, .detail-bullet-list li"):
            t = row.get_text(" ", strip=True)
            if re.search(r"éditeur|editeur|publisher", t, re.I):
                parts = re.split(r"[:：]", t, 1)
                if len(parts) > 1:
                    editeur = parts[1].strip().split(";")[0].strip()
            elif re.search(r"date|parution|publication", t, re.I):
                m = re.search(r"(\d{1,2}\s+\w+\s+\d{4}|\d{4})", t)
                if m:
                    annee_m = re.search(r"\d{4}", m.group(0))
                    if annee_m: annee = annee_m.group(0)
            elif re.search(r"collection|série|series", t, re.I):
                parts = re.split(r"[:：]", t, 1)
                if len(parts) > 1:
                    collection = parts[1].strip()

        # Résumé
        resume_el = soup2.select_one("#bookDescription_feature_div, #productDescription")
        resume = resume_el.get_text(strip=True)[:400] if resume_el else ""

        # PEGI Amazon (souvent dans les détails)
        pegi_brut = ""
        m_pegi = re.search(r"PEGI\s*(\d+)", texte_page, re.I)
        if m_pegi: pegi_brut = m_pegi.group(1)

        # Tranche d'âge Amazon
        tranche = ""
        m_age = re.search(r"(\d+)\s*[àa-]\s*\d+\s*ans?|à partir de\s*(\d+)\s*ans?|dès\s*(\d+)\s*ans?", texte_page, re.I)
        if m_age: tranche = m_age.group(0)

        public   = detecter_public("", titre, collection, editeur, tranche)
        type_doc = detecter_type(titre, collection, resume, "", "", editeur, public)
        genre    = detecter_genre(titre, collection, resume, "", type_doc)
        tome     = extraire_tome("", titre)
        serie    = extraire_serie(titre, "", collection, tome)
        pegi     = pegi_brut or detecter_pegi(type_doc, public, texte_page)

        return {"titre":titre,"auteur":auteur,"illustrateur":"",
                "editeur":editeur,"annee":annee,"type":type_doc,
                "public":public,"genre":genre,"serie":serie,"tome":tome,
                "pegi":pegi,"collection":collection,"resume":resume,
                "source":"Amazon","statut":"trouvé"}
    except Exception:
        return None

# ─────────────────────────────────────────────────────────
# SOURCE 7 : Cultura
# ─────────────────────────────────────────────────────────

def _cultura_depuis_titre_page(soup, isbn):
    """Extrait série/tome/titre/auteur depuis la balise <title> de la page de
    résultats Cultura.

    Ajouté le 2026-07-24 : la liste de résultats est rendue en JavaScript, donc
    aucun lien produit n'est visible pour un script -- l'ancienne logique
    abandonnait (return None) et Cultura ne remontait jamais rien. En revanche
    le <title> de la page, lui, est bien dans le HTML et contient TOUT ce qui
    nous manquait, série et tome compris. Exemple réel :
      « Harry Potter Tome 6 : Harry Potter et le Prince de sang-mêlé :
        J. K. Rowling- Livres audio - CD | Cultura »
    Format : « SÉRIE Tome N : TITRE : AUTEUR- catégories | Cultura »
    """
    el = soup.find("title")
    if not el:
        return None
    brut = el.get_text(strip=True)
    if not brut or "Cultura" not in brut:
        return None
    # retire le suffixe « | Cultura »
    corps = brut.split("|")[0].strip()
    parts = [p.strip() for p in corps.split(" : ") if p.strip()]
    if not parts:
        return None

    serie = tome = ""
    titre = parts[0]
    auteur = ""

    # 1re partie de la forme « Série Tome 6 » → série + tome
    m = re.match(r"^(.*?)\s+Tome\s+(\d+)\s*$", parts[0], re.I)
    if m and len(parts) >= 2:
        serie = m.group(1).strip()
        tome = m.group(2)
        if len(parts) >= 3:
            # « Série Tome N : TITRE : AUTEUR- catégories »
            titre = parts[1].strip()
            auteur = parts[2]
        else:
            # « Série Tome N : AUTEUR- catégories » (pas de titre d'album
            # distinct) : ne PAS prendre l'auteur pour un titre -- on retombe
            # sur le nom de la série, qui est correct.
            titre = serie
            auteur = parts[1]
    elif len(parts) >= 2:
        titre = parts[1].strip()
        auteur = parts[2] if len(parts) >= 3 else ""

    # l'auteur est parfois suivi des catégories : « J. K. Rowling- Livres audio - CD »
    if auteur:
        auteur = re.split(r"\s*-\s*(?:Livres?|Romans?|BD|Mangas?|CD|DVD)\b", auteur, 1)[0]
        auteur = auteur.rstrip("- ").strip()

    if not titre:
        return None
    return {"titre": titre, "serie": serie, "tome": tome, "auteur": auteur}


def cultura_lookup(isbn):
    try:
        url = f"https://www.cultura.com/catalogsearch/result/?q={isbn}"
        r = requests.get(url, headers=HEADERS, timeout=(5, 10))
        if r.status_code != 200: return None
        soup = BeautifulSoup(r.text, "html.parser")
        lien = soup.select_one("a.product-item-link,.product-name a,h2 a")
        if not lien:
            # Repli : lire le <title> de la page de résultats (voir fonction
            # ci-dessus). C'est le cas NORMAL aujourd'hui, la liste étant
            # rendue en JavaScript.
            base = _cultura_depuis_titre_page(soup, isbn)
            if not base:
                return None
            titre = base["titre"]
            serie = base["serie"] or extraire_serie(titre, "", "", base["tome"])
            tome = base["tome"] or extraire_tome("", titre)
            public = detecter_public("", titre, "", "", "")
            type_doc = detecter_type(titre, "", "", "", "", "", public)
            genre = detecter_genre(titre, "", "", "", type_doc)
            return {"titre": titre, "auteur": base["auteur"], "illustrateur": "",
                    "editeur": "", "annee": "", "type": type_doc,
                    "public": public, "genre": genre, "serie": serie, "tome": tome,
                    "pegi": "", "collection": "", "resume": "",
                    "source": "Cultura", "statut": "trouvé"}
        href = lien.get("href","")
        if not href.startswith("http"): href = "https://www.cultura.com" + href
        r2 = requests.get(href, headers=HEADERS, timeout=(5, 10))
        if r2.status_code != 200: return None
        soup2 = BeautifulSoup(r2.text, "html.parser")

        titre_el = soup2.select_one("h1[itemprop='name'],h1.product-name,h1.page-title")
        titre = titre_el.get_text(strip=True) if titre_el else ""
        if not titre: return None

        auteur_el  = soup2.select_one("[itemprop='author'],.product-author,.author")
        editeur_el = soup2.select_one("[itemprop='publisher'],.publisher,.editeur")
        annee_el   = soup2.select_one("[itemprop='datePublished'],.date-publication")
        resume_el  = soup2.select_one("[itemprop='description'],.product-description,.description")
        coll_el    = soup2.select_one(".collection,[itemprop='isPartOf']")
        age_el     = soup2.select_one(".age,.public-cible,.age-conseille")

        auteur     = auteur_el.get_text(strip=True)  if auteur_el  else ""
        editeur    = editeur_el.get_text(strip=True) if editeur_el else ""
        annee_r    = annee_el.get_text(strip=True)   if annee_el   else ""
        resume     = resume_el.get_text(strip=True)[:400] if resume_el else ""
        collection = coll_el.get_text(strip=True)    if coll_el    else ""
        tranche    = age_el.get_text(strip=True)     if age_el     else ""

        m = re.search(r"\d{4}", annee_r)
        annee = m.group(0) if m else ""
        public   = detecter_public("", titre, collection, editeur, tranche)
        type_doc = detecter_type(titre, collection, resume, "", "", editeur, public)
        genre    = detecter_genre(titre, collection, resume, "", type_doc)
        tome     = extraire_tome("", titre)
        serie    = extraire_serie(titre, "", collection, tome)
        pegi     = detecter_pegi(type_doc, public, soup2.get_text())

        return {"titre":titre,"auteur":auteur,"illustrateur":"",
                "editeur":editeur,"annee":annee,"type":type_doc,
                "public":public,"genre":genre,"serie":serie,"tome":tome,
                "pegi":pegi,"collection":collection,"resume":resume,
                "source":"Cultura","statut":"trouvé"}
    except Exception:
        return None

# ─────────────────────────────────────────────────────────
# SOURCE 8 : Decitre
# ─────────────────────────────────────────────────────────

def decitre_lookup(isbn):
    try:
        # URL de recherche mise à jour le 2026-07-24 : l'ancienne adresse
        # /livres/{isbn}.html renvoie désormais 404 (Decitre a changé ses URL).
        url = f"https://www.decitre.fr/rechercher/result?q={isbn}"
        r = requests.get(url, headers=HEADERS, timeout=(5, 10))
        if r.status_code != 200: return None
        soup = BeautifulSoup(r.text, "html.parser")
        titre_el = soup.select_one("h1[itemprop='name'],h1.product-name")
        titre = titre_el.get_text(strip=True) if titre_el else ""
        if not titre: return None

        auteur_el  = soup.select_one("[itemprop='author'],.product-author")
        editeur_el = soup.select_one("[itemprop='publisher'],.product-publisher")
        annee_el   = soup.select_one("[itemprop='datePublished'],.product-date")
        coll_el    = soup.select_one(".product-collection,[itemprop='isPartOf']")
        resume_el  = soup.select_one("[itemprop='description'],.product-description")
        age_el     = soup.select_one(".product-age,.age-lecteur")

        auteur     = auteur_el.get_text(strip=True)  if auteur_el  else ""
        editeur    = editeur_el.get_text(strip=True) if editeur_el else ""
        annee_r    = annee_el.get_text(strip=True)   if annee_el   else ""
        collection = coll_el.get_text(strip=True)    if coll_el    else ""
        resume     = resume_el.get_text(strip=True)[:400] if resume_el else ""
        tranche    = age_el.get_text(strip=True)     if age_el     else ""

        m = re.search(r"\d{4}", annee_r)
        annee = m.group(0) if m else ""
        public   = detecter_public("", titre, collection, editeur, tranche)
        type_doc = detecter_type(titre, collection, resume, "", "", editeur, public)
        genre    = detecter_genre(titre, collection, resume, "", type_doc)
        tome     = extraire_tome("", titre)
        serie    = extraire_serie(titre, "", collection, tome)
        pegi     = detecter_pegi(type_doc, public, soup.get_text())

        return {"titre":titre,"auteur":auteur,"illustrateur":"",
                "editeur":editeur,"annee":annee,"type":type_doc,
                "public":public,"genre":genre,"serie":serie,"tome":tome,
                "pegi":pegi,"collection":collection,"resume":resume,
                "source":"Decitre","statut":"trouvé"}
    except Exception:
        return None

# ─────────────────────────────────────────────────────────
# SOURCE 9 : Ricochet-jeunes
# ─────────────────────────────────────────────────────────

def ricochet_lookup(isbn):
    try:
        url = f"https://www.ricochet-jeunes.org/livres/recherche?isbn={isbn}"
        r = requests.get(url, headers=HEADERS, timeout=(5, 10))
        if r.status_code != 200: return None
        soup = BeautifulSoup(r.text, "html.parser")
        lien = soup.select_one("a.livre-title,.livre-item a,h2 a,h3 a")
        if not lien: return None
        href = lien.get("href","")
        if not href.startswith("http"): href = "https://www.ricochet-jeunes.org" + href
        r2 = requests.get(href, headers=HEADERS, timeout=(5, 10))
        if r2.status_code != 200: return None
        soup2 = BeautifulSoup(r2.text, "html.parser")

        titre_el = soup2.select_one("h1.titre,h1[itemprop='name'],h1")
        titre = titre_el.get_text(strip=True) if titre_el else ""
        if not titre: return None

        def gi(label):
            for el in soup2.select("dl dt,.meta,.fiche-info"):
                if label.lower() in el.get_text().lower():
                    nxt = el.find_next_sibling()
                    if nxt: return nxt.get_text(strip=True)
            return ""

        auteur     = gi("auteur") or gi("texte")
        illustr    = gi("illustrateur") or gi("illustration")
        editeur    = gi("éditeur") or gi("editeur")
        annee_r    = gi("date") or gi("parution") or gi("année")
        collection = gi("collection")
        public_r   = gi("âge") or gi("public") or gi("à partir")
        resume_el  = soup2.select_one(".resume,.description,[itemprop='description']")
        resume     = resume_el.get_text(strip=True)[:400] if resume_el else ""

        m = re.search(r"\d{4}", annee_r)
        annee = m.group(0) if m else ""
        public   = detecter_public("", titre, collection, editeur, public_r)
        type_doc = detecter_type(titre, collection, resume, "", "", editeur, public)
        genre    = detecter_genre(titre, collection, resume, "", type_doc)
        tome     = extraire_tome("", titre)
        serie    = extraire_serie(titre, "", collection, tome)
        pegi     = detecter_pegi(type_doc, public, "")

        return {"titre":titre,"auteur":auteur,"illustrateur":illustr,
                "editeur":editeur,"annee":annee,"type":type_doc,
                "public":public,"genre":genre,"serie":serie,"tome":tome,
                "pegi":pegi,"collection":collection,"resume":resume,
                "source":"Ricochet","statut":"trouvé"}
    except Exception:
        return None

# ─────────────────────────────────────────────────────────
# SOURCE 10 : Leslibraires.fr
# ─────────────────────────────────────────────────────────

def leslibraires_lookup(isbn):
    try:
        url = f"https://www.leslibraires.fr/livre/{isbn}"
        r = requests.get(url, headers=HEADERS, timeout=(5, 10))
        if r.status_code != 200: return None
        soup = BeautifulSoup(r.text, "html.parser")

        titre_el = soup.select_one("h1[itemprop='name'],h1.product-title")
        titre = titre_el.get_text(strip=True) if titre_el else ""
        if not titre: return None

        auteur_el  = soup.select_one("[itemprop='author'],.product-author")
        editeur_el = soup.select_one("[itemprop='publisher'],.product-publisher")
        annee_el   = soup.select_one("[itemprop='datePublished'],.product-date")
        resume_el  = soup.select_one("[itemprop='description'],.product-description")
        coll_el    = soup.select_one(".product-collection")
        age_el     = soup.select_one(".product-age,.public")

        auteur     = auteur_el.get_text(strip=True)  if auteur_el  else ""
        editeur    = editeur_el.get_text(strip=True) if editeur_el else ""
        annee_r    = annee_el.get_text(strip=True)   if annee_el   else ""
        resume     = resume_el.get_text(strip=True)[:400] if resume_el else ""
        collection = coll_el.get_text(strip=True)    if coll_el    else ""
        tranche    = age_el.get_text(strip=True)     if age_el     else ""

        m = re.search(r"\d{4}", annee_r)
        annee = m.group(0) if m else ""
        public   = detecter_public("", titre, collection, editeur, tranche)
        type_doc = detecter_type(titre, collection, resume, "", "", editeur, public)
        genre    = detecter_genre(titre, collection, resume, "", type_doc)
        tome     = extraire_tome("", titre)
        serie    = extraire_serie(titre, "", collection, tome)
        pegi     = detecter_pegi(type_doc, public, "")

        return {"titre":titre,"auteur":auteur,"illustrateur":"",
                "editeur":editeur,"annee":annee,"type":type_doc,
                "public":public,"genre":genre,"serie":serie,"tome":tome,
                "pegi":pegi,"collection":collection,"resume":resume,
                "source":"LesLibraires","statut":"trouvé"}
    except Exception:
        return None

# ─────────────────────────────────────────────────────────
# SOURCE 11 : Mollat
# ─────────────────────────────────────────────────────────

def mollat_lookup(isbn):
    try:
        # URL de recherche mise à jour le 2026-07-24 (l'ancienne renvoyait 404).
        url = f"https://www.mollat.com/recherche?q={isbn}"
        r = requests.get(url, headers=HEADERS, timeout=(5, 10))
        if r.status_code != 200: return None
        soup = BeautifulSoup(r.text, "html.parser")
        lien = soup.select_one("a.book-title,.book-item a,h2.title a,h3 a")
        if not lien: return None
        href = lien.get("href","")
        if not href.startswith("http"): href = "https://www.mollat.com" + href
        r2 = requests.get(href, headers=HEADERS, timeout=(5, 10))
        if r2.status_code != 200: return None
        soup2 = BeautifulSoup(r2.text, "html.parser")

        titre_el = soup2.select_one("h1[itemprop='name'],h1.book-title,h1")
        titre = titre_el.get_text(strip=True) if titre_el else ""
        if not titre: return None

        auteur_el  = soup2.select_one("[itemprop='author'],.book-author")
        editeur_el = soup2.select_one("[itemprop='publisher'],.book-publisher")
        annee_el   = soup2.select_one("[itemprop='datePublished'],.book-date")
        resume_el  = soup2.select_one("[itemprop='description'],.book-description,.resume")
        coll_el    = soup2.select_one(".book-collection,.collection")
        age_el     = soup2.select_one(".book-age,.public,.age")

        auteur     = auteur_el.get_text(strip=True)  if auteur_el  else ""
        editeur    = editeur_el.get_text(strip=True) if editeur_el else ""
        annee_r    = annee_el.get_text(strip=True)   if annee_el   else ""
        resume     = resume_el.get_text(strip=True)[:400] if resume_el else ""
        collection = coll_el.get_text(strip=True)    if coll_el    else ""
        tranche    = age_el.get_text(strip=True)     if age_el     else ""

        m = re.search(r"\d{4}", annee_r)
        annee = m.group(0) if m else ""
        public   = detecter_public("", titre, collection, editeur, tranche)
        type_doc = detecter_type(titre, collection, resume, "", "", editeur, public)
        genre    = detecter_genre(titre, collection, resume, "", type_doc)
        tome     = extraire_tome("", titre)
        serie    = extraire_serie(titre, "", collection, tome)
        pegi     = detecter_pegi(type_doc, public, "")

        return {"titre":titre,"auteur":auteur,"illustrateur":"",
                "editeur":editeur,"annee":annee,"type":type_doc,
                "public":public,"genre":genre,"serie":serie,"tome":tome,
                "pegi":pegi,"collection":collection,"resume":resume,
                "source":"Mollat","statut":"trouvé"}
    except Exception:
        return None

# ─────────────────────────────────────────────────────────
# SOURCE 12 : Booknode
# ─────────────────────────────────────────────────────────

def booknode_lookup(isbn):
    try:
        url = f"https://booknode.com/recherche?q={isbn}"
        r = requests.get(url, headers=HEADERS, timeout=(5, 10))
        if r.status_code != 200: return None
        soup = BeautifulSoup(r.text, "html.parser")
        lien = soup.select_one("a.book-title,.book-item a,.book_title a,h2 a,h3 a")
        if not lien: return None
        href = lien.get("href","")
        if not href.startswith("http"): href = "https://booknode.com" + href
        r2 = requests.get(href, headers=HEADERS, timeout=(5, 10))
        if r2.status_code != 200: return None
        soup2 = BeautifulSoup(r2.text, "html.parser")

        titre_el = soup2.select_one("h1.book-title,h1[itemprop='name'],h1")
        titre = titre_el.get_text(strip=True) if titre_el else ""
        if not titre: return None

        auteur_el  = soup2.select_one("[itemprop='author'],.author-name,.book-author")
        editeur_el = soup2.select_one("[itemprop='publisher'],.publisher")
        resume_el  = soup2.select_one(".book-synopsis,[itemprop='description'],.synopsis")
        serie_el   = soup2.select_one(".book-serie,.serie-name,.collection-name")

        auteur  = auteur_el.get_text(strip=True)  if auteur_el  else ""
        editeur = editeur_el.get_text(strip=True) if editeur_el else ""
        resume  = resume_el.get_text(strip=True)[:400] if resume_el else ""
        serie_r = serie_el.get_text(strip=True)   if serie_el   else ""

        # Tome souvent dans le titre sur Booknode
        tome  = extraire_tome("", titre)
        serie = extraire_serie(titre, serie_r, "", tome)
        public   = detecter_public("", titre, "", editeur, "")
        type_doc = detecter_type(titre, "", resume, "", "", editeur, public)
        genre    = detecter_genre(titre, "", resume, "", type_doc)
        pegi     = detecter_pegi(type_doc, public, "")

        return {"titre":titre,"auteur":auteur,"illustrateur":"",
                "editeur":editeur,"annee":"","type":type_doc,
                "public":public,"genre":genre,"serie":serie,"tome":tome,
                "pegi":pegi,"collection":"","resume":resume,
                "source":"Booknode","statut":"trouvé"}
    except Exception:
        return None

# ─────────────────────────────────────────────────────────
# SOURCE 13 : Google Books
# ─────────────────────────────────────────────────────────

def google_books_lookup(isbn):
    try:
        url = f"https://www.googleapis.com/books/v1/volumes?q=isbn:{isbn}"
        r = requests.get(url, headers=HEADERS, timeout=(5, 10))
        if r.status_code != 200: return None
        data = r.json()
        if data.get("totalItems",0) == 0: return None
        info    = data["items"][0]["volumeInfo"]
        titre   = info.get("title","")
        sub     = info.get("subtitle","")
        if sub: titre += f" : {sub}"
        auteur  = " / ".join(info.get("authors",[])[:2])
        editeur = info.get("publisher","")
        annee   = info.get("publishedDate","")[:4]
        resume  = info.get("description","")[:400]
        cats    = " ".join(info.get("categories",[]))
        public   = detecter_public("", titre, "", editeur, "")
        type_doc = detecter_type(titre, "", resume, cats, "", editeur, public)
        genre    = detecter_genre(titre, "", resume, cats, type_doc)
        tome     = extraire_tome("", titre)
        serie    = extraire_serie(titre, "", "", tome)
        pegi     = detecter_pegi(type_doc, public, "")
        return {"titre":titre,"auteur":auteur,"illustrateur":"",
                "editeur":editeur,"annee":annee,"type":type_doc,
                "public":public,"genre":genre,"serie":serie,"tome":tome,
                "pegi":pegi,"collection":"","resume":resume,
                "source":"Google Books","statut":"trouvé"}
    except Exception:
        return None

# ─────────────────────────────────────────────────────────
# SOURCE 14 : Open Library
# ─────────────────────────────────────────────────────────

def placedeslibraires_lookup(isbn):
    """Place des Libraires — réseau national des libraires indépendants
    (dont la Librairie Générale d'Arcachon, notre fournisseur).

    Ajoutée le 2026-07-25 après vérification en conditions réelles. C'est
    aujourd'hui la MEILLEURE source du moteur, pour trois raisons :
      - ses données viennent de Dilicom (fichier professionnel du livre
        français), donc série et tome sont normalisés, pas devinés ;
      - la fiche est adressable DIRECTEMENT par ISBN (/livre/{isbn}/), sans
        page de recherche intermédiaire ;
      - le contenu est rendu côté serveur, donc lisible sans JavaScript --
        contrairement à Mollat, Booknode, Decitre ou Ricochet, tous devenus
        illisibles pour un script.

    Format de <title> observé :
      « NARUTO Tome 19 - Masashi Kishimoto - Kana - Poche - Place des Libraires »
      « Prince de sang-mêlé - J. K. Rowling - Gallimard Jeunesse - CD Audio - ... »
    Structure : « [SÉRIE Tome N |] TITRE - AUTEUR - ÉDITEUR - SUPPORT - Place des Libraires »
    """
    try:
        url = f"https://www.placedeslibraires.fr/livre/{isbn}/"
        r = requests.get(url, headers=HEADERS, timeout=(5, 10), allow_redirects=True)
        if r.status_code != 200:
            return None
        soup = BeautifulSoup(r.text, "html.parser")
        el = soup.find("title")
        if not el:
            return None
        brut = el.get_text(strip=True)
        if not brut or "Place des Libraires" not in brut:
            return None

        # retire le suffixe « - Place des Libraires »
        corps = re.sub(r"\s*-\s*Place des Libraires\s*$", "", brut).strip()
        parts = [p.strip() for p in corps.split(" - ") if p.strip()]
        if not parts:
            return None

        titre = parts[0]
        auteur = parts[1] if len(parts) >= 2 else ""
        editeur = parts[2] if len(parts) >= 3 else ""
        support = parts[3] if len(parts) >= 4 else ""

        # « NARUTO Tome 19 » -> série + tome
        serie = tome = ""
        m = re.match(r"^(.*?)\s+Tome\s+(\d+)\s*$", titre, re.I)
        if m:
            serie = m.group(1).strip()
            tome = m.group(2)
            if not serie:
                serie = ""
        if not tome:
            tome = extraire_tome("", titre)
        if not serie:
            serie = extraire_serie(titre, "", "", tome)

        m_annee = re.search(r"\b(19|20)\d{2}\b", soup.get_text()[:4000])
        annee = m_annee.group(0) if m_annee else ""

        public = detecter_public("", titre, "", editeur, "")
        type_doc = detecter_type(titre, "", "", support, "", editeur, public)
        genre = detecter_genre(titre, "", "", support, type_doc)
        pegi = detecter_pegi(type_doc, public, "")

        if not titre:
            return None
        return {"titre": titre, "auteur": auteur, "illustrateur": "",
                "editeur": editeur, "annee": annee, "type": type_doc,
                "public": public, "genre": genre, "serie": serie, "tome": tome,
                "pegi": pegi, "collection": "", "resume": "",
                "source": "Place des Libraires", "statut": "trouvé"}
    except Exception:
        return None


def openlibrary_lookup(isbn):
    try:
        url = f"https://openlibrary.org/api/books?bibkeys=ISBN:{isbn}&format=json&jscmd=data"
        r   = requests.get(url, headers=HEADERS, timeout=(5, 10))
        if r.status_code != 200: return None
        data = r.json()
        key  = f"ISBN:{isbn}"
        if key not in data: return None
        book    = data[key]
        titre   = book.get("title","")
        auteur  = " / ".join(a.get("name","") for a in book.get("authors",[])[:2])
        pubs    = book.get("publishers",[])
        editeur = pubs[0].get("name","") if pubs else ""
        annee   = str(book.get("publish_date",""))[-4:]
        sujets  = " ".join(
            s.get("name","") if isinstance(s,dict) else str(s)
            for s in book.get("subjects",[])[:6]
        )
        if not titre: return None
        public   = detecter_public("", titre, "", editeur, "")
        type_doc = detecter_type(titre, "", "", sujets, "", editeur, public)
        genre    = detecter_genre(titre, "", "", sujets, type_doc)
        tome     = extraire_tome("", titre)
        serie    = extraire_serie(titre, "", "", tome)
        pegi     = detecter_pegi(type_doc, public, "")
        return {"titre":titre,"auteur":auteur,"illustrateur":"",
                "editeur":editeur,"annee":annee,"type":type_doc,
                "public":public,"genre":genre,"serie":serie,"tome":tome,
                "pegi":pegi,"collection":"","resume":"",
                "source":"Open Library","statut":"trouvé"}
    except Exception:
        return None

# ─────────────────────────────────────────────────────────
# MOTEUR PRINCIPAL
# ─────────────────────────────────────────────────────────

SOURCES = [
    # ── Sources vérifiées fonctionnelles le 2026-07-24/25 ──
    # Place des Libraires en TÊTE : données Dilicom (fichier professionnel du
    # livre), fiche adressable par ISBN, rendue côté serveur, série et tome
    # normalisés. C'est la source la plus fiable dont nous disposons.
    ("Place des Libraires", placedeslibraires_lookup),
    ("Cultura",       cultura_lookup),   # série+tome via le <title> de la page
    ("Open Library",  openlibrary_lookup),
    ("BnF",           bnf_lookup),       # notices officielles françaises
    # ── Sources conservées mais actuellement muettes (diagnostic 2026-07-24) ──
    # Elles ne coûtent qu'un appel qui échoue vite ; laissées en fin de liste
    # au cas où les sites redeviendraient lisibles.
    ("Decitre",       decitre_lookup),
    ("Mollat",        mollat_lookup),
    ("Manga News",    manganews_lookup),
    ("Booknode",      booknode_lookup),
    # ── Bloquées (HTTP 403 / anti-robot) : conservées pour mémoire ──
    ("Amazon",        amazon_lookup),
    ("Fnac",          fnac_lookup),
    ("BDfugue",       bdfugue_lookup),
    ("LesLibraires",  leslibraires_lookup),
]
# Babelio retirée : son robots.txt interdit explicitement l'accès automatisé.
# Google Books retirée le 2026-07-22 : API instable (503 reproductibles,
# avec ou sans clé). google_books_lookup() reste définie plus haut dans ce
# fichier si besoin de la réactiver un jour.

# Priorité par catégorie de fichier (isbn_mangas.txt, isbn_bd.txt, ...) :
# les sources listées ici sont interrogées EN PREMIER pour cette catégorie,
# le reste de SOURCES suit dans son ordre habituel.
PRIORITE_PAR_CATEGORIE = {
    "manga":           ["Manga News"],
    "bd":              ["BDfugue"],
    # Pas de source "spécialiste roman/documentaire" dans la liste (contrairement
    # à BDfugue/Manga News) — choix de départ basé sur la nature généraliste de
    # ces sites, à ajuster avec vous selon ce que vous observerez en pratique.
    "roman_jeunesse":  ["Booknode", "Decitre"],
    "roman_ado":       ["Booknode", "Decitre"],
    "documentaire":    ["BnF", "Google Books"],
}

def sources_pour_categorie(categorie):
    noms_prio = PRIORITE_PAR_CATEGORIE.get(categorie, [])
    if not noms_prio:
        return SOURCES
    prio  = [s for s in SOURCES if s[0] in noms_prio]
    reste = [s for s in SOURCES if s[0] not in noms_prio]
    return prio + reste

# Champs sur lesquels repose l'arrêt anticipé de la recherche
CHAMPS_ARRET = ("titre", "tome", "serie", "auteur", "illustrateur", "editeur", "annee")

import concurrent.futures

_executeur_sources = concurrent.futures.ThreadPoolExecutor(max_workers=4)
DELAI_MAX_SOURCE = 20  # généreux par rapport aux 12s internes à chaque requests.get,
# pour laisser le temps à une page lente normale, mais strict par rapport à un blocage
# DNS/réseau silencieux que le timeout de requests ne couvre pas toujours.


def _appeler_source_avec_delai(fn, isbn):
    futur = _executeur_sources.submit(fn, isbn)
    return futur.result(timeout=DELAI_MAX_SOURCE)


def chercher_isbn(isbn, categorie=None):
    """
    Interroge les sources (dans l'ordre adapté à la catégorie du fichier),
    collecte les résultats, choisit le meilleur titre par vote majoritaire,
    puis enrichit champ par champ.
    Les corrections manuelles (table CORRECTIONS) ont priorité absolue.

    Arrêt anticipé : dès que titre/tome/série/auteur/illustrateur/éditeur/
    année sont tous renseignés ET qu'au moins 2 sources distinctes sont
    d'accord sur le titre, on arrête d'interroger les sources restantes.
    """
    # ── Corrections manuelles prioritaires
    correction = CORRECTIONS.get(isbn)

    def normaliser_titre(t):
        return re.sub(r'[^\w\s]', '', t.lower()).strip()

    tous_resultats = []

    for nom, fn in sources_pour_categorie(categorie):
        try:
            res = _appeler_source_avec_delai(fn, isbn)
            if res and res.get("titre"):
                res["_source"] = nom
                tous_resultats.append(res)
        except Exception:
            pass
        time.sleep(0.1)

        # ── Vérifier si on peut arrêter la recherche plus tôt
        if len(tous_resultats) >= 2:
            votes = Counter(normaliser_titre(r["titre"]) for r in tous_resultats)
            _, nb_accord = votes.most_common(1)[0]
            if nb_accord >= 2:
                fusion = {}
                for r in tous_resultats:
                    for champ in CHAMPS_ARRET:
                        if not fusion.get(champ) and r.get(champ):
                            fusion[champ] = r[champ]
                # L'illustrateur ne bloque pas indéfiniment : beaucoup de BD n'ont
                # qu'un seul auteur (texte + dessin) et n'auront jamais ce champ.
                # Après 4 sources sans illustrateur trouvé, on considère le champ clos.
                champs_obligatoires = [c for c in CHAMPS_ARRET if c != "illustrateur"]
                illustrateur_ok = bool(fusion.get("illustrateur")) or len(tous_resultats) >= 4
                if all(fusion.get(champ) for champ in champs_obligatoires) and illustrateur_ok:
                    break

    if not tous_resultats:
        return {**EMPTY}

    # ── Vote majoritaire sur le titre
    def normaliser(t):
        return re.sub(r'[^\w\s]', '', t.lower()).strip()

    def score_titre(titre):
        """
        Score de qualité d'un titre :
        - Pénalise les titres commençant par un chiffre seul (ex: "7 - Archibald...")
        - Pénalise les titres trop longs (doublons avec numéro inclus)
        - Favorise les titres propres et complets
        """
        score = 0
        # Pénalité si commence par chiffre + tiret (titre mal formaté)
        if re.match(r'^\d{1,3}\s*[-–]', titre):
            score -= 10
        # Pénalité proportionnelle à la longueur (plus court = plus propre)
        score -= len(titre) * 0.01
        return score

    # Regrouper les titres similaires
    votes = {}
    for r in tous_resultats:
        tn = normaliser(r["titre"])
        votes[tn] = votes.get(tn, 0) + 1

    # Choisir le titre avec le meilleur score combiné (fréquence + qualité)
    def score_combine(tn):
        freq = votes[tn]
        # Retrouver le meilleur titre original pour ce groupe
        titres_groupe = [r["titre"] for r in tous_resultats if normaliser(r["titre"]) == tn]
        meilleur = max(titres_groupe, key=score_titre)
        return (freq, score_titre(meilleur))

    titre_gagnant_norm = max(votes, key=score_combine)
    titres_candidats = [
        r["titre"] for r in tous_resultats
        if normaliser(r["titre"]) == titre_gagnant_norm
    ]
    # Parmi les candidats, prendre le mieux scoré
    meilleur_titre = max(titres_candidats, key=score_titre)

    # ── Partir du résultat le plus complet
    tous_resultats.sort(
        key=lambda r: sum(1 for c in CHAMPS_CLES if r.get(c)),
        reverse=True
    )
    resultat = tous_resultats[0].copy()
    resultat["titre"] = meilleur_titre
    sources_ok = [resultat["_source"]]

    # ── Enrichir avec les autres sources
    for res in tous_resultats[1:]:
        enrichi = False
        for champ in ("auteur","illustrateur","editeur","annee",
                      "type","public","genre","serie","tome",
                      "pegi","collection","resume"):
            if not resultat.get(champ) and res.get(champ):
                resultat[champ] = res[champ]
                enrichi = True
        if enrichi and res["_source"] not in sources_ok:
            sources_ok.append(res["_source"])

    # ── Nettoyage titre et série
    tome  = resultat.get("tome", "")
    serie = resultat.get("serie", "")
    if resultat.get("titre"):
        resultat["titre"] = nettoyer_titre(resultat["titre"], tome, serie)
    if resultat.get("serie"):
        s = re.sub(r'\s*\d{1,3}\s*$', '', resultat["serie"]).strip(" -\u2013:,")
        s = re.sub(r'\s*[-\u2013,]\s*[Tt]ome\s*\d{1,3}.*$', '', s).strip()
        if len(s) > 2:
            resultat["serie"] = s

    # ── Fallbacks champs vides — appliqués systématiquement
    titre      = resultat.get("titre", "")
    type_doc   = resultat.get("type", "")
    public     = resultat.get("public", "")
    collection = resultat.get("collection", "").lower()
    editeur    = resultat.get("editeur", "")

    # ── PUBLIC : 4 niveaux de fallback + défaut ultime
    if not public:
        # Niveau 1 : depuis les infos du résultat principal
        public = detecter_public("", titre, collection, editeur, "")
    if not public:
        # Niveau 2 : chercher dans toutes les sources (résumés, tranches d'âge)
        for r in tous_resultats:
            texte = " ".join([r.get("titre",""), r.get("collection",""),
                              r.get("resume",""), r.get("editeur","")])
            m = re.search(r'(\d+)\s*[àa]\s*\d+\s*ans?|dès\s*(\d+)\s*ans?|à partir de\s*(\d+)', texte, re.I)
            tranche = m.group(0) if m else ""
            p = detecter_public("", r.get("titre",""), r.get("collection","").lower(),
                                r.get("editeur",""), tranche)
            if p:
                public = p
                break
    if not public:
        # Niveau 3 : depuis le type détecté
        if type_doc in ("Manga",):
            public = "Ado (12+)"
        elif type_doc in ("Album", "Première lecture"):
            public = "Dès 6 ans"
        elif type_doc == "Roman ado / YA":
            public = "Ado (12+)"
        elif type_doc == "Documentaire":
            public = "8-12 ans"
    if not public:
        # Niveau 4 : défaut ultime — tout document jeunesse a un public
        public = "Jeunesse"
    resultat["public"] = public

    # ── TYPE : 4 niveaux de fallback + défaut ultime
    if not type_doc:
        # Niveau 1 : depuis les infos du résultat principal
        type_doc = detecter_type(titre, collection, "", "", "", editeur, public)
    if not type_doc:
        # Niveau 2 : chercher dans toutes les sources
        for r in tous_resultats:
            t = detecter_type(r.get("titre",""), r.get("collection",""),
                              r.get("resume",""), "", "", r.get("editeur",""), public)
            if t:
                type_doc = t
                break
    if not type_doc:
        # Niveau 3 : depuis le public
        if public in ("Dès 3 ans",):
            type_doc = "Album"
        elif public in ("Dès 6 ans",):
            type_doc = "Première lecture"
        elif public in ("Ado (12+)",):
            type_doc = "Roman ado / YA"
    if not type_doc:
        # Niveau 4 : défaut ultime
        type_doc = "Roman jeunesse"
    resultat["type"] = type_doc

    # ── GENRE : 3 niveaux de fallback + défaut ultime
    if not resultat.get("genre"):
        # Niveau 1 : depuis toutes les sources
        for r in tous_resultats:
            g = detecter_genre(r.get("titre",""), r.get("collection",""),
                               r.get("resume",""), "", type_doc)
            if g:
                resultat["genre"] = g
                break
    if not resultat.get("genre"):
        # Niveau 2 : depuis le résumé agrégé de toutes les sources
        resume_total = " ".join(r.get("resume","") for r in tous_resultats)
        g = detecter_genre(titre, collection, resume_total, "", type_doc)
        if g:
            resultat["genre"] = g
    if not resultat.get("genre"):
        # Niveau 3 : défaut par type
        GENRE_DEFAUT = {
            "Manga":             "Aventure",
            "BD":                "Aventure / Humour",
            "Album":             "Vie quotidienne",
            "Première lecture":  "Aventure",
            "Roman jeunesse":    "Aventure",
            "Roman ado / YA":    "Aventure",
            "Documentaire":      "Sciences",
            "Conte / Poésie":    "Conte / Mythe",
            "Livre-jeu / Activités": "Cuisine / Activités",
        }
        resultat["genre"] = GENRE_DEFAUT.get(type_doc, "Vie quotidienne")

    # ── PEGI
    if not resultat.get("pegi"):
        resultat["pegi"] = detecter_pegi(type_doc, resultat.get("public",""), "")


    # PEGI : recalcul
    if not resultat.get("pegi"):
        resultat["pegi"] = detecter_pegi(type_doc, resultat.get("public",""), "")

    resultat["source"] = sources_ok[0] if len(sources_ok) == 1 \
                         else f"{sources_ok[0]} +{len(sources_ok)-1}"
    resultat["statut"] = "trouvé"

    # ── Année manquante → "N/C"
    if not resultat.get("annee"):
        resultat["annee"] = "N/C"

    # ── Normaliser l'éditeur
    if resultat.get("editeur"):
        resultat["editeur"] = normaliser_editeur(resultat["editeur"])

    # ── Appliquer les corrections manuelles (priorité absolue)
    if correction:
        for champ, val in correction.items():
            if val:  # ne pas écraser avec une valeur vide
                resultat[champ] = val
        if "source" in resultat:
            resultat["source"] = resultat["source"] + " [corrigé]"

    return resultat


