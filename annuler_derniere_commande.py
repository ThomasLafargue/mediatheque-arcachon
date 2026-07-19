#!/usr/bin/env python3
"""
Annule la dernière proposition de commande ET la proposition de désherbage
qui va avec (les deux onglets créés ensemble par nouvelle_acquisition.py).

Important : ce script supprime uniquement les onglets de PROPOSITION
("Commande du ..." et "Désherbage du ..."). Il ne touche JAMAIS à
l'Inventaire ni à l'onglet Pilon — si des livres ont déjà été reçus ou
sortis du fonds, ils restent tels quels, seul le suivi de la commande
elle-même est supprimé.

Usage : python3 annuler_derniere_commande.py
"""

import os, sys
from openpyxl import load_workbook

script_dir = os.path.dirname(os.path.abspath(__file__))
output_file = os.path.join(script_dir, "inventaire_mediatheque.xlsx")

print("="*65)
print("  Annulation de la dernière commande — Médiathèque d'Arcachon")
print("="*65)

if not os.path.exists(output_file):
    print(f"\n⚠  Fichier introuvable : {output_file}")
    input("\nEntrée pour quitter...")
    sys.exit(1)

wb = load_workbook(output_file)

feuilles_commande = [n for n in wb.sheetnames if n.startswith("Commande du ")]
feuilles_desherbage = [n for n in wb.sheetnames if n.startswith("Désherbage du ") or n.startswith("Proposition desherbage du ")]

if not feuilles_commande and not feuilles_desherbage:
    print("\n  Aucun onglet de commande ou de désherbage trouvé. Rien à annuler.")
    input("\nEntrée pour quitter...")
    sys.exit(0)

derniere_commande = feuilles_commande[-1] if feuilles_commande else None
dernier_desherbage = feuilles_desherbage[-1] if feuilles_desherbage else None

print(f"\nDernière commande trouvée    : {derniere_commande or '(aucune)'}")
print(f"Dernier désherbage trouvé    : {dernier_desherbage or '(aucun)'}")

def compter_coches(nom_feuille, colonne_index):
    if not nom_feuille:
        return 0
    feuille = wb[nom_feuille]
    return sum(
        1 for row in feuille.iter_rows(min_row=1, values_only=True)
        if row and len(row) > colonne_index and row[colonne_index] == "Oui"
    )

nb_recus = compter_coches(derniere_commande, 7)   # colonne H = "Reçu ?"
nb_sortis = compter_coches(dernier_desherbage, 7)  # colonne H = "Sorti ?"

if nb_recus:
    print(f"\n  ⚠  {nb_recus} tome(s) de cette commande sont déjà marqués 'Reçu' — "
          f"ils resteront dans l'Inventaire, mais ce suivi sera perdu si vous continuez.")
if nb_sortis:
    print(f"  ⚠  {nb_sortis} document(s) de ce désherbage sont déjà marqués 'Sorti' — "
          f"ils resteront retirés de l'Inventaire, mais ce suivi sera perdu si vous continuez.")

print("\nCeci supprimera uniquement ces deux onglets de proposition.")
print("L'Inventaire et l'onglet Pilon ne seront PAS modifiés.")
reponse = input("\nConfirmer la suppression ? (oui/non) : ").strip().lower()

if reponse not in ("oui", "o", "yes", "y"):
    print("\nAnnulé : rien n'a été supprimé.")
    input("\nEntrée pour quitter...")
    sys.exit(0)

if derniere_commande:
    del wb[derniere_commande]
    print(f"  🗑  Onglet « {derniere_commande} » supprimé.")
if dernier_desherbage:
    del wb[dernier_desherbage]
    print(f"  🗑  Onglet « {dernier_desherbage} » supprimé.")

wb.save(output_file)
print("\n✓ Terminé.")
print("="*65)
print(f"\n  Commandes utiles :")
print(f"    python3 recherche_isbn.py        → lancer une nouvelle recherche")
print(f"    python3 maj_statistiques.py      → mettre à jour les stats sans recherche")
print(f"    python3 nouvelle_acquisition.py  → préparer une nouvelle commande")
print(f"    python3 annuler_derniere_commande.py → annuler la dernière commande/désherbage")
print(f"    caffeinate                       → empêcher le Mac de dormir (autre onglet)")
print("="*65)
input("\nEntrée pour quitter...")
