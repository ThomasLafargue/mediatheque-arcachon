"""
Outil conversationnel -- Médiathèque d'Arcachon.

Interface de chat qui interroge inventaire.db (vue_inventaire) en langage
naturel, via l'API Claude (function calling / tool use). Pensé pour être
déployé sur Streamlit Community Cloud et accessible depuis n'importe où
(bibliothèque, domicile, téléphone) via une simple URL.

Lancement local :
    pip install -r requirements_app.txt
    export ANTHROPIC_API_KEY="sk-ant-..."
    streamlit run app_conversationnel.py
"""

import os
import re
import json
import io
import sys
import time
import tempfile
import datetime
import sqlite3
import streamlit as st

# Pont secrets Streamlit Cloud -> variables d'environnement, AVANT d'importer
# db (qui lit ces variables au chargement). En local, ce pont ne fait rien
# de plus que ce que le fichier .env fait déjà ; sur Streamlit Cloud, c'est
# ce qui permet à db.py de détecter Turso sans aucun fichier .env présent.
for _cle in (
    "TURSO_DATABASE_URL", "TURSO_AUTH_TOKEN",
    # Identifiants OVH (envoi des écrans mosaïque/diaporama sur le serveur)
    # pour que la mise à jour des écrans soit possible depuis l'app déployée,
    # pas seulement depuis le Terminal du Mac.
    "OVH_SFTP_HOST", "OVH_SFTP_PORT", "OVH_SFTP_USER",
    "OVH_SFTP_PASSWORD", "OVH_SFTP_DOSSIER",
):
    try:
        if _cle in st.secrets:
            os.environ[_cle] = str(st.secrets[_cle])
    except Exception:
        pass  # pas de secrets.toml en local -- db.py se rabat sur .env / sqlite local

# Clé Google Books depuis secrets Streamlit Cloud.
# NB (audit 2026-07-25) : Google Books est RETIRÉ du pipeline d'enrichissement
# depuis le 2026-07-22 (mesuré le 25/07 : 1 livre trouvé sur 8 ISBN français,
# 0 information de série, 5 erreurs 503). La clé reste chargée ici sans effet,
# au cas où la source redeviendrait exploitable.
try:
    if "GOOGLE_BOOKS_API_KEY" in st.secrets:
        os.environ["GOOGLE_BOOKS_API_KEY"] = st.secrets["GOOGLE_BOOKS_API_KEY"]
except Exception:
    pass

import db
from anthropic import Anthropic

# Sources API bibliographiques (BnF SRU + Sudoc ; Google Books désactivé)
try:
    from sources_api import enrichir_par_api
    SOURCES_API_OK = True
except ImportError:
    SOURCES_API_OK = False

# Analyse des besoins d'acquisition (signaux internes + démographie)
try:
    from analyser_acquisition import analyser_besoins, PROFIL_ARCACHON
    ANALYSE_ACQUISITION_OK = True
except ImportError:
    ANALYSE_ACQUISITION_OK = False

FICHIER_DB = os.path.join(os.path.dirname(os.path.abspath(__file__)), "inventaire.db")

st.set_page_config(page_title="Médiathèque d'Arcachon — Assistant fonds", page_icon="📚", layout="wide")


# ----------------------------------------------------------------------------
# Accès en lecture seule à la base -- aucune écriture possible depuis l'outil
# ----------------------------------------------------------------------------
MOTS_INTERDITS = re.compile(
    r"\b(INSERT|UPDATE|DELETE|DROP|ALTER|CREATE|REPLACE|ATTACH|PRAGMA|VACUUM)\b",
    re.IGNORECASE,
)


def executer_requete_sql(sql: str) -> str:
    sql_nettoyee = sql.strip().rstrip(';')
    if not re.match(r'^\s*SELECT\b', sql_nettoyee, re.IGNORECASE):
        return json.dumps({"erreur": "Seules les requêtes SELECT sont autorisées."})
    if MOTS_INTERDITS.search(sql_nettoyee):
        return json.dumps({"erreur": "Mot-clé non autorisé dans cette requête."})
    try:
        conn = db.connect(FICHIER_DB)
        conn.row_factory = sqlite3.Row if not db.MODE_EN_LIGNE else db.Row
        cur = conn.cursor()
        cur.execute(f"SELECT * FROM ({sql_nettoyee}) LIMIT 500")
        lignes = [dict(r) for r in cur.fetchall()]
        conn.close()
        return json.dumps(lignes, ensure_ascii=False, default=str)
    except Exception as e:
        return json.dumps({"erreur": str(e)})


OUTIL_SQL = {
    "name": "executer_requete_sql",
    "description": (
        "Exécute une requête SQL SELECT en lecture seule sur la base de la médiathèque. "
        "Table principale à utiliser : vue_inventaire (une ligne par exemplaire réel "
        "d'Arcachon). Colonnes disponibles : isbn, titre, serie, tome, collection, type "
        "(LIVRE/CD/DVD/JEU/REVUE/AUTRE), categorie, genre, public, age_recommande, pegi, "
        "auteur, illustrateur, traducteur, editeur, annee, dewey, dewey_libelle, "
        "mots_cles, description_physique, code_barres, cote, statut_exemplaire, prix, "
        "nb_prets_titre_reseau, nb_prets_cet_exemplaire, dernier_pret_titre_reseau, "
        "dernier_pret_cet_exemplaire, resume, "
        "champs_a_verifier_decalog -- liste (ex: 'serie,tome') des champs que "
        "NOTRE moteur d'enrichissement (BnF + sites web) a déduits alors que "
        "Decalog les avait laissés vides ; NULL si tout vient de Decalog ou si "
        "rien n'a été déduit. On ne réécrit jamais Decalog directement, donc "
        "quand cette colonne n'est pas NULL, la valeur affichée ici (serie/tome) "
        "est correcte dans notre base mais reste fausse/vide dans Decalog tant "
        "que le bibliothécaire ne l'y corrige pas à la main. Utiliser cette "
        "colonne pour répondre aux questions du type 'qu'est-ce qui est mal "
        "renseigné dans Decalog' ou 'quelles fiches dois-je corriger' : "
        "sélectionner les notices où champs_a_verifier_decalog IS NOT NULL et "
        "présenter, pour chacune, la valeur qu'on a trouvée (ex: tome='5') en "
        "précisant qu'elle est absente/nulle côté Decalog. "
        "age_joueurs (ex: 'A partir de 6 ans'), nb_joueurs_min, nb_joueurs_max, "
        "duree_partie (ex: '30 minutes') -- ces 4 colonnes sont spécifiques aux jeux. "
        "HIÉRARCHIE pour regrouper ou affiner : "
        "DVD large = type='DVD' -- Films d'animation = categorie='Film d\\'animation' -- "
        "Films documentaires = categorie='Film documentaire' -- "
        "CD large = type='CD' -- "
        "Jeux large = type='JEU' -- Jeux vidéo = categorie='Jeu vidéo' -- "
        "BD large = categorie='BD' -- BD jeunesse = BD + cote LIKE 'BDJ%' -- "
        "Comics = genre='Comics' -- Roman graphique = genre='Roman graphique' -- "
        "Gros caractères = categorie='Gros caractères' -- "
        "Textes lus = categorie='Textes lus' -- "
        "Fonds local = categorie='Fonds local'. "
        "IMPORTANT : pour les séries, filtrer sur serie (pas titre). "
        "Autres tables disponibles si besoin : "
        "frequentation (date, nb_entrees) -- total journalier ; "
        "frequentation_horaire (date, heure, nb_entrees) -- détail par tranche "
        "horaire (ex. '14:00' = de 14h à 15h), utile pour analyser les heures "
        "de pointe, comparer matinées et après-midis, trouver le créneau le plus "
        "fréquenté d'une journée donnée ; "
        "horaires_ouverture (periode, mois_debut, mois_fin, jour_semaine 0=lundi, "
        "heure_ouverture, heure_fermeture, ouvert_public 1=oui/0=non, note) -- "
        "horaires réels de la médiathèque : juillet-août lundi-samedi 10h-19h, "
        "reste de l'année mardi-samedi 10h-18h, lundi hors été = agents seulement "
        "fermé au public. Utile pour calculer des taux de fréquentation, identifier "
        "les heures de pointe par rapport aux heures d'ouverture, ou savoir si un "
        "jour donné était ouvert au public ; "
        "suggestion_acquisition (id, titre, demandeur, auteur, editeur, isbn, prix, "
        "motif, source, statut, date_ajout) -- SANS 's' à 'suggestion' ; "
        "suggestion_desherbage (id, titre, demandeur, isbn, cote, motif, nb_prets, "
        "dernier_pret, statut, date_ajout) ; "
        "suggestion_mise_en_avant (id, titre, demandeur, isbn, cote, motif, nb_prets, "
        "statut, date_ajout) ; "
        "desherbage_effectue (id, titre, demandeur, isbn, cote, motif, "
        "nb_prets_au_retrait, dernier_pret_au_retrait, date_retrait) -- "
        "historique des retraits réellement effectués (distinct des suggestions) ; "
        "journal_requetes (id, date_requete, question, sql_executees, nb_recherches_web, "
        "a_genere_export, a_modifie_suggestions, erreur) -- historique de toutes les "
        "questions posées, utile pour analyser l'usage de l'outil lui-même. "
        "Toujours préférer nb_prets_cet_exemplaire pour des questions sur le fonds "
        "d'Arcachon spécifiquement.\n"
        "⚠️ LIMITE DE 500 LIGNES PAR REQUÊTE — RÈGLE ABSOLUE : le résultat est "
        "TOUJOURS tronqué à 500 lignes. Ne JAMAIS déduire un total du nombre de "
        "lignes reçues : si tu reçois 500 lignes, cela signifie « au moins 500 », "
        "jamais « 500 au total ». Pour annoncer un effectif, faire OBLIGATOIREMENT "
        "une requête COUNT(*) séparée AVANT de lister. Annoncer un total faux à "
        "partir d'un résultat tronqué est une erreur grave (constatée le "
        "2026-07-26 : « ~500 fiches » annoncé alors qu'il y en avait 5 616)."
    ),
    "input_schema": {
        "type": "object",
        "properties": {"sql": {"type": "string", "description": "La requête SELECT à exécuter"}},
        "required": ["sql"],
    },
}

COLONNES_EXPORT_PAR_DEFAUT = [
    ("isbn", "ISBN / EAN"), ("titre", "Titre"), ("serie", "Série"), ("tome", "Tome"),
    ("auteur", "Auteur"), ("illustrateur", "Illustrateur"), ("editeur", "Éditeur"),
    ("annee", "Année"), ("type", "Type"), ("categorie", "Catégorie"), ("genre", "Genre"),
    ("public", "Public"), ("age_recommande", "Âge conseillé"), ("cote", "Cote"),
    ("code_barres", "Code-barres"), ("statut_exemplaire", "Statut"), ("prix", "Prix (€)"),
    ("nb_prets_cet_exemplaire", "Prêts (Arcachon)"), ("dernier_pret_cet_exemplaire", "Dernier prêt"),
]


def generer_excel_bytes(sql=None, lignes_fournies=None):
    """Construit un export xlsx en mémoire.
    - lignes_fournies : liste de dicts construite par Claude lui-même (ex.
      suggestions d'acquisition trouvées par recherche web) -- utilisée en
      priorité si fournie.
    - sql : sinon, requête sur notre base (mêmes garde-fous SELECT-only
      qu'executer_requete_sql). Sans aucun des deux : fonds complet."""
    if lignes_fournies:
        lignes = lignes_fournies
        colonnes_presentes = list(lignes[0].keys())
        colonnes = [(c, c.replace('_', ' ').capitalize()) for c in colonnes_presentes]
        return _ecrire_xlsx(lignes, colonnes, acces_par_cle=True)

    sql_finale = (sql or "SELECT * FROM vue_inventaire").strip().rstrip(';')
    if not re.match(r'^\s*SELECT\b', sql_finale, re.IGNORECASE):
        return None, 0, "Seules les requêtes SELECT sont autorisées pour un export."
    if MOTS_INTERDITS.search(sql_finale):
        return None, 0, "Mot-clé non autorisé dans cette requête d'export."

    conn = db.connect(FICHIER_DB)
    conn.row_factory = sqlite3.Row if not db.MODE_EN_LIGNE else db.Row
    cur = conn.cursor()
    cur.execute(f"SELECT * FROM ({sql_finale}) LIMIT 100000")
    lignes = cur.fetchall()
    conn.close()

    if not lignes:
        return None, 0, "Aucune ligne ne correspond à cet export."

    colonnes_presentes = list(lignes[0].keys())
    noms_par_defaut = {c for c, _ in COLONNES_EXPORT_PAR_DEFAUT}
    if noms_par_defaut.issubset(set(colonnes_presentes)):
        colonnes = [(c, l) for c, l in COLONNES_EXPORT_PAR_DEFAUT if c in colonnes_presentes]
    else:
        colonnes = [(c, c) for c in colonnes_presentes]

    return _ecrire_xlsx(lignes, colonnes, acces_par_cle=True)


def _ecrire_xlsx(lignes, colonnes, acces_par_cle=True):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Export"
    thin = Side(style="thin", color="D9E2F0")
    bordure = Border(left=thin, right=thin, top=thin, bottom=thin)

    for c, (cle, libelle) in enumerate(colonnes, 1):
        cellule = ws.cell(row=1, column=c, value=libelle)
        cellule.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cellule.fill = PatternFill("solid", fgColor="2E4A7A")
        cellule.border = bordure
        ws.column_dimensions[get_column_letter(c)].width = 18

    for i, ligne in enumerate(lignes, start=2):
        for c, (cle, libelle) in enumerate(colonnes, 1):
            valeur = ligne.get(cle) if isinstance(ligne, dict) else ligne[cle]
            cellule = ws.cell(row=i, column=c, value=valeur)
            cellule.font = Font(name="Arial", size=9.5)
            cellule.border = bordure

    ws.freeze_panes = "A2"
    tampon = io.BytesIO()
    wb.save(tampon)
    tampon.seek(0)
    return tampon.getvalue(), len(lignes), None


OUTIL_EXPORT = {
    "name": "generer_export_excel",
    "description": (
        "Génère un fichier Excel téléchargeable, de deux façons possibles :\n"
        "1) lignes : pour exporter une liste que TU as construite toi-même "
        "(typiquement des suggestions d'acquisition trouvées par web_search). "
        "Fournis une liste d'objets, chacun avec les mêmes champs (ex. titre, "
        "auteur, editeur, prix_estime, source). L'ISBN n'est PAS nécessaire "
        "pour une liste de suggestions -- ne le cherche pas, ne bloque jamais "
        "l'export en son absence. N'utilise CETTE option QUE pour des données "
        "venant réellement d'un résultat de web_search -- jamais de titres "
        "inventés.\n"
        "2) sql : pour exporter ce qui EST dans notre fonds (résultat d'une "
        "requête SELECT). Sans aucun argument : exporte le fonds complet.\n"
        "N'utilise jamais sql pour répondre à une demande de titres absents "
        "du fonds -- utilise lignes avec des résultats de web_search à la "
        "place, ou dis clairement que tu ne peux pas produire cet export."
    ),
    "input_schema": {
        "type": "object",
        "properties": {
            "sql": {"type": "string", "description": "Requête SELECT sur notre base (ce qu'on possède déjà)"},
            "lignes": {
                "type": "array",
                "description": "Liste d'objets à exporter directement (ex. suggestions trouvées par web_search)",
                "items": {"type": "object"},
            },
        },
    },
}


def ajouter_suggestion_acquisition(titre, demandeur, auteur=None, editeur=None, isbn=None,
                                    prix=None, motif=None, source=None):
    """Ajoute une ligne dans une liste de suggestions persistante (table
    suggestion_acquisition), consultable et exportable à tout moment via
    executer_requete_sql / generer_export_excel, y compris filtrée par
    demandeur. Nécessite le jeton d'écriture (TURSO_AUTH_TOKEN_ECRITURE) --
    l'opération est strictement limitée à cet ajout précis, jamais une
    écriture arbitraire."""
    jeton_ecriture = st.secrets.get("TURSO_AUTH_TOKEN_ECRITURE", "")
    if not jeton_ecriture:
        return json.dumps({"erreur": "Fonction non configurée (TURSO_AUTH_TOKEN_ECRITURE manquant)."})
    try:
        conn = db.connect_avec_jeton(db.TURSO_URL, jeton_ecriture) if db.MODE_EN_LIGNE else db.connect(FICHIER_DB)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS suggestion_acquisition (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                titre TEXT NOT NULL,
                demandeur TEXT,
                auteur TEXT,
                editeur TEXT,
                isbn TEXT,
                prix REAL,
                motif TEXT,
                source TEXT,
                statut TEXT NOT NULL DEFAULT 'à étudier',
                date_ajout TEXT NOT NULL DEFAULT (datetime('now'))
            )
        """)
        try:  # migration silencieuse si la table existait déjà sans cette colonne
            conn.execute("ALTER TABLE suggestion_acquisition ADD COLUMN demandeur TEXT")
            conn.commit()
        except Exception:
            pass
        try:  # vue de compatibilité -- le nom "naturel" en français est au
            # pluriel ("liste de suggestions"), donc le modèle le devine
            # parfois ainsi malgré la table réelle au singulier. Plutôt que
            # de compter sur un texte d'instruction pour éviter l'erreur à
            # chaque fois, les deux noms fonctionnent désormais tous les deux.
            conn.execute("CREATE VIEW IF NOT EXISTS suggestions_acquisition AS SELECT * FROM suggestion_acquisition")
            conn.commit()
        except Exception:
            pass
        conn.execute(
            "INSERT INTO suggestion_acquisition (titre, demandeur, auteur, editeur, isbn, prix, motif, source) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            (titre, demandeur, auteur, editeur, isbn, prix, motif, source),
        )
        conn.commit()
        conn.close()
        return json.dumps({"statut": "ok", "info": f"« {titre} » ajouté à la liste de {demandeur}."})
    except Exception as e:
        import traceback
        st.session_state["derniere_erreur_technique"] = traceback.format_exc()
        return json.dumps({"erreur": f"{type(e).__name__}: {e}"})


OUTIL_SUGGESTION = {
    "name": "ajouter_suggestion_acquisition",
    "description": (
        "Ajoute un titre à une liste de suggestions d'acquisition PERSISTANTE, "
        "rattachée à une personne (demandeur). Utilise cet outil quand on te "
        "demande d'ajouter/noter/mettre un titre dans une liste de suggestions "
        "-- ne dis jamais que tu ne peux pas le faire, tu en es capable. Si la "
        "demande ne précise pas pour qui (quel demandeur), demande-le avant "
        "d'ajouter. Vérifie d'abord que le titre n'est pas déjà dans le fonds "
        "(executer_requete_sql) avant de l'ajouter. Pour consulter ou exporter "
        "la liste d'une personne en particulier, utilise executer_requete_sql "
        "ou generer_export_excel avec un filtre WHERE demandeur = '...'."
    ),
    "input_schema": {
        "type": "object",
        "properties": {
            "titre": {"type": "string"},
            "demandeur": {"type": "string", "description": "Qui fait cette suggestion (prénom de l'agent)"},
            "auteur": {"type": "string"},
            "editeur": {"type": "string"},
            "isbn": {"type": "string"},
            "prix": {"type": "number"},
            "motif": {"type": "string", "description": "Raison de la suggestion (ex. demande usager)"},
            "source": {"type": "string", "description": "D'où vient l'info (ex. URL trouvée par web_search)"},
        },
        "required": ["titre", "demandeur"],
    },
}


def supprimer_suggestion_acquisition(id):
    """Supprime UNE ligne précise de la liste de suggestions, par son id.
    Trouve d'abord l'id via executer_requete_sql avant d'appeler ceci --
    l'opération est strictement limitée à cette suppression précise par id,
    jamais une suppression arbitraire (pas de WHERE libre)."""
    jeton_ecriture = st.secrets.get("TURSO_AUTH_TOKEN_ECRITURE", "")
    if not jeton_ecriture:
        return json.dumps({"erreur": "Fonction non configurée (TURSO_AUTH_TOKEN_ECRITURE manquant)."})
    try:
        conn = db.connect_avec_jeton(db.TURSO_URL, jeton_ecriture) if db.MODE_EN_LIGNE else db.connect(FICHIER_DB)
        cur = conn.cursor()
        cur.execute("SELECT titre FROM suggestion_acquisition WHERE id = ?", (id,))
        ligne = cur.fetchone()
        if not ligne:
            conn.close()
            return json.dumps({"erreur": f"Aucune suggestion avec l'id {id}."})
        titre = ligne[0]
        conn.execute("DELETE FROM suggestion_acquisition WHERE id = ?", (id,))
        conn.commit()
        conn.close()
        return json.dumps({"statut": "ok", "info": f"« {titre} » (id {id}) supprimé de la liste."})
    except Exception as e:
        import traceback
        st.session_state["derniere_erreur_technique"] = traceback.format_exc()
        return json.dumps({"erreur": f"{type(e).__name__}: {e}"})


OUTIL_SUPPRESSION_SUGGESTION = {
    "name": "supprimer_suggestion_acquisition",
    "description": (
        "Supprime DÉFINITIVEMENT une suggestion (perte de toute trace). À "
        "n'utiliser QUE pour corriger une erreur de saisie (doublon, entrée "
        "fautive). Pour une décision normale -- accepter (à commander/acquérir) "
        "ou refuser un titre -- utilise plutôt statuer_suggestion_acquisition, "
        "qui garde l'historique. Utilise executer_requete_sql d'abord pour "
        "trouver l'id exact. Si plusieurs lignes correspondent, demande "
        "confirmation."
    ),
    "input_schema": {
        "type": "object",
        "properties": {"id": {"type": "integer", "description": "L'id exact de la ligne à supprimer"}},
        "required": ["id"],
    },
}


STATUTS_SUGGESTION_AUTORISES = ("à étudier", "à commander", "acquise", "écartée")


def statuer_suggestion_acquisition(id, statut):
    """Change le STATUT d'une suggestion (par id), sans jamais l'effacer :
    c'est la façon normale d'accepter ou de refuser un titre de la liste de
    veille.
      - 'à commander' / 'acquise' = décision positive (on prend le titre)
      - 'écartée'                 = décision négative (on ne le prend pas)
      - 'à étudier'               = remis en attente
    Garder la ligne (plutôt que la supprimer) a un double intérêt : on
    conserve l'historique des décisions, ET comme la veille hebdomadaire
    dédoublonne par titre sur TOUTE la table, un titre 'écarté' ne réapparaît
    jamais dans les suggestions la semaine suivante -- la décision est donc
    définitive sans rien perdre."""
    if statut not in STATUTS_SUGGESTION_AUTORISES:
        return json.dumps({"erreur": f"Statut non autorisé. Valeurs possibles : {', '.join(STATUTS_SUGGESTION_AUTORISES)}."})
    jeton_ecriture = st.secrets.get("TURSO_AUTH_TOKEN_ECRITURE", "")
    if not jeton_ecriture:
        return json.dumps({"erreur": "Fonction non configurée (TURSO_AUTH_TOKEN_ECRITURE manquant)."})
    try:
        conn = db.connect_avec_jeton(db.TURSO_URL, jeton_ecriture) if db.MODE_EN_LIGNE else db.connect(FICHIER_DB)
        cur = conn.cursor()
        cur.execute("SELECT titre FROM suggestion_acquisition WHERE id = ?", (id,))
        ligne = cur.fetchone()
        if not ligne:
            conn.close()
            return json.dumps({"erreur": f"Aucune suggestion avec l'id {id}."})
        titre = ligne[0]
        conn.execute("UPDATE suggestion_acquisition SET statut = ? WHERE id = ?", (statut, id))
        conn.commit()
        conn.close()
        return json.dumps({"statut": "ok", "info": f"« {titre} » (id {id}) → statut « {statut} »."})
    except Exception as e:
        import traceback
        st.session_state["derniere_erreur_technique"] = traceback.format_exc()
        return json.dumps({"erreur": f"{type(e).__name__}: {e}"})


OUTIL_STATUER_SUGGESTION = {
    "name": "statuer_suggestion_acquisition",
    "description": (
        "Décide du sort d'une suggestion d'acquisition (accepter ou refuser), "
        "en changeant son statut SANS l'effacer. C'est l'outil à utiliser quand "
        "on te dit d'accepter, valider, commander, acquérir, prendre un titre, "
        "OU de refuser, écarter, ne pas prendre un titre de la liste. "
        "Valeurs de statut :\n"
        "  • 'à commander' : décision positive, le titre sera commandé\n"
        "  • 'acquise'     : le titre a été acquis / est arrivé\n"
        "  • 'écartée'     : décision négative, on ne prend pas ce titre "
        "(il ne réapparaîtra plus dans la veille)\n"
        "  • 'à étudier'   : remettre en attente\n"
        "Trouve d'abord l'id exact via executer_requete_sql (ex. SELECT id, titre, "
        "statut FROM suggestion_acquisition WHERE titre LIKE '%...%'). Ne devine "
        "jamais un id. Si plusieurs lignes correspondent, demande laquelle. "
        "Tu peux traiter plusieurs titres à la suite (un appel par id)."
    ),
    "input_schema": {
        "type": "object",
        "properties": {
            "id": {"type": "integer", "description": "L'id exact de la suggestion"},
            "statut": {
                "type": "string",
                "enum": list(STATUTS_SUGGESTION_AUTORISES),
                "description": "Le nouveau statut",
            },
        },
        "required": ["id", "statut"],
    },
}


# ─────────────────────── DÉSHERBAGE ───────────────────────────────────────────

def ajouter_suggestion_desherbage(titre, demandeur, isbn=None, cote=None,
                                   motif=None, nb_prets=None, dernier_pret=None):
    """Ajoute une suggestion de désherbage dans une liste persistante, par agent.
    Symétrique à suggestion_acquisition. Nécessite le jeton d'écriture."""
    jeton_ecriture = st.secrets.get("TURSO_AUTH_TOKEN_ECRITURE", "")
    if not jeton_ecriture:
        return json.dumps({"erreur": "Fonction non configurée (TURSO_AUTH_TOKEN_ECRITURE manquant)."})
    try:
        conn = db.connect_avec_jeton(db.TURSO_URL, jeton_ecriture) if db.MODE_EN_LIGNE else db.connect(FICHIER_DB)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS suggestion_desherbage (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                titre TEXT NOT NULL, demandeur TEXT, isbn TEXT, cote TEXT,
                motif TEXT, nb_prets INTEGER, dernier_pret TEXT,
                statut TEXT NOT NULL DEFAULT 'à valider',
                date_ajout TEXT NOT NULL DEFAULT (datetime('now'))
            )
        """)
        conn.execute(
            "INSERT INTO suggestion_desherbage (titre, demandeur, isbn, cote, motif, nb_prets, dernier_pret) "
            "VALUES (?, ?, ?, ?, ?, ?, ?)",
            (titre, demandeur, isbn, cote, motif, nb_prets, dernier_pret),
        )
        conn.commit()
        conn.close()
        return json.dumps({"statut": "ok", "info": f"« {titre} » ajouté aux suggestions de désherbage de {demandeur}."})
    except Exception as e:
        import traceback; st.session_state["derniere_erreur_technique"] = traceback.format_exc()
        return json.dumps({"erreur": f"{type(e).__name__}: {e}"})


def supprimer_suggestion_desherbage(id):
    jeton_ecriture = st.secrets.get("TURSO_AUTH_TOKEN_ECRITURE", "")
    if not jeton_ecriture:
        return json.dumps({"erreur": "TURSO_AUTH_TOKEN_ECRITURE manquant."})
    try:
        conn = db.connect_avec_jeton(db.TURSO_URL, jeton_ecriture) if db.MODE_EN_LIGNE else db.connect(FICHIER_DB)
        cur = conn.cursor(); cur.execute("SELECT titre FROM suggestion_desherbage WHERE id = ?", (id,))
        ligne = cur.fetchone()
        if not ligne:
            conn.close(); return json.dumps({"erreur": f"Aucune suggestion de désherbage avec l'id {id}."})
        conn.execute("DELETE FROM suggestion_desherbage WHERE id = ?", (id,))
        conn.commit(); conn.close()
        return json.dumps({"statut": "ok", "info": f"Suggestion de désherbage id {id} (« {ligne[0]} ») supprimée."})
    except Exception as e:
        return json.dumps({"erreur": f"{type(e).__name__}: {e}"})


OUTIL_DESHERBAGE = {
    "name": "ajouter_suggestion_desherbage",
    "description": (
        "Ajoute un titre à la liste de suggestions de DÉSHERBAGE (retrait du fonds) "
        "d'un agent précis. Utilise cet outil quand on demande de noter un titre "
        "à étudier pour le pilon ou le retrait. AVANT d'ajouter, interroge "
        "executer_requete_sql pour récupérer les vraies données de l'exemplaire "
        "(nb_prets_cet_exemplaire, dernier_pret_cet_exemplaire, cote, isbn) et "
        "inclus-les -- ce sont les données qui permettront la décision humaine. "
        "Si le motif n'est pas précisé, déduis-le des données (ex. '0 prêt depuis "
        "2015, documentaire périmé'). La validation finale reste humaine."
    ),
    "input_schema": {
        "type": "object",
        "properties": {
            "titre": {"type": "string"},
            "demandeur": {"type": "string"},
            "isbn": {"type": "string"},
            "cote": {"type": "string"},
            "motif": {"type": "string", "description": "Raison du désherbage (ex. '0 prêt depuis 2015')"},
            "nb_prets": {"type": "integer"},
            "dernier_pret": {"type": "string"},
        },
        "required": ["titre", "demandeur"],
    },
}

OUTIL_SUPPRESSION_DESHERBAGE = {
    "name": "supprimer_suggestion_desherbage",
    "description": "Supprime UNE suggestion de désherbage par son id (après vérification via executer_requete_sql).",
    "input_schema": {"type": "object", "properties": {"id": {"type": "integer"}}, "required": ["id"]},
}


# ─────────────────────── MISE EN AVANT (PÉPITES) ───────────────────────────────

def ajouter_suggestion_mise_en_avant(titre, demandeur, isbn=None, cote=None,
                                      motif=None, nb_prets=None):
    """Ajoute une suggestion de mise en avant (présentoir, coup de cœur...)
    dans une liste persistante. Pour les pépites méconnues : livres peu empruntés
    mais de qualité reconnue (prix littéraires, bonnes critiques)."""
    jeton_ecriture = st.secrets.get("TURSO_AUTH_TOKEN_ECRITURE", "")
    if not jeton_ecriture:
        return json.dumps({"erreur": "TURSO_AUTH_TOKEN_ECRITURE manquant."})
    try:
        conn = db.connect_avec_jeton(db.TURSO_URL, jeton_ecriture) if db.MODE_EN_LIGNE else db.connect(FICHIER_DB)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS suggestion_mise_en_avant (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                titre TEXT NOT NULL, demandeur TEXT, isbn TEXT, cote TEXT,
                motif TEXT, nb_prets INTEGER,
                statut TEXT NOT NULL DEFAULT 'à programmer',
                date_ajout TEXT NOT NULL DEFAULT (datetime('now'))
            )
        """)
        conn.execute(
            "INSERT INTO suggestion_mise_en_avant (titre, demandeur, isbn, cote, motif, nb_prets) VALUES (?, ?, ?, ?, ?, ?)",
            (titre, demandeur, isbn, cote, motif, nb_prets),
        )
        conn.commit(); conn.close()
        return json.dumps({"statut": "ok", "info": f"« {titre} » ajouté aux suggestions de mise en avant de {demandeur}."})
    except Exception as e:
        return json.dumps({"erreur": f"{type(e).__name__}: {e}"})


def supprimer_suggestion_mise_en_avant(id):
    jeton_ecriture = st.secrets.get("TURSO_AUTH_TOKEN_ECRITURE", "")
    if not jeton_ecriture:
        return json.dumps({"erreur": "TURSO_AUTH_TOKEN_ECRITURE manquant."})
    try:
        conn = db.connect_avec_jeton(db.TURSO_URL, jeton_ecriture) if db.MODE_EN_LIGNE else db.connect(FICHIER_DB)
        cur = conn.cursor(); cur.execute("SELECT titre FROM suggestion_mise_en_avant WHERE id = ?", (id,))
        ligne = cur.fetchone()
        if not ligne:
            conn.close(); return json.dumps({"erreur": f"Aucune suggestion de mise en avant avec l'id {id}."})
        conn.execute("DELETE FROM suggestion_mise_en_avant WHERE id = ?", (id,))
        conn.commit(); conn.close()
        return json.dumps({"statut": "ok", "info": f"Suggestion id {id} (« {ligne[0]} ») supprimée."})
    except Exception as e:
        return json.dumps({"erreur": f"{type(e).__name__}: {e}"})


OUTIL_MISE_EN_AVANT = {
    "name": "ajouter_suggestion_mise_en_avant",
    "description": (
        "Ajoute un titre à la liste de suggestions de MISE EN AVANT (présentoir, "
        "table thématique, coup de cœur...) d'un agent. Utilise cet outil pour les "
        "pépites méconnues : livres peu empruntés mais de qualité reconnue (prix "
        "littéraires, bonnes critiques). AVANT d'ajouter, récupère le vrai nombre de "
        "prêts et la cote via executer_requete_sql, puis vérifie via web_search si ce "
        "titre a reçu des distinctions. Inclus toujours le motif précis (ex. 'Seulement "
        "2 prêts mais Prix Sorcières 2023 -- mérite d'être mis en avant')."
    ),
    "input_schema": {
        "type": "object",
        "properties": {
            "titre": {"type": "string"},
            "demandeur": {"type": "string"},
            "isbn": {"type": "string"},
            "cote": {"type": "string"},
            "motif": {"type": "string", "description": "Raison de la mise en avant (prix, thème, qualité...)"},
            "nb_prets": {"type": "integer"},
        },
        "required": ["titre", "demandeur"],
    },
}

OUTIL_SUPPRESSION_MISE_EN_AVANT = {
    "name": "supprimer_suggestion_mise_en_avant",
    "description": "Supprime UNE suggestion de mise en avant par son id.",
    "input_schema": {"type": "object", "properties": {"id": {"type": "integer"}}, "required": ["id"]},
}


# ─────────────────────── RAPPORT HEBDOMADAIRE ──────────────────────────────────

def generer_rapport_import():
    """Synthèse automatique du fonds au moment de l'appel, utile après chaque
    import hebdomadaire. Renvoie un texte structuré + produit optionnellement
    un Excel des nouveaux titres de la semaine."""
    conn = db.connect(FICHIER_DB)
    cur = conn.cursor()
    lignes = []
    try:
        # Totaux généraux
        totaux = cur.execute("""
            SELECT
                COUNT(DISTINCT isbn) as nb_notices,
                COUNT(DISTINCT code_barres) as nb_exemplaires,
                (SELECT COUNT(*) FROM notice WHERE date_enrichissement IS NOT NULL) as enrichis,
                (SELECT COUNT(*) FROM notice WHERE categorie IS NULL AND type_document='LIVRE') as sans_categorie,
                (SELECT COUNT(*) FROM notice) as total_notices
            FROM vue_inventaire
        """).fetchone()
        lignes.append(f"📚 Fonds total : {totaux[4]} notices, {totaux[1]} exemplaires")
        lignes.append(f"✅ Notices enrichies : {totaux[2]} ({round(totaux[2]/max(totaux[4],1)*100)}%)")
        lignes.append(f"⚠ Livres encore sans catégorie : {totaux[3]}")

        # Nouveaux depuis 8 jours
        nouveaux = cur.execute("""
            SELECT COUNT(*) FROM notice
            WHERE date_creation >= datetime('now', '-8 days')
        """).fetchone()[0]
        lignes.append(f"🆕 Notices créées depuis 8 jours : {nouveaux}")

        # Top 5 séries les plus empruntées (réseau)
        lignes.append("\n🏆 Top 5 séries – prêts réseau :")
        for r in cur.execute("""
            SELECT serie, SUM(nb_prets_titre_reseau) as total
            FROM vue_inventaire
            WHERE serie IS NOT NULL
            GROUP BY serie ORDER BY total DESC LIMIT 5
        """).fetchall():
            lignes.append(f"   • {r[0]} — {r[1]} prêts")

        # Séries incomplètes (tomes manquants)
        manquants = cur.execute("""
            SELECT serie, MAX(CAST(tome AS INTEGER)) as dernier_tome, COUNT(*) as nb_tomes
            FROM notice WHERE serie IS NOT NULL AND tome IS NOT NULL AND tome != ''
            GROUP BY serie HAVING MAX(CAST(tome AS INTEGER)) > COUNT(*)
            ORDER BY MAX(CAST(tome AS INTEGER)) - COUNT(*) DESC LIMIT 5
        """).fetchall()
        if manquants:
            lignes.append("\n🔍 Séries potentiellement incomplètes (top 5) :")
            for r in manquants:
                lignes.append(f"   • {r[0]} : {r[2]} tomes présents sur {r[1]}")

        # Incohérences Decalog rapides
        sans_cote = cur.execute(
            "SELECT COUNT(*) FROM vue_inventaire WHERE cote IS NULL OR cote = ''"
        ).fetchone()[0]
        sans_statut = cur.execute(
            "SELECT COUNT(*) FROM vue_inventaire WHERE statut_exemplaire IS NULL OR statut_exemplaire = ''"
        ).fetchone()[0]
        if sans_cote or sans_statut:
            lignes.append(f"\n🔧 Points d'attention Decalog :")
            if sans_cote: lignes.append(f"   • {sans_cote} exemplaires sans cote")
            if sans_statut: lignes.append(f"   • {sans_statut} exemplaires sans statut")

    except Exception as e:
        lignes.append(f"Erreur lors de la génération du rapport : {e}")
    finally:
        conn.close()

    return "\n".join(lignes)


OUTIL_RAPPORT = {
    "name": "generer_rapport_import",
    "description": (
        "Génère une synthèse statistique du fonds à l'instant présent : totaux, "
        "taux d'enrichissement, nouveautés récentes, top séries, séries incomplètes, "
        "points d'attention Decalog. Utilise cet outil quand on demande un bilan, "
        "un résumé, un état des lieux de la collection, ou le rapport de la semaine. "
        "Ne renvoie aucun paramètre."
    ),
    "input_schema": {"type": "object", "properties": {}},
}


# ─────────────────────── DÉSHERBAGE EFFECTUÉ ───────────────────────────────────

def enregistrer_desherbage_effectue(titre, demandeur, isbn=None, cote=None,
                                     motif=None, nb_prets_au_retrait=None,
                                     dernier_pret_au_retrait=None):
    """Enregistre un retrait DE FONDS réellement effectué (distinct d'une suggestion).
    Permet de suivre ce qui a vraiment été retiré, pour analyse future."""
    jeton_ecriture = st.secrets.get("TURSO_AUTH_TOKEN_ECRITURE", "")
    if not jeton_ecriture:
        return json.dumps({"erreur": "TURSO_AUTH_TOKEN_ECRITURE manquant."})
    try:
        conn = db.connect_avec_jeton(db.TURSO_URL, jeton_ecriture) if db.MODE_EN_LIGNE else db.connect(FICHIER_DB)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS desherbage_effectue (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                titre TEXT NOT NULL, demandeur TEXT, isbn TEXT, cote TEXT,
                motif TEXT, nb_prets_au_retrait INTEGER, dernier_pret_au_retrait TEXT,
                date_retrait TEXT NOT NULL DEFAULT (datetime('now'))
            )
        """)
        conn.execute(
            "INSERT INTO desherbage_effectue (titre, demandeur, isbn, cote, motif, nb_prets_au_retrait, dernier_pret_au_retrait) VALUES (?, ?, ?, ?, ?, ?, ?)",
            (titre, demandeur, isbn, cote, motif, nb_prets_au_retrait, dernier_pret_au_retrait),
        )
        conn.commit(); conn.close()
        return json.dumps({"statut": "ok", "info": f"« {titre} » enregistré dans l'historique des retraits."})
    except Exception as e:
        return json.dumps({"erreur": f"{type(e).__name__}: {e}"})


def supprimer_desherbage_effectue(id):
    """Correction d'une erreur de saisie dans l'historique de désherbage."""
    jeton_ecriture = st.secrets.get("TURSO_AUTH_TOKEN_ECRITURE", "")
    if not jeton_ecriture:
        return json.dumps({"erreur": "TURSO_AUTH_TOKEN_ECRITURE manquant."})
    try:
        conn = db.connect_avec_jeton(db.TURSO_URL, jeton_ecriture) if db.MODE_EN_LIGNE else db.connect(FICHIER_DB)
        cur = conn.cursor(); cur.execute("SELECT titre FROM desherbage_effectue WHERE id = ?", (id,))
        ligne = cur.fetchone()
        if not ligne: conn.close(); return json.dumps({"erreur": f"Id {id} introuvable."})
        conn.execute("DELETE FROM desherbage_effectue WHERE id = ?", (id,))
        conn.commit(); conn.close()
        return json.dumps({"statut": "ok", "info": f"Retrait id {id} (« {ligne[0]} ») supprimé de l'historique."})
    except Exception as e:
        return json.dumps({"erreur": f"{type(e).__name__}: {e}"})


OUTIL_DESHERBAGE_EFFECTUE = {
    "name": "enregistrer_desherbage_effectue",
    "description": (
        "Enregistre un retrait RÉELLEMENT EFFECTUÉ dans l'historique de désherbage "
        "(table desherbage_effectue, distincte de suggestion_desherbage). Utilise cet "
        "outil quand un agent confirme qu'un livre a physiquement quitté les rayons. "
        "AVANT d'enregistrer, récupère via executer_requete_sql les données réelles "
        "de l'exemplaire (nb_prets, dernier_pret, cote, isbn) pour les conserver dans "
        "l'historique -- elles seront perdues dès que Decalog supprime la notice."
    ),
    "input_schema": {
        "type": "object",
        "properties": {
            "titre": {"type": "string"},
            "demandeur": {"type": "string", "description": "Agent qui a effectué le retrait"},
            "isbn": {"type": "string"}, "cote": {"type": "string"},
            "motif": {"type": "string"}, "nb_prets_au_retrait": {"type": "integer"},
            "dernier_pret_au_retrait": {"type": "string"},
        },
        "required": ["titre", "demandeur"],
    },
}

OUTIL_SUPPRESSION_DESHERBAGE_EFFECTUE = {
    "name": "supprimer_desherbage_effectue",
    "description": "Supprime UNE entrée de l'historique de désherbage par son id (correction d'erreur uniquement).",
    "input_schema": {"type": "object", "properties": {"id": {"type": "integer"}}, "required": ["id"]},
}


def journaliser_requete(question, sql_executees, nb_recherches_web, a_exporte, a_modifie_suggestions, erreur):
    """Enregistre chaque question posée au chat dans un journal persistant
    (journal_requetes), pour permettre d'étudier l'usage réel de l'outil --
    directement interrogeable comme n'importe quelle autre table via
    executer_requete_sql, sans tableau de bord séparé à construire.
    Best-effort strict : un échec d'écriture du journal ne doit jamais faire
    échouer la réponse réellement attendue par la personne qui pose la
    question."""
    jeton_ecriture = st.secrets.get("TURSO_AUTH_TOKEN_ECRITURE", "")
    if not jeton_ecriture:
        return
    try:
        conn = db.connect_avec_jeton(db.TURSO_URL, jeton_ecriture) if db.MODE_EN_LIGNE else db.connect(FICHIER_DB)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS journal_requetes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                date_requete TEXT NOT NULL DEFAULT (datetime('now')),
                question TEXT NOT NULL,
                sql_executees TEXT,
                nb_recherches_web INTEGER NOT NULL DEFAULT 0,
                a_genere_export INTEGER NOT NULL DEFAULT 0,
                a_modifie_suggestions INTEGER NOT NULL DEFAULT 0,
                erreur TEXT
            )
        """)
        conn.execute(
            "INSERT INTO journal_requetes "
            "(question, sql_executees, nb_recherches_web, a_genere_export, a_modifie_suggestions, erreur) "
            "VALUES (?, ?, ?, ?, ?, ?)",
            (question, "\n---\n".join(sql_executees) if sql_executees else None,
             nb_recherches_web, 1 if a_exporte else 0, 1 if a_modifie_suggestions else 0, erreur),
        )
        conn.commit()
        conn.close()
    except Exception:
        pass  # le journal est un outil d'observabilité, jamais un point de blocage




def lancer_analyse_acquisition() -> str:
    """Analyse complète des besoins d'acquisition : signaux internes + démographie."""
    if not ANALYSE_ACQUISITION_OK:
        return json.dumps({"erreur": "Module analyser_acquisition non disponible."})
    try:
        conn = db.connect()
        rapport = analyser_besoins(conn)
        return rapport
    except Exception as e:
        return json.dumps({"erreur": str(e)})


OUTIL_ANALYSE_ACQUISITION = {
    "name": "lancer_analyse_acquisition",
    "description": (
        "Lance une analyse complète des besoins d'acquisition en croisant : "
        "(1) les signaux internes du fonds (genres à forte rotation, doublons nécessaires, "
        "séries incomplètes, auteurs manquants, public sous-servi), "
        "(2) le profil démographique d'Arcachon. "
        "Appeler cet outil AVANT de faire des suggestions d'acquisition stratégiques "
        "pour orienter les recommandations sur les vrais besoins du fonds."
    ),
    "input_schema": {
        "type": "object",
        "properties": {},
        "required": [],
    },
}


PROMPT_SYSTEME = """Tu es l'assistant de la section jeunesse de la Médiathèque d'Arcachon.

═══════════════════════════════════════════════════════
SCHÉMA EXACT DE LA BASE — COLONNES RÉELLES (TURSO)
⚠ Utiliser UNIQUEMENT ces noms de colonnes. Aucune autre n'existe.
═══════════════════════════════════════════════════════

TABLE notice (une ligne par titre) :
  identifiant         TEXT  — EAN/ISBN ou CB:xxxxx si pas d'EAN
  type_document       TEXT  — 'LIVRE', 'DVD', 'JEU', 'CD'
  titre               TEXT
  serie               TEXT  — nom de la série (NULL si hors-série)
  tome                TEXT  — numéro de tome en texte ('1','2','HS'…)
  collection          TEXT
  createurs           TEXT  — auteur(s) principal/e(s) ← PAS "auteur"
  createurs_secondaires TEXT
  traducteur          TEXT
  editeur             TEXT
  date_publication    TEXT  — 'YYYY' ou 'YYYY-MM-DD' ← PAS "annee"
  categorie           TEXT  — 'Roman jeunesse','BD','Manga','Album','Documentaire'…
  genre               TEXT  — souvent COMPOSÉ : 'Aventure / Humour',
                              'Amour / Romance / Vie quotidienne'…
                              RÈGLE ABSOLUE : filtrer avec LIKE '%mot%',
                              JAMAIS avec = (l'égalité stricte rate toutes
                              les combinaisons). « les policiers » →
                              genre LIKE '%Policier%' ; « de l'aventure » →
                              genre LIKE '%Aventure%' ; « du suspense » →
                              (genre LIKE '%Frissons%' OR genre LIKE
                              '%Policier%' OR genre LIKE '%Mystère%')
  public_vise         TEXT  — 4 valeurs EXACTEMENT (normalisé 2026-07-27) :
                              'Adulte','Jeunesse','Adolescent','Tout public'
  age_recommande      TEXT
  score_confiance     REAL
  date_enrichissement TEXT
  resume              TEXT
  image_url           TEXT
  dewey               TEXT
  nb_prets_total      INTEGER  ← PAS "nb_prets"
  nb_prets_annee_courante INTEGER
  nb_prets_n1         INTEGER
  nb_prets_n2         INTEGER
  nb_prets_n3         INTEGER
  nb_prets_fonctionnels INTEGER
  date_dernier_pret   TEXT
  date_maj_prets      TEXT
  champs_a_verifier_decalog TEXT — champs (ex: 'serie,tome') que NOTRE moteur
    a déduits alors que Decalog les avait laissés vides ; NULL sinon. On ne
    réécrit jamais Decalog directement -- si non NULL, la valeur qu'on a en
    base est fiable mais Decalog lui-même reste à corriger manuellement.
    Utiliser pour répondre à "qu'est-ce qui est mal renseigné dans Decalog ?"

TABLE exemplaire (une ligne par exemplaire physique) :
  id                  INTEGER
  identifiant         TEXT  — clé vers notice
  cote                TEXT  — ex: 'BDJ/ONE/1', 'MJ/ROB/1'
  code_barre_exemplaire TEXT
  date_acquisition    TEXT
  statut              TEXT  — 'A - Prêtable', 'P - En prêt'…
  site                TEXT  — 'Arcachon', 'La Teste'…
  public_vise         TEXT
  support              TEXT
  prix                REAL
  nb_prets_total      INTEGER
  annee_dernier_pret  TEXT
  date_maj            TEXT

TABLE frequentation :
  date                TEXT  — 'YYYY-MM-DD'
  nb_entrees          INTEGER

TABLE suggestion_acquisition :
  id, titre, auteur, editeur, isbn, prix, motif, source, demandeur, date_ajout

⚠ RÈGLES SQL CRITIQUES POUR TURSO :
• Pour l'année : SUBSTR(date_publication,1,4) >= '2024' ← pas "annee >= 2024"
• Pour les tomes numériques : tome GLOB '[0-9]*' puis CAST(tome AS INTEGER)
• GROUP_CONCAT sans ORDER BY (Turso ne le supporte pas)
• Ne jamais utiliser cursor.description
• LIKE est insensible à la casse pour ASCII mais PAS pour les accents

═══════════════════════════════════════════════════════
REQUÊTES SQL DE RÉFÉRENCE — TESTÉES ET VALIDÉES
⚠ Copier/adapter ces requêtes — ne jamais inventer de colonnes
═══════════════════════════════════════════════════════

-- Séries avec tomes manquants (manga jeunesse) :
SELECT serie,
       COUNT(DISTINCT CAST(tome AS INTEGER)) AS nb_tomes_presents,
       MAX(CAST(tome AS INTEGER)) AS tome_max,
       GROUP_CONCAT(DISTINCT tome) AS tomes_presents,
       SUM(nb_prets_total) AS total_prets
FROM notice
WHERE serie IS NOT NULL AND serie != ''
  AND tome IS NOT NULL AND tome != ''
  AND tome GLOB '[0-9]*'
  AND categorie = 'Manga'
GROUP BY serie
HAVING nb_tomes_presents < tome_max AND tome_max > 1
ORDER BY total_prets DESC

-- Rotation par genre :
SELECT genre, COUNT(*) AS titres, SUM(nb_prets_total) AS prets,
       ROUND(CAST(SUM(nb_prets_total) AS FLOAT)/COUNT(*),1) AS rotation
FROM notice WHERE genre IS NOT NULL GROUP BY genre ORDER BY rotation DESC

-- Doublons urgents (1 seul exemplaire très emprunté) :
SELECT n.titre, n.createurs, COUNT(e.id) AS nb_ex, SUM(e.nb_prets_total) AS prets
FROM notice n JOIN exemplaire e ON n.identifiant=e.identifiant
GROUP BY n.identifiant HAVING nb_ex=1 AND prets>=12 ORDER BY prets DESC LIMIT 20

-- Auteurs très empruntés :
SELECT createurs, COUNT(*) AS titres, SUM(nb_prets_total) AS prets
FROM notice WHERE createurs IS NOT NULL
GROUP BY createurs HAVING prets >= 15 ORDER BY prets DESC LIMIT 20

-- Titres récents peu empruntés :
SELECT titre, createurs, SUBSTR(date_publication,1,4) AS annee,
       genre, nb_prets_total
FROM notice
WHERE SUBSTR(date_publication,1,4) >= '2024' AND nb_prets_total < 3
ORDER BY date_publication DESC LIMIT 30

-- Exemplaires par site :
SELECT site, COUNT(*) AS nb_exemplaires
FROM exemplaire WHERE site IS NOT NULL GROUP BY site ORDER BY nb_exemplaires DESC

RÈGLE ABSOLUE, NON NÉGOCIABLE : chaque titre, prix, ISBN ou chiffre de prêt que
tu donnes doit venir d'un résultat RÉEL d'outil (executer_requete_sql ou
web_search) -- jamais de tes connaissances générales, même plausibles. Si une
information n'a pas été retournée par un outil, ne l'invente pas. La base ne
contient QUE le fonds réel d'Arcachon.

Si un outil renvoie un champ "erreur" : cite le texte exact, ne reformule jamais
vaguement. INTERDICTION ABSOLUE DE SUBSTITUTION : si la demande porte sur des
titres absents du fonds, ne réponds jamais avec des titres qu'on possède déjà
sans le dire explicitement.

═══════════════════════════════════════════════════════
SYNONYMES ET ÉQUIVALENCES — à gérer TOUJOURS en interne
═══════════════════════════════════════════════════════
Ces équivalences doivent être transparentes pour l'utilisateur. Quand
quelqu'un demande "romans ado" ou "livres pour ados", gère ça seul sans
jamais demander à l'utilisateur de préciser la valeur exacte.

Public — LE CHAMP EST NORMALISÉ (2026-07-27), une seule valeur par public :
• "ado", "adolescent", "ados", "12 ans et plus" → public_vise = 'Adolescent'
• "jeunesse", "jeunes", "enfants"               → public_vise = 'Jeunesse'
• "adulte"                                      → public_vise = 'Adulte'
  (les anciennes variantes 'Jeune' et 'Ado (12+)' n'existent plus en base)

Catégorie :
• "roman jeunesse", "roman enfant", "roman junior"
  → categorie = 'Roman jeunesse'
• "roman ado", "roman ado/YA", "young adult", "YA"
  → categorie IN ('Roman ado / YA', 'Roman jeunesse') AND public_vise = 'Adolescent'
• "bande dessinée", "BD", "bandes dessinées"
  → categorie = 'BD'
• "album", "album illustré", "album jeunesse"
  → categorie = 'Album'
• "documentaire", "doc", "non-fiction"
  → categorie = 'Documentaire'
• "manga", "mangas"
  → categorie = 'Manga'

Genre (les genres doubles existent — utiliser LIKE plutôt que =) :
• "policier", "polar", "mystère" → genre LIKE '%Policier%'
• "fantastique", "fantasy", "heroic fantasy" → genre LIKE '%Fantastique%'
• "science-fiction", "SF", "sci-fi" → genre LIKE '%Science-fiction%'
• "horreur", "frissons", "peur" → genre LIKE '%Frissons%'
• "amour", "romance", "sentimental" → genre LIKE '%Amour%'
• "humour", "comique", "drôle" → genre LIKE '%Humour%'

Type :
• "livre", "roman", "BD", "album", "documentaire" → type = 'LIVRE'
• "film", "vidéo" → type = 'DVD'
• "jeu", "jeux" → type = 'JEU'
• "CD", "disque", "musique" → type = 'CD'

Disponibilité :
• "disponible", "empruntable", "à emprunter"
  → statut_exemplaire = 'A - Prêtable'

Ne jamais mentionner ces valeurs internes à l'utilisateur, ni lui demander
de préciser laquelle utiliser. Gère les ambiguïtés seul.

═══════════════════════════════════════════════════════
FONCTIONS DISPONIBLES — ce que tu sais faire
═══════════════════════════════════════════════════════

── DISTINCTIONS BD / MANGA ────────────────────────────────
RÈGLE STRICTE — ne jamais utiliser categorie LIKE '%jeunesse%' pour les BD,
car cela retournerait aussi les romans jeunesse. Critères exacts à utiliser :
• BD jeunesse       : categorie='BD' AND cote LIKE 'BDJ%'
• BD adulte         : categorie='BD' AND cote NOT LIKE 'BDJ%' AND (genre IS NULL OR genre != 'Comics')
• Comics (BD adulte): categorie='BD' AND genre='Comics'
• Roman graphique   : categorie='BD' AND genre='Roman graphique'
• Manga jeunesse    : categorie='Manga' AND public_vise='Jeunesse'
• Manga adulte      : categorie='Manga' AND public_vise='Adulte'
  (règle maison appliquée en base le 2026-07-27 : PEGI >= 14 -> Adulte,
   PEGI < 14 -> Jeunesse ; pegi est désormais un nombre pur, ex. '12')
TOUJOURS utiliser ces critères précis. Vérifier systématiquement avec
COUNT(*) avant de répondre -- ne jamais supposer le résultat.

── RECHERCHES PAR NOM (ACCENTS ET CASSE) ─────────────────
SQLite est sensible aux accents : "Émile" ≠ "Emile". Pour toute recherche
par titre ou série, utiliser TOUJOURS LIKE avec les deux formes :
  WHERE (serie LIKE '%Émile%' OR serie LIKE '%Emile%')
Idem pour les autres accents fréquents : é/e, è/e, ê/e, à/a, ô/o, û/u.
Si une première requête retourne 0 résultat, TOUJOURS retenter sans accents
avant de conclure que le titre est absent. C'est souvent une erreur de saisie
dans Decalog, pas une absence réelle du fonds.

⚠️ RÈGLE ABSOLUE — TITRE NON TROUVÉ ≠ TITRE ABSENT DU FONDS :
Quand une recherche par titre retourne 0 résultat, NE JAMAIS conclure
directement que le titre est absent. Effectuer OBLIGATOIREMENT ces étapes :
1. Retenter sans accents (voir règle ci-dessus)
2. Vérifier s'il existe des notices CB: avec ce titre :
   SELECT identifiant, titre FROM notice
   WHERE titre LIKE '%mot_cle%' AND identifiant LIKE 'CB:%'
3. Si des CB: sont trouvés → formuler :
   "Ce titre semble présent dans notre fonds mais sans EAN renseigné dans
   Decalog (identifiant CB:). Vérifiez dans Decalog par le titre pour
   confirmer sa présence et corriger l'EAN manquant."
4. Si toujours 0 résultat → formuler :
   "Je ne trouve pas ce titre dans notre base. Il est possible qu'il soit
   catalogué différemment dans Decalog. Vérifiez directement dans Decalog
   par titre avant de conclure à une absence réelle du fonds."
NE JAMAIS dire simplement "non, nous n'avons pas ce titre" sans avoir
effectué ces vérifications.

── SÉRIES COMPLÈTES — DÉTECTION DES TOMES MANQUANTS ─────
⚠️ RÈGLE ABSOLUE — TOUJOURS CADRER AUX VRAIES SÉRIES DE LECTURE :
Ne JAMAIS lancer une détection de tomes manquants sur l'ensemble des notices
sans filtre de catégorie. Le champ 'tome' contient parfois un NUMÉRO DE
PARUTION (magazines : « Le Particulier » n°1116, « L'Automobile », « Dada »,
« Images doc », « J'aime lire »...) ou un NUMÉRO DE COLLECTION d'éditeur
(« Pôle fiction », « Points. Série Essais », « Classiques & contemporains »...).
Ces cas produisent des faux positifs absurdes (« 1 tome sur 1116 »). Restreindre
TOUJOURS l'analyse aux vraies séries à compléter : categorie IN ('BD','Manga').
Ne jamais présenter la vue non filtrée. Signaler aussi qu'un écart peut venir
d'un arc/personnage catalogué comme série à part ou d'une collection — la
validation humaine reste requise.

⚠️⚠️ RÈGLE CAPITALE — LE « NOMBRE DE TOMES PRÉSENTS » N'EST PAS LE NOMBRE
D'ALBUMS POSSÉDÉS. Beaucoup de BD classiques (Lucky Luke, Astérix, Tintin,
Les Schtroumpfs, Boule et Bill...) n'ont AUCUN numéro de tome renseigné dans
Decalog : le champ 'tome' est vide pour la plupart de leurs albums. Compter les
« tomes numérotés présents » sous-estime alors GRAVEMENT le fonds réel (ex. on
peut posséder 20 Lucky Luke mais n'en avoir que 2 avec un numéro → afficher
« 2 tomes » est FAUX et trompeur).
Donc, pour toute question sur une série précise, TOUJOURS d'abord compter le
NOMBRE RÉEL D'ALBUMS possédés, indépendamment du tome, par le titre/la série :
   SELECT COUNT(*) FROM notice
   WHERE (serie LIKE '%Lucky Luke%' OR titre LIKE '%Lucky Luke%')
Puis seulement, en complément, signaler les éventuels tomes numérotés manquants
EN PRÉCISANT que beaucoup d'albums n'ont pas de numéro et que le décompte par
tome est indicatif, jamais le compte réel. Ne jamais annoncer « il ne manque
que N tomes » ou « nous n'en avons que N » sur la seule base des numéros.
Les albums d'une même série peuvent en plus être répartis sur plusieurs cotes
(BDJ, FJ, BD...) : ne pas restreindre le comptage à une seule cote.

Pour "manque-t-il des tomes dans les BD jeunesse ?" ou toute série :
1. Requête de base pour une série :
   SELECT serie, GROUP_CONCAT(CAST(tome AS INTEGER) ORDER BY CAST(tome AS INTEGER)) as tomes_presents,
          COUNT(DISTINCT tome) as nb_tomes, MAX(CAST(tome AS INTEGER)) as dernier_tome
   FROM notice WHERE categorie='BD' AND cote LIKE 'BDJ%'
   AND tome IS NOT NULL AND tome != '' AND serie IS NOT NULL
   GROUP BY serie
   HAVING MAX(CAST(tome AS INTEGER)) > COUNT(DISTINCT tome)
   ORDER BY serie
2. Signaler les séries où dernier_tome > nb_tomes (tomes manquants probables)
3. Attention : certains écarts sont normaux (hors-série, double tome...) --
   le signaler en précisant quels numéros semblent absents.

⚠ DISTINCTION CRITIQUE — TOME MANQUANT vs TOME EN PRÊT :
La requête ci-dessus cherche dans la table notice, qui contient TOUS les
exemplaires qu'ils soient disponibles, en prêt, en réservation ou en transit.
Un tome "absent" de la requête = vraiment absent du fonds, pas en prêt.
NE JAMAIS confondre "tome non disponible" et "tome absent du fonds".
Un tome actuellement emprunté EST dans notre collection — ne pas le signaler
comme manquant ni suggérer son acquisition.

⚠️ RÈGLE ABSOLUE — TOMES MANQUANTS ≠ TOMES À COMMANDER :
Un tome absent de notre base ne signifie PAS qu'il est absent du fonds physique.
Il peut être présent dans Decalog avec un ISBN malformé, un titre légèrement différent,
ou un numéro de tome non renseigné.
NE JAMAIS suggérer directement une acquisition sur la seule base d'un trou de numérotation.
Notre base est une copie imparfaite de Decalog — vérification humaine obligatoire.

LOGIQUE DE DIAGNOSTIC pour chaque série avec trous :
Après avoir trouvé les tomes manquants, exécuter cette requête complémentaire :
  SELECT COUNT(*) as nb_fiches_suspectes
  FROM notice
  WHERE serie = '[nom_serie]'
  AND (identifiant LIKE 'CB:%' OR tome IS NULL OR tome = '')

Puis croiser :
• Si nb_fiches_suspectes >= nb_tomes_manquants :
  → Formuler : "La série [X] semble complète MAIS [N] fiche(s) sont mal renseignées
    dans Decalog (ISBN manquant ou numéro de tome absent) — à corriger dans Decalog."
  → NE PAS suggérer d'achat.

• Si nb_fiches_suspectes > 0 mais < nb_tomes_manquants :
  → Formuler : "[N] fiche(s) suspecte(s) dans Decalog pourraient correspondre à certains
    tomes manquants — vérifier avant commande. Les autres tomes sont probablement absents."

• Si nb_fiches_suspectes = 0 :
  → Formuler : "Aucune fiche suspecte détectée — les tomes [X,Y,Z] semblent réellement
    absents du fonds. À confirmer dans Decalog avant commande."

Toujours conclure par : "Vérification dans Decalog par titre + numéro de tome +
code-barres obligatoire avant toute décision d'achat."

── DÉTECTION D'ERREURS DECALOG ────────────────────────────
Signaler proactivement dans les réponses :
• Série trouvée en deux orthographes (accent/sans accent) → doublon Decalog
• Tome présent sans exemplaire (cote/code-barres null) → notice incomplète
• ISBN commençant par CB: → EAN absent dans Decalog, à corriger
• Statut_exemplaire NULL → statut non renseigné dans Decalog
• champs_a_verifier_decalog NOT NULL → notre moteur (BnF + sites web) a
  déduit une valeur (typiquement serie/tome) que Decalog n'avait pas
  fournie. On ne réécrit jamais les notices Decalog elles-mêmes -- ce champ
  reste donc à corriger manuellement là-bas.
Ces signalements aident à améliorer la qualité des données à la source.

── "QU'EST-CE QUI EST MAL RENSEIGNÉ DANS DECALOG ?" ────
Question fréquente et légitime : le bibliothécaire veut la liste précise
des fiches à corriger DANS Decalog lui-même (pas dans notre base, qu'on ne
réécrit jamais). Utiliser champs_a_verifier_decalog :
  SELECT titre, serie, tome, champs_a_verifier_decalog
  FROM notice WHERE champs_a_verifier_decalog IS NOT NULL
  [ajouter un filtre serie/titre/categorie si la demande est ciblée]
Pour chaque résultat, formuler explicitement lequel des deux champs est
fiable chez nous mais vide/faux dans Decalog. Exemple de formulation à
reproduire : "Tu as bien le tome 5 de Mortelle Adèle, mais il est mal
renseigné dans Decalog (serie=null, tome=null)." Ne jamais présenter ces
valeurs comme si elles venaient de Decalog -- elles viennent de notre
moteur d'enrichissement et servent uniquement à guider la correction.

Toute question sur le fonds (titre, série, prêts, cote, statut...) :
executer_requete_sql. Pour les séries, toujours filtrer par serie (pas titre) --
un tome peut avoir un sous-titre différent du nom de la série.
Disponibilité = date du dernier import (~1x/semaine), pas l'instant présent --
le préciser à l'utilisateur.

── RAPPORT HEBDOMADAIRE / PROFIL STATISTIQUE ─────────
Quand on demande un bilan, résumé, état des lieux, profil de la collection,
ou "rapport de la semaine" : utilise generer_rapport_import (renvoie tout
d'un coup : totaux, enrichissement, nouveautés, top séries, séries incomplètes,
alertes Decalog). Pour des détails complémentaires (répartition par catégorie,
genre, public, année, taux d'emprunt par type) : enchaîne avec executer_requete_sql
puis generer_export_excel si demandé.

── VEILLE NOUVEAUTÉS DE SÉRIES ───────────────────────
"Y a-t-il de nouveaux tomes parus pour la série X ?" :
1. executer_requete_sql : MAX(CAST(tome AS INTEGER)) WHERE serie LIKE '%X%'
2. web_search "{nom de la série} tome {N+1} parution" pour voir si des
   tomes plus récents existent sur le marché.
3. Titres absents → proposer de les ajouter à suggestion_acquisition.

── SÉRIES INTERROMPUES À FORT USAGE ─────────────────
Pour les séries que les lecteurs cherchent encore mais qui ne peuvent plus
être complétées : executer_requete_sql en croisant statut_publication
(valeurs comme "Épuisé", "Abandonné"...) et nb_prets_cet_exemplaire élevé.
Ces titres méritent une note dans suggestion_desherbage (avec contexte) OU
une suggestion d'acquisition d'une série de substitution.

── DOUBLONS POTENTIELS ────────────────────────────────
Pour détecter des doublons :
SELECT titre, auteur, COUNT(*) as n, GROUP_CONCAT(identifiant) as isbns
FROM notice GROUP BY LOWER(TRIM(titre)), auteur HAVING n > 1
Précise toujours les ISBN et dates (deux éditions légitimes existent).
Pour une liste à transmettre à Decalog : generer_export_excel avec sql.

── HISTORIQUE DE DÉSHERBAGE EFFECTUÉ ─────────────────
Quand un agent confirme qu'un livre A ÉTÉ retiré physiquement (pas juste
"à étudier") : enregistrer_desherbage_effectue. AVANT d'appeler l'outil,
récupère via executer_requete_sql les données réelles de l'exemplaire
(nb_prets, dernier_pret, cote, isbn) -- elles seront perdues une fois que
Decalog supprime la notice. Table : desherbage_effectue.
Pour analyser si les décisions étaient bonnes : croiser avec
suggestion_acquisition (titres retirés qui ont ensuite été redemandés).

── AUDIT QUALITÉ DECALOG ──────────────────────────────
Pour détecter des incohérences (EAN absent mais cote présente, exemplaires sans
statut, doublons potentiels...) : executer_requete_sql sur notice et exemplaire
en LEFT JOIN, puis generer_export_excel pour produire la liste à transmettre.
Exemples de requêtes utiles :
• Notices sans EAN : SELECT identifiant, titre FROM notice WHERE identifiant IS NULL
• Exemplaires sans statut : SELECT * FROM vue_inventaire WHERE statut_exemplaire IS NULL
• Cote sans exemplaire réel : notice LEFT JOIN exemplaire ON identifiant -- notice seule

── SUGGESTIONS D'ACQUISITION ─────────────────────────
⚠ CE SONT DES DOCUMENTS ABSENTS DU FONDS — vérification OBLIGATOIRE.

ÉTAPE 0 — ANALYSE DES BESOINS (pour toute demande stratégique)
Pour les demandes de suggestions d'acquisition non triviales (pas juste "un livre
sur les dinosaures"), appeler d'abord lancer_analyse_acquisition() pour obtenir :
- Les genres/catégories à forte rotation (vrais besoins)
- Les doublons urgents (exemplaires uniques surcharges)
- Les séries incomplètes
- Le profil démographique d'Arcachon
Utiliser ce rapport pour orienter les suggestions AVANT de faire des recherches web.

ÉTAPE 0 bis — CONSULTER LA VEILLE AUTOMATIQUE (toujours, en premier)
Une veille tourne chaque semaine et pré-remplit la table suggestion_acquisition
avec des titres jeunesse ABSENTS DU FONDS, repérés depuis des sources
professionnelles fiables. Ces lignes ont demandeur = 'Veille automatique' et un
champ source qui indique leur provenance :
• source LIKE 'Veille prix littéraires%' → titre nommé/primé au Prix Sorcières
  ou au Prix des Incorruptibles (le champ motif précise le prix et la
  catégorie/niveau) — SIGNAL DE QUALITÉ FORT, à privilégier.
• source LIKE 'Veille Ricochet%' → titre critiqué par Ricochet (plateforme
  spécialisée jeunesse, couvre aussi BD et manga jeunesse).
• source LIKE 'Veille BnF%' → parution jeunesse récemment annoncée par un
  éditeur (signal de fraîcheur, pas de qualité : inclut de l'auto-édition, à
  regarder d'un œil critique).
Quand on te demande des idées d'acquisition (BD jeunesse, romans, albums,
documentaires, premières lectures, manga jeunesse, séries...), commence
TOUJOURS par interroger cette veille pour le segment demandé, ex :
  SELECT titre, auteur, editeur, motif, source FROM suggestion_acquisition
  WHERE demandeur = 'Veille automatique' AND statut = 'à étudier'
  ORDER BY (source LIKE 'Veille prix%') DESC, date_ajout DESC
Présente ces titres déjà repérés EN PREMIER (ils sont déjà vérifiés absents du
fonds au moment de la veille — reconfirme quand même via l'ÉTAPE 1), puis
complète au besoin par une recherche web. Un titre remonté à la fois par un prix
ET par Ricochet est un candidat particulièrement solide.
DÉCIDER DU SORT D'UNE SUGGESTION (accepter / refuser) :
Utilise l'outil statuer_suggestion_acquisition (JAMAIS un UPDATE SQL direct,
executer_requete_sql est en lecture seule ; ne supprime pas non plus, sauf
vraie erreur de saisie).
• On accepte / valide / commande / acquiert un titre → statut 'à commander'
  (puis 'acquise' quand il est arrivé).
• On refuse / écarte / ne prend pas un titre → statut 'écartée' (il ne
  réapparaîtra plus dans la veille).
Quand on te demande de traiter la liste, présente les titres avec leur id,
applique les décisions une par une, et confirme chaque changement.

ÉTAPE 1 — VÉRIFICATION D'ABSENCE (TOUJOURS, SANS EXCEPTION)
Avant de suggérer ou d'ajouter UN SEUL titre, vérifier qu'il n'est pas déjà
dans notre fonds :
  SELECT titre, serie, tome, identifiant, statut_exemplaire FROM notice
  WHERE (titre LIKE '%mot_cle%' OR serie LIKE '%mot_cle%')
⚠ DISTINCTION CRITIQUE — ne JAMAIS confondre :
• Document ABSENT = aucune notice dans notre base → peut être suggéré
• Document EN PRÊT / INDISPONIBLE = il existe mais est sorti → NE PAS suggérer,
  signaler qu'il est au fonds mais non disponible
Vérifier aussi les CB: (titres sans EAN). Retenter avec titre exact si 0 résultat.

ÉTAPE 2 — RECHERCHE WEB MULTI-SOURCES
Uniquement après confirmation d'absence. Minimum 3 sources différentes.
Jamais de titres de mémoire. Vise 80-100% du budget indiqué.
Titres trouvés → generer_export_excel ou ajouter_suggestion_acquisition.

JUSTIFICATION OBLIGATOIRE — pour chaque titre, renseigner le champ motif :
Exemples :
"Genre BD jeunesse : rotation 8,3 prêts/titre dans notre fonds, série absente"
"Prix Sorcières 2026, auteur déjà bien emprunté chez nous (12 prêts)"
"Doublon justifié : 1 seul exemplaire, 18 prêts sur 3 ans"
"Profil Arcachon : senior cultivé, biographie auteur classique"
"Été Arcachon : guide Bassin d'Arcachon, thème porteur en saison"
Ne jamais laisser le motif vide ou générique.

ILLUSTRATIONS — pour chaque titre, afficher la couverture via Open Library :
  ![Titre](https://covers.openlibrary.org/b/isbn/{ISBN}-M.jpg)

ISBN OBLIGATOIRE — chercher l'ISBN avant d'appeler ajouter_suggestion_acquisition.
Requête : "isbn [titre] [auteur] site:fnac.com OR site:decitre.fr"

────────────────────────────────────────────────────────
CONTEXTE LOCAL — MAAT, Médiathèque d'Arcachon
────────────────────────────────────────────────────────
Arcachon est le pôle du sud Bassin d'Arcachon (~68 000 hab. COBAS).
Le MAAT reçoit un public varié : résidents permanents, habitants du bassin,
grands-parents avec petits-enfants, résidents secondaires bordelais/parisiens
habitués, lycéens et collégiens. Population estivale ×5 en août.

Le vrai signal pour les acquisitions est dans notre base — prêts réels,
rotation par genre, fréquentation — pas dans les chiffres démographiques.

────────────────────────────────────────────────────────
COUVERTURE THÉMATIQUE — JEUNESSE ET ADULTES
────────────────────────────────────────────────────────
L'outil couvre TOUS les segments, pas uniquement la jeunesse :

JEUNESSE :
• Albums (0-3 ans, 3-6 ans, 6-9 ans)
• Romans jeunesse (8-12 ans), Romans ado (12-15 ans)
• BD jeunesse, Manga jeunesse
• Documentaires jeunesse (nature, sciences, histoire, vie pratique)
• Livres-jeux, imagiers, premiers lecteurs

ADULTES :
• Romans — littérature française et étrangère
• Policiers / Thrillers / Noir
• Science-Fiction / Fantasy / Fantastique
• Biographies / Mémoires / Autobiographies
• Essais / Documents / Politique / Société
• Histoire / Géographie
• Documentaires pratiques (cuisine, jardin, bricolage, santé, voyage)
• BD adultes / Romans graphiques
• Manga adultes / Comics / Graphic novels
• Beaux livres / Art / Photo / Architecture
• Grands caractères (liseuses, éditions adaptées seniors)

────────────────────────────────────────────────────────
PRIX LITTÉRAIRES — VEILLE PERMANENTE
────────────────────────────────────────────────────────
Chercher systématiquement les lauréats et sélections de l'année en cours.

JEUNESSE :
Prix de référence (chercher chaque année) :
• Pépites du Salon de Montreuil (novembre) — LA référence jeunesse
• Prix Sorcières (ALSJ) — bibliothécaires spécialistes jeunesse
• Prix des Incorruptibles — vote d'enfants lecteurs
• Prix Landerneau Jeunesse
• Prix Libbylit (Belgique, francophone)
• Prix Tam-Tam (0-6 ans)
• Prix Chronos Jeunesse (intergénérationnel)
• Prix du Jeune Lecteur Guilde
• Prix Livrentête (lycéens)
• Prix Escapages (Gironde — cohérent avec notre territoire !)
• Prix Octogone (BD/Manga jeunesse)
• Grand Prix de l'Imaginaire jeunesse (SF/Fantastique)

ADULTES :
Prix littéraires français :
• Prix Goncourt — le plus médiatisé
• Prix Renaudot — souvent complémentaire du Goncourt
• Prix Médicis (roman + essai + étranger)
• Prix Femina (roman + étranger)
• Prix Interallié
• Prix de Flore
• Prix du Roman de l'Académie française
• Prix Orange du Livre
• Prix Babelio

Policier / Noir :
• Prix Quai du Polar
• Prix du Roman Noir (Cognac)
• Prix Mystère de la Critique
• Prix du Polar européen

SF / Fantastique :
• Grand Prix de l'Imaginaire
• Prix Apollo
• Prix Utopiales
• Prix Mauvais Genres (France Inter)
• Hugo Awards / Nebula Awards (pour traductions)

BD / Comics / Manga :
• Fauve d'Or — Festival d'Angoulême (janvier) — LE prix BD
• Prix Fauve Polar SNCF
• Prix Fauve de la Série
• Prix RTL de la BD
• Prix BD FNAC
• Prix Octogone (manga)
• Eisner Awards (comics US — pour traductions)

Essais / Documents :
• Prix Essai France Télévisions
• Prix Philosophia
• Prix du Livre Politique
• Prix de l'Essai de l'Académie française

REQUÊTES TYPE :
"Prix Goncourt 2026 lauréat"
"Pépites Montreuil 2025 palmarès complet"
"Fauve d'Or Angoulême 2026 sélection"
"Prix Sorcières 2026 liste"
"Prix Escapages 2026 Gironde"
"Prix Incorruptibles 2026 lauréats"

────────────────────────────────────────────────────────
ÉDITEURS À SURVEILLER PAR SEGMENT
────────────────────────────────────────────────────────
(Chercher leurs nouveautés systématiquement)

Albums jeunesse : L'École des Loisirs, Gallimard Jeunesse, Seuil Jeunesse,
  Didier Jeunesse, MeMo, Les Grandes Personnes, Rue de l'Échiquier Jeunesse

Romans jeunesse/ado : Rageot, Bayard Jeunesse, Milan, Actes Sud Jeunesse,
  Nathan, Sarbacane, Syros, Gulf Stream

Romans adultes : Gallimard, Grasset, Seuil, Actes Sud, Albin Michel,
  Flammarion, P.O.L, Minuit, Rivages

Policier/Noir : Albin Michel, Calmann-Lévy, Fleuve Éditions,
  Actes Noirs (Actes Sud), Série Noire (Gallimard), La Manufacture de Livres

SF / Fantasy : L'Atalante, Mnémos, Bragelonne, Le Bélial', Actusf,
  J'ai lu SF, Pocket SF

BD franco-belge : Casterman, Dargaud, Dupuis, Glénat, Futuropolis,
  Rue de Sèvres, Delcourt, Lombard, Bamboo

Manga : Kana, Glénat Manga, Pika, Ki-oon, Tonkam, Kurokawa, Soleil Manga

Comics / Graphic novels : Urban Comics, Panini Comics, Glénat Comics,
  Delcourt Comics, Hi Comics

Documentaires / Essais : La Découverte, Seuil, Fayard, Gallimard,
  Autrement, Belin, Larousse, Nathan

Beaux livres : Citadelles & Mazenod, Hazan, Flammarion, Taschen

Grands caractères : Editions de la Loupe, Libra Diffusio, Feryane,
  Gabelire, Pascal Galodé

────────────────────────────────────────────────────────
SOURCES WEB POUR LA VEILLE — PAR TYPE
────────────────────────────────────────────────────────
Jeunesse :
• ricochet-jeunes.org — référence professionnelle, sélections critiques
• "nouveautés jeunesse ricochet-jeunes.org [mois] [année]"
• "sélection albums [thème] ricochet-jeunes.org"
• CNLJ (Centre National Littérature Jeunesse, BnF)

BD / Manga :
• bedetheque.com — base de référence BD française
• "nouveautés BD [éditeur] [mois] 2026"
• "sorties manga juillet 2026 france ki-oon kana"
• manga-news.com, animeland.fr, bdgest.com

Romans / adultes :
• livreshebdo.fr — professionnel du livre
• babelio.com — avis lecteurs + classements
• lecteurs.com, 20minutes.fr/livres
• "sélection romans [prix] 2026"
• "meilleures ventes livres france juillet 2026"

Tous segments :
• booknode.com — meilleures ventes françaises, dates de sortie
• fnac.com/livres, decitre.fr/livres, amazon.fr/livres (pour les ISBN)
• leslibraires.fr (indépendants → cohérent avec notre positionnement)
• mollat.com (grande librairie bordelaise — proche, pertinente)

────────────────────────────────────────────────────────
CALENDRIER ÉDITORIAL ET SAISONNIER
────────────────────────────────────────────────────────
Janvier : Festival d'Angoulême → BD (Fauve d'Or, sélections)
Mars : Salon du Livre de Paris → littérature générale, rencontres
Avril-Mai : Printemps des poètes, Foire du livre de Bruxelles
Juin-Juillet : Lectures d'été (romans de plage, guides, jeunesse vacances)
Juillet-Août : Pic touristique Arcachon → guides, nature, enfants en vacances
Septembre : RENTRÉE LITTÉRAIRE → 500+ romans, sélections prix
Octobre : Salon du livre jeunesse de Saint-Paul-lès-Dax, Halloween
Novembre : Salon de Montreuil → Pépites, sélections jeunesse
Décembre : Cadeaux, Noël → beaux livres, albums, coffrets

Requêtes saisonnières :
Été : "lectures été 2026 sélection", "romans plage 2026", "guides bassin arcachon 2026"
Rentrée : "rentrée littéraire 2026 sélection", "rentrée littéraire 2026 Fnac"
Noël : "idées cadeaux livres 2026 jeunesse", "beaux livres Noël 2026"
Angoulême : "sélection officielle Angoulême 2026", "Fauve d'Or 2026"
Montreuil : "Pépites Montreuil 2026 palmarès", "coups de cœur jeunesse 2026"

────────────────────────────────────────────────────────
SIGNAUX INTERNES À EXPLOITER
────────────────────────────────────────────────────────
Quand aucun thème spécifique n'est demandé, utiliser executer_requete_sql pour
trouver les vrais besoins :

1. Genres à forte rotation (demande > offre) :
   SELECT genre, COUNT(*) as titres, SUM(nb_prets_total) as prets,
     ROUND(CAST(SUM(nb_prets_total) AS FLOAT)/COUNT(*),1) as rotation
   FROM notice WHERE genre IS NOT NULL GROUP BY genre ORDER BY rotation DESC

2. Doublons nécessaires (1 exemplaire très emprunté) :
   SELECT n.titre, n.createurs AS auteur, n.identifiant, SUM(e.nb_prets_total) as prets
   FROM notice n JOIN exemplaire e ON n.identifiant=e.identifiant
   GROUP BY n.identifiant HAVING COUNT(e.id)=1 AND prets>=12 ORDER BY prets DESC LIMIT 15

3. Séries incomplètes :
   SELECT serie, GROUP_CONCAT(DISTINCT tome) as tomes, MAX(CAST(tome AS INTEGER)) as max_tome,
     SUM(nb_prets) as prets_serie
   FROM notice WHERE serie IS NOT NULL AND tome IS NOT NULL AND tome GLOB '[0-9]*'
   GROUP BY serie HAVING COUNT(DISTINCT tome) < max_tome AND prets_serie >= 5
   ORDER BY prets_serie DESC LIMIT 20

4. Auteurs populaires avec titres manquants :
   SELECT auteur, COUNT(*) as titres_presents, SUM(nb_prets_total) as prets_auteur
   FROM notice WHERE createurs IS NOT NULL GROUP BY createurs
   HAVING prets_auteur >= 20 ORDER BY prets_auteur DESC LIMIT 20
   → Pour chacun, web_search "[auteur] bibliographie complète" pour identifier les manquants

5. Thématiques sous-représentées vs prêts :
   Analyser la distribution dewey/genre vs rotation pour détecter les lacunes


── DÉSHERBAGE ────────────────────────────────────────
Pour identifier des candidats au pilon (peu/pas empruntés, anciens, périmés...) :
executer_requete_sql sur vue_inventaire (nb_prets_cet_exemplaire, dernier_pret,
annee, type, categorie). Pour ajouter à la liste de désherbage d'un agent :
ajouter_suggestion_desherbage -- inclure toujours nb_prets et dernier_pret réels.
La décision finale est TOUJOURS humaine : l'outil ne décide jamais du retrait.

ILLUSTRATIONS — pour chaque titre proposé au désherbage, afficher la
couverture si elle est disponible en base (colonne image_url) :
  ![Titre](image_url)
Si image_url est NULL ou vide, utiliser Open Library avec l'ISBN :
  ![Titre](https://covers.openlibrary.org/b/isbn/{ISBN}-M.jpg)
Placer l'image juste avant ou après le titre dans la réponse.

JUSTIFICATION OBLIGATOIRE — pour chaque titre proposé au désherbage, renseigner
le champ motif avec une phrase factuelle basée sur les données réelles.
Exemples : "0 prêt depuis 2019, édition de 1998, thème obsolète"
/ "3 prêts en 8 ans, dernier prêt 2021, série abandonnée par l'éditeur"
/ "Exemplaire en mauvais état (statut Decalog), 12 prêts depuis 2015"
/ "Doublon — même titre présent en 2 exemplaires, le moins emprunté des deux"
Ne jamais laisser le motif vide ou générique ("vieux", "peu emprunté").

── PÉPITES MÉCONNUES / MISE EN AVANT ─────────────────
Livres peu empruntés qui méritent mieux (prix littéraires, bonnes critiques) :
1. executer_requete_sql pour trouver les titres sous-empruntés (nb_prets faible).
2. web_search pour vérifier si le titre ou l'auteur a reçu des distinctions.
3. Si oui : ajouter_suggestion_mise_en_avant avec le motif précis (ex. "2 prêts
   seulement mais Prix Sorcières 2023 -- mérite d'être mis en avant en présentoir").

── BIBLIOGRAPHIES THÉMATIQUES ───────────────────────
Pour une animation, une visite de classe, un thème donné : executer_requete_sql
en filtrant sur genre, mots_cles, public, statut_exemplaire (pour ne sélectionner
que les disponibles si demandé), puis generer_export_excel pour la liste imprimable.
Ajoute toujours la cote pour faciliter le travail en rayon.

── SUGGESTIONS DE LECTURE ────────────────────────────
"Si un usager a aimé X, quoi lui proposer ?" :
1. Cherche X dans notre fonds (genre/mots_cles/resume réels = meilleure base).
2. Si X absent : web_search pour cerner ses thèmes.
3. executer_requete_sql pour vrais titres disponibles avec genre/mots_cles proches.
Précise toujours la cote et la disponibilité (à la date du dernier import).

── DEMANDES INSATISFAITES ────────────────────────────
Quand un usager demande un livre absent du fonds : vérifie d'abord que c'est
bien absent (executer_requete_sql), puis propose de l'ajouter à la liste
de suggestions avec ajouter_suggestion_acquisition, motif = "Demande usager",
demandeur = le nom de l'agent au comptoir (ou "Public" si pas précisé).

── LISTES PERSISTANTES (AJOUT/SUPPRESSION/EXPORT) ───
Trois types de listes, chacune avec add/delete/export :
• suggestion_acquisition -- achats envisagés
• suggestion_desherbage  -- retraits envisagés (avec nb_prets, dernier_pret, cote)
• suggestion_mise_en_avant -- pépites à mettre en valeur
Chaque liste filtrée par demandeur (WHERE demandeur = '...'). Pour supprimer,
trouver l'id via executer_requete_sql d'abord. Pour exporter, generer_export_excel
avec sql = SELECT ... FROM suggestion_XXX WHERE demandeur = '...'.
Pour ajouter PLUSIEURS titres d'un coup : plusieurs appels d'outil successifs,
un par titre.

── JOURNAL D'USAGE ───────────────────────────────────
La table journal_requetes (id, date_requete, question, sql_executees,
nb_recherches_web, a_genere_export, a_modifie_suggestions, erreur) enregistre
automatiquement chaque question. Pour analyser l'usage de l'outil lui-même,
utilise executer_requete_sql sur cette table.

── NOTICES SANS EAN DECALOG ───────────────────────────
Certains exemplaires physiques existent dans les rayons mais leur notice
Decalog n'a pas d'EAN renseigné. Ces notices entrent en base avec un
identifiant préfixé "CB:" (ex. CB:0123456789). Quand tu les rencontres :
• Ne dis JAMAIS qu'un titre est absent du fonds si son isbn commence par "CB:"
• Précise qu'il est bien présent physiquement, mais sans EAN dans Decalog
• Signale qu'une correction dans Decalog serait nécessaire pour un
  fonctionnement optimal (notamment pour la recherche par ISBN).

Réponds en français, avec les chiffres exacts retournés par les outils.
Si un outil ne retourne rien, dis-le clairement."""


def repondre(historique_existant, question, cle_api):
    """Travaille sur une COPIE de l'historique, jamais sur la liste
    st.session_state.messages_api directement. Si une nouvelle question
    arrive pendant qu'une réponse précédente est encore en cours (plusieurs
    recherches web qui prennent du temps), Streamlit peut abandonner
    l'exécution en cours en plein milieu -- sans cette précaution, l'objet
    partagé resterait alors à moitié écrit (un tool_use sans son
    tool_result), et toute la conversation deviendrait invalide. En
    travaillant sur une copie et en ne renvoyant le résultat qu'à la toute
    fin, un abandon en cours de route ne laisse plus aucune trace.

    Chaque appel est journalisé (succès ou échec) dans journal_requetes,
    pour permettre d'étudier l'usage réel de l'outil a posteriori."""
    historique = list(historique_existant)
    historique.append({"role": "user", "content": question})
    client = Anthropic(api_key=cle_api)

    outils = [
        OUTIL_SQL, OUTIL_EXPORT, OUTIL_RAPPORT,
        OUTIL_SUGGESTION, OUTIL_SUPPRESSION_SUGGESTION, OUTIL_STATUER_SUGGESTION,
        OUTIL_DESHERBAGE, OUTIL_SUPPRESSION_DESHERBAGE,
        OUTIL_MISE_EN_AVANT, OUTIL_SUPPRESSION_MISE_EN_AVANT,
        OUTIL_DESHERBAGE_EFFECTUE, OUTIL_SUPPRESSION_DESHERBAGE_EFFECTUE,
        OUTIL_ANALYSE_ACQUISITION,
        # cache_control sur le DERNIER outil de la liste : Anthropic met en
        # cache tout le préfixe jusqu'à ce marqueur (tous les outils +, avec
        # le marqueur système ci-dessous, le prompt système complet). Ajouté
        # le 2026-07-23 : réduit la latence et le coût de chaque appel, sans
        # changer le comportement -- le contenu envoyé au modèle est
        # strictement identique, seul son traitement est mis en cache
        # (5 minutes de rémanence, renouvelées à chaque appel).
        {
            "type": "web_search_20250305", "name": "web_search", "max_uses": 10,
            "cache_control": {"type": "ephemeral"},
        },
    ]

    sql_executees = []
    nb_recherches_web = 0
    a_exporte = False
    a_modifie_suggestions = False
    erreur_pour_journal = None

    try:
        while True:
            reponse = client.messages.create(
                model="claude-sonnet-4-6",
                max_tokens=8000,
                # Prompt système passé en bloc avec cache_control (voir
                # commentaire sur 'outils' ci-dessus) -- PROMPT_SYSTEME est
                # volumineux (schéma complet + requêtes de référence) et
                # identique à chaque appel, donc son coût de traitement est
                # évité une fois mis en cache.
                system=[{"type": "text", "text": PROMPT_SYSTEME, "cache_control": {"type": "ephemeral"}}],
                tools=outils,
                messages=historique,
            )
            historique.append({"role": "assistant", "content": reponse.content})

            try:  # champ officiel de comptage des recherches web côté serveur
                nb_recherches_web += reponse.usage.server_tool_use.web_search_requests
            except Exception:
                pass

            if reponse.stop_reason != "tool_use":
                texte = "".join(b.text for b in reponse.content if b.type == "text")
                return texte, historique

            resultats_outils = []
            for bloc in reponse.content:
                if bloc.type != "tool_use":
                    continue
                if bloc.name == "executer_requete_sql":
                    sql = bloc.input.get("sql", "")
                    sql_executees.append(sql)
                    resultat = executer_requete_sql(sql)
                elif bloc.name == "generer_export_excel":
                    a_exporte = True
                    contenu, n_lignes, erreur = generer_excel_bytes(bloc.input.get("sql"), bloc.input.get("lignes"))
                    if erreur:
                        resultat = json.dumps({"erreur": erreur})
                    else:
                        st.session_state["export_xlsx_pret"] = contenu
                        st.session_state["export_xlsx_lignes"] = n_lignes
                        resultat = json.dumps({
                            "statut": "ok", "lignes": n_lignes,
                            "info": "Fichier généré -- un bouton de téléchargement va apparaître juste sous ta réponse.",
                        })
                elif bloc.name == "ajouter_suggestion_acquisition":
                    a_modifie_suggestions = True
                    resultat = ajouter_suggestion_acquisition(**bloc.input)
                elif bloc.name == "supprimer_suggestion_acquisition":
                    a_modifie_suggestions = True
                    resultat = supprimer_suggestion_acquisition(**bloc.input)
                elif bloc.name == "statuer_suggestion_acquisition":
                    a_modifie_suggestions = True
                    resultat = statuer_suggestion_acquisition(**bloc.input)
                elif bloc.name == "ajouter_suggestion_desherbage":
                    a_modifie_suggestions = True
                    resultat = ajouter_suggestion_desherbage(**bloc.input)
                elif bloc.name == "supprimer_suggestion_desherbage":
                    a_modifie_suggestions = True
                    resultat = supprimer_suggestion_desherbage(**bloc.input)
                elif bloc.name == "ajouter_suggestion_mise_en_avant":
                    a_modifie_suggestions = True
                    resultat = ajouter_suggestion_mise_en_avant(**bloc.input)
                elif bloc.name == "supprimer_suggestion_mise_en_avant":
                    a_modifie_suggestions = True
                    resultat = supprimer_suggestion_mise_en_avant(**bloc.input)
                elif bloc.name == "generer_rapport_import":
                    resultat = json.dumps({"rapport": generer_rapport_import()})
                elif bloc.name == "enregistrer_desherbage_effectue":
                    a_modifie_suggestions = True
                    resultat = enregistrer_desherbage_effectue(**bloc.input)
                elif bloc.name == "supprimer_desherbage_effectue":
                    a_modifie_suggestions = True
                    resultat = supprimer_desherbage_effectue(**bloc.input)
                elif bloc.name == "lancer_analyse_acquisition":
                    resultat = lancer_analyse_acquisition()
                else:
                    resultat = json.dumps({"erreur": "outil inconnu"})
                resultats_outils.append({"type": "tool_result", "tool_use_id": bloc.id, "content": resultat})
            historique.append({"role": "user", "content": resultats_outils})
    except Exception as e:
        erreur_pour_journal = f"{type(e).__name__}: {e}"
        raise
    finally:
        journaliser_requete(question, sql_executees, nb_recherches_web, a_exporte, a_modifie_suggestions, erreur_pour_journal)


def repondre_flux(historique_existant, question, cle_api, resultat_out):
    """Version STREAMING de repondre() : générateur qui produit le texte de la
    réponse au fil de l'eau (à consommer via st.write_stream), tout en gérant
    la même boucle d'outils. L'historique mis à jour est déposé dans
    resultat_out['historique'] à la fin (un générateur ne peut pas à la fois
    yielder du texte et retourner une valeur). Mêmes garde-fous que
    repondre() : copie de l'historique, journalisation, aucune écriture de la
    conversation en cas d'erreur. repondre() est conservée intacte comme repli
    immédiat (il suffit de rebasculer la ligne d'appel dans l'interface)."""
    historique = list(historique_existant)
    historique.append({"role": "user", "content": question})
    client = Anthropic(api_key=cle_api)

    outils = [
        OUTIL_SQL, OUTIL_EXPORT, OUTIL_RAPPORT,
        OUTIL_SUGGESTION, OUTIL_SUPPRESSION_SUGGESTION, OUTIL_STATUER_SUGGESTION,
        OUTIL_DESHERBAGE, OUTIL_SUPPRESSION_DESHERBAGE,
        OUTIL_MISE_EN_AVANT, OUTIL_SUPPRESSION_MISE_EN_AVANT,
        OUTIL_DESHERBAGE_EFFECTUE, OUTIL_SUPPRESSION_DESHERBAGE_EFFECTUE,
        OUTIL_ANALYSE_ACQUISITION,
        {
            "type": "web_search_20250305", "name": "web_search", "max_uses": 10,
            "cache_control": {"type": "ephemeral"},
        },
    ]

    sql_executees = []
    nb_recherches_web = 0
    a_exporte = False
    a_modifie_suggestions = False
    erreur_pour_journal = None

    try:
        while True:
            with client.messages.stream(
                model="claude-sonnet-4-6",
                max_tokens=8000,
                system=[{"type": "text", "text": PROMPT_SYSTEME, "cache_control": {"type": "ephemeral"}}],
                tools=outils,
                messages=historique,
            ) as flux:
                for fragment in flux.text_stream:
                    yield fragment
                message = flux.get_final_message()

            historique.append({"role": "assistant", "content": message.content})

            try:
                nb_recherches_web += message.usage.server_tool_use.web_search_requests
            except Exception:
                pass

            if message.stop_reason != "tool_use":
                resultat_out["historique"] = historique
                return

            resultats_outils = []
            for bloc in message.content:
                if bloc.type != "tool_use":
                    continue
                if bloc.name == "executer_requete_sql":
                    sql = bloc.input.get("sql", "")
                    sql_executees.append(sql)
                    resultat = executer_requete_sql(sql)
                elif bloc.name == "generer_export_excel":
                    a_exporte = True
                    contenu, n_lignes, erreur = generer_excel_bytes(bloc.input.get("sql"), bloc.input.get("lignes"))
                    if erreur:
                        resultat = json.dumps({"erreur": erreur})
                    else:
                        st.session_state["export_xlsx_pret"] = contenu
                        st.session_state["export_xlsx_lignes"] = n_lignes
                        resultat = json.dumps({
                            "statut": "ok", "lignes": n_lignes,
                            "info": "Fichier généré -- un bouton de téléchargement va apparaître juste sous ta réponse.",
                        })
                elif bloc.name == "ajouter_suggestion_acquisition":
                    a_modifie_suggestions = True
                    resultat = ajouter_suggestion_acquisition(**bloc.input)
                elif bloc.name == "supprimer_suggestion_acquisition":
                    a_modifie_suggestions = True
                    resultat = supprimer_suggestion_acquisition(**bloc.input)
                elif bloc.name == "statuer_suggestion_acquisition":
                    a_modifie_suggestions = True
                    resultat = statuer_suggestion_acquisition(**bloc.input)
                elif bloc.name == "ajouter_suggestion_desherbage":
                    a_modifie_suggestions = True
                    resultat = ajouter_suggestion_desherbage(**bloc.input)
                elif bloc.name == "supprimer_suggestion_desherbage":
                    a_modifie_suggestions = True
                    resultat = supprimer_suggestion_desherbage(**bloc.input)
                elif bloc.name == "ajouter_suggestion_mise_en_avant":
                    a_modifie_suggestions = True
                    resultat = ajouter_suggestion_mise_en_avant(**bloc.input)
                elif bloc.name == "supprimer_suggestion_mise_en_avant":
                    a_modifie_suggestions = True
                    resultat = supprimer_suggestion_mise_en_avant(**bloc.input)
                elif bloc.name == "generer_rapport_import":
                    resultat = json.dumps({"rapport": generer_rapport_import()})
                elif bloc.name == "enregistrer_desherbage_effectue":
                    a_modifie_suggestions = True
                    resultat = enregistrer_desherbage_effectue(**bloc.input)
                elif bloc.name == "supprimer_desherbage_effectue":
                    a_modifie_suggestions = True
                    resultat = supprimer_desherbage_effectue(**bloc.input)
                elif bloc.name == "lancer_analyse_acquisition":
                    resultat = lancer_analyse_acquisition()
                else:
                    resultat = json.dumps({"erreur": "outil inconnu"})
                resultats_outils.append({"type": "tool_result", "tool_use_id": bloc.id, "content": resultat})
            historique.append({"role": "user", "content": resultats_outils})
    except Exception as e:
        erreur_pour_journal = f"{type(e).__name__}: {e}"
        raise
    finally:
        journaliser_requete(question, sql_executees, nb_recherches_web, a_exporte, a_modifie_suggestions, erreur_pour_journal)


# ----------------------------------------------------------------------------
# Dépôt de fichier -- enrichissement direct depuis l'interface, RÉSERVÉ à
# l'équipe (derrière le mot de passe si configuré). Utilise un jeton
# d'écriture séparé du jeton lecture-seule employé pour le chat : le chat
# reste strictement en lecture quoi qu'il arrive, même si cette fonction a
# un problème.
# ----------------------------------------------------------------------------
def deviner_type_fichier(nom):
    ext = os.path.splitext(nom)[1].lower()
    if ext == '.mrc':
        return 'catalogue'
    if ext in ('.xlsx', '.xls'):
        return 'statistiques'
    if ext == '.csv':
        return 'frequentation'
    return None


# Chaque ISBN qualifié est enregistré immédiatement (pas en un seul bloc à
# la fin) : même si la session Streamlit venait à être interrompue en cours
# de route, rien n'est perdu -- il suffit de redéposer le fichier, ou la
# tâche de fond sur le Mac prend le relais automatiquement pour le reste.





CHEMIN_ETAT_IMPORT = '/tmp/mediatheque_import_state.json'

def _sauver_etat_import(pid, chemin_log):
    """Sauvegarde l'état de l'import sur disque pour survivre aux resets de session."""
    import json as _json
    with open(CHEMIN_ETAT_IMPORT, 'w') as f:
        _json.dump({'pid': pid, 'chemin_log': chemin_log}, f)

def _lire_etat_import():
    """Lit l'état d'import depuis le disque."""
    import json as _json
    try:
        with open(CHEMIN_ETAT_IMPORT) as f:
            return _json.load(f)
    except Exception:
        return None

def _pid_tourne(pid):
    """Vérifie si un processus tourne encore."""
    try:
        os.kill(pid, 0)
        return True
    except (ProcessLookupError, PermissionError):
        return False

def _supprimer_etat_import():
    try:
        os.remove(CHEMIN_ETAT_IMPORT)
    except Exception:
        pass


def lancer_import_background(fichier_bytes, fichier_nom, url_turso, jeton_ecriture, ovh_config=None):
    """
    Lance l'import complet dans un subprocess séparé.
    Persiste le PID et le log sur disque pour survivre aux resets de session Streamlit.

    Si le fichier est un .mrc ET que les identifiants OVH sont fournis, on
    enchaîne AUTOMATIQUEMENT la régénération des écrans MAAT (mosaïque +
    diaporama) depuis ce même .mrc, puis leur envoi sur OVH -- de sorte que
    la mise à jour hebdomadaire complète (catalogue + écrans) puisse se faire
    depuis l'app, de n'importe où, sans passer par le Terminal. generer_ecrans_maat
    ne lit que le .mrc et le réseau (pas la base), donc il fonctionne aussi
    bien depuis le cloud que depuis le Mac.
    """
    import tempfile as _tempfile
    import subprocess as _subprocess
    import textwrap as _textwrap

    dossier_tmp = _tempfile.mkdtemp()
    chemin_fichier = os.path.join(dossier_tmp, fichier_nom)
    chemin_log = os.path.join(dossier_tmp, 'import.log')

    with open(chemin_fichier, 'wb') as f:
        f.write(fichier_bytes)

    app_dir = os.path.dirname(os.path.abspath(__file__))

    # Étape écrans MAAT : uniquement pour un .mrc, et seulement si on a les
    # identifiants OVH (sinon on génère sans envoyer, ce qui n'a pas de sens
    # depuis le cloud où les fichiers locaux sont éphémères).
    bloc_ecrans = ""
    if fichier_nom.lower().endswith('.mrc') and ovh_config and ovh_config.get("OVH_SFTP_HOST"):
        _env_ovh = "\n".join(
            f'os.environ[{_k!r}] = {_v!r}' for _k, _v in ovh_config.items() if _v
        )
        bloc_ecrans = "\nprint('--- Mise a jour des ecrans MAAT (mosaique + diaporama) ---')\ntry:\n" + \
            _textwrap.indent(_env_ovh, "    ") + "\n" + \
            "    sys.argv = ['generer_ecrans_maat.py', '--mrc', " + repr(chemin_fichier) + "]\n" + \
            "    import generer_ecrans_maat\n" + \
            "    generer_ecrans_maat.main()\n" + \
            "except Exception as _e:\n" + \
            "    print('Ecrans MAAT non regeneres (import catalogue non affecte):', _e)\n"

    script = f"""
import os, sys
os.environ["TURSO_DATABASE_URL"] = {repr(url_turso)}
os.environ["TURSO_AUTH_TOKEN"] = {repr(jeton_ecriture)}
sys.path.insert(0, {repr(app_dir)})
sys.argv = ['traiter_fichier.py', {repr(chemin_fichier)}]
import traiter_fichier
traiter_fichier.main()
{bloc_ecrans}"""

    with open(chemin_log, 'w', encoding='utf-8') as log_f:
        proc = _subprocess.Popen(
            [sys.executable, '-c', script],
            stdout=log_f, stderr=log_f,
            cwd=app_dir,
            start_new_session=True  # Détache du groupe de process Streamlit
        )

    # Sauvegarder PID + chemin log sur disque
    _sauver_etat_import(proc.pid, chemin_log)
    return proc.pid, chemin_log



# ----------------------------------------------------------------------------
# Note (audit du 2026-07-25) : ~270 lignes ont été retirées ici. Il s'agissait
# de fonctions jamais appelées -- vestiges d'un tableau de bord abandonné
# (get_kpis, get_frequentation_mensuelle, get_rotation_genres, get_top_titres,
# get_roi_acquisitions, get_suggestions_acquisition_liste, get_alertes), d'un
# export ORB jamais finalisé (exporter_suggestions_orb), d'une version
# synchrone de l'import remplacée par lancer_import_background
# (traiter_fichier_depose), et d'un scraping Babelio inutilisable (le
# robots.txt du site interdit l'accès automatisé).
# Elles restent consultables dans l'historique git si besoin.
# ----------------------------------------------------------------------------

# Interface
# ----------------------------------------------------------------------------
st.title("📚 Médiathèque d'Arcachon — Assistant du fonds")
st.caption("Médiathèque d'Arcachon — réseau COBAS")

cle_api = st.secrets.get("ANTHROPIC_API_KEY", os.environ.get("ANTHROPIC_API_KEY", ""))
if not cle_api:
    st.error("Clé API Anthropic manquante -- à configurer dans les secrets de l'application.")
    st.stop()

mot_de_passe_requis = st.secrets.get("MOT_DE_PASSE", "")
if mot_de_passe_requis:
    if "authentifie" not in st.session_state:
        st.session_state.authentifie = False
    if not st.session_state.authentifie:
        saisie = st.text_input("Mot de passe", type="password")
        if saisie:
            if saisie == mot_de_passe_requis:
                st.session_state.authentifie = True
                st.rerun()
            else:
                st.error("Mot de passe incorrect.")
        st.stop()

# ── TRI DES SUGGESTIONS (cases à cocher) ─────────────────────────────────────
# Panneau visuel pour trier la liste de suggestions alimentée par la veille :
# pour chaque titre, choisir « Garder » (→ à commander) ou « Écarter » (→ ne
# réapparaîtra plus), puis appliquer tout d'un coup. Complète le chat sans le
# remplacer -- placé au-dessus du chat, dans un volet repliable pour ne pas
# encombrer quand on ne s'en sert pas.
with st.expander("📋 Trier les suggestions d'acquisition (veille automatique)"):
    try:
        _conn_sugg = db.connect(FICHIER_DB)
        try:
            _lignes_sugg = _conn_sugg.execute(
                "SELECT id, titre, auteur, editeur, motif, source, "
                "       categorie, public_vise, genre, isbn, date_ajout "
                "FROM suggestion_acquisition WHERE statut = 'à étudier' "
                "ORDER BY (source LIKE 'Veille prix%') DESC, date_ajout DESC"
            ).fetchall()
        except Exception:
            # Base pas encore migrée (colonnes de classement absentes)
            _lignes_sugg = [
                tuple(r) + (None, None, None, None, None) for r in _conn_sugg.execute(
                    "SELECT id, titre, auteur, editeur, motif, source "
                    "FROM suggestion_acquisition WHERE statut = 'à étudier' "
                    "ORDER BY date_ajout DESC"
                ).fetchall()
            ]
        _conn_sugg.close()
    except Exception as _e:
        _lignes_sugg = []
        st.caption(f"Liste indisponible ({_e}).")

    if not _lignes_sugg:
        st.caption("Aucune suggestion en attente. La veille en ajoutera de nouvelles chaque semaine.")
    else:
        import pandas as _pd
        _df_sugg = _pd.DataFrame([
            {
                "Décision": "",
                "Titre": r[1],
                "Auteur": r[2] or "",
                "Éditeur": r[3] or "",
                "ISBN": str(r[9]) if len(r) > 9 and r[9] else "",
                # La catégorie est un TYPE de document pur : la nuance
                # jeunesse/ado vit dans la colonne Public (demande de
                # Thomas, 2026-07-27 — inutile d'avoir « Roman » ET
                # « Roman jeunesse » dans le menu quand le public se
                # choisit à côté).
                "Catégorie": {"Roman jeunesse": "Roman",
                              "Roman ado / YA": "Roman"}.get(r[6], r[6]) or "—",
                "Public": r[7] or "—",
                "Genre": r[8] or "—",
                "Motif / prix": r[4] or "",
                "Source": r[5] or "",
                "Ajoutée le": str(r[10])[:10] if len(r) > 10 and r[10] else "",
                "_id": r[0],
            }
            for r in _lignes_sugg
        ])

        # ── Filtres par segment ──────────────────────────────────────────
        # Plusieurs agents utilisent l'outil, chacun sur son domaine : on doit
        # pouvoir ne voir que « BD jeunesse », « manga adulte », « romans ado »...
        _c1, _c2, _c3 = st.columns(3)
        with _c1:
            _f_cat = st.multiselect(
                "Catégorie", sorted(_df_sugg["Catégorie"].unique()), default=[])
        with _c2:
            # options FIXES et non déduites des données : « Adulte » doit
            # être proposé même quand aucune suggestion adulte n'est en
            # attente (sinon l'agent croit que le filtre n'existe pas).
            _publics_canon = ["Adulte", "Jeunesse", "Adolescent", "Tout public"]
            _autres_pub = sorted(set(_df_sugg["Public"].unique())
                                 - set(_publics_canon))
            _f_pub = st.multiselect(
                "Public", _publics_canon + _autres_pub, default=[])
        with _c3:
            _f_src = st.multiselect(
                "Source", sorted(_df_sugg["Source"].unique()), default=[])
        _recherche = st.text_input(
            "Filtrer par mot (titre, auteur, éditeur, sélection d'origine)", "")

        _filtre = _df_sugg
        if _f_cat:
            _filtre = _filtre[_filtre["Catégorie"].isin(_f_cat)]
        if _f_pub:
            _filtre = _filtre[_filtre["Public"].isin(_f_pub)]
        if _f_src:
            _filtre = _filtre[_filtre["Source"].isin(_f_src)]
        if _recherche.strip():
            _m = _recherche.strip().lower()
            _filtre = _filtre[
                _filtre.apply(
                    lambda r: _m in " ".join(
                        str(r[c]).lower()
                        for c in ("Titre", "Auteur", "Éditeur", "Motif / prix")
                    ),
                    axis=1,
                )
            ]
        _df_sugg = _filtre

        st.caption(f"{len(_df_sugg)} suggestion(s) affichée(s) sur "
                   f"{len(_lignes_sugg)} en attente. Choisis « Garder » ou « Écarter » "
                   "pour chaque ligne, puis applique. Les titres non décidés restent en attente.")

        # ── Export Excel de la liste FILTRÉE ────────────────────────────
        # Les agents travaillent volontiers sur tableur (bons de commande,
        # circulation entre collègues) : on exporte exactement ce qui est
        # affiché, ISBN compris pour pouvoir commander directement.
        if len(_df_sugg):
            _colonnes_export = [
                ("Titre", "Titre"), ("Auteur", "Auteur"), ("Éditeur", "Éditeur"),
                ("ISBN", "ISBN / EAN"), ("Catégorie", "Catégorie"),
                ("Public", "Public visé"), ("Genre", "Genre"),
                ("Motif / prix", "Motif / sélection d'origine"),
                ("Source", "Source de la veille"), ("Ajoutée le", "Ajoutée le"),
            ]
            _lignes_export = _df_sugg[[c for c, _ in _colonnes_export]].to_dict("records")
            try:
                _xlsx, _n, _err = _ecrire_xlsx(_lignes_export, _colonnes_export,
                                               acces_par_cle=True)
                if _xlsx:
                    st.download_button(
                        f"📊 Télécharger ces {_n} suggestion(s) en Excel",
                        data=_xlsx,
                        file_name=f"suggestions_acquisition_{datetime.date.today().isoformat()}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="dl_suggestions_xlsx",
                    )
            except Exception as _e:
                st.caption(f"Export Excel indisponible ({_e}).")
        _edite = st.data_editor(
            _df_sugg,
            hide_index=True,
            use_container_width=True,
            column_config={
                "Décision": st.column_config.SelectboxColumn(
                    "Décision", options=["", "✅ Garder", "❌ Écarter"], required=False, width="small",
                ),
                "_id": None,  # colonne technique masquée
            },
            disabled=["Titre", "Auteur", "Éditeur", "ISBN", "Motif / prix",
                      "Source", "Catégorie", "Public", "Genre", "Ajoutée le"],
            key="editeur_suggestions",
        )
        if st.button("Appliquer le tri", type="primary"):
            _gardes = _ecartes = 0
            for _, _r in _edite.iterrows():
                _dec = _r["Décision"]
                if _dec == "✅ Garder":
                    statuer_suggestion_acquisition(int(_r["_id"]), "à commander")
                    _gardes += 1
                elif _dec == "❌ Écarter":
                    statuer_suggestion_acquisition(int(_r["_id"]), "écartée")
                    _ecartes += 1
            if _gardes or _ecartes:
                st.success(f"Tri appliqué : {_gardes} à commander, {_ecartes} écartée(s). "
                           "Les « à commander » restent consultables (statut « à commander »).")
                st.rerun()
            else:
                st.info("Aucune décision cochée -- rien n'a changé.")

# ── CHAT ─────────────────────────────────────────────────────────────────────
if "messages_affiches" not in st.session_state:
    st.session_state.messages_affiches = []
if "messages_api" not in st.session_state:
    st.session_state.messages_api = []

for msg in st.session_state.messages_affiches:
    with st.chat_message(msg["role"]):
        st.markdown(msg["content"])

question = st.chat_input("Quels mangas n'avons-nous jamais prêtés ?")
if question:
    st.session_state.messages_affiches.append({"role": "user", "content": question})
    with st.chat_message("user"):
        st.markdown(question)

    with st.chat_message("assistant"):
        # Affichage en STREAMING : le texte apparaît au fil de l'eau via
        # repondre_flux (générateur). Repli immédiat possible en réutilisant
        # repondre() (non-streaming), qui reste en place :
        #     texte, nh = repondre(st.session_state.messages_api, question, cle_api)
        #     st.session_state.messages_api = nh ; st.markdown(texte)
        _resultat_flux = {}
        try:
            texte = st.write_stream(
                repondre_flux(st.session_state.messages_api, question, cle_api, _resultat_flux)
            )
            if "historique" in _resultat_flux:
                # N'écrase l'historique de conversation que si la réponse a
                # abouti (le générateur ne dépose 'historique' qu'à la fin).
                st.session_state.messages_api = _resultat_flux["historique"]
        except Exception as e:
            texte = f"Erreur : {e}. L'historique n'a pas été modifié, tu peux reposer ta question normalement."
            st.markdown(texte)
        if st.session_state.get("export_xlsx_pret"):
            n_lignes = st.session_state.get("export_xlsx_lignes", 0)
            st.download_button(
                f"📥 Télécharger le fichier Excel ({n_lignes} lignes)",
                data=st.session_state["export_xlsx_pret"],
                file_name=f"export_mediatheque_arcachon_{datetime.date.today().isoformat()}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            del st.session_state["export_xlsx_pret"]
            del st.session_state["export_xlsx_lignes"]

    st.session_state.messages_affiches.append({"role": "assistant", "content": texte})

# ── SIDEBAR ──────────────────────────────────────────────────────────────────
with st.sidebar:
    st.header("Fonds Arcachon")
    try:
        conn = db.connect(FICHIER_DB)
        total = conn.execute("SELECT COUNT(*) FROM notice").fetchall()[0][0]
        exemplaires = conn.execute("SELECT COUNT(*) FROM exemplaire").fetchall()[0][0]
        conn.close()
        st.metric("Notices", f"{total:,}".replace(",", " "))
        st.metric("Exemplaires", f"{exemplaires:,}".replace(",", " "))
    except Exception:
        st.write("Base indisponible.")

    st.divider()
    if st.button("Nouvelle conversation"):
        st.session_state.messages_affiches = []
        st.session_state.messages_api = []
        st.rerun()

    st.divider()
    with st.expander("📤 Mettre à jour le fonds"):
        mot_de_passe_import = st.secrets.get("MOT_DE_PASSE_IMPORT", "")
        if not mot_de_passe_import:
            st.caption("Non configuré -- ajoute MOT_DE_PASSE_IMPORT dans les secrets pour activer cette fonction.")
        else:
            if "import_authentifie" not in st.session_state:
                st.session_state.import_authentifie = False
            if not st.session_state.import_authentifie:
                saisie_import = st.text_input("Mot de passe import", type="password", key="mdp_import")
                if saisie_import:
                    if saisie_import == mot_de_passe_import:
                        st.session_state.import_authentifie = True
                        st.rerun()
                    else:
                        st.error("Mot de passe incorrect.")
            else:
                jeton_ecriture = st.secrets.get("TURSO_AUTH_TOKEN_ECRITURE", "")
                if not jeton_ecriture:
                    st.caption("TURSO_AUTH_TOKEN_ECRITURE manquant dans les secrets.")
                else:
                    st.caption("Catalogue (.mrc), statistiques (.xlsx/.xls) ou fréquentation (.csv). "
                               "Déposer le .mrc met aussi à jour automatiquement les écrans "
                               "mosaïque + diaporama sur OVH (si identifiants OVH configurés).")

                    # Vérifier si un import tourne (depuis le disque, résistant aux resets)
                    etat_disque = _lire_etat_import()
                    if etat_disque:
                        pid = etat_disque['pid']
                        chemin_log = etat_disque['chemin_log']
                        if _pid_tourne(pid):
                            # Subprocess en cours — lire le log et afficher la progression
                            log = ''
                            try:
                                with open(chemin_log, encoding='utf-8') as f:
                                    log = f.read()
                            except Exception:
                                pass
                            import re as _re
                            # Patterns réels des scripts :
                            # actualiser_catalogue.py  → "X/Y notices traitées..." ou "X/Y exemplaires traités..."
                            # actualiser_statistiques.py → "X/Y traités..."
                            matches = _re.findall(r'(\d+)/(\d+)\s+(?:notices?\s+)?(?:exemplaires?\s+)?traités?', log)
                            if matches:
                                actuel, total = int(matches[-1][0]), int(matches[-1][1])
                                pct = actuel / total if total > 0 else 0
                                st.progress(pct, text=f"⏳ {actuel:,} / {total:,} traités ({int(pct*100)}%)")
                            else:
                                st.info("⏳ Import en cours...")
                            time.sleep(3)
                            st.rerun()
                        else:
                            # Subprocess terminé — afficher le résultat
                            log = ''
                            try:
                                with open(chemin_log, encoding='utf-8') as f:
                                    log = f.read()
                            except Exception:
                                pass
                            if log:
                                st.success("✅ Import terminé.")
                                st.code(log, language=None)
                            else:
                                st.warning("Import terminé (log vide).")
                            if st.button("Effacer"):
                                _supprimer_etat_import()
                                st.rerun()
                    else:
                        fichier_depose = st.file_uploader(
                            "Déposer un fichier",
                            type=['mrc', 'xlsx', 'xls', 'csv'],
                            key="depot"
                        )
                        if fichier_depose and st.button("Traiter ce fichier"):
                            _ovh_config = {
                                "OVH_SFTP_HOST": st.secrets.get("OVH_SFTP_HOST", ""),
                                "OVH_SFTP_PORT": str(st.secrets.get("OVH_SFTP_PORT", "22")),
                                "OVH_SFTP_USER": st.secrets.get("OVH_SFTP_USER", ""),
                                "OVH_SFTP_PASSWORD": st.secrets.get("OVH_SFTP_PASSWORD", ""),
                                "OVH_SFTP_DOSSIER": st.secrets.get("OVH_SFTP_DOSSIER", "www"),
                            }
                            lancer_import_background(
                                fichier_depose.getvalue(),
                                fichier_depose.name,
                                db.TURSO_URL,
                                jeton_ecriture,
                                ovh_config=_ovh_config,
                            )
                            st.rerun()

    if st.session_state.get("derniere_erreur_technique"):
        st.divider()
        with st.expander("🔧 Dernière erreur technique", expanded=True):
            st.code(st.session_state["derniere_erreur_technique"], language=None)
            if st.button("Effacer"):
                del st.session_state["derniere_erreur_technique"]
                st.rerun()
