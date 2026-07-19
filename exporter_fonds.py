#!/usr/bin/env python3
"""
Génère un tableau xlsx propre et lisible du fonds complet -- utilisable par
n'importe quel agent de la médiathèque, sans connaissance technique.

Usage :
    python3 exporter_fonds.py
    (écrit "Fonds Arcachon - AAAA-MM-JJ.xlsx" dans le même dossier)
"""

import os
import sqlite3
import db
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

FICHIER_DB = os.path.join(os.path.dirname(os.path.abspath(__file__)), "inventaire.db")

COLONNES = [
    ("isbn", "ISBN / EAN", 16),
    ("titre", "Titre", 32),
    ("serie", "Série", 20),
    ("tome", "Tome", 7),
    ("auteur", "Auteur", 18),
    ("illustrateur", "Illustrateur", 16),
    ("editeur", "Éditeur", 16),
    ("annee", "Année", 9),
    ("type", "Type", 8),
    ("categorie", "Catégorie", 14),
    ("genre", "Genre", 16),
    ("public", "Public", 11),
    ("age_recommande", "Âge conseillé", 13),
    ("cote", "Cote", 16),
    ("code_barres", "Code-barres", 14),
    ("statut_exemplaire", "Statut", 16),
    ("prix", "Prix (€)", 9),
    ("nb_prets_cet_exemplaire", "Prêts (Arcachon)", 14),
    ("dernier_pret_cet_exemplaire", "Dernier prêt", 12),
    ("dewey", "Dewey", 8),
    ("mots_cles", "Mots-clés", 28),
]

C_HEADER = "2E4A7A"
C_ALT_ROW = "F2F6FB"


def exporter():
    conn = db.connect(FICHIER_DB)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    champs = ", ".join(c[0] for c in COLONNES)
    cur.execute(f"SELECT {champs} FROM vue_inventaire ORDER BY titre, code_barres")
    lignes = cur.fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Fonds Arcachon"

    thin = Side(style="thin", color="D9E2F0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for c, (cle, libelle, largeur) in enumerate(COLONNES, 1):
        cell = ws.cell(row=1, column=c, value=libelle)
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=C_HEADER)
        cell.alignment = Alignment(vertical="center")
        cell.border = border
        ws.column_dimensions[get_column_letter(c)].width = largeur

    for i, r in enumerate(lignes, start=2):
        fond = C_ALT_ROW if i % 2 == 0 else "FFFFFF"
        for c, (cle, libelle, largeur) in enumerate(COLONNES, 1):
            val = r[cle]
            cell = ws.cell(row=i, column=c, value=val)
            cell.font = Font(name="Arial", size=9.5)
            cell.fill = PatternFill("solid", fgColor=fond)
            cell.border = border

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLONNES))}{len(lignes)+1}"

    nom_fichier = f"Fonds Arcachon - {datetime.date.today().isoformat()}.xlsx"
    chemin = os.path.join(os.path.dirname(os.path.abspath(__file__)), nom_fichier)
    wb.save(chemin)
    print(f"Export écrit : {chemin} ({len(lignes)} lignes)")
    return chemin


if __name__ == "__main__":
    exporter()
