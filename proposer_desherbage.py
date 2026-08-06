#!/usr/bin/env python3
"""
proposer_desherbage.py — Candidats au désherbage selon les formules du
document ENSSIB « Désherbage » (BDP Sarthe) transmis par Marie (2026-08-02).
STRICTEMENT EN LECTURE ; produit un Excel de travail pour l'équipe.

⚠ CE N'EST PAS UNE LISTE DE PILON : ce sont des CANDIDATS répondant aux
critères mécaniques (âge d'édition / années sans prêt). La décision finale
exige l'examen en rayon (état physique, critères IOUPI : Incorrect,
Ordinaire, Usé, Périmé, Inadéquat) et, avant tout retrait effectif, la
délibération municipale (modèles pages 12-13 du document ENSSIB).

FORMULES APPLIQUÉES (âge édition / années sans prêt, par rayon) :
  Documentaires (Dewey, adulte et jeunesse — mêmes critères selon ENSSIB) :
    00x informatique                    5 / 2
    03x encyclopédies, dictionnaires    5 / 2
    32x politique (actualité)           5 / 2
    34x droit, codes, législation       3 / 1
    autres 1xx-6xx                     10 / 3
    91x géographie, guides de voyage   10 / 3
    7xx et 9xx (art, histoire)         15 / 5
  Fiction :
    Romans adultes        sans prêt depuis 5 ans (sauf classiques — à valider)
    Romans/albums jeunesse sans prêt depuis 5 ans
    BD (tous publics)      sans prêt depuis 10 ans
  Doublons : notices à 3 exemplaires ou plus dont un sans prêt depuis 3 ans.

EXCLUSIONS : Fonds local (ne se désherbe JAMAIS), périodiques (REVUE),
supports non-livre.

Usage :  python3 proposer_desherbage.py
"""
import datetime
import os
import re
import sys
from collections import defaultdict

DOSSIER = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, DOSSIER)
import db  # noqa: E402

ANNEE = datetime.date.today().year


def regle_documentaire(dewey):
    """Renvoie (âge max, années sans prêt max, libellé) selon la classe Dewey."""
    if not dewey:
        return None
    m = re.match(r"^(\d{1,3})", str(dewey).strip())
    if not m:
        return None
    d = int(m.group(1).ljust(3, "0"))
    if d < 100:
        if 30 <= d < 40:
            return (5, 2, "03x encyclopédies/dictionnaires : 5 ans / 2 sans prêt")
        return (5, 2, "00x informatique/généralités : 5 ans / 2 sans prêt")
    if 320 <= d < 330:
        return (5, 2, "32x politique/actualité : 5 ans / 2 sans prêt")
    if 340 <= d < 350:
        return (3, 1, "34x droit/législation : 3 ans / 1 sans prêt")
    if 910 <= d < 920:
        return (10, 3, "91x guides et géographie : 10 ans / 3 sans prêt")
    if d >= 700:
        return (15, 5, "7xx-9xx art/histoire : 15 ans / 5 sans prêt")
    return (10, 3, "1xx-6xx documentaires : 10 ans / 3 sans prêt")


def main():
    conn = db.connect()
    lignes = conn.execute("""
        SELECT n.identifiant, n.titre, n.createurs, n.editeur,
               n.date_publication, n.dewey, n.categorie, n.public_vise,
               n.serie, n.tome, e.code_barre_exemplaire, e.cote, e.statut,
               e.prix, COALESCE(e.nb_prets_total, 0), e.annee_dernier_pret
        FROM notice n JOIN exemplaire e ON e.identifiant = n.identifiant
        WHERE n.type_document = 'LIVRE'
          AND COALESCE(n.categorie, '') != 'Fonds local'
    """).fetchall()
    conn.close()
    print(f"{len(lignes)} exemplaires LIVRE examinés (hors Fonds local).")

    nb_ex_par_notice = defaultdict(int)
    for l in lignes:
        nb_ex_par_notice[l[0]] += 1

    resultats = defaultdict(list)   # onglet -> lignes
    for (ident, titre, createurs, editeur, date_pub, dewey, cat, public,
         serie, tome, cb, cote, statut, prix, nb_prets, an_pret) in lignes:

        m = re.match(r"^(\d{4})", str(date_pub or ""))
        age = ANNEE - int(m.group(1)) if m else None
        sans_pret = (ANNEE - int(an_pret)) if an_pret else None
        jamais = (nb_prets == 0 and not an_pret)

        base = (cb, cote or "", titre or "", (createurs or "").split(";")[0],
                editeur or "", m.group(1) if m else "?",
                str(an_pret or ("jamais" if jamais else "?")), nb_prets,
                prix, statut or "")

        est_doc = (cat == "Documentaire") or (dewey and cat not in (
            "BD", "Manga", "Album"))
        # 1) documentaires : formule âge/usage par Dewey
        if est_doc and dewey:
            r = regle_documentaire(dewey)
            if r and age is not None and age >= r[0] and (
                    jamais or (sans_pret is not None and sans_pret >= r[1])):
                resultats["Documentaires"].append(base + (r[2],))
                continue

        # 2) fiction : usage seul
        if cat in ("BD", "Manga"):
            if jamais and age is not None and age >= 10:
                resultats["BD-manga jamais prêtées"].append(
                    base + ("BD/manga : jamais prêté et > 10 ans",))
            elif sans_pret is not None and sans_pret >= 10:
                resultats["BD-manga sans prêt 10 ans"].append(
                    base + ("BD/manga : sans prêt depuis 10 ans",))
            continue
        if cat in ("Roman jeunesse", "Album", "Première lecture",
                   "Roman ado / YA"):
            if sans_pret is not None and sans_pret >= 5:
                resultats["Jeunesse sans prêt 5 ans"].append(
                    base + ("Jeunesse : sans prêt depuis 5 ans",))
            continue
        if public == "Adulte" and not est_doc:
            if sans_pret is not None and sans_pret >= 5:
                resultats["Romans adultes sans prêt 5 ans"].append(
                    base + ("Roman adulte : sans prêt depuis 5 ans "
                            "(classiques à conserver !)",))
            continue

    # 3) doublons : ≥ 3 exemplaires dont un dormant depuis 3 ans
    for (ident, titre, createurs, editeur, date_pub, dewey, cat, public,
         serie, tome, cb, cote, statut, prix, nb_prets, an_pret) in lignes:
        if nb_ex_par_notice[ident] >= 3 and an_pret and ANNEE - int(an_pret) >= 3:
            m = re.match(r"^(\d{4})", str(date_pub or ""))
            resultats["Doublons dormants"].append(
                (cb, cote or "", titre or "", (createurs or "").split(";")[0],
                 editeur or "", m.group(1) if m else "?", str(an_pret),
                 nb_prets, prix, statut or "",
                 f"{nb_ex_par_notice[ident]} exemplaires, celui-ci sans prêt "
                 f"depuis {ANNEE - int(an_pret)} ans"))

    # ── Excel ────────────────────────────────────────────────────────────
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    mode = wb.active
    mode.title = "Mode d'emploi"
    consignes = [
        "CANDIDATS AU DÉSHERBAGE — méthode ENSSIB-IOUPI (document transmis par Marie)",
        "",
        "Ces listes appliquent UNIQUEMENT les critères mécaniques (âge d'édition,",
        "années sans prêt). Avant toute décision, examiner chaque livre EN RAYON :",
        "  I = Incorrect (fausse information)    O = Ordinaire, médiocre",
        "  U = Usé, détérioré                    P = Périmé",
        "  I = Inadéquat (ne correspond pas au fonds / au public)",
        "",
        "Destinations possibles : pilon / don / vente / rachat / mise en réserve.",
        "Le fonds local n'est jamais désherbé (déjà exclu de ces listes).",
        "AVANT tout retrait effectif : délibération municipale + procès-verbal",
        "(modèles pages 12-13 du document ENSSIB).",
        "",
        "Les 'classiques et valeurs sûres' se conservent même sans prêt : à",
        "signaler dans la colonne Décision.",
    ]
    for i, t in enumerate(consignes, 1):
        mode.cell(row=i, column=1, value=t).font = Font(
            name="Arial", size=10, bold=(i == 1))
    mode.column_dimensions["A"].width = 90

    entetes = ["Code-barres", "Cote", "Titre", "Auteur", "Éditeur",
               "Année éd.", "Dernier prêt", "Prêts (ex.)", "Prix (€)",
               "Statut", "Règle", "Décision (équipe)"]
    thin = Side(style="thin", color="D9E2F0")
    bord = Border(left=thin, right=thin, top=thin, bottom=thin)
    total = 0
    for onglet in sorted(resultats):
        lst = sorted(resultats[onglet], key=lambda x: (x[1], x[2]))
        ws = wb.create_sheet(re.sub(r"[/\\?*\[\]:]", "-", onglet)[:31])
        for c, nomc in enumerate(entetes, 1):
            cell = ws.cell(row=1, column=c, value=nomc)
            cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2E4A7A")
            cell.border = bord
            ws.column_dimensions[get_column_letter(c)].width = (
                38 if nomc in ("Titre", "Règle") else 14)
        for i, lg in enumerate(lst, start=2):
            for c, val in enumerate(lg, 1):
                cell = ws.cell(row=i, column=c, value=val)
                cell.font = Font(name="Arial", size=9.5)
                cell.border = bord
        ws.freeze_panes = "A2"
        total += len(lst)
        print(f"  {len(lst):>6}  {onglet}")

    sortie = os.path.join(
        DOSSIER, f"Desherbage_candidats_{datetime.date.today().isoformat()}.xlsx")
    wb.save(sortie)
    print(f"  {total:>6}  TOTAL candidats")
    print(f"\n✓ {os.path.basename(sortie)}")


if __name__ == "__main__":
    main()
