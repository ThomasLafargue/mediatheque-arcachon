#!/usr/bin/env python3
"""
retrouver_ean_manquants.py — Retrouve les EAN/ISBN absents de Decalog, par
recherche titre + auteur sur Place des Libraires (données Dilicom).

CONTEXTE (2026-07-26) : 1 190 livres imprimés publiés depuis 2000 n'ont pas
d'EAN dans Decalog. Ils reçoivent chez nous un identifiant de substitution
« CB:xxxx » et sont donc INVISIBLES à toute recherche par ISBN : jamais
enrichis, susceptibles d'être suggérés à l'achat alors qu'on les possède
déjà, et absents des écrans faute de couverture.
Le trou se concentre sur 2000-2017 (pic en 2012-2016) ; à partir de 2018 le
catalogage est propre (0,2 % de lacunes).

CE QUE FAIT CE SCRIPT
  1. liste les notices concernées (identifiant CB:, imprimé, date >= seuil) ;
  2. pour chacune, interroge la recherche de Place des Libraires par
     titre + auteur ;
  3. compare le titre trouvé au titre Decalog et attribue un INDICE DE
     CONFIANCE ;
  4. produit un fichier Excel trié, prêt pour une correction dans Decalog.

IL N'ÉCRIT RIEN EN BASE. La correction reste manuelle et humaine : un ISBN
erroné serait bien pire que pas d'ISBN du tout. L'indice de confiance sert
justement à trier ce qui peut être saisi rapidement de ce qui demande
vérification.

Usage :
    python3 retrouver_ean_manquants.py --test 15     (essai sur 15 notices)
    python3 retrouver_ean_manquants.py               (tout, long)
    python3 retrouver_ean_manquants.py --depuis 2010 (restreindre la période)
"""

import datetime
import os
import re
import sys
import time
import unicodedata
import urllib.parse

DOSSIER = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, DOSSIER)
import db  # noqa: E402 — correctif SSL (certifi) inclus

import requests  # noqa: E402
from bs4 import BeautifulSoup  # noqa: E402

BASE = "https://www.placedeslibraires.fr"
EN_TETES = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Accept-Language": "fr-FR,fr;q=0.9",
}
PAUSE = 0.6          # courtoisie envers le site
ANNEE_MIN = 2000     # avant, l'ISBN n'était pas généralisé


def _normaliser(t):
    if not t:
        return ""
    t = unicodedata.normalize("NFKD", t)
    t = "".join(c for c in t if not unicodedata.combining(c)).lower()
    t = re.sub(r"[^a-z0-9 ]", " ", t)
    return re.sub(r"\s+", " ", t).strip()


def _similarite(a, b):
    """Part de mots communs (0 à 1). Simple et lisible : on ne cherche pas la
    finesse d'un algorithme de distance, mais un indice de confiance."""
    ma, mb = set(_normaliser(a).split()), set(_normaliser(b).split())
    if not ma or not mb:
        return 0.0
    return len(ma & mb) / max(len(ma), len(mb))


def notices_sans_ean(annee_min):
    """Notices imprimées récentes dont l'identifiant est un substitut CB:."""
    conn = db.connect()
    try:
        lignes = conn.execute(
            "SELECT n.identifiant, n.titre, n.createurs, n.date_publication, "
            "       n.editeur, e.cote "
            "FROM notice n LEFT JOIN exemplaire e ON n.identifiant = e.identifiant "
            "WHERE n.identifiant LIKE 'CB:%' "
            "  AND n.type_document = 'LIVRE' "
            "  AND CAST(SUBSTR(n.date_publication, 1, 4) AS INTEGER) >= ? "
            "ORDER BY n.date_publication DESC",
            (annee_min,),
        ).fetchall()
    finally:
        conn.close()
    # dédoublonnage par identifiant (plusieurs exemplaires possibles)
    vues, res = set(), []
    for ident, titre, auteur, date_pub, editeur, cote in lignes:
        if ident in vues:
            continue
        vues.add(ident)
        res.append({"identifiant": ident, "titre": titre, "auteur": auteur,
                    "annee": (date_pub or "")[:4], "editeur": editeur, "cote": cote})
    return res


def chercher_ean(titre, auteur):
    """Cherche le livre sur Place des Libraires. Renvoie (isbn, titre_trouve)
    ou (None, None)."""
    requete = " ".join(x for x in (titre, (auteur or "").split(",")[0]) if x)[:120]
    url = f"{BASE}/listeliv.php?mots_recherche=" + urllib.parse.quote(requete)
    try:
        r = requests.get(url, headers=EN_TETES, timeout=20, allow_redirects=True)
        if r.status_code != 200:
            return None, None
        soup = BeautifulSoup(r.text, "html.parser")
        for a in soup.select("a[href*='/livre/']"):
            m = re.search(r"/livre/(\d{9,13})", a.get("href", ""))
            if not m:
                continue
            isbn = m.group(1)
            # le libellé du lien est souvent vide (image) : on lit alors le
            # slug de l'URL, qui reprend le titre
            libelle = re.sub(r"\s+", " ", a.get_text(strip=True))
            if not libelle:
                ms = re.search(r"/livre/\d+/([^/?]+)", a.get("href", ""))
                libelle = ms.group(1).replace("-", " ") if ms else ""
            return isbn, libelle
    except Exception:
        return None, None
    return None, None


def main():
    limite = None
    annee_min = ANNEE_MIN
    if "--test" in sys.argv:
        try:
            limite = int(sys.argv[sys.argv.index("--test") + 1])
        except (IndexError, ValueError):
            limite = 15
    if "--depuis" in sys.argv:
        try:
            annee_min = int(sys.argv[sys.argv.index("--depuis") + 1])
        except (IndexError, ValueError):
            pass

    print("═══ Recherche des EAN manquants (Place des Libraires) ═══\n")
    notices = notices_sans_ean(annee_min)
    print(f"{len(notices)} notice(s) sans EAN publiées depuis {annee_min}.")
    if limite:
        notices = notices[:limite]
        print(f"Mode test : {len(notices)} traitée(s).")
    print()

    resultats = []
    trouves = 0
    for i, n in enumerate(notices, 1):
        isbn, titre_trouve = chercher_ean(n["titre"], n["auteur"])
        conf = _similarite(n["titre"], titre_trouve) if isbn else 0.0
        if isbn:
            trouves += 1
        niveau = ("élevée" if conf >= 0.8 else
                  "moyenne" if conf >= 0.5 else
                  "faible" if isbn else "—")
        resultats.append({
            "Cote": n["cote"] or "",
            "Titre (Decalog)": n["titre"] or "",
            "Auteur": n["auteur"] or "",
            "Année": n["annee"],
            "Éditeur": n["editeur"] or "",
            "ISBN trouvé": isbn or "",
            "Titre trouvé (Place des Libraires)": titre_trouve or "",
            "Confiance": niveau,
            "Identifiant actuel": n["identifiant"],
        })
        marque = {"élevée": "✓", "moyenne": "~", "faible": "?"}.get(niveau, "∅")
        print(f"  {marque} [{i}/{len(notices)}] {str(n['titre'])[:40]:40} "
              f"→ {isbn or 'introuvable'}  ({niveau})")
        time.sleep(PAUSE)

    print(f"\n→ {trouves}/{len(notices)} ISBN retrouvés")
    eleves = sum(1 for r in resultats if r["Confiance"] == "élevée")
    print(f"→ dont {eleves} en confiance ÉLEVÉE (saisissables rapidement)")

    # Export Excel
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        wb = Workbook(); ws = wb.active; ws.title = "EAN à corriger"
        colonnes = list(resultats[0].keys()) if resultats else []
        thin = Side(style="thin", color="D9E2F0")
        bordure = Border(left=thin, right=thin, top=thin, bottom=thin)
        for c, nom in enumerate(colonnes, 1):
            cell = ws.cell(row=1, column=c, value=nom)
            cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2E4A7A")
            cell.border = bordure
            ws.column_dimensions[get_column_letter(c)].width = 22
        # trié : confiance élevée d'abord (le plus vite corrigeable)
        ordre = {"élevée": 0, "moyenne": 1, "faible": 2, "—": 3}
        for i, r in enumerate(sorted(resultats, key=lambda x: ordre[x["Confiance"]]), start=2):
            for c, nom in enumerate(colonnes, 1):
                cell = ws.cell(row=i, column=c, value=r[nom])
                cell.font = Font(name="Arial", size=9.5)
                cell.border = bordure
        ws.freeze_panes = "A2"
        sortie = os.path.join(
            DOSSIER, f"EAN_a_corriger_{datetime.date.today().isoformat()}.xlsx")
        wb.save(sortie)
        print(f"\n✓ Export : {os.path.basename(sortie)}")
        print("  À vérifier puis saisir dans Decalog. Rien n'a été écrit en base.")
    except Exception as e:
        print(f"\n⚠ Export Excel impossible : {e}")


if __name__ == "__main__":
    main()
