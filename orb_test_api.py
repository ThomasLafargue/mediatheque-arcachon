#!/usr/bin/env python3
"""
orb_test_api.py — Test GRATUIT de l'API ORB (Decitre) avec un compte démo.

Compte démo à demander à chargeclientele@decitre.fr (ou nboyer@decitre.fr).
La liste des EAN autorisés en démo est publique :
    https://doc.api.base-orb.fr/demo_products.txt

Ce script interroge /products sur les EAN de démo et produit :
  - un aperçu console de la richesse des notices (champs remplis) ;
  - ORB_test_notices.xlsx : une ligne par notice, les champs utiles à plat ;
  - si --marc : un échantillon MarcXchange via /products/updates (si le
    compte démo y a droit), enregistré dans ORB_test_marcxchange.xml.

Usage :
    python3 orb_test_api.py IDENTIFIANT CLE_API
    python3 orb_test_api.py IDENTIFIANT CLE_API --marc
"""
import sys
import json
import urllib.request

BASE = "https://api.base-orb.fr/v1"
URL_EANS_DEMO = "https://doc.api.base-orb.fr/demo_products.txt"


def appel(chemin, user, cle):
    req = urllib.request.Request(BASE + chemin)
    import base64
    req.add_header("Authorization", "Basic " +
                   base64.b64encode(f"{user}:{cle}".encode()).decode())
    req.add_header("Accept-Encoding", "gzip")
    r = urllib.request.urlopen(req, timeout=30)
    donnees = r.read()
    if r.headers.get("Content-Encoding") == "gzip":
        import gzip
        donnees = gzip.decompress(donnees)
    return json.loads(donnees)


def plat(p):
    """Aplatit les champs utiles d'un produit ORB."""
    def g(*chemin, defaut=""):
        v = p
        for c in chemin:
            v = (v or {}).get(c) if isinstance(v, dict) else None
        return v if v not in (None, {}) else defaut
    auteurs = " ; ".join(
        f"{a.get('last_name','')}, {a.get('first_name','')}".strip(", ")
        for a in (p.get("authors") or []))
    return {
        "ean13": p.get("ean13", ""),
        "titre": p.get("title", ""),
        "label": p.get("label", ""),
        "sous_titre": p.get("subtitle", ""),
        "serie": g("series", "label"),
        "tome": g("series_sequence", "label"),
        "auteurs": auteurs,
        "editeur": g("publisher", "label"),
        "collection": g("collection", "label"),
        "n_collection": g("collection_sequence", "label"),
        "date_parution": p.get("published_at", ""),
        "prix_ttc": p.get("price_with_taxes", ""),
        "pages": p.get("number_of_pages", ""),
        "dewey": g("thesaurus", "dewey", "code"),
        "theso_decitre": g("thesaurus", "decitre", "family", "label"),
        "publics": " ; ".join(t.get("label", "") for t in (p.get("targeted_audiences") or [])),
        "age": g("age_requirement", "label"),
        "dispo": g("availability_orb", "label"),
        "resume_present": "oui" if p.get("summary") else "",
        "image_present": "oui" if g("images", "front") else "",
        "prix_litteraires": " ; ".join(
            f"{la.get('award',{}).get('label','')} {la.get('year','')}"
            for la in (p.get("literary_awards") or [])),
    }


def main():
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)
    user, cle = sys.argv[1], sys.argv[2]
    avec_marc = "--marc" in sys.argv

    eans = urllib.request.urlopen(URL_EANS_DEMO, timeout=20).read().decode().split()
    print(f"{len(eans)} EAN de démo.")

    produits = []
    for i in range(0, len(eans), 25):
        lot = eans[i:i + 25]
        rep = appel(f"/products?eans={','.join(lot)}&sort=ean_asc&limit=100",
                    user, cle)
        produits += rep.get("data", [])
        print(f"  ... {len(produits)} notices reçues", end="\r")
    print(f"\n{len(produits)} notices récupérées.")

    lignes = [plat(p) for p in produits]
    # taux de remplissage par champ
    print("\nTaux de remplissage des champs utiles :")
    for champ in lignes[0].keys():
        n = sum(1 for l in lignes if l[champ])
        print(f"  {100*n//len(lignes):>3} %  {champ}")

    from openpyxl import Workbook
    from openpyxl.styles import Font
    wb = Workbook(); ws = wb.active; ws.title = "Notices démo ORB"
    for c, nom in enumerate(lignes[0].keys(), 1):
        ws.cell(row=1, column=c, value=nom).font = Font(bold=True)
    for i, l in enumerate(lignes, start=2):
        for c, v in enumerate(l.values(), 1):
            ws.cell(row=i, column=c, value=str(v)[:500])
    wb.save("ORB_test_notices.xlsx")
    print("\n✓ ORB_test_notices.xlsx")

    if avec_marc:
        rep = appel("/products/updates?since=1&limit=20&format=marcxchange",
                    user, cle)
        contenu = rep.get("data")
        with open("ORB_test_marcxchange.xml", "w", encoding="utf-8") as f:
            f.write(contenu if isinstance(contenu, str) else json.dumps(contenu, ensure_ascii=False))
        print("✓ ORB_test_marcxchange.xml (échantillon UNIMARC)")


if __name__ == "__main__":
    main()
