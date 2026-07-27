#!/usr/bin/env python3
"""
exporter_corrections_decalog.py — Tableau Excel des fiches Decalog mal
renseignées, avec le champ concerné et la valeur à saisir.

PRINCIPE : notre moteur d'enrichissement complète les champs que Decalog a
laissés vides (série, tome, catégorie, genre, public visé...). Ces déductions
vivent dans NOTRE base, mais Decalog, lui, reste incomplet. Ce script produit
la liste de travail pour corriger Decalog à la source — ce qui bénéficie
ensuite à tout le réseau COBAS, pas seulement à notre outil.

DEUX FAMILLES DE CORRECTIONS
  1. Champs déduits par notre moteur (colonne champs_a_verifier_decalog) :
     on connaît la valeur à saisir, il n'y a qu'à la recopier.
  2. Champs manquants sans déduction possible : on les signale quand même,
     car ce sont des trous à combler manuellement.

Une feuille Excel par type de correction, pour pouvoir répartir le travail
entre agents.

N'ÉCRIT RIEN EN BASE ni dans Decalog : c'est une liste de travail.

Usage :
    python3 exporter_corrections_decalog.py
    python3 exporter_corrections_decalog.py --jeunesse   (fonds jeunesse seul)
    python3 exporter_corrections_decalog.py --max 500
"""

import datetime
import os
import sys

DOSSIER = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, DOSSIER)
import db  # noqa: E402

LIBELLES = {
    "serie": "Série",
    "tome": "Tome",
    "categorie": "Catégorie",
    "genre": "Genre",
    "public_vise": "Public visé",
    "pegi": "PEGI",
    "dewey": "Dewey",
}


TAILLE_TRANCHE = 300


def _executer_avec_reprise(sql, essais=4):
    """Exécute une requête en la découpant en TRANCHES.

    Turso (protocole Hrana) coupe les réponses volumineuses avec
    « unexpected EOF during chunk size line » : sur ~4 300 lignes contenant
    titres et auteurs, l'échec est systématique et non passager (constaté le
    2026-07-26, y compris après plusieurs tentatives). On lit donc par
    tranches de 300 lignes, avec une nouvelle connexion et quelques essais
    par tranche. Plus lent, mais fiable -- et cela laisse respirer la base
    quand l'enrichissement tourne en parallèle.

    La requête reçue NE DOIT PAS contenir de LIMIT/OFFSET : ils sont ajoutés
    ici, avec un ORDER BY stable indispensable à la pagination.
    """
    import time
    lignes, decalage = [], 0
    while True:
        tranche = None
        derniere = None
        for tentative in range(essais):
            conn = None
            try:
                conn = db.connect()
                tranche = conn.execute(
                    f"{sql} ORDER BY n.identifiant "
                    f"LIMIT {TAILLE_TRANCHE} OFFSET {decalage}"
                ).fetchall()
                break
            except Exception as e:
                derniere = e
                if tentative < essais - 1:
                    time.sleep(2 * (tentative + 1))
            finally:
                if conn is not None:
                    try:
                        conn.close()
                    except Exception:
                        pass
        if tranche is None:
            raise derniere
        lignes.extend(tranche)
        print(f"  ... {len(lignes)} ligne(s) lues", end="\r", flush=True)
        if len(tranche) < TAILLE_TRANCHE:
            break
        decalage += TAILLE_TRANCHE
    print(f"  ... {len(lignes)} ligne(s) lues")
    return lignes


def recuperer(jeunesse_seule=False, maxi=None):
    filtre_jeunesse = (
        " AND n.public_vise IN ('Jeunesse','Adolescent') "
        if jeunesse_seule else ""
    )
    # NB (2026-07-26) : une sous-requête corrélée sur `exemplaire` (44 000
    # lignes) s'est révélée bien PLUS coûteuse que la jointure -- chaque ligne
    # déclenchait un balayage. On revient donc à la jointure, en dédoublonnant
    # côté Python. La cote reste facultative : si la base est trop sollicitée,
    # l'option --sans-cote permet de s'en passer.
    cote_sq = "e.cote AS cote"
    depuis = ("FROM notice n LEFT JOIN exemplaire e "
              "ON e.identifiant = n.identifiant ")
    if "--sans-cote" in sys.argv:
        cote_sq = "'' AS cote"
        depuis = "FROM notice n "

    # 1. Champs DÉDUITS par notre moteur : la valeur à saisir est connue
    deduits = _executer_avec_reprise(
        "SELECT n.identifiant, n.titre, n.createurs, n.date_publication, "
        "       n.champs_a_verifier_decalog, n.serie, n.tome, n.categorie, "
        "       n.genre, n.public_vise, " + cote_sq + " "
        + depuis +
        "WHERE n.champs_a_verifier_decalog IS NOT NULL "
        "  AND n.champs_a_verifier_decalog != ''" + filtre_jeunesse
    )

    # 2. Notices sans EAN (identifiant substitut CB:) : invisibles à toute
    #    recherche par ISBN tant que Decalog n'est pas corrigé
    sans_ean = _executer_avec_reprise(
        "SELECT n.identifiant, n.titre, n.createurs, n.date_publication, "
        + cote_sq + " "
        + depuis +
        "WHERE n.identifiant LIKE 'CB:%' AND n.type_document = 'LIVRE' "
        "  AND CAST(SUBSTR(n.date_publication,1,4) AS INTEGER) >= 2000"
        + filtre_jeunesse
    )

    from collections_editeur import qualifier

    lignes_deduits, lignes_douteuses, vus = [], [], set()
    for (ident, titre, auteur, date_pub, champs, serie, tome,
         categorie, genre, public, cote) in deduits:
        if ident in vus:
            continue
        vus.add(ident)
        valeurs = {"serie": serie, "tome": tome, "categorie": categorie,
                   "genre": genre, "public_vise": public}
        # Une « série » qui est en fait une collection d'éditeur (Folio,
        # Aire Libre, 10-18...) ne doit PAS être saisie telle quelle dans
        # Decalog : on l'isole dans une feuille à part plutôt que de la
        # supprimer, le jugement final revenant au bibliothécaire.
        fiable, raison = qualifier(serie, tome)
        for champ in [c.strip() for c in (champs or "").split(",") if c.strip()]:
            ligne = {
                "Cote": cote or "",
                "ISBN / identifiant": ident,
                "Titre": titre or "",
                "Auteur": auteur or "",
                "Année": (date_pub or "")[:4],
                "Champ à corriger": LIBELLES.get(champ, champ),
                "Valeur à saisir dans Decalog": valeurs.get(champ) or "",
            }
            # le doute ne porte que sur la série (un tome numéroté est sûr)
            if champ == "serie" and not fiable:
                ligne["Pourquoi à vérifier"] = raison
                lignes_douteuses.append(ligne)
            else:
                lignes_deduits.append(ligne)

    lignes_sans_ean, vus2 = [], set()
    for ident, titre, auteur, date_pub, cote in sans_ean:
        if ident in vus2:
            continue
        vus2.add(ident)
        lignes_sans_ean.append({
            "Cote": cote or "",
            "Identifiant actuel": ident,
            "Titre": titre or "",
            "Auteur": auteur or "",
            "Année": (date_pub or "")[:4],
            "Champ à corriger": "EAN / ISBN",
            "Valeur à saisir dans Decalog": "(à retrouver — voir "
                                            "retrouver_ean_manquants.py)",
        })

    # Tri par cote côté Python (le tri SQL imposait une jointure trop lourde) :
    # les agents corrigent en suivant l'ordre des rayons.
    lignes_deduits.sort(key=lambda l: (l["Cote"] or "zzz", l["Titre"] or ""))
    lignes_douteuses.sort(key=lambda l: (l["Cote"] or "zzz", l["Titre"] or ""))
    lignes_sans_ean.sort(key=lambda l: (l["Cote"] or "zzz", l["Titre"] or ""))

    if maxi:
        lignes_deduits = lignes_deduits[:maxi]
        lignes_douteuses = lignes_douteuses[:maxi]
        lignes_sans_ean = lignes_sans_ean[:maxi]
    return lignes_deduits, lignes_douteuses, lignes_sans_ean


def ecrire_feuille(ws, lignes, titre_feuille):
    from openpyxl.styles import Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    if not lignes:
        ws.cell(row=1, column=1, value="(aucune ligne)")
        return
    colonnes = list(lignes[0].keys())
    thin = Side(style="thin", color="D9E2F0")
    bordure = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c, nom in enumerate(colonnes, 1):
        cell = ws.cell(row=1, column=c, value=nom)
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2E4A7A")
        cell.border = bordure
        largeur = 40 if "Titre" in nom or "Valeur" in nom else 20
        ws.column_dimensions[get_column_letter(c)].width = largeur
    for i, ligne in enumerate(lignes, start=2):
        for c, nom in enumerate(colonnes, 1):
            cell = ws.cell(row=i, column=c, value=ligne[nom])
            cell.font = Font(name="Arial", size=9.5)
            cell.border = bordure
    ws.freeze_panes = "A2"


def main():
    jeunesse = "--jeunesse" in sys.argv
    maxi = None
    if "--max" in sys.argv:
        try:
            maxi = int(sys.argv[sys.argv.index("--max") + 1])
        except (IndexError, ValueError):
            pass

    print("═══ Fiches Decalog à corriger ═══\n")
    deduits, douteuses, sans_ean = recuperer(jeunesse, maxi)
    print(f"  {len(deduits)} correction(s) SÛRES (série narrative ou tome)")
    print(f"  {len(douteuses)} à VÉRIFIER (collection d'éditeur prise pour une série)")
    print(f"  {len(sans_ean)} notice(s) sans EAN à retrouver\n")

    if deduits:
        from collections import Counter
        c = Counter(d["Champ à corriger"] for d in deduits)
        print("  Répartition par champ :")
        for champ, n in c.most_common():
            print(f"    {n:6}  {champ}")

    from openpyxl import Workbook
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "1 - Corrections sûres"
    ecrire_feuille(ws1, deduits, "Corrections sûres")
    ws2 = wb.create_sheet("2 - À vérifier")
    ecrire_feuille(ws2, douteuses, "À vérifier")
    ws3 = wb.create_sheet("3 - EAN manquant")
    ecrire_feuille(ws3, sans_ean, "EAN manquant")

    suffixe = "_jeunesse" if jeunesse else ""
    sortie = os.path.join(
        DOSSIER,
        f"Corrections_Decalog{suffixe}_{datetime.date.today().isoformat()}.xlsx")
    wb.save(sortie)
    print(f"\n✓ Export : {os.path.basename(sortie)}")
    print("  Feuille 1 « Corrections sûres » : à recopier tel quel dans Decalog.")
    print("  Feuille 2 « À vérifier » : la valeur trouvée est une collection")
    print("             d'éditeur (Folio, Aire Libre...) — NE PAS saisir comme")
    print("             série sans vérification par un bibliothécaire.")
    print("  Feuille 3 « EAN manquant » : ISBN à retrouver.")


if __name__ == "__main__":
    main()
