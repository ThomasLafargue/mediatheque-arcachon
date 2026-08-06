#!/usr/bin/env python3
"""
complement_bd_jeunesse.py — Complément de commande : ~1 500 € de BD JEUNESSE
récentes (one-shots et débuts de séries) depuis le rayon BD jeunesse de
Place des Libraires, trié par date de parution décroissante.

DEMANDE (Thomas, 2026-08-05) : la commande jeunesse manquait de nouveautés
BD. 1 500 € de BD jeunesse récentes, SANS comics ni mangas (gérés à part).

GARDE-FOUS : mêmes règles que preparer_commande_jeunesse.py (réutilisé en
module) — jamais un document du fonds ni de la commande en cours ni des
scans « à cataloguer », pas de coffrets/présentoirs, prix plafonné, cache
partagé (_commande_cache.json). ⚠ Un seul scraper Place des Libraires à la
fois.

SORTIES : Complement_BD_AAAA-MM-JJ.xlsx + complement_orb_bd_AAAA-MM-JJ.txt/.csv
(fichiers SÉPARÉS de la commande principale, à importer en plus dans ORB).

Usage :  python3 -u complement_bd_jeunesse.py > complement_bd.log 2>&1 &
"""
import csv
import datetime
import glob
import os
import re
import sys
import unicodedata

DOSSIER = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, DOSSIER)
import db  # noqa: E402
import preparer_commande_jeunesse as base  # noqa: E402  (briques réutilisées)

BUDGET_BD = 1500.0
base.BUDGET = BUDGET_BD          # le Panier s'arrête à cette enveloppe
PRIX_MAX_BD = 30.0
MAX_PAGES = 40                   # ~20 titres/page côté PdL

# éditeurs/labels à écarter : mangas et comics
EXCLUS_EDITEUR = ("ki-oon", "kioon", "nobi nobi", "kana", "pika", "kurokawa",
                  "soleil manga", "kaze", "crunchyroll", "tonkam", "meian",
                  "doki", "ototo", "taifu", "akata", "panini", "mangetsu",
                  "naban", "urban comics", "marvel", "dc comics", "hi comics",
                  "404 comics", "komics", "vestron", "graph zeppelin")
MOTS_EXCLUS = ("manga", "comics", "manhwa", "webtoon")

# Rayon BD de Place des Libraires. L'URL du rayon complet est attestée dans
# carte_placedeslibraires.md ; les sous-rayons jeunesse sont découverts
# dynamiquement depuis la page /bd/ssh-9019.
URL_RAYON_BD_COMPLET = (
    base.BASE + "/listeliv.php?refgtl=home&base=allbooks"
    "&select_tri_recherche=dateparution_decroissant&codegtl1=3000000"
    "&codegtl2=3020000&rayon=Bandes+dessin%26eacute%3Bes+%2F+Comics+%2F+"
    "Mangas%7CBandes+dessin%26eacute%3Bes")

# éditeurs à dominante BD jeunesse : acceptés d'office
EDITEURS_BD_JEUNESSE = (
    "bd kids", "jungle", "gouttiere", "kennes", "auzou", "biscoto", "makaka",
    "glenat jeunesse", "p'tit louis", "petit a petit", "little urban",
    "frimousse", "tourbillon", "bayard", "milan")


def decouvrir_sous_rayons_jeunesse():
    """Cherche sur la page rayon BD des liens listeliv « jeunesse »."""
    html = base._get(base.BASE + "/bd/ssh-9019")
    urls = []
    if not html:
        return urls
    from bs4 import BeautifulSoup
    soup = BeautifulSoup(html, "html.parser")
    for a in soup.find_all("a", href=True):
        libelle = _sa(a.get_text(strip=True))
        href = a["href"]
        if "listeliv.php" not in href:
            continue
        if "jeunesse" in libelle or "jeunesse" in _sa(href):
            url = href if href.startswith("http") else base.BASE + href
            url = url.split("#")[0]
            if url not in urls:
                urls.append((a.get_text(strip=True)[:60], url))
    return urls


def _sa(t):
    t = unicodedata.normalize("NFKD", t or "")
    return "".join(c for c in t if not unicodedata.combining(c)).lower()


def exclu(f):
    ed, ti = _sa(f.get("editeur")), _sa(f.get("titre"))
    if any(e in ed for e in EXCLUS_EDITEUR):
        return True
    return any(m in ti or m in ed for m in MOTS_EXCLUS)


def est_jeunesse(f):
    """BD jeunesse ? éditeur spécialisé accepté d'office, sinon détection
    du moteur (titre/éditeur)."""
    ed = _sa(f.get("editeur"))
    if any(e in ed for e in EDITEURS_BD_JEUNESSE):
        return True
    try:
        import moteur_recherche as mr
        return mr.detecter_public("", f.get("titre", ""), "",
                                  f.get("editeur", ""), "") in (
            "Jeunesse", "Adolescent")
    except Exception:
        return False


def main():
    base.charger_cache()
    conn = db.connect()
    print("Chargement du fonds (dédoublonnage)...")
    isbns, series, titres = base.charger_fonds(conn)
    conn.close()

    # + commande en cours + scans « à cataloguer » + exclusions manuelles
    for chemin in (glob.glob(os.path.join(DOSSIER, "commande_orb_*.txt"))
                   + glob.glob(os.path.join(DOSSIER, "a cataloguer*.txt"))
                   + glob.glob(os.path.join(DOSSIER, "exclusions_*.txt"))):
        avec = {l.strip() for l in open(chemin) if l.strip() and l.strip().isdigit()}
        isbns |= avec
        print(f"  + {len(avec)} EAN exclus via {os.path.basename(chemin)}")

    panier = base.Panier(isbns, series, titres)
    vus = 0

    def parcourir(nom, url_base, filtrer_jeunesse):
        nonlocal vus
        sep = "&" if "?" in url_base else "?"
        for page in range(1, MAX_PAGES + 1):
            if panier.total >= BUDGET_BD * 0.99:
                return
            html = base._get(f"{url_base}{sep}page={page}")
            if not html:
                print(f"  [{nom}] page {page} : inaccessible.")
                return
            eans = []
            for m in re.finditer(r"/livre/(\d{9,13})", html):
                if m.group(1) not in eans:
                    eans.append(m.group(1))
            if not eans:
                print(f"  [{nom}] page {page} : fin du rayon.")
                return
            for ean in eans:
                if panier.total >= BUDGET_BD * 0.99:
                    return
                f = base.fiche(ean)
                vus += 1
                if not f or exclu(f):
                    continue
                if filtrer_jeunesse and not est_jeunesse(f):
                    continue
                if f.get("prix") and f["prix"] > PRIX_MAX_BD:
                    continue
                panier.ajouter("BD", f, ean, "BD nouveauté",
                               f"Nouveautés BD jeunesse Place des Libraires ({nom})")
            print(f"  [{nom}] page {page:>2} : {len(panier.lignes)} retenues — "
                  f"{panier.total:.2f} € — {vus} fiches vues")

    print(f"\n─── BD JEUNESSE, nouveautés d'abord (objectif {BUDGET_BD:.0f} €) ───")
    try:
        sous_rayons = decouvrir_sous_rayons_jeunesse()
        if sous_rayons:
            print(f"Sous-rayons jeunesse découverts : {[n for n, _ in sous_rayons]}")
            for nom, url in sous_rayons[:4]:
                parcourir(nom, url, filtrer_jeunesse=False)
        if panier.total < BUDGET_BD * 0.99:
            print("Complément via le rayon BD général (filtre jeunesse actif) :")
            parcourir("rayon BD général", URL_RAYON_BD_COMPLET,
                      filtrer_jeunesse=True)
    finally:
        base.sauver_cache()

    if not panier.lignes:
        print("Rien retenu — vérifier le code rayon dans URL_RAYON.")
        return

    # ── sorties dédiées ──────────────────────────────────────────────────
    jour = datetime.date.today().isoformat()
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    ws = wb.active
    ws.title = "Complément BD"
    entetes = ["Série", "Tome", "Titre", "Auteur", "Éditeur", "EAN",
               "Prix (€)", "Motif"]
    thin = Side(style="thin", color="D9E2F0")
    bord = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c, nom in enumerate(entetes, 1):
        cell = ws.cell(row=1, column=c, value=nom)
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2E4A7A")
        cell.border = bord
        ws.column_dimensions[get_column_letter(c)].width = (
            40 if nom in ("Titre", "Motif") else 16)
    lignes = sorted(panier.lignes, key=lambda l: (l["serie"], str(l["tome"]).zfill(3), l["titre"]))
    for i, l in enumerate(lignes, start=2):
        for c, cle in enumerate(("serie", "tome", "titre", "auteur",
                                 "editeur", "ean", "prix", "motif"), 1):
            cell = ws.cell(row=i, column=c, value=l[cle])
            cell.font = Font(name="Arial", size=9.5)
            cell.border = bord
    ws.freeze_panes = "A2"
    ws.cell(row=len(lignes) + 3, column=6, value="TOTAL").font = Font(bold=True)
    ws.cell(row=len(lignes) + 3, column=7, value=round(panier.total, 2)).font = Font(bold=True)
    wb.save(os.path.join(DOSSIER, f"Complement_BD_{jour}.xlsx"))

    with open(os.path.join(DOSSIER, f"complement_orb_bd_{jour}.txt"), "w") as f:
        for l in lignes:
            f.write(l["ean"] + "\n")
    with open(os.path.join(DOSSIER, f"complement_orb_bd_{jour}.csv"), "w",
              newline="") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(["EAN", "Quantité"])
        for l in lignes:
            w.writerow([l["ean"], 1])

    print(f"\n✓ {len(lignes)} BD, {panier.total:.2f} € "
          f"(one-shots et tomes 1 en tête des nouveautés)")
    print(f"✓ Complement_BD_{jour}.xlsx + complement_orb_bd_{jour}.txt/.csv")


if __name__ == "__main__":
    main()
