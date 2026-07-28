#!/usr/bin/env python3
"""
comparer_recolement.py — Récolement en continu, UN SEUL fichier de scans.

PRINCIPE (choisi par Thomas le 2026-07-28) : un seul fichier à gérer,
`recolement_scans.txt`, qu'on alimente à la scanette au fil des semaines —
rayon après rayon, retours au fil de l'eau, zones mélangées, peu importe.
Le script fait le tri tout seul :
  - chaque code N'EST COMPTÉ QU'UNE FOIS, avec la DATE de son premier scan ;
  - re-scanner un document déjà vu est sans effet (aucune annulation,
    aucun double comptage) ;
  - la comparaison se fait avec le stock issu de l'export Decalog
    (table exemplaire, rafraîchie par l'import hebdomadaire).

LES ZONES SONT DÉDUITES des cotes : le rapport ne considère « en cours de
récolement » que les zones où au moins un document a été scanné. Les rayons
pas encore commencés n'encombrent donc pas la liste des manquants.

FONCTIONNEMENT
    python3 comparer_recolement.py            -> versement + rapport
  1. lit recolement_scans.txt (un code-barres par ligne, scanette en mode
     clavier ; lignes # ignorées) ;
  2. verse les codes nouveaux dans la mémoire _recolement_memoire.txt
     (code <TAB> date de premier scan — lisible, éditable, sauvegardable) ;
  3. régénère Recolement_etat.xlsx :
       1 - Manquants   attendus des zones en cours, jamais scannés
                       (⚠ inclut les documents EN PRÊT : la liste ne
                       devient significative qu'en fin de campagne)
       2 - Scannés     chaque document vu, avec sa date de scan
       3 - Inconnus    codes absents de la base (étiquette orpheline,
                       notice supprimée, autre site)
       4 - Résumé      progression zone par zone

  Lancé automatiquement chaque lundi matin par launchd
  (com.maat.recolement.plist) : il suffit d'alimenter le fichier de scans,
  le rapport se met à jour tout seul. Relançable à la main à tout moment.

REMISE À ZÉRO d'une campagne : supprimer (ou archiver) la mémoire
_recolement_memoire.txt. Le fichier de scans peut alors être vidé aussi.

STRICTEMENT EN LECTURE sur la base.
"""
import datetime
import os
import re
import sys

DOSSIER = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, DOSSIER)
import db  # noqa: E402

FICHIER_SCANS = os.path.join(DOSSIER, "recolement_scans.txt")
FICHIER_MEMOIRE = os.path.join(DOSSIER, "_recolement_memoire.txt")
FICHIER_RAPPORT = os.path.join(DOSSIER, "Recolement_etat.xlsx")


def zone_de(cote):
    """La zone = le préfixe alphabétique de la cote : 'BDJ 741' -> 'BDJ',
    'R DIC' -> 'R', 'RP ROT' -> 'RP'. Fuste mais stable, et surtout déduit
    des données réelles plutôt que d'une liste à maintenir."""
    m = re.match(r"^([A-Za-z]+)", (cote or "").strip())
    return m.group(1).upper() if m else "(sans cote)"


def lire_scans():
    if not os.path.exists(FICHIER_SCANS):
        return []
    codes = []
    with open(FICHIER_SCANS, encoding="utf-8", errors="replace") as f:
        for ligne in f:
            if ligne.lstrip().startswith("#"):
                continue
            code = re.sub(r"\s+", "", ligne.split("#")[0])
            if code:
                codes.append(code)
    return codes


def lire_memoire():
    """code -> (nb_occurrences_comptées, [dates de scan]).

    Format : code <TAB> nb <TAB> date1;date2;...
    Un document peut être scanné PLUSIEURS fois au fil de la campagne :
    parti en prêt puis revenu, il repasse sous la scanette. Chaque passage
    garde sa date — c'est une information de circulation, pas une erreur
    (précision de Thomas, 2026-07-28). L'ancien format à 2 colonnes
    (code <TAB> date) est lu comme « 1 scan »."""
    memoire = {}
    if os.path.exists(FICHIER_MEMOIRE):
        with open(FICHIER_MEMOIRE, encoding="utf-8") as f:
            for ligne in f:
                if ligne.lstrip().startswith("#") or not ligne.strip():
                    continue
                m = ligne.rstrip("\n").split("\t")
                if len(m) >= 3:
                    memoire[m[0]] = (int(m[1]),
                                     [d for d in m[2].split(";") if d])
                elif len(m) == 2:      # ancien format
                    memoire[m[0]] = (1, [m[1]])
                else:
                    memoire[m[0]] = (1, ["?"])
    return memoire


def ecrire_memoire(memoire):
    """Réécrit la mémoire complète (petite : un document par ligne)."""
    with open(FICHIER_MEMOIRE, "w", encoding="utf-8") as f:
        f.write("# code-barres <TAB> nb de scans comptés <TAB> dates\n")
        for code, (nb, dates) in memoire.items():
            f.write(f"{code}\t{nb}\t{';'.join(dates)}\n")


def main():
    aujourdhui = datetime.date.today().isoformat()
    memoire = lire_memoire()
    scans = lire_scans()

    # Occurrences dans le fichier de scans : si un code y figure PLUS de
    # fois que ce que la mémoire a déjà compté, ce sont de nouveaux
    # passages sous la scanette -> une date de plus par passage. (Le
    # fichier de scans grossit sans jamais être vidé : on ne compte que
    # la différence, jamais deux fois la même ligne.)
    from collections import Counter
    occurrences = Counter(scans)
    nouveaux = re_scannes = 0
    for code, occ in occurrences.items():
        nb_connu, dates = memoire.get(code, (0, []))
        if occ > nb_connu:
            supplement = occ - nb_connu
            memoire[code] = (occ, dates + [aujourdhui] * supplement)
            if nb_connu == 0:
                nouveaux += 1
            else:
                re_scannes += 1
    if nouveaux or re_scannes:
        ecrire_memoire(memoire)
    print(f"Scans lus : {len(scans)} — premiers scans : {nouveaux} — "
          f"documents revus (prêt/retour ?) : {re_scannes} — "
          f"total suivi : {len(memoire)}")

    if not memoire:
        print("Rien à comparer (fichier recolement_scans.txt vide ou absent).")
        return

    # ── le stock, zone par zone ──────────────────────────────────────────
    conn = db.connect()
    stock = {}          # code -> (identifiant, cote, zone, statut, titre, auteur)
    for cb, ident, cote, statut, titre, auteur in conn.execute(
            "SELECT e.code_barre_exemplaire, e.identifiant, e.cote, "
            "       e.statut, n.titre, n.createurs "
            "FROM exemplaire e LEFT JOIN notice n "
            "  ON n.identifiant = e.identifiant "
            "WHERE e.code_barre_exemplaire IS NOT NULL").fetchall():
        stock[str(cb)] = (ident, cote or "", zone_de(cote), statut or "",
                          titre or "", auteur or "")
    conn.close()

    scannes_connus = {c: stock[c] for c in memoire if c in stock}
    inconnus = sorted(c for c in memoire if c not in stock)
    zones_en_cours = sorted({v[2] for v in scannes_connus.values()})

    manquants = sorted(
        [(cb, v[0], v[1], v[3], v[4], v[5])
         for cb, v in stock.items()
         if v[2] in zones_en_cours and cb not in memoire],
        key=lambda x: (x[2], x[4]))

    # progression par zone
    lignes_zones = []
    for z in zones_en_cours:
        attendus_z = sum(1 for v in stock.values() if v[2] == z)
        vus_z = sum(1 for v in scannes_connus.values() if v[2] == z)
        pct = 100.0 * vus_z / attendus_z if attendus_z else 0.0
        lignes_zones.append((z, attendus_z, vus_z, attendus_z - vus_z,
                             f"{pct:.1f} %"))

    # ── rapport Excel ────────────────────────────────────────────────────
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    def feuille(ws, entetes, lignes):
        thin = Side(style="thin", color="D9E2F0")
        bordure = Border(left=thin, right=thin, top=thin, bottom=thin)
        for c, nom in enumerate(entetes, 1):
            cell = ws.cell(row=1, column=c, value=nom)
            cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2E4A7A")
            cell.border = bordure
            ws.column_dimensions[get_column_letter(c)].width = (
                34 if nom in ("Titre", "Auteur") else 16)
        for i, lg in enumerate(lignes, start=2):
            for c, val in enumerate(lg, 1):
                cell = ws.cell(row=i, column=c, value=val)
                cell.font = Font(name="Arial", size=9.5)
                cell.border = bordure
        ws.freeze_panes = "A2"
        if lignes:
            ws.auto_filter.ref = (
                f"A1:{get_column_letter(len(entetes))}{len(lignes)+1}")

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "1 - Manquants"
    feuille(ws1, ["Code-barres", "ISBN-identifiant", "Cote", "Statut",
                  "Titre", "Auteur"], manquants)
    ws2 = wb.create_sheet("2 - Scannés")
    feuille(ws2, ["Code-barres", "Nb de scans", "Dates de scan",
                  "Cote", "Titre", "Auteur"],
            sorted(((cb, memoire[cb][0], " ; ".join(memoire[cb][1]),
                     v[1], v[4], v[5])
                    for cb, v in scannes_connus.items()),
                   key=lambda x: (-x[1], x[3], x[4])))
    ws3 = wb.create_sheet("3 - Inconnus")
    feuille(ws3, ["Code-barres scanné (absent de la base)", "Dates de scan"],
            [(c, " ; ".join(memoire[c][1])) for c in inconnus])
    ws4 = wb.create_sheet("4 - Résumé")
    feuille(ws4, ["Zone (préfixe de cote)", "Attendus", "Scannés",
                  "Manquants", "Progression"],
            lignes_zones + [
                ("", "", "", "", ""),
                ("Rapport du", aujourdhui, "", "", ""),
                ("Codes inconnus de la base", len(inconnus), "", "", ""),
                ("⚠ Les manquants incluent les documents en prêt",
                 "", "", "", ""),
            ])
    wb.save(FICHIER_RAPPORT)

    print(f"Zones en cours : {', '.join(zones_en_cours)}")
    for z, att, vus, man, pct in lignes_zones:
        print(f"  {z:12} {vus:>6}/{att:<6} scannés ({pct}) — {man} manquants")
    print(f"Inconnus : {len(inconnus)}")
    print(f"✓ Rapport : {os.path.basename(FICHIER_RAPPORT)}")


if __name__ == "__main__":
    main()
