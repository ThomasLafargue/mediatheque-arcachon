#!/usr/bin/env python3
"""
charger_ean_en_base.py — Charge le dernier EAN_a_corriger_*.xlsx dans la
table `ean_retrouve`, pour que le CHAT puisse servir les résultats aux
agents depuis n'importe quel poste (« quels EAN en confiance élevée reste-t-il
à saisir ? », exports par rayon...).

POURQUOI (2026-07-30) : les résultats de la recherche d'EAN ne vivaient que
dans un fichier Excel sur le Mac mini. Les agents travaillent depuis le
MAAT, via le chat : la base est leur seul point d'accès commun.

La table est ENTIÈREMENT REMPLACÉE à chaque chargement (les résultats d'une
nouvelle recherche annulent les anciens). Une colonne `statut_saisie` est
préservée d'un chargement à l'autre quand l'identifiant se retrouve : les
lignes déjà marquées « saisie » par l'équipe ne redeviennent pas « à faire ».

Usage :  python3 charger_ean_en_base.py
"""
import glob
import os
import sys

DOSSIER = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, DOSSIER)
import db  # noqa: E402


def main():
    candidats = sorted(glob.glob(os.path.join(DOSSIER, "EAN_a_corriger_*.xlsx")))
    if not candidats:
        print("Aucun fichier EAN_a_corriger_*.xlsx trouvé.")
        return
    chemin = candidats[-1]
    print(f"Chargement de {os.path.basename(chemin)}...")

    from openpyxl import load_workbook
    ws = load_workbook(chemin, read_only=True).active
    lignes = list(ws.iter_rows(min_row=1, values_only=True))
    entetes = [str(c) for c in lignes[0]]
    idx = {nom: i for i, nom in enumerate(entetes)}

    conn = db.connect()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS ean_retrouve (
            identifiant_actuel TEXT PRIMARY KEY,  -- l'ancien CB:xxxx
            cote TEXT, titre TEXT, auteur TEXT, annee TEXT, editeur TEXT,
            isbn_trouve TEXT, titre_trouve TEXT, auteur_trouve TEXT,
            confiance TEXT,           -- élevée / moyenne / faible / —
            statut_saisie TEXT NOT NULL DEFAULT 'à faire',
            date_chargement TEXT DEFAULT (date('now'))
        )
    """)
    # mémoriser les statuts déjà travaillés par l'équipe
    statuts = dict(conn.execute(
        "SELECT identifiant_actuel, statut_saisie FROM ean_retrouve "
        "WHERE statut_saisie != 'à faire'").fetchall())
    conn.execute("DELETE FROM ean_retrouve")

    n = 0
    for l in lignes[1:]:
        ident = l[idx["Identifiant actuel"]]
        if not ident:
            continue
        conn.execute(
            "INSERT OR REPLACE INTO ean_retrouve "
            "(identifiant_actuel, cote, titre, auteur, annee, editeur, "
            " isbn_trouve, titre_trouve, auteur_trouve, confiance, "
            " statut_saisie) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
            (ident, l[idx["Cote"]], l[idx["Titre (Decalog)"]],
             l[idx["Auteur"]], str(l[idx["Année"]] or ""),
             l[idx["Éditeur"]], str(l[idx["ISBN trouvé"]] or ""),
             l[idx["Titre trouvé (Place des Libraires)"]],
             l[idx["Auteur trouvé"]], l[idx["Confiance"]],
             statuts.get(ident, "à faire")))
        n += 1
    conn.commit()

    for conf, nb in conn.execute(
            "SELECT confiance, COUNT(*) FROM ean_retrouve "
            "GROUP BY confiance ORDER BY 2 DESC").fetchall():
        print(f"  {nb:>5}  confiance {conf}")
    print(f"✓ {n} ligne(s) chargées dans la table ean_retrouve.")
    conn.close()


if __name__ == "__main__":
    main()
